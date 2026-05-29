"""
JAA Daily Timesheet Application v3.0 — Single File
Just copy-paste this one file and run it.
Requirements: pip install openpyxl
Optional:     pip install pandas  (for Work Mapping import only)
Default admin PIN: 2580

v3.0 Enhancements:
  - Submission tracking (On Time / Late / Partial / Not Submitted)
  - Admin Dashboard with KPI cards and color indicators
  - Time overlap validation
  - Email notifications (SMTP/Gmail)
  - Auto popup reminders at configurable time
  - Approval workflow (Draft → Submitted → Approved/Rejected)
  - Enhanced reports (productivity, client-wise, late submission)
  - "Others" work category added
"""

import sqlite3, os, sys, re, csv, shutil, hashlib, threading, json, smtplib, logging
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from datetime import datetime, date, timedelta
from contextlib import contextmanager
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR   = Path(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = BASE_DIR / "data"
DOCS_DIR   = BASE_DIR / "documents"
T1_DIR     = DOCS_DIR / "t1_forms"
T5_DIR     = DOCS_DIR / "t5_forms"
BACKUP_DIR = BASE_DIR / "backups"
LOGS_DIR   = BASE_DIR / "logs"
DB_PATH     = DATA_DIR / "jaa_timesheet.db"
MASTER_XLSX = DATA_DIR / "JAA_Master_Timesheet.xlsx"
LOCK_FILE   = DATA_DIR / "JAA_Master.lock"
CONFIG_FILE = DATA_DIR / "jaa_config.json"
MAX_PDF_MB  = 10
LEARN_DIR     = BASE_DIR / "learning"
LEARN_PPTS    = LEARN_DIR / "ppts"
LEARN_PDFS    = LEARN_DIR / "pdfs"
LEARN_DOCS    = LEARN_DIR / "docs"
LEARN_EXCEL   = LEARN_DIR / "excel"
LEARN_MISC    = LEARN_DIR / "misc"
LEARN_SOPS    = LEARN_DIR / "sops"
for _d in [DATA_DIR, DOCS_DIR, T1_DIR, T5_DIR, BACKUP_DIR, LOGS_DIR,
           LEARN_DIR, LEARN_PPTS, LEARN_PDFS, LEARN_DOCS, LEARN_EXCEL,
           LEARN_MISC, LEARN_SOPS]:
    _d.mkdir(parents=True, exist_ok=True)

# ─── Logging Setup ────────────────────────────────────────────────────────────
logging.basicConfig(
    filename=str(LOGS_DIR / "jaa_app.log"),
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger("JAA")

# ─── Config helpers ───────────────────────────────────────────────────────────
_DEFAULT_CONFIG = {
    "submission_cutoff": "20:30",          # HH:MM — after this = Late
    "reminder_time":     "19:00",          # HH:MM — popup reminder
    "reminder_enabled":  True,
    "smtp_host":         "smtp.gmail.com",
    "smtp_port":         587,
    "smtp_user":         "",
    "smtp_pass":         "",
    "manager_email":     "",
    "email_enabled":     False,
}

def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            cfg = dict(_DEFAULT_CONFIG)
            cfg.update(data)
            return cfg
        except Exception:
            pass
    return dict(_DEFAULT_CONFIG)

def save_config(cfg: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)


# ─── TIMESHEET LOCK RULE ─────────────────────────────────────────────────────
# Timesheets lock on the 27th of every month for ALL entries in that month.
# Admins are never locked.

LOCK_DAY     = 27   # lock kicks in from this day of the month onwards
GRACE_DAYS   = 7    # employees can edit entries up to this many days old

def is_timesheet_locked(entry_date: str = None, role: str = "user") -> bool:
    """
    Returns True if the timesheet entry should be locked for editing.

    Rules (employees only — admins are never locked):
    - Entries within the last GRACE_DAYS (7) days are always editable.
    - Entries older than GRACE_DAYS are locked.
    - From the 27th of the current month onwards, ALL entries
      including those within the grace window are locked.
    - Previous months are locked once their 27th has passed.
    """
    if role == "admin":
        return False
    today     = date.today()
    grace_cutoff = today - timedelta(days=GRACE_DAYS)

    if entry_date:
        try:
            ed = date.fromisoformat(entry_date)
        except Exception:
            return True   # can't parse → safe to lock

        # 27th lock: on/after LOCK_DAY → everything in current+previous months locked
        if today.day >= LOCK_DAY:
            if ed.year < today.year or ed.month <= today.month:
                return True
            return False   # future month (shouldn't happen)

        # Before 27th: previous months always locked
        if ed.year < today.year or ed.month < today.month:
            return True

        # Same month — allow if within GRACE_DAYS
        return ed < grace_cutoff

    else:
        # No specific date — check if saving is blanket-locked right now
        return today.day >= LOCK_DAY


# ════════════════════════════════════════════════════════════════════════════
# DATABASE
# ════════════════════════════════════════════════════════════════════════════





@contextmanager
def get_conn():
    conn = sqlite3.connect(str(DB_PATH), timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ─── SCHEMA CREATION ────────────────────────────────────────────────────────────

def init_db():
    """Create all tables if they don't exist. Safe to call on every startup."""
    with get_conn() as conn:
        conn.executescript("""
        -- ── EMPLOYEES ──────────────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS employees (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id              TEXT    UNIQUE NOT NULL,
            full_name           TEXT    NOT NULL,
            role                TEXT    NOT NULL,
            join_date           TEXT,                        -- YYYY-MM-DD
            batch               TEXT,
            join_year           INTEGER,
            seq_in_batch        INTEGER DEFAULT 0,
            status              TEXT    NOT NULL DEFAULT 'Active',
            reporting_to        TEXT,
            t1_path             TEXT,
            t5_path             TEXT,
            t5_upload_date      TEXT,
            pin_hash            TEXT,
            internship_end_date TEXT,                        -- YYYY-MM-DD override
            extension_note      TEXT,                        -- reason for extension
            created_at          TEXT    DEFAULT (datetime('now')),
            updated_at          TEXT    DEFAULT (datetime('now'))
        );
        """)

        # ── Migrate: add new columns to existing DB if they don't exist ──────────
        existing_cols = {row[1] for row in
                         conn.execute("PRAGMA table_info(employees)").fetchall()}
        if "internship_end_date" not in existing_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN internship_end_date TEXT")
        if "extension_note" not in existing_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN extension_note TEXT")

        conn.executescript("""
        -- ── COMPANIES ──────────────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS companies (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            short_code    TEXT    UNIQUE NOT NULL,
            full_name     TEXT    NOT NULL,
            unique_code   TEXT    UNIQUE NOT NULL,
            is_active     INTEGER DEFAULT 1,
            created_at    TEXT    DEFAULT (datetime('now'))
        );

        -- ── WORK CATEGORY MASTER ───────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS work_categories (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT    UNIQUE NOT NULL,
            description   TEXT,
            sort_order    INTEGER DEFAULT 0,
            is_active     INTEGER DEFAULT 1
        );

        -- ── OPERATIONAL AREAS ─────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS operational_areas (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id   INTEGER NOT NULL REFERENCES work_categories(id),
            name          TEXT    NOT NULL,
            sort_order    INTEGER DEFAULT 0,
            is_active     INTEGER DEFAULT 1,
            UNIQUE(category_id, name)
        );

        -- ── SUB-CATEGORIES ────────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS sub_categories (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            area_id       INTEGER NOT NULL REFERENCES operational_areas(id),
            name          TEXT    NOT NULL,
            sort_order    INTEGER DEFAULT 0,
            is_active     INTEGER DEFAULT 1,
            UNIQUE(area_id, name)
        );

        -- ── TIMESHEET ENTRIES ─────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS timesheet_entries (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_date      TEXT    NOT NULL,               -- YYYY-MM-DD
            day_name        TEXT,
            emp_id          TEXT    NOT NULL REFERENCES employees(emp_id),
            reporting_to    TEXT,
            company_code    TEXT    REFERENCES companies(unique_code),
            client_name     TEXT,
            work_category   TEXT,
            operational_area TEXT,
            sub_category    TEXT,
            start_time      TEXT,                           -- HH:MM
            end_time        TEXT,                           -- HH:MM
            break_mins      INTEGER DEFAULT 0,
            total_hrs       REAL    DEFAULT 0,
            task_desc       TEXT,
            notes           TEXT,
            work_location   TEXT,
            status          TEXT    DEFAULT 'In Progress',
            created_at      TEXT    DEFAULT (datetime('now')),
            updated_at      TEXT    DEFAULT (datetime('now'))
        );

        -- ── SESSION / ATTENDANCE LOG ──────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS attendance_log (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id        TEXT    NOT NULL REFERENCES employees(emp_id),
            login_dt      TEXT    NOT NULL,                 -- ISO datetime
            logout_dt     TEXT,                             -- NULL until logout
            duration_mins REAL,                             -- computed on logout
            session_date  TEXT    NOT NULL,                 -- YYYY-MM-DD
            ip_info       TEXT                              -- optional metadata
        );

        -- ── LEAVE RECORDS ─────────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS leave_records (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id        TEXT    NOT NULL REFERENCES employees(emp_id),
            start_date    TEXT    NOT NULL,
            end_date      TEXT    NOT NULL,
            reason        TEXT,
            leave_type    TEXT    DEFAULT 'Casual',
            status        TEXT    NOT NULL DEFAULT 'Pending',
            approved_by   TEXT,
            reviewed_at   TEXT,
            admin_note    TEXT,
            created_at    TEXT    DEFAULT (datetime('now'))
        );

        -- ── ADMIN PINS ────────────────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS admin_pins (
            key           TEXT    PRIMARY KEY,              -- '__admin__' or emp name lower
            pin_hash      TEXT    NOT NULL,
            updated_at    TEXT    DEFAULT (datetime('now'))
        );

        -- ── INDEXES ──────────────────────────────────────────────────────────
        CREATE INDEX IF NOT EXISTS idx_entries_date    ON timesheet_entries(entry_date);
        CREATE INDEX IF NOT EXISTS idx_entries_emp     ON timesheet_entries(emp_id);
        CREATE INDEX IF NOT EXISTS idx_attendance_emp  ON attendance_log(emp_id);
        CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_log(session_date);

        -- ── SUBMISSION TRACKING ───────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS submission_log (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id          TEXT    NOT NULL REFERENCES employees(emp_id),
            submit_date     TEXT    NOT NULL,               -- YYYY-MM-DD date being submitted for
            submitted_at    TEXT,                           -- ISO datetime of actual submission
            submission_status TEXT  DEFAULT 'Not Submitted',
            -- 'On Time' | 'Late' | 'Partially Filled' | 'Not Submitted'
            approval_status TEXT    DEFAULT 'Draft',
            -- 'Draft' | 'Submitted' | 'Approved' | 'Rejected' | 'Revision Required'
            approved_by     TEXT,
            approved_at     TEXT,
            reviewer_comment TEXT,
            total_hrs       REAL    DEFAULT 0,
            entry_count     INTEGER DEFAULT 0,
            created_at      TEXT    DEFAULT (datetime('now')),
            updated_at      TEXT    DEFAULT (datetime('now')),
            UNIQUE(emp_id, submit_date)
        );

        -- ── EMAIL NOTIFICATION LOG ────────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS email_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            sent_at     TEXT    NOT NULL,
            recipient   TEXT    NOT NULL,
            subject     TEXT,
            status      TEXT,                               -- 'Sent' | 'Failed'
            error_msg   TEXT
        );

        -- ── INDEXES for new tables ────────────────────────────────────────────
        CREATE INDEX IF NOT EXISTS idx_submission_emp  ON submission_log(emp_id);
        CREATE INDEX IF NOT EXISTS idx_submission_date ON submission_log(submit_date);

        -- ── LEARNING HUB — CATEGORIES ─────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS learning_categories (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    UNIQUE NOT NULL,
            icon        TEXT    DEFAULT '📂',
            sort_order  INTEGER DEFAULT 0,
            is_active   INTEGER DEFAULT 1,
            created_at  TEXT    DEFAULT (datetime('now'))
        );

        -- ── LEARNING HUB — MATERIALS ──────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS learning_materials (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            title           TEXT    NOT NULL,
            category        TEXT,
            material_type   TEXT    NOT NULL,
            -- 'PDF' | 'PPTX' | 'DOCX' | 'Excel' | 'YouTube' | 'Article' | 'Notice' | 'Other'
            file_path       TEXT,
            youtube_url     TEXT,
            description     TEXT,
            tags            TEXT,
            uploaded_by     TEXT,
            upload_date     TEXT    DEFAULT (date('now')),
            expiry_date     TEXT,                           -- NULL = never expires
            is_pinned       INTEGER DEFAULT 0,              -- 1 = pinned to top
            priority        TEXT    DEFAULT 'Normal',       -- 'High'|'Normal'|'Low'
            is_active       INTEGER DEFAULT 1,
            view_count      INTEGER DEFAULT 0,
            created_at      TEXT    DEFAULT (datetime('now')),
            updated_at      TEXT    DEFAULT (datetime('now'))
        );

        -- ── LEARNING HUB — ARTICLES (inline rich text) ────────────────────────
        CREATE TABLE IF NOT EXISTS learning_articles (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            material_id INTEGER REFERENCES learning_materials(id) ON DELETE CASCADE,
            body        TEXT,
            created_at  TEXT    DEFAULT (datetime('now')),
            updated_at  TEXT    DEFAULT (datetime('now'))
        );

        -- ── INDEXES ───────────────────────────────────────────────────────────
        CREATE INDEX IF NOT EXISTS idx_lm_category ON learning_materials(category);
        CREATE INDEX IF NOT EXISTS idx_lm_type     ON learning_materials(material_type);
        """)

    # ── company_assignments table ─────────────────────────────────────────────
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS company_assignments (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            company_code      TEXT    NOT NULL,
            partner_emp_id    TEXT    NOT NULL,
            member_emp_id     TEXT    NOT NULL,
            reports_to_emp_id TEXT,
            created_at        TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (company_code)      REFERENCES companies(unique_code),
            FOREIGN KEY (partner_emp_id)    REFERENCES employees(emp_id),
            FOREIGN KEY (member_emp_id)     REFERENCES employees(emp_id),
            FOREIGN KEY (reports_to_emp_id) REFERENCES employees(emp_id)
        );
        CREATE INDEX IF NOT EXISTS idx_ca_partner  ON company_assignments(partner_emp_id);
        CREATE INDEX IF NOT EXISTS idx_ca_company  ON company_assignments(company_code);
        CREATE INDEX IF NOT EXISTS idx_ca_member   ON company_assignments(member_emp_id);
        """)

    # ── reimbursements table ──────────────────────────────────────────────────
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS reimbursements (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id            TEXT    NOT NULL,
            company_code      TEXT    NOT NULL,
            category          TEXT    NOT NULL,
            sub_category      TEXT,
            amount            REAL    NOT NULL,
            description       TEXT,
            receipt_path      TEXT,
            status            TEXT    DEFAULT 'Pending',
            rejection_note    TEXT,
            billed            INTEGER DEFAULT 0,
            billing_advice_id INTEGER,
            submitted_at      TEXT    DEFAULT (datetime('now','localtime')),
            reviewed_at       TEXT,
            reviewed_by       TEXT,
            FOREIGN KEY (emp_id)       REFERENCES employees(emp_id),
            FOREIGN KEY (company_code) REFERENCES companies(unique_code)
        );
        CREATE INDEX IF NOT EXISTS idx_reimb_emp    ON reimbursements(emp_id);
        CREATE INDEX IF NOT EXISTS idx_reimb_co     ON reimbursements(company_code);
        CREATE INDEX IF NOT EXISTS idx_reimb_status ON reimbursements(status);
        CREATE INDEX IF NOT EXISTS idx_reimb_billed ON reimbursements(billed, status);
        """)

        # ── notifications table ───────────────────────────────────────────────────
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS notifications (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id         TEXT    NOT NULL,
            title          TEXT    NOT NULL,
            message        TEXT    NOT NULL,
            type           TEXT    DEFAULT 'general',
            reference_id   INTEGER,
            reference_type TEXT,
            is_read        INTEGER DEFAULT 0,
            created_at     TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (emp_id) REFERENCES employees(emp_id)
        );
        CREATE INDEX IF NOT EXISTS idx_notif_emp    ON notifications(emp_id);
        CREATE INDEX IF NOT EXISTS idx_notif_read   ON notifications(emp_id, is_read);
        CREATE INDEX IF NOT EXISTS idx_notif_ref    ON notifications(reference_id, reference_type);
        """)

    # ── tasks table ───────────────────────────────────────────────────────────
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS tasks (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            company_code        TEXT    NOT NULL,
            title               TEXT    NOT NULL,
            description         TEXT,
            assigned_by_emp_id  TEXT    NOT NULL,
            assigned_to_emp_id  TEXT    NOT NULL,
            due_date            TEXT    NOT NULL,
            status              TEXT    DEFAULT 'Pending',
            task_type           TEXT    DEFAULT 'Regular',
            created_at          TEXT    DEFAULT (datetime('now','localtime')),
            updated_at          TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (company_code)         REFERENCES companies(unique_code),
            FOREIGN KEY (assigned_by_emp_id)   REFERENCES employees(emp_id),
            FOREIGN KEY (assigned_to_emp_id)   REFERENCES employees(emp_id)
        );
        CREATE INDEX IF NOT EXISTS idx_tasks_to    ON tasks(assigned_to_emp_id);
        CREATE INDEX IF NOT EXISTS idx_tasks_by    ON tasks(assigned_by_emp_id);
        CREATE INDEX IF NOT EXISTS idx_tasks_co    ON tasks(company_code);
        CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
        """)

    # Migrate existing DBs: add new columns if absent
    _migrate_learning_columns()


# ─── NOTIFICATION FUNCTIONS ──────────────────────────────────────────────────

def create_notification(emp_id: str, title: str, message: str,
                        notif_type: str = "general",
                        reference_id: int = None,
                        reference_type: str = None) -> int:
    """Insert a new notification row. Returns new id."""
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO notifications
              (emp_id, title, message, type, reference_id, reference_type)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (emp_id, title, message, notif_type,
              reference_id, reference_type))
        return cur.lastrowid


def get_notifications(emp_id: str, limit: int = 20) -> list[dict]:
    """Latest notifications for this employee — unread first, then by date desc."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT * FROM notifications
            WHERE  emp_id = ?
            ORDER  BY is_read ASC, created_at DESC
            LIMIT  ?
        """, (emp_id, limit)).fetchall()
        return [dict(r) for r in rows]


def get_unread_count(emp_id: str) -> int:
    """Count of unread notifications — used for badge display."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM notifications WHERE emp_id=? AND is_read=0",
            (emp_id,)
        ).fetchone()
        return row[0] if row else 0


def mark_notification_read(notification_id: int, emp_id: str) -> bool:
    """Mark a single notification as read — only if it belongs to emp_id."""
    with get_conn() as conn:
        conn.execute(
            "UPDATE notifications SET is_read=1 WHERE id=? AND emp_id=?",
            (notification_id, emp_id)
        )
        return True


def mark_all_read(emp_id: str) -> bool:
    """Mark all notifications as read for this employee."""
    with get_conn() as conn:
        conn.execute(
            "UPDATE notifications SET is_read=1 WHERE emp_id=?",
            (emp_id,)
        )
        return True


def create_task_notifications(task: dict, event_type: str) -> None:
    """Central dispatcher for all task-related notifications.

    event_type values:
      task_assigned       — new task created
      status_in_progress  — assignee started work (In Progress)
      status_done         — assignee marked Done

    task dict must include:
      id, title, company_code, company_name (or fetched),
      assigned_by_emp_id, assigned_to_emp_id,
      due_date, task_type,
      assigned_to_name, assigned_by_name  (optional — fetched if absent)
    """
    task_id       = task.get("id")
    title         = task.get("title", "")
    task_type     = task.get("task_type", "Regular")
    due_date      = task.get("due_date", "")
    by_emp_id     = task.get("assigned_by_emp_id", "")
    to_emp_id     = task.get("assigned_to_emp_id", "")
    company_code  = task.get("company_code", "")

    # Fetch company name if not supplied
    company_name = task.get("company_name") or ""
    if not company_name and company_code:
        try:
            co = get_company_by_code(company_code)
            company_name = co["full_name"] if co else company_code
        except Exception:
            company_name = company_code

    # Fetch employee names if not supplied
    to_name = task.get("assigned_to_name") or ""
    by_name = task.get("assigned_by_name") or ""
    if not to_name and to_emp_id:
        try:
            emp = get_employee(to_emp_id)
            to_name = emp["full_name"] if emp else to_emp_id
        except Exception:
            to_name = to_emp_id
    if not by_name and by_emp_id:
        try:
            emp = get_employee(by_emp_id)
            by_name = emp["full_name"] if emp else by_emp_id
        except Exception:
            by_name = by_emp_id

    # Emergency prefix for title
    emergency_prefix = "🔴 " if task_type == "Emergency" else ""

    if event_type == "task_assigned":
        notif_title = f"{emergency_prefix}{'Emergency ' if task_type=='Emergency' else (task_type+' ' if task_type!='Regular' else '')}Task Assigned"
        message = (f"{task_type} task assigned: {title} "
                   f"for {company_name} — due {due_date}")
        create_notification(
            emp_id=to_emp_id, title=notif_title, message=message,
            notif_type="task_assigned",
            reference_id=task_id, reference_type="task"
        )

    elif event_type == "status_in_progress":
        notif_title = f"Task Started: {title}"
        message = f"{to_name} has started: {title} — {task_type}"
        create_notification(
            emp_id=by_emp_id, title=notif_title, message=message,
            notif_type="task_updated",
            reference_id=task_id, reference_type="task"
        )

    elif event_type == "status_done":
        notif_title = f"Task Completed: {title}"
        message = f"{to_name} has completed: {title} for {company_name}"
        # Notify assigned_by always
        create_notification(
            emp_id=by_emp_id, title=notif_title, message=message,
            notif_type="task_done",
            reference_id=task_id, reference_type="task"
        )
        # Walk the full reports_to chain upward and notify everyone above the assignee
        # Covers: Intern → Employee → Associate → Partner (any depth)
        try:
            with get_conn() as conn:
                notified = {by_emp_id}   # don't double-notify assigned_by
                current_emp = to_emp_id
                for _ in range(5):       # max 5 hops to avoid infinite loop
                    row = conn.execute("""
                        SELECT ca.reports_to_emp_id
                        FROM   company_assignments ca
                        WHERE  ca.company_code = ? AND ca.member_emp_id = ?
                        LIMIT 1
                    """, (company_code, current_emp)).fetchone()
                    if not row or not row["reports_to_emp_id"]:
                        break
                    supervisor_id = row["reports_to_emp_id"]
                    if supervisor_id == current_emp:  # safety: self-loop guard
                        break
                    if supervisor_id not in notified:
                        create_notification(
                            emp_id=supervisor_id,
                            title=notif_title, message=message,
                            notif_type="task_done",
                            reference_id=task_id, reference_type="task"
                        )
                        notified.add(supervisor_id)
                    current_emp = supervisor_id
        except Exception:
            pass


def check_overdue_notifications() -> int:
    """Scan tasks for overdue items and create notifications if not already done today.
    Returns count of new notifications created.
    Call once per branch at startup."""
    from datetime import date as _date
    today_s = _date.today().isoformat()
    created = 0
    try:
        with get_conn() as conn:
            overdue = conn.execute("""
                SELECT t.id, t.title, t.company_code, t.due_date,
                       t.task_type, t.assigned_to_emp_id, t.assigned_by_emp_id,
                       co.full_name AS company_name,
                       ab.full_name AS assigned_by_name,
                       at_.full_name AS assigned_to_name
                FROM   tasks t
                JOIN   companies  co  ON co.unique_code = t.company_code
                JOIN   employees  ab  ON ab.emp_id = t.assigned_by_emp_id
                JOIN   employees  at_ ON at_.emp_id = t.assigned_to_emp_id
                WHERE  t.due_date < ? AND t.status != 'Done'
            """, (today_s,)).fetchall()

        for t in overdue:
            task_type = t["task_type"] or "Regular"
            prefix    = "🔴 " if task_type == "Emergency" else ("🟡 " if task_type == "Specific" else "")
            notif_title = f"{prefix}Overdue Task: {t['title']}"
            message     = (f"Overdue {task_type} task: {t['title']} "
                           f"for {t['company_name']} was due on {t['due_date']}")
            ref_id = t["id"]

            # Check both recipients
            for emp_id in {t["assigned_to_emp_id"], t["assigned_by_emp_id"]}:
                if not emp_id:
                    continue
                # Avoid duplicate: if a task_overdue notification for this task was created today
                with get_conn() as conn:
                    already = conn.execute("""
                        SELECT id FROM notifications
                        WHERE  emp_id = ?
                          AND  reference_id = ?
                          AND  reference_type = 'task'
                          AND  type = 'task_overdue'
                          AND  DATE(created_at) = ?
                    """, (emp_id, ref_id, today_s)).fetchone()
                if not already:
                    create_notification(
                        emp_id=emp_id, title=notif_title, message=message,
                        notif_type="task_overdue",
                        reference_id=ref_id, reference_type="task"
                    )
                    created += 1
    except Exception as e:
        log.warning(f"check_overdue_notifications error: {e}")
    return created


# ─── COMPANY ASSIGNMENT FUNCTIONS ────────────────────────────────────────────

def assign_company_to_partner(company_code: str, partner_emp_id: str) -> bool:
    """Create an anchor row so the partner 'owns' this company.
    One anchor row: member = partner themselves, reports_to = NULL.
    Safe to call multiple times — does nothing if already exists."""
    with get_conn() as conn:
        exists = conn.execute(
            "SELECT id FROM company_assignments "
            "WHERE company_code=? AND partner_emp_id=? AND member_emp_id=?",
            (company_code, partner_emp_id, partner_emp_id)
        ).fetchone()
        if exists:
            return False
        conn.execute(
            "INSERT INTO company_assignments (company_code, partner_emp_id, member_emp_id, reports_to_emp_id) "
            "VALUES (?, ?, ?, NULL)",
            (company_code, partner_emp_id, partner_emp_id)
        )
        # NOTIFICATION HOOK: trigger assignment notification here
        return True


def add_company_assignment(company_code: str, partner_emp_id: str,
                           member_emp_id: str, reports_to_emp_id: str = None) -> bool:
    """Add a team member to a company under a partner.
    Prevents duplicate rows for the same (company, member) pair."""
    with get_conn() as conn:
        exists = conn.execute(
            "SELECT id FROM company_assignments "
            "WHERE company_code=? AND member_emp_id=?",
            (company_code, member_emp_id)
        ).fetchone()
        if exists:
            return False
        conn.execute(
            "INSERT INTO company_assignments "
            "(company_code, partner_emp_id, member_emp_id, reports_to_emp_id) "
            "VALUES (?, ?, ?, ?)",
            (company_code, partner_emp_id, member_emp_id, reports_to_emp_id or None)
        )
        # NOTIFICATION HOOK: trigger assignment notification here
        try:
            co = get_company_by_code(company_code)
            co_name = co["full_name"] if co else company_code
            partner = get_employee(partner_emp_id)
            p_name  = partner["full_name"] if partner else partner_emp_id
            create_notification(
                emp_id=member_emp_id,
                title="Company Assignment",
                message=f"You have been assigned to {co_name} by {p_name}",
                notif_type="company_assigned",
                reference_id=None, reference_type="assignment"
            )
        except Exception:
            pass
        return True


def remove_company_assignment(assignment_id: int) -> bool:
    """Delete a single assignment row by id.
    Does NOT delete the partner's own anchor row via this path."""
    # Fetch details BEFORE deleting — needed for notification
    removed_emp_id  = None
    removed_co_code = None
    try:
        with get_conn() as conn:
            row = conn.execute(
                "SELECT member_emp_id, company_code FROM company_assignments WHERE id=?",
                (assignment_id,)
            ).fetchone()
            if row:
                removed_emp_id  = row["member_emp_id"]
                removed_co_code = row["company_code"]
    except Exception:
        pass

    with get_conn() as conn:
        conn.execute("DELETE FROM company_assignments WHERE id=?", (assignment_id,))

    # NOTIFICATION HOOK: trigger assignment notification here
    if removed_emp_id and removed_co_code:
        try:
            co = get_company_by_code(removed_co_code)
            co_name = co["full_name"] if co else removed_co_code
            create_notification(
                emp_id=removed_emp_id,
                title="Company Assignment Removed",
                message=f"You have been removed from {co_name}",
                notif_type="company_removed",
                reference_id=None, reference_type="assignment"
            )
        except Exception:
            pass
    return True


def get_partner_companies(partner_emp_id: str = None) -> list[dict]:
    """Return ALL companies in the branch. All partners have equal access.
    partner_emp_id is accepted for signature compatibility but not used as a filter.
    Returns same column shape as get_all_companies() so templates work consistently."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM companies WHERE is_active = 1 ORDER BY full_name"
        ).fetchall()
        return [dict(r) for r in rows]


def get_company_team(company_code: str) -> list[dict]:
    """Return all members assigned to a company with names and reporting structure."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT ca.id, ca.company_code, ca.partner_emp_id, ca.member_emp_id,
                   ca.reports_to_emp_id, ca.created_at,
                   e.full_name  AS member_name,  e.role  AS member_role,
                   r.full_name  AS reports_to_name,
                   p.full_name  AS partner_name
            FROM   company_assignments ca
            JOIN   employees e ON e.emp_id = ca.member_emp_id
            LEFT JOIN employees r ON r.emp_id = ca.reports_to_emp_id
            JOIN   employees p ON p.emp_id = ca.partner_emp_id
            WHERE  ca.company_code = ?
            ORDER  BY ca.reports_to_emp_id NULLS FIRST, e.full_name
        """, (company_code,)).fetchall()
        return [dict(r) for r in rows]


def get_employee_assigned_companies(emp_id: str) -> list[dict]:
    """Return all companies an employee is assigned to (for timesheet dropdown)."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT DISTINCT co.unique_code, co.short_code, co.full_name, co.is_active
            FROM   company_assignments ca
            JOIN   companies co ON co.unique_code = ca.company_code
            WHERE  ca.member_emp_id = ?
              AND  co.is_active = 1
            ORDER  BY co.full_name
        """, (emp_id,)).fetchall()
        return [dict(r) for r in rows]


def get_all_assignments_for_admin() -> list[dict]:
    """Return full assignment list — for admin read-only view."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT ca.id, ca.company_code, ca.partner_emp_id, ca.member_emp_id,
                   ca.reports_to_emp_id, ca.created_at,
                   co.full_name  AS company_name, co.short_code,
                   p.full_name   AS partner_name,
                   m.full_name   AS member_name,  m.role AS member_role,
                   r.full_name   AS reports_to_name
            FROM   company_assignments ca
            JOIN   companies  co ON co.unique_code = ca.company_code
            JOIN   employees  p  ON p.emp_id  = ca.partner_emp_id
            JOIN   employees  m  ON m.emp_id  = ca.member_emp_id
            LEFT JOIN employees r ON r.emp_id = ca.reports_to_emp_id
            ORDER  BY co.full_name, p.full_name, ca.reports_to_emp_id NULLS FIRST, m.full_name
        """).fetchall()
        return [dict(r) for r in rows]


def get_all_companies_with_teams() -> list[dict]:
    """Return every active company with a pre-built list of their assigned members.
    One DB round-trip. Used by the Partner Assignments page to avoid N+1 queries.

    Each item:
      { unique_code, short_code, full_name,
        members: [ { member_emp_id, member_name, member_role,
                     partner_emp_id, reports_to_emp_id, reports_to_name,
                     id (assignment row id) } ] }
    members list excludes partner anchor rows (member_emp_id == partner_emp_id).
    """
    with get_conn() as conn:
        companies = conn.execute("""
            SELECT unique_code, short_code, full_name
            FROM   companies
            WHERE  is_active = 1
            ORDER  BY full_name
        """).fetchall()

        assignments = conn.execute("""
            SELECT ca.id, ca.company_code,
                   ca.partner_emp_id, ca.member_emp_id,
                   ca.reports_to_emp_id,
                   m.full_name  AS member_name,
                   m.role       AS member_role,
                   r.full_name  AS reports_to_name
            FROM   company_assignments ca
            JOIN   employees m ON m.emp_id = ca.member_emp_id
            LEFT JOIN employees r ON r.emp_id = ca.reports_to_emp_id
            WHERE  ca.member_emp_id != ca.partner_emp_id
            ORDER  BY ca.company_code, ca.reports_to_emp_id NULLS FIRST, m.full_name
        """).fetchall()

    # Index assignments by company_code
    from collections import defaultdict
    by_co = defaultdict(list)
    for a in assignments:
        by_co[a["company_code"]].append(dict(a))

    result = []
    for co in companies:
        result.append({
            "unique_code": co["unique_code"],
            "short_code":  co["short_code"],
            "full_name":   co["full_name"],
            "members":     by_co.get(co["unique_code"], []),
        })
    return result


# ─── TASK FUNCTIONS ───────────────────────────────────────────────────────────

def create_task(company_code: str, title: str, description: str,
                assigned_by_emp_id: str, assigned_to_emp_id: str,
                due_date: str, task_type: str = "Regular") -> int:
    """Create a new task with status Pending. Returns new task id.
    task_type values: Regular | Emergency | Specific"""
    if task_type not in ("Regular", "Emergency", "Specific"):
        task_type = "Regular"
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO tasks
              (company_code, title, description, assigned_by_emp_id,
               assigned_to_emp_id, due_date, status, task_type)
            VALUES (?, ?, ?, ?, ?, ?, 'Pending', ?)
        """, (company_code, title, description or "",
              assigned_by_emp_id, assigned_to_emp_id, due_date, task_type))
        task_id = cur.lastrowid

    # NOTIFICATION HOOK: trigger new task notification here
    # Message format: "New {task_type} task assigned: {title} for {company_name} — due {due_date}"
    try:
        create_task_notifications({
            "id": task_id, "title": title, "task_type": task_type,
            "company_code": company_code, "due_date": due_date,
            "assigned_by_emp_id": assigned_by_emp_id,
            "assigned_to_emp_id": assigned_to_emp_id,
        }, event_type="task_assigned")
    except Exception:
        pass
    return task_id


def get_tasks_assigned_to(emp_id: str) -> list[dict]:
    """All tasks assigned TO this employee, with company + assigner names."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT t.*, co.full_name AS company_name,
                   ab.full_name AS assigned_by_name
            FROM   tasks t
            JOIN   companies co ON co.unique_code = t.company_code
            JOIN   employees ab ON ab.emp_id = t.assigned_by_emp_id
            WHERE  t.assigned_to_emp_id = ?
            ORDER  BY t.due_date ASC, t.created_at DESC
        """, (emp_id,)).fetchall()
        return [dict(r) for r in rows]


def get_tasks_assigned_by(emp_id: str) -> list[dict]:
    """All tasks created BY this employee, with company + assignee names."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT t.*, co.full_name AS company_name,
                   at_.full_name AS assigned_to_name
            FROM   tasks t
            JOIN   companies  co  ON co.unique_code = t.company_code
            JOIN   employees  at_ ON at_.emp_id = t.assigned_to_emp_id
            WHERE  t.assigned_by_emp_id = ?
            ORDER  BY t.due_date ASC, t.created_at DESC
        """, (emp_id,)).fetchall()
        return [dict(r) for r in rows]


def get_all_tasks_for_company(company_code: str) -> list[dict]:
    """All tasks for a specific company with full details."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT t.*, co.full_name AS company_name,
                   ab.full_name AS assigned_by_name,
                   at_.full_name AS assigned_to_name,
                   at_.role      AS assigned_to_role
            FROM   tasks t
            JOIN   companies  co  ON co.unique_code = t.company_code
            JOIN   employees  ab  ON ab.emp_id = t.assigned_by_emp_id
            JOIN   employees  at_ ON at_.emp_id = t.assigned_to_emp_id
            WHERE  t.company_code = ?
            ORDER  BY t.due_date ASC
        """, (company_code,)).fetchall()
        return [dict(r) for r in rows]


def update_task_status(task_id: int, new_status: str, emp_id: str) -> bool:
    """Update task status — only if emp_id matches assigned_to_emp_id.
    Valid statuses: Pending, In Progress, Done."""
    if new_status not in ("Pending", "In Progress", "Done"):
        return False
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM tasks WHERE id=? AND assigned_to_emp_id=?",
            (task_id, emp_id)
        ).fetchone()
        if not row:
            return False
        task = dict(row)
        conn.execute("""
            UPDATE tasks
            SET    status = ?, updated_at = datetime('now','localtime')
            WHERE  id = ?
        """, (new_status, task_id))

    # NOTIFICATION HOOK: trigger status update notification here
    try:
        task["status"] = new_status
        if new_status == "In Progress":
            create_task_notifications(task, event_type="status_in_progress")
        elif new_status == "Done":
            create_task_notifications(task, event_type="status_done")
    except Exception:
        pass
    return True


def get_all_tasks_for_admin() -> list[dict]:
    """All tasks across all companies — for admin read-only view."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT t.*, co.full_name AS company_name,
                   ab.full_name  AS assigned_by_name,
                   at_.full_name AS assigned_to_name,
                   at_.role      AS assigned_to_role
            FROM   tasks t
            JOIN   companies  co  ON co.unique_code = t.company_code
            JOIN   employees  ab  ON ab.emp_id = t.assigned_by_emp_id
            JOIN   employees  at_ ON at_.emp_id = t.assigned_to_emp_id
            ORDER  BY t.due_date ASC, co.full_name, t.created_at DESC
        """).fetchall()
        return [dict(r) for r in rows]


def get_overdue_tasks() -> list[dict]:
    """Tasks whose due_date has passed and status is not Done."""
    today = date.today().isoformat()
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT t.*, co.full_name AS company_name,
                   ab.full_name  AS assigned_by_name,
                   at_.full_name AS assigned_to_name
            FROM   tasks t
            JOIN   companies  co  ON co.unique_code = t.company_code
            JOIN   employees  ab  ON ab.emp_id = t.assigned_by_emp_id
            JOIN   employees  at_ ON at_.emp_id = t.assigned_to_emp_id
            WHERE  t.due_date < ?
              AND  t.status   != 'Done'
            ORDER  BY t.due_date ASC
        """, (today,)).fetchall()
        return [dict(r) for r in rows]


# ─── REIMBURSEMENT FUNCTIONS ──────────────────────────────────────────────────

REIMB_CATEGORIES  = ("Travel/Conveyance", "Refreshment/Lunch", "Others")
REIMB_SUBCATEGORIES = ("Bus", "Train", "Rickshaw/Taxi", "Auto", "Metro")


def submit_reimbursement(emp_id: str, company_code: str, category: str,
                         sub_category: str, amount: float,
                         description: str, receipt_path: str = None) -> int:
    """Create a new reimbursement entry with status Pending. Returns new id."""
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO reimbursements
              (emp_id, company_code, category, sub_category, amount,
               description, receipt_path, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'Pending')
        """, (emp_id, company_code, category, sub_category or None,
              amount, description or "", receipt_path or None))
        return cur.lastrowid


def get_my_reimbursements(emp_id: str) -> list[dict]:
    """All reimbursements for this employee with company name and status."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT r.*, co.full_name AS company_name
            FROM   reimbursements r
            JOIN   companies co ON co.unique_code = r.company_code
            WHERE  r.emp_id = ?
            ORDER  BY r.submitted_at DESC
        """, (emp_id,)).fetchall()
        return [dict(r) for r in rows]


def get_pending_reimbursements() -> list[dict]:
    """All Pending reimbursements for Admin with employee and company names."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT r.*, co.full_name AS company_name,
                   e.full_name  AS emp_name, e.role AS emp_role
            FROM   reimbursements r
            JOIN   companies  co ON co.unique_code = r.company_code
            JOIN   employees  e  ON e.emp_id = r.emp_id
            WHERE  r.status = 'Pending'
            ORDER  BY r.submitted_at ASC
        """).fetchall()
        return [dict(r) for r in rows]


def get_approved_reimbursements_for_company(company_code: str) -> list[dict]:
    """Approved unbilled reimbursements for a company — used in Billing Advice."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT r.*, e.full_name AS emp_name
            FROM   reimbursements r
            JOIN   employees e ON e.emp_id = r.emp_id
            WHERE  r.company_code = ?
              AND  r.status  = 'Approved'
              AND  r.billed  = 0
            ORDER  BY r.category, r.submitted_at ASC
        """, (company_code,)).fetchall()
        return [dict(r) for r in rows]


def approve_reimbursement(reimbursement_id: int, reviewed_by: str) -> bool:
    """Approve a reimbursement. Returns False if not found."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM reimbursements WHERE id=?", (reimbursement_id,)
        ).fetchone()
        if not row:
            return False
        conn.execute("""
            UPDATE reimbursements
            SET    status='Approved',
                   reviewed_at=datetime('now','localtime'),
                   reviewed_by=?
            WHERE  id=?
        """, (reviewed_by, reimbursement_id))
        return True


def reject_reimbursement(reimbursement_id: int, reviewed_by: str,
                         rejection_note: str) -> bool:
    """Reject a reimbursement with a mandatory note."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM reimbursements WHERE id=?", (reimbursement_id,)
        ).fetchone()
        if not row:
            return False
        conn.execute("""
            UPDATE reimbursements
            SET    status='Rejected',
                   rejection_note=?,
                   reviewed_at=datetime('now','localtime'),
                   reviewed_by=?
            WHERE  id=?
        """, (rejection_note, reviewed_by, reimbursement_id))
        return True


def mark_reimbursements_billed(reimbursement_ids: list, billing_advice_id: int) -> int:
    """Set billed=1 and billing_advice_id for a list of reimbursement ids.
    Returns count of rows updated."""
    if not reimbursement_ids:
        return 0
    placeholders = ",".join("?" * len(reimbursement_ids))
    with get_conn() as conn:
        cur = conn.execute(
            f"UPDATE reimbursements SET billed=1, billing_advice_id=? "
            f"WHERE id IN ({placeholders})",
            [billing_advice_id] + list(reimbursement_ids)
        )
        return cur.rowcount


def get_reimbursement_summary_for_company(company_code: str) -> dict:
    """Return total approved unbilled amount per category for a company.
    Used for Billing Advice auto-pull.
    Returns: {category: {total: float, ids: [int], items: [dict]}}"""
    rows = get_approved_reimbursements_for_company(company_code)
    summary = {}
    for r in rows:
        cat = r["category"]
        if cat not in summary:
            summary[cat] = {"total": 0.0, "ids": [], "items": []}
        summary[cat]["total"]  += r["amount"]
        summary[cat]["ids"].append(r["id"])
        summary[cat]["items"].append(r)
    return summary


def get_intern_oop_for_company(company_code: str) -> list[dict]:
    """Return per-employee approved+unbilled reimbursement totals for a company.
    Used by billing advice auto-pull. Returns list of dicts with emp details + category totals."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT r.emp_id,
                   e.full_name AS emp_name,
                   e.role      AS emp_role,
                   SUM(CASE WHEN r.category='Travel/Conveyance' THEN r.amount ELSE 0 END) AS travel,
                   SUM(CASE WHEN r.category='Refreshment/Lunch' THEN r.amount ELSE 0 END) AS refreshment,
                   SUM(CASE WHEN r.category='Others'           THEN r.amount ELSE 0 END) AS others,
                   SUM(r.amount) AS total,
                   GROUP_CONCAT(r.id) AS reimb_ids
            FROM   reimbursements r
            JOIN   employees e ON e.emp_id = r.emp_id
            WHERE  r.company_code = ?
              AND  r.status = 'Approved'
              AND  r.billed = 0
            GROUP  BY r.emp_id
            ORDER  BY e.full_name
        """, (company_code,)).fetchall()
        result = []
        for row in rows:
            d = dict(row)
            d['reimb_ids'] = [int(x) for x in d['reimb_ids'].split(',')] if d['reimb_ids'] else []
            result.append(d)
        return result


def create_billing_advice(emp_id: str, company_code: str, fy: str,
                           form_data: dict, oop_rows: list) -> int:
    """Save a billing advice draft. Returns the new id."""
    import json as _json
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO billing_advice
              (emp_id, company_code, fy, firm_name, service_desc, audit_date,
               num_items, udin, first_time, xbrl, partner1, partner2,
               trainee1, trainee2, trainee3, months_oop,
               cra2_by, cra4_by, srn_cra2, srn_cra4, cra2_cost, cra4_cost,
               contact_name, contact_desig, additional_info,
               submitted_by, status, form_data_json)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'draft',?)
        """, (
            emp_id, company_code, fy,
            form_data.get('firm',''), form_data.get('serviceDesc',''),
            form_data.get('auditDate',''), form_data.get('numItems',''),
            form_data.get('udin',''), form_data.get('firstTime',''),
            form_data.get('xbrl','0'),
            form_data.get('partner1',''), form_data.get('partner2',''),
            form_data.get('trainee1',''), form_data.get('trainee2',''), form_data.get('trainee3',''),
            form_data.get('months',''),
            form_data.get('cra2By','Company'), form_data.get('cra4By','Company'),
            form_data.get('srnCRA2',''), form_data.get('srnCRA4',''),
            float(form_data.get('cra2Cost') or 0), float(form_data.get('cra4Cost') or 0),
            form_data.get('contactName',''), form_data.get('contactDesig',''),
            form_data.get('additionalInfo',''),
            emp_id, _json.dumps(form_data)
        ))
        ba_id = cur.lastrowid
        # Save per-intern OOP rows
        for row in oop_rows:
            conn.execute("""
                INSERT INTO billing_advice_oop
                  (billing_advice_id, emp_id, person_name, role,
                   travel, refreshment, others, total, reimb_ids_json)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                ba_id,
                row.get('emp_id',''),
                row.get('emp_name',''),
                row.get('emp_role',''),
                float(row.get('travel') or 0),
                float(row.get('refreshment') or 0),
                float(row.get('others') or 0),
                float(row.get('total') or 0),
                _json.dumps(row.get('reimb_ids', []))
            ))
        return ba_id


def get_billing_advice_for_emp(emp_id: str) -> list[dict]:
    """All billing advice drafts submitted by this employee."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT ba.*, co.full_name AS company_name
            FROM   billing_advice ba
            JOIN   companies co ON co.unique_code = ba.company_code
            WHERE  ba.emp_id = ?
            ORDER  BY ba.submitted_at DESC
        """, (emp_id,)).fetchall()
        return [dict(r) for r in rows]


def get_billing_advice_by_id(ba_id: int) -> 'dict | None':
    """Full billing advice row by id."""
    with get_conn() as conn:
        row = conn.execute("""
            SELECT ba.*, co.full_name AS company_name,
                   e.full_name AS emp_full_name
            FROM   billing_advice ba
            JOIN   companies co ON co.unique_code = ba.company_code
            JOIN   employees e  ON e.emp_id = ba.emp_id
            WHERE  ba.id = ?
        """, (ba_id,)).fetchone()
        return dict(row) if row else None


def get_billing_advice_oop_rows(ba_id: int) -> list[dict]:
    """Per-person OOP rows for a billing advice."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM billing_advice_oop WHERE billing_advice_id=? ORDER BY id",
            (ba_id,)
        ).fetchall()
        return [dict(r) for r in rows]


def get_all_billing_advice_drafts() -> list[dict]:
    """All billing advice rows for partner view, ordered by company then date."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT ba.*, co.full_name AS company_name,
                   e.full_name AS submitter_name
            FROM   billing_advice ba
            JOIN   companies co ON co.unique_code = ba.company_code
            JOIN   employees e  ON e.emp_id = ba.emp_id
            ORDER  BY co.full_name ASC, ba.submitted_at DESC
        """).fetchall()
        return [dict(r) for r in rows]


def finalize_billing_advice(ba_id: int, partner_emp_id: str,
                             professional_fee: float,
                             partner_oop1: float, partner_oop2: float,
                             invoice_no: str, invoice_date: str,
                             tax_type: str) -> bool:
    """Partner finalizes a billing advice — marks it finalized and marks all linked reimbs billed."""
    import json as _json
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM billing_advice WHERE id=? AND status='draft'", (ba_id,)
        ).fetchone()
        if not row:
            return False
        # Update billing_advice
        conn.execute("""
            UPDATE billing_advice
               SET status='finalized', professional_fee=?,
                   partner_oop1=?, partner_oop2=?,
                   invoice_no=?, invoice_date=?, tax_type=?,
                   finalized_by=?, finalized_at=datetime('now','localtime')
             WHERE id=?
        """, (professional_fee, partner_oop1, partner_oop2,
              invoice_no, invoice_date, tax_type, partner_emp_id, ba_id))
        # Mark all linked reimbursements as billed
        oop_rows = conn.execute(
            "SELECT reimb_ids_json FROM billing_advice_oop WHERE billing_advice_id=?",
            (ba_id,)
        ).fetchall()
        all_ids = []
        for oor in oop_rows:
            try:
                all_ids.extend(_json.loads(oor['reimb_ids_json'] or '[]'))
            except Exception:
                pass
        if all_ids:
            placeholders = ','.join('?' * len(all_ids))
            conn.execute(
                f"UPDATE reimbursements SET billed=1, billing_advice_id=? WHERE id IN ({placeholders})",
                [ba_id] + all_ids
            )
    return True


def check_active_billing_draft(emp_id: str, company_code: str, fy: str) -> 'dict | None':
    """Return existing draft for same emp+company+fy, or None."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, status FROM billing_advice WHERE emp_id=? AND company_code=? AND fy=? AND status='draft'",
            (emp_id, company_code, fy)
        ).fetchone()
        return dict(row) if row else None


def delete_billing_advice(ba_id: int) -> bool:
    """Delete a draft billing advice (only if still draft)."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM billing_advice WHERE id=? AND status='draft'", (ba_id,)
        ).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM billing_advice_oop WHERE billing_advice_id=?", (ba_id,))
        conn.execute("DELETE FROM billing_advice WHERE id=?", (ba_id,))
    return True


def get_all_reimbursements_for_admin() -> list[dict]:
    """All reimbursements across all employees and companies with full details."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT r.*, co.full_name AS company_name,
                   e.full_name AS emp_name, e.role AS emp_role
            FROM   reimbursements r
            JOIN   companies  co ON co.unique_code = r.company_code
            JOIN   employees  e  ON e.emp_id = r.emp_id
            ORDER  BY r.submitted_at DESC
        """).fetchall()
        return [dict(r) for r in rows]


# ─── LEARNING HUB MIGRATION ──────────────────────────────────────────────────

def _migrate_learning_columns():
    """Safe ALTER TABLE additions for v3.0 → v3.1 upgrade on existing databases."""
    new_cols = [
        ("expiry_date", "TEXT"),
        ("is_pinned",   "INTEGER DEFAULT 0"),
        ("priority",    "TEXT DEFAULT 'Normal'"),
    ]
    with get_conn() as conn:
        existing = {row[1] for row in conn.execute(
            "PRAGMA table_info(learning_materials)"
        ).fetchall()}
        for col, coldef in new_cols:
            if col not in existing:
                try:
                    conn.execute(
                        f"ALTER TABLE learning_materials ADD COLUMN {col} {coldef}"
                    )
                    log.info(f"Migration: added column learning_materials.{col}")
                except Exception as e:
                    log.warning(f"Migration skip {col}: {e}")


# ─── LEAVE RECORDS MIGRATION ─────────────────────────────────────────────────

def _migrate_leave_columns():
    """Safe ALTER TABLE for leave_records — adds status/review cols to existing DBs."""
    new_cols = [
        ("leave_type",  "TEXT DEFAULT 'Casual'"),
        ("status",      "TEXT NOT NULL DEFAULT 'Pending'"),
        ("reviewed_at", "TEXT"),
        ("admin_note",  "TEXT"),
    ]
    with get_conn() as conn:
        existing = {row[1] for row in conn.execute(
            "PRAGMA table_info(leave_records)"
        ).fetchall()}
        for col, coldef in new_cols:
            if col not in existing:
                try:
                    conn.execute(
                        f"ALTER TABLE leave_records ADD COLUMN {col} {coldef}"
                    )
                    # Back-fill: anything already in the table was admin-added = Approved
                    if col == "status":
                        conn.execute(
                            "UPDATE leave_records SET status='Approved' WHERE status IS NULL OR status=''"
                        )
                    log.info(f"Migration: added column leave_records.{col}")
                except Exception as e:
                    log.warning(f"Migration skip leave_records.{col}: {e}")


# ─── EMPLOYEE OPERATIONS ────────────────────────────────────────────────────────

def get_all_employees(include_inactive=False) -> list[dict]:
    with get_conn() as conn:
        if include_inactive:
            rows = conn.execute("SELECT * FROM employees ORDER BY full_name").fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM employees WHERE status='Active' ORDER BY full_name"
            ).fetchall()
        return [dict(r) for r in rows]


def get_employee(emp_id: str) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM employees WHERE emp_id=?", (emp_id,)
        ).fetchone()
        return dict(row) if row else None


def get_employee_by_name(name: str) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM employees WHERE LOWER(full_name)=LOWER(?)", (name,)
        ).fetchone()
        return dict(row) if row else None


def add_employee(emp_id, full_name, role, join_date=None, batch=None,
                 join_year=None, seq=0, reporting_to=None) -> bool:
    try:
        with get_conn() as conn:
            conn.execute("""
                INSERT INTO employees
                (emp_id, full_name, role, join_date, batch, join_year, seq_in_batch, reporting_to)
                VALUES (?,?,?,?,?,?,?,?)
            """, (emp_id, full_name, role, join_date, batch, join_year, seq, reporting_to))
        return True
    except sqlite3.IntegrityError:
        return False


def update_employee(emp_id: str, **kwargs) -> bool:
    allowed = {"full_name", "role", "status", "reporting_to", "t1_path",
               "t5_path", "t5_upload_date", "pin_hash", "batch",
               "join_year", "join_date",
               "internship_end_date", "extension_note"}   # ← extension fields
    fields = {k: v for k, v in kwargs.items() if k in allowed}

    # Handle EMP ID rename separately
    new_emp_id = kwargs.get("new_emp_id", "").strip().upper()

    if not fields and not new_emp_id:
        return False

    with get_conn() as conn:
        if fields:
            fields["updated_at"] = datetime.now().isoformat()
            set_clause = ", ".join(f"{k}=?" for k in fields)
            values = list(fields.values()) + [emp_id]
            conn.execute(f"UPDATE employees SET {set_clause} WHERE emp_id=?", values)

        if new_emp_id and new_emp_id != emp_id.strip().upper():
            # Update the primary key — also cascade to related tables
            conn.execute("UPDATE employees        SET emp_id=? WHERE emp_id=?",
                         (new_emp_id, emp_id))
            conn.execute("UPDATE timesheet_entries SET emp_id=? WHERE emp_id=?",
                         (new_emp_id, emp_id))
            conn.execute("UPDATE attendance_log    SET emp_id=? WHERE emp_id=?",
                         (new_emp_id, emp_id))
            conn.execute("UPDATE leave_records     SET emp_id=? WHERE emp_id=?",
                         (new_emp_id, emp_id))
            # Update the PIN key
            conn.execute("UPDATE admin_pins SET key=? WHERE key=?",
                         (new_emp_id.lower(), emp_id.lower()))
    return True


def remove_employee(emp_id: str):
    with get_conn() as conn:
        conn.execute("DELETE FROM employees WHERE emp_id=?", (emp_id,))


def set_employee_status(emp_id: str, status: str):
    """status: Active / Inactive / Completed"""
    update_employee(emp_id, status=status)


def _db_set_t1_path(emp_id: str, file_path: str):
    update_employee(emp_id, t1_path=file_path)


def _db_set_t5_path(emp_id: str, file_path: str):
    today = date.today().isoformat()
    update_employee(emp_id, t5_path=file_path, t5_upload_date=today, status="Completed")


def is_active(emp_id: str) -> bool:
    emp = get_employee(emp_id)
    return emp is not None and emp["status"] == "Active"


# ─── COMPANY OPERATIONS ─────────────────────────────────────────────────────────

def get_all_companies() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM companies WHERE is_active=1 ORDER BY full_name"
        ).fetchall()
        return [dict(r) for r in rows]


def get_company_by_code(unique_code: str) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM companies WHERE unique_code=?", (unique_code,)
        ).fetchone()
        return dict(row) if row else None


def get_company_by_shortcode(short_code: str) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM companies WHERE UPPER(short_code)=UPPER(?)", (short_code,)
        ).fetchone()
        return dict(row) if row else None


def get_company_by_name(name: str) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM companies WHERE LOWER(full_name)=LOWER(?)", (name,)
        ).fetchone()
        return dict(row) if row else None


def search_companies(query: str, limit=10) -> list[dict]:
    q = f"%{query}%"
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT * FROM companies
            WHERE is_active=1 AND (
                LOWER(full_name)   LIKE LOWER(?)
             OR UPPER(short_code)  LIKE UPPER(?)
             OR UPPER(unique_code) LIKE UPPER(?)
            )
            ORDER BY full_name LIMIT ?
        """, (q, q, q, limit)).fetchall()
        return [dict(r) for r in rows]


def add_company(short_code, full_name, unique_code) -> bool:
    try:
        with get_conn() as conn:
            conn.execute(
                "INSERT INTO companies (short_code, full_name, unique_code) VALUES (?,?,?)",
                (short_code.upper(), full_name, unique_code.upper())
            )
        return True
    except sqlite3.IntegrityError:
        return False


def update_company(unique_code: str, short_code=None, full_name=None) -> bool:
    updates = {}
    if short_code: updates["short_code"] = short_code.upper()
    if full_name:  updates["full_name"]  = full_name
    if not updates: return False
    set_clause = ", ".join(f"{k}=?" for k in updates)
    values = list(updates.values()) + [unique_code]
    with get_conn() as conn:
        conn.execute(f"UPDATE companies SET {set_clause} WHERE unique_code=?", values)
    return True


def remove_company(unique_code: str):
    with get_conn() as conn:
        conn.execute("UPDATE companies SET is_active=0 WHERE unique_code=?", (unique_code,))


def next_company_number() -> int:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT unique_code FROM companies ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if not row:
            return 1
        import re
        m = re.search(r"(\d+)$", row["unique_code"])
        return int(m.group(1)) + 1 if m else 1


# ─── WORK CATEGORY / OPERATIONAL AREA / SUB-CATEGORY ───────────────────────────

def get_work_categories() -> list[str]:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT name FROM work_categories WHERE is_active=1 ORDER BY sort_order, name"
        ).fetchall()
        return [r["name"] for r in rows]


def get_operational_areas(category_name: str) -> list[str]:
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT oa.name FROM operational_areas oa
            JOIN work_categories wc ON oa.category_id = wc.id
            WHERE wc.name=? AND oa.is_active=1
            ORDER BY oa.sort_order, oa.name
        """, (category_name,)).fetchall()
        return [r["name"] for r in rows]


def get_sub_categories(category_name: str, area_name: str) -> list[str]:
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT sc.name FROM sub_categories sc
            JOIN operational_areas oa ON sc.area_id = oa.id
            JOIN work_categories wc ON oa.category_id = wc.id
            WHERE wc.name=? AND oa.name=? AND sc.is_active=1
            ORDER BY sc.sort_order, sc.name
        """, (category_name, area_name)).fetchall()
        return [r["name"] for r in rows]


def add_work_category(name: str, sort_order=0) -> bool:
    try:
        with get_conn() as conn:
            conn.execute(
                "INSERT INTO work_categories (name, sort_order) VALUES (?,?)",
                (name, sort_order)
            )
        return True
    except sqlite3.IntegrityError:
        return False


def add_operational_area(category_name: str, area_name: str, sort_order=0) -> bool:
    with get_conn() as conn:
        cat = conn.execute(
            "SELECT id FROM work_categories WHERE name=?", (category_name,)
        ).fetchone()
        if not cat:
            return False
        try:
            conn.execute(
                "INSERT INTO operational_areas (category_id, name, sort_order) VALUES (?,?,?)",
                (cat["id"], area_name, sort_order)
            )
            return True
        except sqlite3.IntegrityError:
            return False


def add_sub_category(category_name: str, area_name: str, sub_name: str, sort_order=0) -> bool:
    with get_conn() as conn:
        area = conn.execute("""
            SELECT oa.id FROM operational_areas oa
            JOIN work_categories wc ON oa.category_id=wc.id
            WHERE wc.name=? AND oa.name=?
        """, (category_name, area_name)).fetchone()
        if not area:
            return False
        try:
            conn.execute(
                "INSERT INTO sub_categories (area_id, name, sort_order) VALUES (?,?,?)",
                (area["id"], sub_name, sort_order)
            )
            return True
        except sqlite3.IntegrityError:
            return False


# ─── WORK MAPPING — RENAME / DEACTIVATE / DELETE ─────────────────────────────

def rename_work_category(old_name: str, new_name: str) -> tuple[bool, str]:
    """Rename a work category. Returns (ok, message)."""
    new_name = new_name.strip()
    if not new_name:
        return False, "Name cannot be empty."
    try:
        with get_conn() as conn:
            conn.execute(
                "UPDATE work_categories SET name=? WHERE name=?", (new_name, old_name)
            )
        return True, f"Renamed to '{new_name}'"
    except sqlite3.IntegrityError:
        return False, f"'{new_name}' already exists."
    except Exception as e:
        return False, str(e)


def rename_operational_area(area_id: int, new_name: str) -> tuple[bool, str]:
    new_name = new_name.strip()
    if not new_name:
        return False, "Name cannot be empty."
    try:
        with get_conn() as conn:
            # Get category_id to check uniqueness within same category
            row = conn.execute(
                "SELECT category_id FROM operational_areas WHERE id=?", (area_id,)
            ).fetchone()
            if not row:
                return False, "Area not found."
            conn.execute(
                "UPDATE operational_areas SET name=? WHERE id=?", (new_name, area_id)
            )
        return True, f"Renamed to '{new_name}'"
    except sqlite3.IntegrityError:
        return False, f"'{new_name}' already exists in this category."
    except Exception as e:
        return False, str(e)


def rename_sub_category(sub_id: int, new_name: str) -> tuple[bool, str]:
    new_name = new_name.strip()
    if not new_name:
        return False, "Name cannot be empty."
    try:
        with get_conn() as conn:
            conn.execute(
                "UPDATE sub_categories SET name=? WHERE id=?", (new_name, sub_id)
            )
        return True, f"Renamed to '{new_name}'"
    except sqlite3.IntegrityError:
        return False, f"'{new_name}' already exists in this area."
    except Exception as e:
        return False, str(e)


def toggle_work_category(name: str, active: bool):
    with get_conn() as conn:
        conn.execute(
            "UPDATE work_categories SET is_active=? WHERE name=?", (1 if active else 0, name)
        )

def toggle_operational_area(area_id: int, active: bool):
    with get_conn() as conn:
        conn.execute(
            "UPDATE operational_areas SET is_active=? WHERE id=?", (1 if active else 0, area_id)
        )

def toggle_sub_category(sub_id: int, active: bool):
    with get_conn() as conn:
        conn.execute(
            "UPDATE sub_categories SET is_active=? WHERE id=?", (1 if active else 0, sub_id)
        )

def delete_work_category_cascade(name: str) -> tuple[bool, str]:
    """Hard-delete a category and all its areas/subs."""
    try:
        with get_conn() as conn:
            cat = conn.execute(
                "SELECT id FROM work_categories WHERE name=?", (name,)
            ).fetchone()
            if not cat:
                return False, "Category not found."
            cat_id = cat["id"]
            # Collect area IDs
            area_ids = [r["id"] for r in conn.execute(
                "SELECT id FROM operational_areas WHERE category_id=?", (cat_id,)
            ).fetchall()]
            # Delete subs first, then areas, then category
            for aid in area_ids:
                conn.execute("DELETE FROM sub_categories WHERE area_id=?", (aid,))
            conn.execute("DELETE FROM operational_areas WHERE category_id=?", (cat_id,))
            conn.execute("DELETE FROM work_categories WHERE id=?", (cat_id,))
        return True, f"Deleted '{name}' and all its areas/sub-categories."
    except Exception as e:
        return False, str(e)

def delete_operational_area_cascade(area_id: int) -> tuple[bool, str]:
    try:
        with get_conn() as conn:
            row = conn.execute(
                "SELECT name FROM operational_areas WHERE id=?", (area_id,)
            ).fetchone()
            name = row["name"] if row else str(area_id)
            conn.execute("DELETE FROM sub_categories WHERE area_id=?", (area_id,))
            conn.execute("DELETE FROM operational_areas WHERE id=?", (area_id,))
        return True, f"Deleted area '{name}' and its sub-categories."
    except Exception as e:
        return False, str(e)

def delete_sub_category_by_id(sub_id: int) -> tuple[bool, str]:
    try:
        with get_conn() as conn:
            row = conn.execute(
                "SELECT name FROM sub_categories WHERE id=?", (sub_id,)
            ).fetchone()
            name = row["name"] if row else str(sub_id)
            conn.execute("DELETE FROM sub_categories WHERE id=?", (sub_id,))
        return True, f"Deleted sub-category '{name}'."
    except Exception as e:
        return False, str(e)

# ID-based fast lookups (used by refactored timesheet dropdowns)
def get_areas_by_cat_id(cat_id: int) -> list[dict]:
    """Return [{id, name}] for active areas under a category id."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name FROM operational_areas "
            "WHERE category_id=? AND is_active=1 ORDER BY sort_order, name",
            (cat_id,)
        ).fetchall()
        return [dict(r) for r in rows]

def get_subs_by_area_id(area_id: int) -> list[dict]:
    """Return [{id, name}] for active sub-categories under an area id."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name FROM sub_categories "
            "WHERE area_id=? AND is_active=1 ORDER BY sort_order, name",
            (area_id,)
        ).fetchall()
        return [dict(r) for r in rows]

def get_work_categories_full() -> list[dict]:
    """Return [{id, name}] for active categories."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, name FROM work_categories "
            "WHERE is_active=1 ORDER BY sort_order, name"
        ).fetchall()
        return [dict(r) for r in rows]

def get_operational_areas_full(category_name: str) -> list[dict]:
    """Return [{id, name}] for active areas under a named category."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT oa.id, oa.name FROM operational_areas oa
            JOIN work_categories wc ON oa.category_id = wc.id
            WHERE wc.name=? AND oa.is_active=1
            ORDER BY oa.sort_order, oa.name
        """, (category_name,)).fetchall()
        return [dict(r) for r in rows]

def get_sub_categories_full(category_name: str, area_name: str) -> list[dict]:
    """Return [{id, name}] for active subs under named category+area."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT sc.id, sc.name FROM sub_categories sc
            JOIN operational_areas oa ON sc.area_id = oa.id
            JOIN work_categories wc ON oa.category_id = wc.id
            WHERE wc.name=? AND oa.name=? AND sc.is_active=1
            ORDER BY sc.sort_order, sc.name
        """, (category_name, area_name)).fetchall()
        return [dict(r) for r in rows]


def get_full_mapping() -> dict:
    """Return complete {category: {area: [sub, ...]}} mapping for UI."""
    with get_conn() as conn:
        cats = conn.execute(
            "SELECT id, name FROM work_categories WHERE is_active=1 ORDER BY sort_order, name"
        ).fetchall()
        mapping = {}
        for cat in cats:
            areas = conn.execute(
                "SELECT id, name FROM operational_areas WHERE category_id=? AND is_active=1 ORDER BY sort_order, name",
                (cat["id"],)
            ).fetchall()
            mapping[cat["name"]] = {}
            for area in areas:
                subs = conn.execute(
                    "SELECT name FROM sub_categories WHERE area_id=? AND is_active=1 ORDER BY sort_order, name",
                    (area["id"],)
                ).fetchall()
                mapping[cat["name"]][area["name"]] = [s["name"] for s in subs]
    return mapping


# ─── TIMESHEET ENTRY OPERATIONS ─────────────────────────────────────────────────

def save_timesheet_entries(entries: list[dict]) -> tuple:
    """
    Save entries. Returns (True, "") on success or (False, error_msg) on failure.
    company_code is validated against the companies table; unrecognised codes are
    stored as NULL so the FK constraint never fires unexpectedly.
    """
    try:
        with get_conn() as conn:
            valid_codes = {
                row[0] for row in
                conn.execute("SELECT unique_code FROM companies WHERE is_active=1").fetchall()
            }
            for e in entries:
                raw_code = (e.get("company_code") or "").strip().upper()
                safe_code = raw_code if raw_code in valid_codes else None
                conn.execute("""
                    INSERT INTO timesheet_entries
                    (entry_date, day_name, emp_id, reporting_to, company_code, client_name,
                     work_category, operational_area, sub_category, start_time, end_time,
                     break_mins, total_hrs, task_desc, notes, work_location, status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    e.get("date"), e.get("day"), e.get("emp_id"),
                    e.get("reporting_to"), safe_code, e.get("client"),
                    e.get("work_category"), e.get("operational_area"), e.get("sub_category"),
                    e.get("start_time"), e.get("end_time"),
                    e.get("break_mins", 0), e.get("total_hrs", 0),
                    e.get("task_desc"), e.get("notes"),
                    e.get("work_location"), e.get("status", "In Progress")
                ))
        return True, ""
    except Exception as exc:
        msg = str(exc)
        log.error(f"DB save error: {msg}")
        print(f"DB save error: {msg}")
        return False, msg


def load_entries(emp_id: str = None, filter_date: date = None,
                 limit=500) -> list[dict]:
    conditions = []
    params = []
    if emp_id:
        conditions.append("te.emp_id=?"); params.append(emp_id)
    if filter_date:
        conditions.append("te.entry_date=?"); params.append(filter_date.isoformat())
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    params.append(limit)
    with get_conn() as conn:
        rows = conn.execute(f"""
            SELECT te.*, e.full_name, e.role, e.emp_id as employee_id
            FROM timesheet_entries te
            JOIN employees e ON te.emp_id = e.emp_id
            {where}
            ORDER BY te.entry_date DESC, te.id DESC
            LIMIT ?
        """, params).fetchall()
        return [dict(r) for r in rows]


def update_timesheet_entry(entry_id: int, **kwargs) -> bool:
    allowed = {"company_code", "client_name", "work_category", "operational_area",
               "sub_category", "start_time", "end_time", "break_mins", "total_hrs",
               "task_desc", "notes", "work_location", "status", "reporting_to"}
    fields = {k: v for k, v in kwargs.items() if k in allowed}
    if not fields:
        return False
    fields["updated_at"] = datetime.now().isoformat()
    set_clause = ", ".join(f"{k}=?" for k in fields)
    values = list(fields.values()) + [entry_id]
    with get_conn() as conn:
        conn.execute(f"UPDATE timesheet_entries SET {set_clause} WHERE id=?", values)
    return True


# ─── ATTENDANCE / SESSION TRACKING ──────────────────────────────────────────────

def record_login(emp_id: str) -> int:
    """Log a login event. Returns the attendance log row id."""
    now = datetime.now()
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO attendance_log (emp_id, login_dt, session_date)
            VALUES (?,?,?)
        """, (emp_id, now.isoformat(), now.date().isoformat()))
        return cur.lastrowid


def record_logout(session_id: int):
    """Log a logout event and compute duration."""
    now = datetime.now()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT login_dt FROM attendance_log WHERE id=?", (session_id,)
        ).fetchone()
        if row:
            login_dt = datetime.fromisoformat(row["login_dt"])
            duration = (now - login_dt).total_seconds() / 60
            conn.execute("""
                UPDATE attendance_log
                SET logout_dt=?, duration_mins=?
                WHERE id=?
            """, (now.isoformat(), round(duration, 1), session_id))


def get_open_session(emp_id: str) -> 'dict | None':
    """Return an open (no logout) session for this employee if one exists."""
    with get_conn() as conn:
        row = conn.execute("""
            SELECT * FROM attendance_log
            WHERE emp_id=? AND logout_dt IS NULL
            ORDER BY id DESC LIMIT 1
        """, (emp_id,)).fetchone()
        return dict(row) if row else None


def get_attendance_log(emp_id: str = None, filter_date: date = None,
                       limit=200) -> list[dict]:
    conditions = []
    params = []
    if emp_id:
        conditions.append("emp_id=?"); params.append(emp_id)
    if filter_date:
        conditions.append("session_date=?"); params.append(filter_date.isoformat())
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    params.append(limit)
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT * FROM attendance_log {where} ORDER BY id DESC LIMIT ?",
            params
        ).fetchall()
        return [dict(r) for r in rows]


# ─── SUBMISSION TRACKING ──────────────────────────────────────────────────────

def _determine_submission_status(submitted_at_iso: str | None,
                                  entry_count: int, total_hrs: float,
                                  cutoff_time: str = "20:30") -> str:
    """Calculate submission status based on time and completeness."""
    if not submitted_at_iso or entry_count == 0:
        return "Not Submitted"
    try:
        submitted_dt = datetime.fromisoformat(submitted_at_iso)
        cutoff_h, cutoff_m = map(int, cutoff_time.split(":"))
        cutoff_dt = submitted_dt.replace(hour=cutoff_h, minute=cutoff_m, second=0)
        if total_hrs < 4 or entry_count < 1:
            return "Partially Filled"
        if submitted_dt <= cutoff_dt:
            return "On Time"
        return "Late"
    except Exception:
        return "Partially Filled"


def upsert_submission(emp_id: str, submit_date_str: str,
                      entry_count: int, total_hrs: float,
                      cutoff_time: str = "20:30") -> dict:
    """Create or update submission log for a date. Returns the record dict."""
    now_iso = datetime.now().isoformat()
    status = _determine_submission_status(now_iso, entry_count, total_hrs, cutoff_time)
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO submission_log
                (emp_id, submit_date, submitted_at, submission_status,
                 approval_status, total_hrs, entry_count, updated_at)
            VALUES (?,?,?,?, 'Submitted', ?,?,?)
            ON CONFLICT(emp_id, submit_date) DO UPDATE SET
                submitted_at     = excluded.submitted_at,
                submission_status= excluded.submission_status,
                approval_status  = CASE WHEN approval_status='Approved'
                                        THEN 'Submitted' ELSE approval_status END,
                total_hrs        = excluded.total_hrs,
                entry_count      = excluded.entry_count,
                updated_at       = excluded.updated_at
        """, (emp_id, submit_date_str, now_iso, status,
              round(total_hrs, 2), entry_count, now_iso))
        row = conn.execute(
            "SELECT * FROM submission_log WHERE emp_id=? AND submit_date=?",
            (emp_id, submit_date_str)
        ).fetchone()
        return dict(row) if row else {}


def get_submission(emp_id: str, submit_date_str: str) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM submission_log WHERE emp_id=? AND submit_date=?",
            (emp_id, submit_date_str)
        ).fetchone()
        return dict(row) if row else None


def get_daily_submission_summary(target_date_str: str) -> dict:
    """Return counts of each status for a given date across all active employees."""
    with get_conn() as conn:
        active_emps = conn.execute(
            "SELECT emp_id FROM employees WHERE status='Active'"
        ).fetchall()
        total_active = len(active_emps)

        submitted = conn.execute("""
            SELECT submission_status, COUNT(*) as cnt
            FROM submission_log
            WHERE submit_date=?
            GROUP BY submission_status
        """, (target_date_str,)).fetchall()

        status_counts = {r["submission_status"]: r["cnt"] for r in submitted}
        submitted_total = sum(status_counts.values())
        not_submitted = total_active - submitted_total

        rows = conn.execute("""
            SELECT sl.*, e.full_name, e.role
            FROM submission_log sl
            JOIN employees e ON sl.emp_id = e.emp_id
            WHERE sl.submit_date=?
            ORDER BY sl.submission_status, e.full_name
        """, (target_date_str,)).fetchall()

        return {
            "total_active":   total_active,
            "on_time":        status_counts.get("On Time", 0),
            "late":           status_counts.get("Late", 0),
            "partial":        status_counts.get("Partially Filled", 0),
            "not_submitted":  not_submitted,
            "records":        [dict(r) for r in rows],
        }


def get_all_submissions(filter_date: str = None, emp_id: str = None,
                        limit=500) -> list[dict]:
    conditions = []
    params = []
    if filter_date:
        conditions.append("sl.submit_date=?"); params.append(filter_date)
    if emp_id:
        conditions.append("sl.emp_id=?"); params.append(emp_id)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    params.append(limit)
    with get_conn() as conn:
        rows = conn.execute(f"""
            SELECT sl.*, e.full_name, e.role, e.reporting_to
            FROM submission_log sl
            JOIN employees e ON sl.emp_id = e.emp_id
            {where}
            ORDER BY sl.submit_date DESC, e.full_name
            LIMIT ?
        """, params).fetchall()
        return [dict(r) for r in rows]


def update_approval(sub_id: int, approval_status: str,
                    approved_by: str = "", comment: str = ""):
    now = datetime.now().isoformat()
    with get_conn() as conn:
        conn.execute("""
            UPDATE submission_log
            SET approval_status=?, approved_by=?, approved_at=?,
                reviewer_comment=?, updated_at=?
            WHERE id=?
        """, (approval_status, approved_by, now, comment, now, sub_id))


def get_employee_productivity(start_date: str, end_date: str) -> list[dict]:
    """Employee-wise total hours summary for a date range."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT te.emp_id, e.full_name, e.role, e.reporting_to,
                   COUNT(*) as entry_count,
                   ROUND(SUM(te.total_hrs), 2) as total_hrs,
                   COUNT(DISTINCT te.entry_date) as days_worked,
                   AVG(te.total_hrs) as avg_hrs_per_entry
            FROM timesheet_entries te
            JOIN employees e ON te.emp_id = e.emp_id
            WHERE te.entry_date BETWEEN ? AND ?
            GROUP BY te.emp_id, e.full_name, e.role, e.reporting_to
            ORDER BY total_hrs DESC
        """, (start_date, end_date)).fetchall()
        return [dict(r) for r in rows]


def get_client_effort(start_date: str, end_date: str) -> list[dict]:
    """Client-wise total hours summary."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT te.client_name,
                   te.company_code,
                   te.work_category,
                   COUNT(*) as entry_count,
                   COUNT(DISTINCT te.emp_id) as staff_count,
                   ROUND(SUM(te.total_hrs), 2) as total_hrs
            FROM timesheet_entries te
            WHERE te.entry_date BETWEEN ? AND ?
              AND te.client_name IS NOT NULL AND te.client_name != ''
            GROUP BY te.client_name, te.company_code, te.work_category
            ORDER BY total_hrs DESC
        """, (start_date, end_date)).fetchall()
        return [dict(r) for r in rows]


def get_missing_timesheets(target_date: str) -> list[dict]:
    """Return active employees who have NO entries for a given date."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT e.emp_id, e.full_name, e.role, e.reporting_to
            FROM employees e
            WHERE e.status = 'Active'
              AND e.emp_id NOT IN (
                  SELECT DISTINCT emp_id FROM timesheet_entries
                  WHERE entry_date = ?
              )
              AND e.emp_id NOT IN (
                  SELECT DISTINCT emp_id FROM leave_records
                  WHERE start_date <= ? AND end_date >= ?
              )
            ORDER BY e.full_name
        """, (target_date, target_date, target_date)).fetchall()
        return [dict(r) for r in rows]


# ─── LEAVE OPERATIONS ────────────────────────────────────────────────────────────

def add_leave(emp_id: str, start: str, end: str, reason="",
              leave_type="Casual", approved_by="Admin",
              status="Approved"):
    """Insert a leave record. Admin-added leaves default to Approved;
    employee requests default to Pending."""
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO leave_records
                (emp_id, start_date, end_date, reason, leave_type,
                 status, approved_by)
            VALUES (?,?,?,?,?,?,?)
        """, (emp_id, start, end, reason, leave_type, status,
              approved_by if status == "Approved" else None))


def request_leave(emp_id: str, start: str, end: str,
                  reason="", leave_type="Casual"):
    """Employee-initiated leave request — always starts as Pending."""
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO leave_records
                (emp_id, start_date, end_date, reason, leave_type, status)
            VALUES (?,?,?,?,?,'Pending')
        """, (emp_id, start, end, reason, leave_type))


def review_leave(leave_id: int, status: str,
                 reviewed_by: str = "Admin", note: str = ""):
    """Approve or Reject a leave request."""
    now = datetime.now().isoformat()
    with get_conn() as conn:
        conn.execute("""
            UPDATE leave_records
            SET status=?, approved_by=?, reviewed_at=?, admin_note=?
            WHERE id=?
        """, (status, reviewed_by, now, note, leave_id))


def remove_leave(leave_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM leave_records WHERE id=?", (leave_id,))


def get_leaves(emp_id: str = None, status_filter: str = None) -> list[dict]:
    conditions = []
    params = []
    if emp_id:
        conditions.append("lr.emp_id=?"); params.append(emp_id)
    if status_filter:
        conditions.append("lr.status=?"); params.append(status_filter)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    with get_conn() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, e.full_name
            FROM leave_records lr
            JOIN employees e ON lr.emp_id = e.emp_id
            {where}
            ORDER BY lr.created_at DESC
        """, params).fetchall()
        return [dict(r) for r in rows]


def get_pending_leave_count() -> int:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM leave_records WHERE status='Pending'"
        ).fetchone()
        return row[0] if row else 0


def is_on_leave(emp_id: str, check_date: date) -> tuple[bool, str]:
    """Only Approved leaves block the timesheet."""
    ds = check_date.isoformat()
    with get_conn() as conn:
        row = conn.execute("""
            SELECT reason FROM leave_records
            WHERE emp_id=? AND status='Approved'
              AND start_date<=? AND end_date>=?
            LIMIT 1
        """, (emp_id, ds, ds)).fetchone()
        return (True, row["reason"] or "") if row else (False, "")


# ─── PIN / AUTH ──────────────────────────────────────────────────────────────────

def _hash(pin: str) -> str:
    return hashlib.sha256(pin.strip().encode()).hexdigest()


def verify_pin(emp_id_or_key: str, pin: str) -> bool:
    key = emp_id_or_key.strip().lower()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT pin_hash FROM admin_pins WHERE key=?", (key,)
        ).fetchone()
        if not row:
            # First login — set PIN
            conn.execute(
                "INSERT INTO admin_pins (key, pin_hash) VALUES (?,?)",
                (key, _hash(pin))
            )
            return True
        return row["pin_hash"] == _hash(pin)


def has_pin(emp_id_or_key: str) -> bool:
    key = emp_id_or_key.strip().lower()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT 1 FROM admin_pins WHERE key=?", (key,)
        ).fetchone()
        return row is not None


def change_pin(emp_id_or_key: str, new_pin: str):
    key = emp_id_or_key.strip().lower()
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO admin_pins (key, pin_hash) VALUES (?,?)
            ON CONFLICT(key) DO UPDATE SET pin_hash=excluded.pin_hash,
                                           updated_at=datetime('now')
        """, (key, _hash(new_pin)))


# ─── IMPORT FROM LEGACY TXT FILES ───────────────────────────────────────────────

def import_from_txt_files(staff_file="staff.txt", companies_file="companies.txt",
                          pins_file="staff_pins.txt"):
    """One-time migration from legacy flat-file format to SQLite."""
    imported = {"employees": 0, "companies": 0, "pins": 0}

    # Fix relative-path issues when app is launched from another folder/exe.
    staff_file = str((BASE_DIR / staff_file) if not os.path.isabs(staff_file) else staff_file)
    companies_file = str((BASE_DIR / companies_file) if not os.path.isabs(companies_file) else companies_file)
    pins_file = str((BASE_DIR / pins_file) if not os.path.isabs(pins_file) else pins_file)

    if os.path.exists(staff_file):
        with open(staff_file, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split("|")
                if len(parts) >= 2:
                    name = parts[0].strip().title()
                    role = parts[1].strip() if len(parts) > 1 else "Intern"
                    yr   = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 2026
                    batch = parts[3].strip() if len(parts) > 3 else ""
                    seq  = int(parts[4]) if len(parts) > 4 and parts[4].isdigit() else 0
                    raw_id = parts[5].strip() if len(parts) > 5 else "0"
                    try:
                        emp_id = f"{int(raw_id):08d}" if raw_id.isdigit() and int(raw_id) > 0 else f"TMP{seq:04d}"
                    except Exception:
                        emp_id = raw_id or f"TMP{seq:04d}"
                    ok = add_employee(emp_id, name, role, batch=batch,
                                      join_year=yr, seq=seq)
                    if ok:
                        imported["employees"] += 1

    if os.path.exists(companies_file):
        with open(companies_file, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split("|", 2)
                if len(parts) == 3:
                    ok = add_company(parts[0].strip(), parts[1].strip(), parts[2].strip())
                    if ok:
                        imported["companies"] += 1

    if os.path.exists(pins_file):
        with open(pins_file, "r", encoding="utf-8") as f:
            with get_conn() as conn:
                for line in f:
                    parts = line.strip().split("|", 1)
                    if len(parts) == 2:
                        key = parts[0].strip().lower()
                        h   = parts[1].strip()
                        conn.execute("""
                            INSERT INTO admin_pins (key, pin_hash) VALUES (?,?)
                            ON CONFLICT(key) DO UPDATE SET pin_hash=excluded.pin_hash
                        """, (key, h))
                        imported["pins"] += 1

    return imported


# ─── SEED DEFAULT WORK MAPPING ───────────────────────────────────────────────────

# ── JAA OFFICIAL WORK MAPPING — sourced from JAA-ERP.xlsx ──────────────────────
DEFAULT_MAPPING = {
    "Concurrent Audit-Bank": {
        "Loans and Advances": [
            "Sanction & Documentation",
            "Post-Disbursement Monitoring",
        ],
        "Deposits and Accounts Opening": [
            "KYC/AML Compliance",
            "Current Account Operations",
            "Interest Calculation",
        ],
        "Compliance and Reporting": [
            "Audit Compliance",
            "Reporting",
        ],
    },
    "Cost Audit/ Record": {
        "Cost Audit/ Records working": [
            "Sales",
            "Purchase",
            "Expenses",
            "Consumption",
            "Movement",
            "Quantitative",
            "Financials",
            "Recipe",
            "Technical Data/ SKF",
        ],
        "Cost Audit Report Reparation": [
            "A4", "B1", "B2", "B2A", "B2B",
            "C1", "C2", "C2A",
            "D1", "D2", "D3", "D4", "D5", "D6",
        ],
    },
    "Internal Audit": {
        "Inventory & Stores": [
            "Stock Movement",
            "Physical Verification",
            "Inventory Valuation",
            "Slow/Non-Moving Stock Review",
        ],
        "Sales & Revenue": [
            "Order Processing",
            "Billing & Dispatch",
            "Collection Follow-up",
            "Credit Limit Monitoring",
        ],
        "Costing & MIS": [
            "Cost Accounting",
            "Cost Sheet Preparation",
            "Product Costing",
            "Budgeting & Budgetary Control",
            "Standard Costing",
            "Variance Analysis",
            "Material Cost Management",
            "Labour Cost Management",
            "Overhead Cost Allocation",
            "Inventory Valuation",
            "Profitability Analysis",
            "Margin Analysis",
            "Cost Records Maintenance",
            "Cost Audit Support",
            "MIS Reporting",
            "Production Cost Analysis",
            "Process Costing",
            "Job / Batch Costing",
            "Activity-Based Costing (ABC)",
            "Cost Centre Accounting",
            "Inter-Unit Cost Allocation",
            "Operational Cost Analysis",
            "Utility Cost Analysis",
            "ERP Costing Integration",
            "Forecasting & Cost Planning",
        ],
    },
    "Projects": {
        "Fixed Assets": [
            "Fixed Asset Accounting",
            "Capital Expenditure (CAPEX) Management",
            "Asset Capitalisation",
            "Asset Register Management",
            "Depreciation Accounting",
            "Asset Valuation",
            "Asset Verification & Control",
            "Asset Compliance & Reporting",
            "Lease Asset Accounting",
            "Impairment & Write-off Management",
            "Asset Budgeting & Planning",
            "Asset Disposal & Scrap Management",
            "ERP & Fixed Asset Module Management",
        ],
        "Costing System on Operational & Process": [
            "Cost Accounting",
            "Product Costing",
            "Process Costing",
            "Standard Costing",
            "Budgeting & Budgetary Control",
            "Material Cost Management",
            "Labour Cost Management",
            "Overhead Accounting",
            "Cost Centre Accounting",
            "Inventory Costing",
            "Variance Analysis",
            "Profitability & Margin Analysis",
            "Cost Audit & Compliance",
            "MIS & Management Reporting",
            "ERP Costing Integration",
            "Activity-Based Costing (ABC)",
            "Utility & Service Cost Allocation",
            "Production Cost Analysis",
            "Financial & Cost Reconciliation",
        ],
    },
    "Stock Audit": {
        "Physical Verification Operations": [
            "Physical Stock Counting",
            "Bin-wise Verification",
            "Location-wise Stock Checking",
            "Identification of Damaged/Obsolete Stock",
        ],
        "Inventory Movement Operations": [
            "GRN Verification",
            "Material Issue & Return Checking",
            "Transfer Entry Verification",
            "Dispatch Verification",
        ],
        "Warehouse Operations": [
            "Storage Condition Review",
            "Stock Segregation Verification",
            "Batch-wise / Lot-wise Monitoring",
        ],
    },
    "Assignment": {
        "Consultancy / Advisory Services": [
            "Data Collection & Verification",
            "Analysis & Review Operations",
            "Audit & Checking Operations",
            "Reporting Operations",
            "System & ERP Operations",
        ],
        "ERP Implementation & Support": [
            "ERP Project Management",
            "Business Process Mapping",
            "ERP Functional Consulting",
        ],
    },
    "Local Content": {
        "Procurement Localization": [
            "Local vs Imported Material Tracking",
            "Purchase Order Verification",
            "Country of Origin Validation",
        ],
        "Costing & Local Value Addition Analysis": [
            "Cost Allocation for Local Content",
            "Import Cost Analysis",
            "Localization Benefit Analysis",
            "Percentage of Local Content Calculation",
        ],
    },
    "Certification": {
        "CAS (Cost Accounting Standards) Compliance Certificate": [
            "Cost Data Collection",
        ],
        "Cost Certificate": [
            "Material Cost Management",
            "Labour Cost Management",
            "Overhead Allocation Management",
            "Capacity & Production Operations",
            "Cost Sheet Verification",
        ],
        "GST Cost Data Certification": [
            "Vendor Invoice Verification",
            "GSTIN Validation",
            "Output GST Verification",
            "Tax Invoice Validation",
            "E-way Bill Verification",
            "Debit/Credit Note Review",
            "ITC Reconciliation with GSTR-2B",
            "Blocked Credit Verification",
            "Apportionment of Common Credits",
            "ITC Utilisation Review",
            "GSTR-1 vs Sales Register Reconciliation",
            "GSTR-3B vs Books Reconciliation",
            "GST Liability Reconciliation",
            "Financial vs GST Data Matching",
        ],
        "Inventory Valuation": [
            "Stock Verification Operations",
            "Material Valuation Operations",
            "Inventory Reconciliation Operations",
            "Costing Operations",
            "WIP & Finished Goods Operations",
            "Documentation & Compliance Operations",
        ],
    },
    "Study": {
        "Inter/ Final": [],
    },
    "Others": {
        "General": [
            "Internal Meeting",
            "Training",
            "Administrative Work",
            "Research",
            "Other",
        ],
    },
}


def seed_work_mapping(overwrite=False):
    """Insert DEFAULT_MAPPING into DB. Skip if data already exists unless overwrite=True."""
    existing = get_work_categories()
    if existing and not overwrite:
        return
    with get_conn() as conn:
        if overwrite:
            conn.execute("DELETE FROM sub_categories")
            conn.execute("DELETE FROM operational_areas")
            conn.execute("DELETE FROM work_categories")
        for sort_c, (cat_name, areas) in enumerate(DEFAULT_MAPPING.items()):
            conn.execute(
                "INSERT OR IGNORE INTO work_categories (name, sort_order) VALUES (?,?)",
                (cat_name, sort_c)
            )
            cat_id = conn.execute(
                "SELECT id FROM work_categories WHERE name=?", (cat_name,)
            ).fetchone()["id"]
            for sort_a, (area_name, subs) in enumerate(areas.items()):
                conn.execute(
                    "INSERT OR IGNORE INTO operational_areas (category_id, name, sort_order) VALUES (?,?,?)",
                    (cat_id, area_name, sort_a)
                )
                area_id = conn.execute(
                    "SELECT id FROM operational_areas WHERE category_id=? AND name=?",
                    (cat_id, area_name)
                ).fetchone()["id"]
                for sort_s, sub in enumerate(subs):
                    conn.execute(
                        "INSERT OR IGNORE INTO sub_categories (area_id, name, sort_order) VALUES (?,?,?)",
                        (area_id, sub, sort_s)
                    )


# ═══════════════════════════════════════════════════════════════════════════════
#  LEARNING HUB — DATABASE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

# ── Default categories seeded on first run ────────────────────────────────────
_DEFAULT_LEARN_CATEGORIES = [
    ("Notices",        "📢"),
    ("SOPs",           "📋"),
    ("Excel Learning", "📗"),
    ("Others",         "📂"),
]

def seed_learning_categories():
    """Seed default learning categories. On first run inserts all defaults.
    On subsequent runs: removes obsolete categories and ensures new ones exist."""
    _KEEP_CATS = {name for name, _ in _DEFAULT_LEARN_CATEGORIES}
    _REMOVE_CATS = {
        "📢 Notices", "GST", "Audit", "Income Tax", "Accounting",
        "ROC / Companies Act", "Internal SOP", "SOP",
        "Client Procedures", "Client Procedure", "Client Instructions",
        "Compliance", "Guidelines", "Cost Audit", "Internal Training",
        "Documentation", "General", "Training", "Excel & Tools",
        "Office Policies", "MIS",
    } - _KEEP_CATS
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM learning_categories").fetchone()[0]
        if count == 0:
            # Fresh DB — insert all defaults
            for i, (name, icon) in enumerate(_DEFAULT_LEARN_CATEGORIES):
                conn.execute(
                    "INSERT OR IGNORE INTO learning_categories (name, icon, sort_order) VALUES (?,?,?)",
                    (name, icon, i)
                )
        else:
            # Existing DB — soft-delete obsolete categories
            for bad in _REMOVE_CATS:
                conn.execute(
                    "UPDATE learning_categories SET is_active=0 WHERE name=?", (bad,)
                )
            # Ensure every default category exists (add if missing)
            existing_names = {
                r[0] for r in conn.execute(
                    "SELECT name FROM learning_categories"
                ).fetchall()
            }
            for i, (name, icon) in enumerate(_DEFAULT_LEARN_CATEGORIES):
                if name not in existing_names:
                    conn.execute(
                        "INSERT INTO learning_categories (name, icon, sort_order) VALUES (?,?,?)",
                        (name, icon, i)
                    )
                else:
                    # Re-activate if it was previously deactivated
                    conn.execute(
                        "UPDATE learning_categories SET is_active=1, sort_order=? WHERE name=?",
                        (i, name)
                    )

def get_learning_categories() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM learning_categories WHERE is_active=1 ORDER BY sort_order, name"
        ).fetchall()
        return [dict(r) for r in rows]

def add_learning_category(name: str, icon: str = "📂") -> bool:
    try:
        with get_conn() as conn:
            conn.execute(
                "INSERT INTO learning_categories (name, icon) VALUES (?,?)", (name, icon)
            )
        return True
    except Exception as e:
        log.error(f"add_learning_category: {e}")
        return False

def delete_learning_category(cat_id: int):
    with get_conn() as conn:
        conn.execute("UPDATE learning_categories SET is_active=0 WHERE id=?", (cat_id,))

def get_learning_materials(category: str = None, material_type: str = None,
                            search: str = None, active_only=True) -> list[dict]:
    conditions = []
    params = []
    if active_only:
        conditions.append("is_active=1")
    if category and category != "All":
        conditions.append("category=?"); params.append(category)
    if material_type and material_type != "All":
        conditions.append("material_type=?"); params.append(material_type)
    if search:
        q = f"%{search}%"
        conditions.append("(LOWER(title) LIKE LOWER(?) OR LOWER(tags) LIKE LOWER(?) OR LOWER(description) LIKE LOWER(?))")
        params.extend([q, q, q])
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT * FROM learning_materials {where} ORDER BY is_pinned DESC, upload_date DESC, title",
            params
        ).fetchall()
        return [dict(r) for r in rows]

def get_learning_material(mat_id: int) -> 'dict | None':
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM learning_materials WHERE id=?", (mat_id,)
        ).fetchone()
        return dict(row) if row else None

def add_learning_material(title: str, category: str, material_type: str,
                          file_path: str = None, youtube_url: str = None,
                          description: str = None, tags: str = None,
                          uploaded_by: str = "", article_body: str = None,
                          is_pinned: int = 0, priority: str = "Normal",
                          expiry_date: str = None) -> int:
    """Insert a material record. Returns new row id."""
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO learning_materials
            (title, category, material_type, file_path, youtube_url,
             description, tags, uploaded_by, upload_date,
             is_pinned, priority, expiry_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (title, category, material_type, file_path, youtube_url,
              description, tags, uploaded_by, date.today().isoformat(),
              is_pinned, priority, expiry_date))
        new_id = cur.lastrowid
        if article_body is not None:
            conn.execute(
                "INSERT INTO learning_articles (material_id, body) VALUES (?,?)",
                (new_id, article_body)
            )
        return new_id

def update_learning_material(mat_id: int, title: str, category: str,
                              description: str, tags: str,
                              article_body: str = None,
                              is_pinned: int = 0, priority: str = "Normal",
                              expiry_date: str = None):
    with get_conn() as conn:
        conn.execute("""
            UPDATE learning_materials
            SET title=?, category=?, description=?, tags=?,
                is_pinned=?, priority=?, expiry_date=?, updated_at=?
            WHERE id=?
        """, (title, category, description, tags,
              is_pinned, priority, expiry_date,
              datetime.now().isoformat(), mat_id))
        if article_body is not None:
            exists = conn.execute(
                "SELECT id FROM learning_articles WHERE material_id=?", (mat_id,)
            ).fetchone()
            if exists:
                conn.execute(
                    "UPDATE learning_articles SET body=?, updated_at=? WHERE material_id=?",
                    (article_body, datetime.now().isoformat(), mat_id)
                )
            else:
                conn.execute(
                    "INSERT INTO learning_articles (material_id, body) VALUES (?,?)",
                    (mat_id, article_body)
                )

def delete_learning_material(mat_id: int, delete_file=True):
    mat = get_learning_material(mat_id)
    if mat and mat.get("file_path") and delete_file:
        try:
            fp = Path(mat["file_path"])
            if fp.exists():
                fp.unlink()
        except Exception as e:
            log.warning(f"delete_learning_material file: {e}")
    with get_conn() as conn:
        conn.execute("UPDATE learning_materials SET is_active=0 WHERE id=?", (mat_id,))

def get_article_body(mat_id: int) -> str:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT body FROM learning_articles WHERE material_id=?", (mat_id,)
        ).fetchone()
        return row["body"] if row and row["body"] else ""

def increment_view_count(mat_id: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE learning_materials SET view_count=view_count+1 WHERE id=?", (mat_id,)
        )

def _copy_to_learning_store(src_path: str, material_type: str) -> str:
    """Copy an uploaded file into the correct learning sub-folder. Returns stored path."""
    src = Path(src_path)
    ext = src.suffix.lower()
    # SOP sub-folder reuses LEARN_DOCS (DOCX) or LEARN_PDFS depending on ext
    sop_dir = LEARN_DIR / "sops"
    sop_dir.mkdir(parents=True, exist_ok=True)
    folder_map = {
        ".pdf":  LEARN_PDFS,
        ".pptx": LEARN_PPTS, ".ppt": LEARN_PPTS,
        ".docx": LEARN_DOCS, ".doc": LEARN_DOCS,
        ".xlsx": LEARN_EXCEL, ".xls": LEARN_EXCEL,
    }
    # SOP type: use the dedicated sops folder
    if material_type == "SOP":
        dest_dir = sop_dir
    else:
        dest_dir = folder_map.get(ext, LEARN_MISC)
    dest_dir.mkdir(parents=True, exist_ok=True)
    # Avoid collisions: prefix with timestamp
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_")
    dest = dest_dir / (ts + src.name)
    shutil.copy2(str(src), str(dest))
    return str(dest)

# ─── EMAIL NOTIFICATION SYSTEM ───────────────────────────────────────────────

def _log_email(recipient: str, subject: str, status: str, error: str = ""):
    with get_conn() as conn:
        conn.execute("""
            INSERT INTO email_log (sent_at, recipient, subject, status, error_msg)
            VALUES (?,?,?,?,?)
        """, (datetime.now().isoformat(), recipient, subject, status, error))


def send_email(cfg: dict, to_addr: str, subject: str, body_html: str) -> tuple[bool, str]:
    """Send an email via SMTP. Returns (success, message)."""
    if not cfg.get("email_enabled"):
        return False, "Email disabled in settings."
    if not cfg.get("smtp_user") or not cfg.get("smtp_pass"):
        return False, "SMTP credentials not configured."
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = cfg["smtp_user"]
        msg["To"]      = to_addr
        msg.attach(MIMEText(body_html, "html", "utf-8"))

        with smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"]), timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.login(cfg["smtp_user"], cfg["smtp_pass"])
            server.sendmail(cfg["smtp_user"], [to_addr], msg.as_string())

        _log_email(to_addr, subject, "Sent")
        log.info(f"Email sent to {to_addr}: {subject}")
        return True, "Sent"
    except Exception as e:
        err = str(e)
        _log_email(to_addr, subject, "Failed", err)
        log.error(f"Email failed to {to_addr}: {err}")
        return False, err


def send_daily_summary_email(cfg: dict, target_date: str = None):
    """Send end-of-day summary to manager email."""
    if not cfg.get("manager_email"):
        return False, "Manager email not set."
    target_date = target_date or date.today().isoformat()
    summary = get_daily_submission_summary(target_date)
    missing = get_missing_timesheets(target_date)

    on_time_records  = [r for r in summary["records"] if r["submission_status"] == "On Time"]
    late_records     = [r for r in summary["records"] if r["submission_status"] == "Late"]
    partial_records  = [r for r in summary["records"] if r["submission_status"] == "Partially Filled"]

    def _rows(records):
        if not records:
            return "<li style='color:#888'>None</li>"
        return "".join(f"<li>{r['full_name']} ({r['role']}) — {r.get('total_hrs',0):.1f} hrs</li>"
                       for r in records)

    body = f"""
    <html><body style="font-family:Calibri,Arial;color:#222;max-width:640px">
    <div style="background:#06355E;color:#E0E9F4;padding:18px 24px;border-radius:6px 6px 0 0">
      <h2 style="margin:0">📋 JAA Daily Timesheet Summary</h2>
      <p style="margin:4px 0 0;opacity:.8">Date: {target_date} &nbsp;|&nbsp; Generated: {datetime.now().strftime('%H:%M')}</p>
    </div>
    <div style="background:#EBF2FA;padding:16px 24px">
      <table style="width:100%;border-collapse:collapse">
        <tr>
          <td style="background:#1A6B45;color:white;padding:12px;border-radius:4px;text-align:center">
            <div style="font-size:28px;font-weight:bold">{summary['on_time']}</div>
            <div>✅ On Time</div>
          </td>
          <td style="width:8px"></td>
          <td style="background:#8A6A1A;color:white;padding:12px;border-radius:4px;text-align:center">
            <div style="font-size:28px;font-weight:bold">{summary['late']}</div>
            <div>⚠ Late</div>
          </td>
          <td style="width:8px"></td>
          <td style="background:#B52A2A;color:white;padding:12px;border-radius:4px;text-align:center">
            <div style="font-size:28px;font-weight:bold">{summary['not_submitted']}</div>
            <div>❌ Not Submitted</div>
          </td>
          <td style="width:8px"></td>
          <td style="background:#4A6080;color:white;padding:12px;border-radius:4px;text-align:center">
            <div style="font-size:28px;font-weight:bold">{summary['partial']}</div>
            <div>📝 Partial</div>
          </td>
        </tr>
      </table>
    </div>
    <div style="padding:16px 24px">
      <h3 style="color:#1A6B45">✅ On Time ({len(on_time_records)})</h3>
      <ul>{_rows(on_time_records)}</ul>
      <h3 style="color:#8A6A1A">⚠ Late Submissions ({len(late_records)})</h3>
      <ul>{_rows(late_records)}</ul>
      <h3 style="color:#B52A2A">❌ Not Submitted ({len(missing)})</h3>
      <ul>{"".join(f"<li>{m['full_name']} ({m['role']})</li>" for m in missing) or "<li>None</li>"}</ul>
    </div>
    <div style="background:#06355E;color:#A8C4DC;padding:10px 24px;font-size:11px;border-radius:0 0 6px 6px">
      JAA Timesheet System v3.0 &nbsp;·&nbsp; Joshi Apte &amp; Associates
    </div>
    </body></html>
    """
    return send_email(cfg, cfg["manager_email"],
                      f"JAA Timesheet Summary — {target_date}", body)


def send_reminder_email(cfg: dict, emp_email: str, emp_name: str):
    """Send reminder to individual employee."""
    body = f"""
    <html><body style="font-family:Calibri,Arial;color:#222;max-width:520px">
    <div style="background:#06355E;color:#E0E9F4;padding:18px 24px;border-radius:6px 6px 0 0">
      <h2 style="margin:0">⏰ Timesheet Reminder</h2>
    </div>
    <div style="padding:20px 24px">
      <p>Hi <strong>{emp_name}</strong>,</p>
      <p>This is a friendly reminder to submit your timesheet for today
         (<strong>{date.today().strftime('%d %B %Y')}</strong>).</p>
      <p>Cutoff time: <strong>{cfg.get('submission_cutoff','20:30')}</strong></p>
      <p>Please log in and save your entries before the cutoff to avoid a late mark.</p>
    </div>
    <div style="background:#06355E;color:#A8C4DC;padding:10px 24px;font-size:11px;border-radius:0 0 6px 6px">
      JAA Timesheet System v3.0 &nbsp;·&nbsp; Joshi Apte &amp; Associates
    </div>
    </body></html>
    """
    return send_email(cfg, emp_email,
                      f"⏰ Reminder: Submit your timesheet — {date.today().isoformat()}", body)


# ════════════════════════════════════════════════════════════════════════════
# DOCUMENT MANAGER
# ════════════════════════════════════════════════════════════════════════════





def _ensure_dirs():
    T1_DIR.mkdir(parents=True, exist_ok=True)
    T5_DIR.mkdir(parents=True, exist_ok=True)


def _validate_pdf(file_path: str) -> tuple[bool, str]:
    """Validate file is PDF and within size limit."""
    if not os.path.exists(file_path):
        return False, "File not found."
    # Check extension
    if not file_path.lower().endswith(".pdf"):
        return False, "Only PDF files are accepted."
    # Check magic bytes (PDF signature: %PDF)
    try:
        with open(file_path, "rb") as f:
            header = f.read(4)
            if header != b"%PDF":
                return False, "File is not a valid PDF."
    except Exception as e:
        return False, f"Cannot read file: {e}"
    # Check size
    size_mb = os.path.getsize(file_path) / (1024 * 1024)
    if size_mb > MAX_PDF_MB:
        return False, f"File too large ({size_mb:.1f} MB). Max allowed: {MAX_PDF_MB} MB."
    return True, "OK"


def _safe_filename(emp_id: str, form_type: str, original_name: str) -> str:
    """Generate a safe, unique filename."""
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = os.path.splitext(original_name)[1].lower() or ".pdf"
    return f"{emp_id}_{form_type}_{ts}{ext}"


def upload_t1(emp_id: str, source_path: str) -> tuple[bool, str]:
    """
    Validate and copy a T1 PDF into the T1 folder.
    Returns (success: bool, message_or_stored_path: str).
    """
    _ensure_dirs()
    ok, msg = _validate_pdf(source_path)
    if not ok:
        return False, msg

    filename = _safe_filename(emp_id, "T1", os.path.basename(source_path))
    dest = T1_DIR / filename
    try:
        shutil.copy2(source_path, dest)
        # Update DB
        _db_set_t1_path(emp_id, str(dest))
        return True, str(dest)
    except Exception as e:
        return False, f"Upload failed: {e}"


def upload_t5(emp_id: str, source_path: str) -> tuple[bool, str]:
    """
    Validate and copy a T5 PDF into the T5 folder.
    Automatically deactivates the employee's timesheet access.
    Returns (success: bool, message_or_stored_path: str).
    """
    _ensure_dirs()
    ok, msg = _validate_pdf(source_path)
    if not ok:
        return False, msg

    filename = _safe_filename(emp_id, "T5", os.path.basename(source_path))
    dest = T5_DIR / filename
    try:
        shutil.copy2(source_path, dest)
        # Update DB — also sets status = Completed
        _db_set_t5_path(emp_id, str(dest))
        return True, str(dest)
    except Exception as e:
        return False, f"Upload failed: {e}"


def get_document_info(emp_id: str) -> dict:
    """Return T1/T5 paths and status for an employee."""
    emp = get_employee(emp_id)
    if not emp:
        return {}
    return {
        "t1_path":        emp.get("t1_path"),
        "t5_path":        emp.get("t5_path"),
        "t5_upload_date": emp.get("t5_upload_date"),
        "status":         emp.get("status"),
        "t1_exists":      bool(emp.get("t1_path") and os.path.exists(emp["t1_path"])),
        "t5_exists":      bool(emp.get("t5_path") and os.path.exists(emp["t5_path"])),
    }


def open_document(file_path: str):
    """Open a PDF in the system default viewer."""
    if not file_path or not os.path.exists(file_path):
        return False, "Document not found."
    import subprocess, sys
    try:
        if sys.platform.startswith("win"):
            os.startfile(file_path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])
        return True, "Opened."
    except Exception as e:
        return False, str(e)

# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════




MASTER_COLS = [
    "Date", "Day", "Employee Name", "Employee ID", "Designation",
    "Reporting To", "Firm Name", "Company Code", "Client / Company",
    "Work Category", "Operational Area", "Sub-Category",
    "Start Time", "End Time", "Break (mins)", "Total Hrs",
    "Task Description", "Notes / Remarks", "Location", "Status"
]

COL_WIDTHS = [12, 10, 22, 12, 14, 18, 24, 14, 30,
              22, 22, 18, 11, 11, 12, 10,
              35, 30, 16, 14]

CLR_HDR_BG = "06355E"
CLR_HDR_FG = "E0E9F4"
CLR_ROW_ODD  = "D2E1F1"
CLR_ROW_EVEN = "FFFFFF"


def _acquire_lock() -> bool:
    try:
        if LOCK_FILE.exists():
            age = (datetime.now().timestamp() - LOCK_FILE.stat().st_mtime)
            if age > 60:
                LOCK_FILE.unlink()
            else:
                return False
        LOCK_FILE.touch()
        return True
    except Exception:
        return False


def _release_lock():
    try:
        if LOCK_FILE.exists():
            LOCK_FILE.unlink()
    except Exception:
        pass


def _backup():
    if not MASTER_XLSX.exists():
        return
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(MASTER_XLSX, BACKUP_DIR / f"JAA_Master_Backup_{ts}.xlsx")


def ensure_master_file():
    """Create master Excel file with headers if it doesn't exist."""
    if MASTER_XLSX.exists():
        return
    MASTER_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Entries"

    hdr_font  = Font(name="Calibri", bold=True, color=CLR_HDR_FG, size=10)
    hdr_fill  = PatternFill("solid", fgColor=CLR_HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (col, w) in enumerate(zip(MASTER_COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary — auto-generated on each save."
    ws2["A1"].font = Font(name="Calibri", bold=True, color=CLR_HDR_BG, size=11)

    wb.save(str(MASTER_XLSX))


def _entry_to_row(e: dict) -> list:
    """Map a DB entry dict to an Excel row."""
    return [
        e.get("entry_date", ""),
        e.get("day_name", ""),
        e.get("full_name", e.get("employee_name", "")),
        e.get("employee_id", e.get("emp_id", "")),
        e.get("role", e.get("designation", "")),
        e.get("reporting_to", ""),
        "Joshi Apte & Associates",
        e.get("company_code", ""),
        e.get("client_name", e.get("client", "")),
        e.get("work_category", ""),
        e.get("operational_area", ""),
        e.get("sub_category", ""),
        e.get("start_time", ""),
        e.get("end_time", ""),
        e.get("break_mins", 0),
        e.get("total_hrs", ""),
        e.get("task_desc", ""),
        e.get("notes", ""),
        e.get("work_location", ""),
        e.get("status", ""),
    ]


def _style_row(ws, row_idx: int, values: list):
    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    fill   = PatternFill("solid", fgColor=CLR_ROW_ODD if row_idx % 2 == 0 else CLR_ROW_EVEN)
    center_cols = {1, 2, 4, 5, 8, 13, 14, 15, 16, 20}
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = center if ci in center_cols else left
        cell.font      = Font(name="Calibri", size=9)


def _rebuild_summary(wb):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws2 = wb.create_sheet("Summary")
    ws  = wb["All Entries"]

    pivot = {}
    all_dates = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        emp = str(row[2] or "").strip()
        dt  = str(row[0] or "").strip()
        hrs = row[15]
        try:
            hrs = float(hrs or 0)
        except Exception:
            hrs = 0
        if emp and dt:
            pivot.setdefault(emp, {})
            pivot[emp][dt] = pivot[emp].get(dt, 0) + hrs
            all_dates.add(dt)

    sorted_dates = sorted(all_dates)
    sorted_emps  = sorted(pivot.keys())

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    hfill = PatternFill("solid", fgColor=CLR_HDR_BG)
    thin  = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws2.cell(row=1, column=1, value="Employee").font = hf
    ws2.cell(row=1, column=1).fill = hfill
    ws2.cell(row=1, column=1).border = border
    ws2.column_dimensions["A"].width = 24

    for ci, dt in enumerate(sorted_dates, 2):
        c = ws2.cell(row=1, column=ci, value=dt)
        c.font = hf; c.fill = hfill; c.border = border
        ws2.column_dimensions[get_column_letter(ci)].width = 12

    total_col = len(sorted_dates) + 2
    tc = ws2.cell(row=1, column=total_col, value="Total Hrs")
    tc.font = hf; tc.fill = hfill; tc.border = border
    ws2.column_dimensions[get_column_letter(total_col)].width = 11

    for ri, emp in enumerate(sorted_emps, 2):
        ws2.cell(row=ri, column=1, value=emp).border = border
        row_total = 0
        for ci, dt in enumerate(sorted_dates, 2):
            hrs = pivot[emp].get(dt, 0)
            row_total += hrs
            cell = ws2.cell(row=ri, column=ci, value=round(hrs, 2) if hrs else "")
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        tot = ws2.cell(row=ri, column=total_col, value=round(row_total, 2))
        tot.border = border
        tot.font = Font(name="Calibri", bold=True, size=9)
        tot.alignment = Alignment(horizontal="center")

    ws2.freeze_panes = "B2"


def append_entries_to_excel(entries: list[dict]) -> bool:
    """Append DB entry dicts to the master Excel file."""
    if not _acquire_lock():
        return False
    try:
        ensure_master_file()
        _backup()
        wb = openpyxl.load_workbook(str(MASTER_XLSX))
        ws = wb["All Entries"]
        next_row = ws.max_row + 1
        for e in entries:
            _style_row(ws, next_row, _entry_to_row(e))
            next_row += 1
        _rebuild_summary(wb)
        tmp = str(MASTER_XLSX) + ".tmp"
        wb.save(tmp)
        os.replace(tmp, str(MASTER_XLSX))
        return True
    except Exception as ex:
        print(f"Excel save error: {ex}")
        return False
    finally:
        _release_lock()


def export_to_csv(filepath: str, emp_id=None, filter_date=None) -> int:
    import csv
    rows = load_entries(emp_id=emp_id, filter_date=filter_date, limit=10000)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=MASTER_COLS)
        writer.writeheader()
        for r in rows:
            writer.writerow({
                "Date": r.get("entry_date"), "Day": r.get("day_name"),
                "Employee Name": r.get("full_name"),
                "Employee ID": r.get("employee_id"),
                "Designation": r.get("role"),
                "Reporting To": r.get("reporting_to"),
                "Firm Name": "Joshi Apte & Associates",
                "Company Code": r.get("company_code"),
                "Client / Company": r.get("client_name"),
                "Work Category": r.get("work_category"),
                "Operational Area": r.get("operational_area"),
                "Sub-Category": r.get("sub_category"),
                "Start Time": r.get("start_time"),
                "End Time": r.get("end_time"),
                "Break (mins)": r.get("break_mins"),
                "Total Hrs": r.get("total_hrs"),
                "Task Description": r.get("task_desc"),
                "Notes / Remarks": r.get("notes"),
                "Location": r.get("work_location"),
                "Status": r.get("status"),
            })
    return len(rows)

# ════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION
# ════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
#  DESIGN SYSTEM
# ═══════════════════════════════════════════════════════════════════════════════
SIDEBAR_BG  = "#06355E"      # Deep Ocean Blue  — sidebar background
SIDEBAR_FG  = "#E0E9F4"      # Very Light Blue  — sidebar text
SIDEBAR_SEL = "#024277"      # Rich Navy        — selected nav item
SIDEBAR_ICN = "#7AAFD4"      # Muted sky blue   — icons / secondary text
ACCENT_GOLD = "#03A3D0"      # Bright Cerulean  — accent stripe / highlights
PAGE_BG     = "#EBF2FA"      # Pale blue-white  — main content background
PANEL_BG    = "#FFFFFF"      # White            — cards / form panels
PANEL_ALT   = "#D2E1F1"      # Light Blue-Grey  — alternate panel bg
TBL_HDR_BG  = "#024277"      # Rich Navy        — table column headers
TBL_HDR_FG  = "#E0E9F4"      # Very Light Blue  — header text
ROW_ODD     = "#D2E1F1"      # Light Blue-Grey  — odd rows
ROW_EVEN    = "#FFFFFF"      # White            — even rows
ACCENT_BLUE = "#06355E"      # Deep Ocean Blue  — primary CTA buttons
BORDER_CLR  = "#A8C4DC"      # Muted blue border
FOCUS_CLR   = "#03A3D0"      # Bright Cerulean  — focus ring
TEXT_DARK   = "#06355E"      # Deep Ocean Blue  — primary body text
TEXT_MID    = "#2A4A6A"      # Mid navy         — secondary labels
TEXT_LIGHT  = "#6A90B0"      # Muted blue-grey  — placeholder / dim text
TEXT_DIM    = TEXT_MID
ENTRY_BG    = "#F5F9FD"      # Near-white blue  — input fields
ENTRY_FG    = "#06355E"      # Deep Ocean Blue  — input text
SUCCESS     = "#1A6B45"      # Deep forest green — success states
WARNING     = "#8A6A1A"      # Dark amber       — warnings
DANGER      = "#B52A2A"      # Deep red         — errors / danger

ADMIN_PIN_KEY   = "__admin__"
DEFAULT_ADMIN   = "2580"
INTERN_PERIOD_DAYS = 455   # ≈ 15 months
REPORTING_OPT   = ["Ashish Thatte", "Supriya Tambe", "Bhakti Pawar", "Arunabha Saha", "Santoshi Dalvi", "Self"]
WORK_LOCATION   = ["Client Place", "Office", "WFH"]
STATUSES        = ["In Progress", "Completed", "Pending", "On Hold"]
DAILY_TARGET    = 7.5   # net billable hrs = complete day (8h in office - 30min break)
APPROVAL_STATUSES = ["Draft", "Submitted", "Approved", "Rejected", "Revision Required"]

# Submission status colours
SUB_COLORS = {
    "On Time":          "#1A6B45",   # green
    "Late":             "#8A6A1A",   # amber
    "Partially Filled": "#03A3D0",   # blue
    "Not Submitted":    "#B52A2A",   # red
}
# Approval status colours
APR_COLORS = {
    "Draft":             "#6A90B0",
    "Submitted":         "#03A3D0",
    "Approved":          "#1A6B45",
    "Rejected":          "#B52A2A",
    "Revision Required": "#8A6A1A",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  STARTUP
# ═══════════════════════════════════════════════════════════════════════════════

def _check_intern_lifecycle():
    """
    Auto-deactivate interns past their internship end date.

    End date priority:
      1. internship_end_date (admin-set extension) — if set, this wins
      2. join_date + 15 months (default rule)

    If today > end_date → set status = 'Completed'
    If internship_end_date is set and in the future → intern stays Active
    regardless of how long ago they joined.
    """
    from datetime import timedelta
    today = date.today()
    default_months_days = INTERN_PERIOD_DAYS   # ≈ 15 months

    with get_conn() as conn:
        rows = conn.execute("""
            SELECT emp_id, full_name, join_date,
                   internship_end_date, extension_note
            FROM employees
            WHERE LOWER(role) = 'intern'
              AND status = 'Active'
        """).fetchall()

        deactivated = []
        for row in rows:
            jd_str  = row["internship_end_date"] or row["join_date"]
            if not jd_str:
                continue
            try:
                if row["internship_end_date"]:
                    # Admin set a custom end date — use it directly
                    end_date = datetime.strptime(
                        row["internship_end_date"], "%Y-%m-%d").date()
                else:
                    # Default: join_date + 15 months
                    jd = datetime.strptime(row["join_date"], "%Y-%m-%d").date()
                    end_date = jd + timedelta(days=default_months_days)

                if today > end_date:
                    conn.execute("""
                        UPDATE employees
                        SET status='Completed', updated_at=datetime('now')
                        WHERE emp_id=?
                    """, (row["emp_id"],))
                    deactivated.append(row["full_name"])
            except Exception:
                continue

    if deactivated:
        print(f"[Lifecycle] Auto-completed {len(deactivated)} interns: "
              f"{', '.join(deactivated)}")
    return deactivated


def get_intern_end_date(emp: dict) -> date | None:
    """Return the effective internship end date for display purposes."""
    from datetime import timedelta
    if emp.get("internship_end_date"):
        try:
            return datetime.strptime(emp["internship_end_date"], "%Y-%m-%d").date()
        except Exception:
            pass
    if emp.get("join_date"):
        try:
            jd = datetime.strptime(emp["join_date"], "%Y-%m-%d").date()
            return jd + timedelta(days=INTERN_PERIOD_DAYS)
        except Exception:
            pass
    return None


# ─── BIRTHDAY SEED DATA ──────────────────────────────────────────────────────
_BIRTHDAY_LIST = [
    ("Aarti Jagtap",             "01-01"),
    ("Aanuj Patil",              "01-01"),
    ("CMA Ayushi Das",           "01-05"),
    ("CMA Arunabha Saha",        "01-31"),
    ("Deshna Gosar",             "02-02"),
    ("Jitesh Khatri",            "02-25"),
    ("Chaitanya",                "03-06"),
    ("Prasanjeet Das",           "03-20"),
    ("Om Gupta",                 "04-05"),
    ("Vruddhi Date",             "05-04"),
    ("Ayushi Kumar",             "05-06"),
    ("Namrata Chalwadi",         "06-09"),
    ("Heet Chandaria",           "06-11"),
    ("Prem Thakur",              "07-14"),
    ("Vruchita Shinde",          "07-31"),
    ("CMA Bhakti Pawar",         "08-13"),
    ("Shravani Kulkarni",        "08-13"),
    ("Kalpesh Chaudhari",        "08-28"),
    ("Dishant Jaiswar",          "08-31"),
    ("Vijayalaxmi Padmashali",   "09-06"),
    ("Aarya Nimkar",             "09-11"),
    ("Aditya Bankar",            "09-29"),
    ("CMA (Dr) Ashish Thatte",   "10-05"),
    ("Vanshika Shetty",          "10-11"),
    ("Suyash Shrikhande",        "10-12"),
    ("Mrunmayee Dongre",         "10-17"),
    ("S.S.Deshpande Sir",        "10-25"),
    ("Spandana Uppin",           "10-25"),
    ("Anjum Sayed",              "11-03"),
    ("Laher Kariya",             "11-05"),
    ("Dinki Jain",               "11-05"),
    ("CMA Santoshi Dalvi",       "11-08"),
    ("Om Bhamre",                "11-17"),
    ("Soham Gole",               "11-18"),
    ("Neha Paimode",             "11-23"),
    ("Omkar Shinde",             "12-09"),
    ("CMA Supriya Tambe",        "12-20"),
    ("Srilakshmi Thevar",        "12-31"),
]

_MONTH_NAMES = ["", "January","February","March","April","May","June",
                "July","August","September","October","November","December"]


def seed_birthdays():
    """Write date_of_birth for employees in the hard-coded list.
    Does NOT overwrite manually-set DOBs."""
    with get_conn() as conn:
        for name, mm_dd in _BIRTHDAY_LIST:
            conn.execute(
                "UPDATE employees SET date_of_birth=? "
                "WHERE LOWER(TRIM(full_name))=LOWER(TRIM(?)) "
                "AND (date_of_birth IS NULL OR date_of_birth='')",
                (mm_dd, name)
            )


def get_todays_birthdays() -> list[dict]:
    """Return employees whose birthday is today (MM-DD match)."""
    today_mm_dd = date.today().strftime("%m-%d")
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM employees WHERE date_of_birth=?",
            (today_mm_dd,)
        ).fetchall()
    return [dict(r) for r in rows]


def get_upcoming_birthdays(days: int = 30) -> list[dict]:
    """Return employees with birthdays in the next `days` days."""
    today = date.today()
    result = []
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM employees "
            "WHERE date_of_birth IS NOT NULL AND date_of_birth != '' "
            "ORDER BY full_name"
        ).fetchall()
    for r in rows:
        emp = dict(r)
        try:
            mm, dd = map(int, emp["date_of_birth"].split("-"))
            bday = date(today.year, mm, dd)
            if bday < today:
                bday = date(today.year + 1, mm, dd)
            delta = (bday - today).days
            if 0 <= delta <= days:
                emp["days_away"] = delta
                emp["bday_date"] = bday
                result.append(emp)
        except Exception:
            continue
    result.sort(key=lambda e: e["days_away"])
    return result


def get_all_birthdays_by_month() -> dict:
    """Return {month_int: [emp_dict, ...]} for all employees with DOB set."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM employees "
            "WHERE date_of_birth IS NOT NULL AND date_of_birth != '' "
            "ORDER BY date_of_birth, full_name"
        ).fetchall()
    by_month: dict = {}
    for r in rows:
        emp = dict(r)
        try:
            mm = int(emp["date_of_birth"].split("-")[0])
            by_month.setdefault(mm, []).append(emp)
        except Exception:
            continue
    return by_month


def auto_monthly_backup() -> str | None:
    """
    Creates a dated backup of the DB once per calendar month.
    Stores a sentinel file so it only runs once per month.
    Returns backup path on success, None if already done this month or failed.
    """
    try:
        sentinel = DATA_DIR / "last_backup_month.txt"
        this_month = date.today().strftime("%Y-%m")
        if sentinel.exists():
            if sentinel.read_text().strip() == this_month:
                return None   # already backed up this month
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        # Keep only last 12 monthly backups
        existing = sorted(BACKUP_DIR.glob("jaa_monthly_backup_*.db"))
        for old in existing[:-11]:
            try: old.unlink()
            except Exception: pass
        ts   = date.today().strftime("%Y_%m")
        dest = BACKUP_DIR / f"jaa_monthly_backup_{ts}.db"
        shutil.copy2(str(DB_PATH), str(dest))
        sentinel.write_text(this_month)
        log.info(f"Monthly backup created: {dest}")
        # Vacuum DB to reclaim space
        with get_conn() as conn:
            conn.execute("VACUUM")
        log.info("DB vacuumed after backup.")
        return str(dest)
    except Exception as e:
        log.warning(f"Monthly backup failed: {e}")
        return None


def clear_trial_data(keep_employees: bool = True,
                     keep_companies: bool = True) -> dict:
    """
    Wipe trial/test data from the DB.
    By default keeps employees and companies (master data), clears everything else.
    Returns counts of deleted rows per table.
    """
    counts = {}
    with get_conn() as conn:
        tables_to_clear = [
            "timesheet_entries",
            "submission_log",
            "attendance_log",
            "email_log",
        ]
        if not keep_employees:
            tables_to_clear += ["admin_pins", "leave_records"]
            conn.execute("DELETE FROM employees")
            counts["employees"] = conn.execute("SELECT changes()").fetchone()[0]
        if not keep_companies:
            tables_to_clear.append("companies")
        for tbl in tables_to_clear:
            try:
                conn.execute(f"DELETE FROM {tbl}")
                counts[tbl] = conn.execute("SELECT changes()").fetchone()[0]
            except Exception as e:
                log.warning(f"clear_trial_data: could not clear {tbl}: {e}")
        conn.execute("VACUUM")
    log.info(f"Trial data cleared: {counts}")
    return counts


# ─── TASKS MIGRATION ──────────────────────────────────────────────────────────

def _migrate_task_columns():
    """Safe ALTER TABLE for tasks — adds task_type to existing DBs."""
    new_cols = [
        ("task_type", "TEXT DEFAULT 'Regular'"),
    ]
    with get_conn() as conn:
        existing = {row[1] for row in conn.execute(
            "PRAGMA table_info(tasks)"
        ).fetchall()}
        for col, coldef in new_cols:
            if col not in existing:
                try:
                    conn.execute(
                        f"ALTER TABLE tasks ADD COLUMN {col} {coldef}"
                    )
                    log.info(f"Migration: added column tasks.{col}")
                except Exception as e:
                    log.warning(f"Migration skip tasks.{col}: {e}")


def startup():
    """Initialise DB, seed defaults, run lifecycle checks."""
    init_db()
    _migrate_leave_columns()
    _migrate_task_columns()
    seed_work_mapping()
    ensure_master_file()
    # One-time migration from legacy text files
    if not get_all_employees():
        result = import_from_txt_files()
        print(f"Migration: {result}")
    # Ensure "Others" category exists (v3.0 addition)
    try:
        if "Others" not in get_work_categories():
            add_work_category("Others")
            add_operational_area("Others", "General")
            for sub in ["Internal Meeting", "Training", "Administrative Work",
                        "Research", "Other"]:
                add_sub_category("Others", "General", sub)
    except Exception as e:
        log.warning(f"Others category seed: {e}")
    # Auto-deactivate interns past 15 months
    _check_intern_lifecycle()
    # Seed Learning Hub categories
    seed_learning_categories()
    # Seed birthday data from BD_List
    seed_birthdays()
    # Monthly auto-backup (runs once per calendar month)
    backup_path = auto_monthly_backup()
    if backup_path:
        log.info(f"Startup: monthly backup saved to {backup_path}")
    log.info("JAA Timesheet v3.0 started.")


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _style_combo(c):
    s = ttk.Style()
    s.theme_use("clam")
    # Configure both the global fallback and a widget-specific named style
    # so that readonly combos placed via .place() don't bleed each other's
    # selection highlight (which causes the "In Progress" ghost on scroll).
    for style_name in ("TCombobox", "JAA.TCombobox"):
        s.configure(style_name,
                    fieldbackground=ENTRY_BG, background=PAGE_BG,
                    foreground=ENTRY_FG, selectbackground=ENTRY_BG,
                    selectforeground=ENTRY_FG, arrowcolor=TEXT_MID,
                    bordercolor=BORDER_CLR, lightcolor=BORDER_CLR, relief="flat",
                    padding=(2, 2, 2, 2))
        s.map(style_name,
              fieldbackground=[("readonly", ENTRY_BG)],
              foreground=[("readonly", ENTRY_FG)],
              selectbackground=[("readonly", ENTRY_BG)],
              selectforeground=[("readonly", ENTRY_FG)])
    try:
        c.configure(style="JAA.TCombobox")
    except Exception:
        pass
    # Prevent the combobox from auto-selecting the first item
    try:
        if not c.get():
            c.set("")
    except Exception:
        pass


def _set_combo_dropdown_width(combo):
    """
    Make the dropdown popup wide enough to show the full text of all values.
    Call this after setting values on a combobox.
    """
    try:
        vals = list(combo["values"])
        if not vals:
            return
        import tkinter.font as tkfont
        fnt = tkfont.Font(family="Calibri", size=9)
        max_px = max(fnt.measure(str(v)) for v in vals) + 32
        # Set the widget width so the popup inherits it
        needed = max(combo["width"] if combo["width"] else 0, max_px // 7)
        combo.configure(width=needed)
    except Exception:
        pass


def _widen_combo_dropdown(combo, min_width=180):
    """
    Post-creation: make the dropdown list wide enough for the longest value.
    Called after values are set or changed.
    """
    def _apply(*_):
        vals = combo["values"]
        if not vals:
            return
        # Measure longest text in pixels using a temporary label
        try:
            import tkinter.font as tkfont
            f = tkfont.Font(family="Calibri", size=9)
            longest = max((f.measure(str(v)) for v in vals), default=0)
            # Add padding for scrollbar + borders
            needed = max(min_width, longest + 30)
            combo.config(width=max(combo["width"], needed // 7))
        except Exception:
            pass
    try:
        combo.bind("<<ComboboxSelected>>", lambda e: None, add="+")
        _apply()
    except Exception:
        pass


def _card(parent, **kw) -> tk.Frame:
    return tk.Frame(parent, bg=PANEL_BG, highlightthickness=1,
                    highlightbackground=BORDER_CLR, **kw)


def _section_hdr(parent, text, icon=""):
    f = tk.Frame(parent, bg=PANEL_BG)
    f.pack(fill="x")
    tk.Frame(f, bg=ACCENT_GOLD, width=4).pack(side="left", fill="y")
    tk.Label(f, text=f"  {icon}  {text}" if icon else f"  {text}",
             font=("Georgia", 11, "bold"), bg=PANEL_BG, fg=TEXT_DARK,
             pady=10, anchor="w").pack(side="left", fill="x", expand=True)


def _hrs_badge(parent, total_hrs: float, target: float = 8.0,
               show_bar: bool = True) -> tk.Frame:
    """
    Returns a frame containing:
      - A colour-coded status badge  🔴 Under Xh  /  🟢 Xh Complete
      - An optional horizontal progress bar (filled proportion of target)
    Colours: green if >= target, amber if >= 50%, red if < 50%.
    """
    done    = total_hrs >= target
    pct     = min(total_hrs / target, 1.0) if target > 0 else 1.0
    bar_clr = "#1A6B45" if done else ("#B45309" if pct >= 0.5 else "#B52A2A")
    icon    = "🟢" if done else "🔴"
    label   = f"{icon}  {total_hrs:.1f} / {target:.0f} h  {'✔ Complete' if done else 'Incomplete'}"

    wrap = tk.Frame(parent, bg=PANEL_BG)

    # Badge row
    badge_row = tk.Frame(wrap, bg=PANEL_BG)
    badge_row.pack(anchor="w", padx=6, pady=(4, 2))
    tk.Label(badge_row, text=label,
             font=("Calibri", 9, "bold"),
             bg=PANEL_BG,
             fg=bar_clr).pack(side="left")

    # Progress bar
    if show_bar:
        bar_outer = tk.Frame(wrap, bg="#D0D8E4",
                             height=6, padx=0, pady=0)
        bar_outer.pack(fill="x", padx=6, pady=(0, 6))
        bar_outer.pack_propagate(False)
        # filled portion rendered after geometry is settled
        def _draw_fill(event, bo=bar_outer, p=pct, c=bar_clr):
            w = bo.winfo_width()
            fill_w = max(int(w * p), 2 if p > 0 else 0)
            for ch in bo.winfo_children():
                ch.destroy()
            if fill_w > 0:
                tk.Frame(bo, bg=c, width=fill_w,
                         height=6).place(x=0, y=0)
        bar_outer.bind("<Configure>", _draw_fill)

    return wrap


def _lbl_field(parent, label, row, padx=(0, 0)):
    tk.Label(parent, text=label, font=("Calibri", 9, "bold"),
             bg=PANEL_BG, fg=TEXT_MID).grid(row=row, column=0, sticky="w",
                                             pady=(8, 2), padx=padx)


def _entry_widget(parent, var, row, col=1, colspan=1, show=None):
    f = tk.Frame(parent, bg=PANEL_BG, highlightthickness=1,
                 highlightbackground=BORDER_CLR)
    f.grid(row=row, column=col, columnspan=colspan, sticky="ew",
           padx=(8, 0), pady=(4, 0))
    kw = dict(textvariable=var, font=("Calibri", 10), bg=ENTRY_BG,
              fg=ENTRY_FG, relief="flat", bd=0, insertbackground=ACCENT_BLUE)
    if show:
        kw["show"] = show
    tk.Entry(f, **kw).pack(fill="x", ipady=7)
    return f


# ═══════════════════════════════════════════════════════════════════════════════
#  AUTOCOMPLETE COMBOBOX WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

class AutocompleteEntry(tk.Frame):
    """Entry with dropdown suggestions filtered as user types."""

    def __init__(self, parent, values=None, on_select=None, bg=ENTRY_BG,
                 font=("Calibri", 9), width=14, **kw):
        super().__init__(parent, bg=bg, highlightthickness=1,
                         highlightbackground=BORDER_CLR, **kw)
        self._all_values = list(values or [])
        self._on_select  = on_select
        self._var        = tk.StringVar()
        self._var.trace_add("write", self._on_type)

        self._entry = tk.Entry(self, textvariable=self._var, font=font,
                               bg=bg, fg=ENTRY_FG, relief="flat", bd=0,
                               insertbackground=ACCENT_BLUE, width=width)
        self._entry.pack(fill="x", ipady=4)
        self._entry.bind("<Down>",    self._focus_list)
        self._entry.bind("<Escape>",  self._hide)
        self._entry.bind("<Return>",  self._on_return)
        self._entry.bind("<FocusOut>", self._on_focus_out)

        self._popup  = None
        self._listbox = None

    def set_values(self, values):
        self._all_values = list(values)

    def get(self):
        return self._var.get()

    def set(self, value):
        self._var.set(value)

    def _on_type(self, *_):
        q = self._var.get().strip().lower()
        if not q:
            self._hide()
            return
        matches = [v for v in self._all_values if q in v.lower()][:12]
        if matches:
            self._show(matches)
        else:
            self._hide()

    def _show(self, matches):
        if self._popup:
            self._popup.destroy()
        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.winfo_height()
        self._popup = tk.Toplevel(self)
        self._popup.wm_overrideredirect(True)
        self._popup.wm_geometry(f"+{x}+{y}")
        self._popup.configure(bg=BORDER_CLR)

        self._listbox = tk.Listbox(self._popup, font=("Calibri", 9),
                                   bg=PANEL_BG, fg=TEXT_DARK, selectmode="single",
                                   selectbackground=ACCENT_BLUE, selectforeground="white",
                                   relief="flat", bd=1, width=max(30, self._entry.winfo_width()//7),
                                   height=min(len(matches), 8), activestyle="none")
        self._listbox.pack(fill="both", padx=1, pady=1)
        for m in matches:
            self._listbox.insert("end", m)
        self._listbox.bind("<ButtonRelease-1>", self._pick)
        self._listbox.bind("<Return>",          self._pick)

    def _pick(self, *_):
        if not self._listbox:
            return
        sel = self._listbox.curselection()
        if sel:
            val = self._listbox.get(sel[0])
            self._var.set(val)
            if self._on_select:
                self._on_select(val)
        self._hide()

    def _focus_list(self, *_):
        if self._listbox:
            self._listbox.focus_set()
            self._listbox.selection_set(0)

    def _on_return(self, *_):
        if self._on_select:
            self._on_select(self._var.get())
        self._hide()

    def _on_focus_out(self, *_):
        self.after(150, self._hide)

    def _hide(self, *_):
        if self._popup:
            self._popup.destroy()
            self._popup = None
            self._listbox = None

    def config(self, **kw):
        if "state" in kw:
            self._entry.config(state=kw.pop("state"))
        super().config(**kw)


class _FakeAC:
    """
    Lightweight autocomplete controller for an existing tk.Entry widget.
    Used when the entry is placed via .place() inside the dual-pane table.
    """
    def __init__(self, entry: tk.Entry, values: list):
        self._entry     = entry
        self._all       = list(values)
        self._on_select = None
        self._popup     = None
        self._listbox   = None

    def get(self): return self._entry.get()
    def set(self, v):
        self._entry.delete(0, "end")
        self._entry.insert(0, v)

    def _on_type(self, *_):
        q = self._entry.get().strip().lower()
        if not q:
            self._hide(); return
        matches = [v for v in self._all if q in v.lower()][:12]
        if matches: self._show(matches)
        else:        self._hide()

    def _show(self, matches):
        if self._popup:
            self._popup.destroy()
        rx = self._entry.winfo_rootx()
        ry = self._entry.winfo_rooty() + self._entry.winfo_height()
        # Measure longest text to size the popup
        try:
            import tkinter.font as tkfont
            fnt = tkfont.Font(family="Calibri", size=9)
            max_px = max(fnt.measure(m) for m in matches) + 24
        except Exception:
            max_px = 220
        pw = max(max_px, 180)
        self._popup = tk.Toplevel(self._entry)
        self._popup.wm_overrideredirect(True)
        self._popup.wm_geometry(f"{pw}x{min(len(matches),8)*22}+{rx}+{ry}")
        self._popup.configure(bg=BORDER_CLR)
        self._listbox = tk.Listbox(
            self._popup, font=("Calibri", 9),
            bg=PANEL_BG, fg=TEXT_DARK, selectmode="single",
            selectbackground=ACCENT_BLUE, selectforeground="white",
            relief="flat", bd=0, activestyle="none")
        self._listbox.pack(fill="both", expand=True, padx=1, pady=1)
        for m in matches:
            self._listbox.insert("end", m)
        self._listbox.bind("<ButtonRelease-1>", self._pick)
        self._listbox.bind("<Return>",          self._pick)

    def _pick(self, *_):
        if not self._listbox: return
        sel = self._listbox.curselection()
        if sel:
            val = self._listbox.get(sel[0])
            self.set(val)
            if self._on_select: self._on_select(val)
        self._hide()

    def _focus_list(self, *_):
        if self._listbox:
            self._listbox.focus_set()
            self._listbox.selection_set(0)

    def _on_return(self, *_):
        if self._on_select: self._on_select(self.get())
        self._hide()

    def _on_focus_out(self, *_):
        self._entry.after(160, self._hide)

    def _hide(self, *_):
        if self._popup:
            self._popup.destroy()
            self._popup = None
            self._listbox = None


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

class TimesheetApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("JAA Timesheet — Joshi Apte & Associates  v2.0")
        self.configure(bg=PAGE_BG)
        self.resizable(True, True)
        self.withdraw()

        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h = 1280, 800
        self.geometry(f"{w}x{h}+{max(0,(sw-w)//2)}+{max(0,(sh-h)//2)}")
        self.minsize(1050, 680)

        # Runtime state
        self.current_emp    = None     # dict from DB
        self.current_role   = "user"
        self.session_id     = None     # attendance_log row id
        self._mapping       = {}       # full 3-level mapping cache

        self._load_mapping()
        self._show_login()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        # Start background reminder thread
        self._start_reminder_thread()

        self.deiconify()
        self.lift()
        self.update()

    def _load_mapping(self):
        """Cache full mapping + fast ID-based lookup dicts for dropdowns."""
        self._mapping = get_full_mapping()
        # Fast lookup: cat_name → [area_name, ...]
        self._cat_areas = {
            cat: list(areas.keys())
            for cat, areas in self._mapping.items()
        }
        # Fast lookup: (cat_name, area_name) → [sub_name, ...]
        self._area_subs = {}
        for cat, areas in self._mapping.items():
            for area, subs in areas.items():
                self._area_subs[(cat, area)] = subs

    def _on_close(self):
        if self.session_id:
            record_logout(self.session_id)
        self.destroy()

    # ── HELPERS ─────────────────────────────────────────────────────────────────

    def _clear(self):
        for w in self.winfo_children():
            w.destroy()

    def _sidebar(self, parent, title, subtitle="") -> tk.Frame:
        """Collapsible sidebar — icon-only (48px) or expanded (220px)."""
        SB_FULL = 220
        SB_MINI = 52

        sb = tk.Frame(parent, bg=SIDEBAR_BG, width=SB_FULL)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        self._sb_expanded  = True
        self._sb_frame     = sb
        self._sb_full      = SB_FULL
        self._sb_mini      = SB_MINI
        self._sb_nav_items = []   # (frame, icon_lbl, text_lbl)

        # Top accent bar
        tk.Frame(sb, bg=ACCENT_GOLD, height=4).pack(fill="x")

        # Toggle button row
        toggle_row = tk.Frame(sb, bg=SIDEBAR_BG)
        toggle_row.pack(fill="x")
        self._toggle_btn = tk.Label(
            toggle_row, text="☰", font=("Calibri", 13),
            bg=SIDEBAR_BG, fg=SIDEBAR_FG,
            cursor="hand2", padx=14, pady=8, anchor="w")
        self._toggle_btn.pack(side="left", fill="x", expand=True)
        self._toggle_btn.bind("<Button-1>", lambda e: self._toggle_sidebar())

        # Logo block (hidden in mini mode)
        self._logo_block = tk.Frame(sb, bg=SIDEBAR_BG, pady=10)
        self._logo_block.pack(fill="x")
        tk.Label(self._logo_block, text="JAA", font=("Georgia", 14, "bold"),
                 bg=SIDEBAR_BG, fg=ACCENT_GOLD).pack()
        self._logo_sub = tk.Label(
            self._logo_block, text=title,
            font=("Calibri", 7, "bold"), bg=SIDEBAR_BG, fg=SIDEBAR_ICN,
            wraplength=190)
        self._logo_sub.pack()
        if subtitle:
            tk.Label(self._logo_block, text=subtitle, font=("Calibri", 7),
                     bg=SIDEBAR_BG, fg=SIDEBAR_ICN, wraplength=190).pack(pady=(2, 0))

        tk.Frame(sb, bg=SIDEBAR_SEL, height=1).pack(fill="x", padx=10, pady=(0, 6))

        # ── Bottom anchor area (sign-out / back) pinned to the base ─────────────
        # Created first so pack(side="bottom") works correctly
        self._sb_bottom = tk.Frame(sb, bg=SIDEBAR_BG)
        self._sb_bottom.pack(side="bottom", fill="x")

        # ── Scrollable nav area ──────────────────────────────────────────────────
        nav_canvas = tk.Canvas(sb, bg=SIDEBAR_BG, highlightthickness=0,
                               yscrollincrement=1)
        nav_canvas.pack(side="top", fill="both", expand=True)

        nav_vsb = tk.Scrollbar(sb, orient="vertical", command=nav_canvas.yview)
        nav_canvas.configure(yscrollcommand=nav_vsb.set)

        nav_inner = tk.Frame(nav_canvas, bg=SIDEBAR_BG)
        nav_win_id = nav_canvas.create_window((0, 0), window=nav_inner, anchor="nw")

        def _sync_scroll(*_):
            nav_canvas.configure(scrollregion=nav_canvas.bbox("all"))
            canvas_h = nav_canvas.winfo_height()
            inner_h  = nav_inner.winfo_reqheight()
            if inner_h > max(canvas_h, 1):
                nav_vsb.pack(side="right", fill="y", before=nav_canvas)
            else:
                nav_vsb.pack_forget()

        def _on_canvas_resize(event):
            nav_canvas.itemconfig(nav_win_id, width=event.width)
            _sync_scroll()

        nav_inner.bind("<Configure>", _sync_scroll)
        nav_canvas.bind("<Configure>", _on_canvas_resize)

        def _on_mousewheel(event):
            nav_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        nav_canvas.bind("<MouseWheel>", _on_mousewheel)
        nav_inner.bind("<MouseWheel>", _on_mousewheel)

        self._sb_nav_inner = nav_inner
        self._sb_nav_canvas = nav_canvas

        return sb

    def _toggle_sidebar(self):
        self._sb_expanded = not self._sb_expanded
        new_w = self._sb_full if self._sb_expanded else self._sb_mini
        self._sb_frame.config(width=new_w)

        # Show/hide logo and text labels
        if self._sb_expanded:
            self._logo_block.pack(fill="x", after=self._toggle_btn.master)
        else:
            self._logo_block.pack_forget()

        # Update each nav item: show/hide text label
        for frame, icon_lbl, text_lbl in self._sb_nav_items:
            if self._sb_expanded:
                text_lbl.pack(side="left", fill="x", expand=True)
                icon_lbl.config(padx=8)
            else:
                text_lbl.pack_forget()
                icon_lbl.config(padx=0, anchor="center")

    def _nav_btn(self, sidebar, text, icon, tab_id, nav_btns, switch_fn, is_active=False):
        bg = SIDEBAR_SEL if is_active else SIDEBAR_BG
        fg = SIDEBAR_FG  if is_active else SIDEBAR_ICN

        # Pack into the scrollable inner frame when available
        nav_parent = getattr(self, "_sb_nav_inner", sidebar)
        f = tk.Frame(nav_parent, bg=bg, cursor="hand2")
        f.pack(fill="x")

        # Accent stripe
        stripe = tk.Frame(f, bg=ACCENT_GOLD if is_active else SIDEBAR_BG, width=4)
        stripe.pack(side="left", fill="y")

        # Icon label (always visible)
        icon_lbl = tk.Label(f, text=icon, font=("Calibri", 12),
                            bg=bg, fg=fg, pady=12, padx=8, anchor="center")
        icon_lbl.pack(side="left")

        # Text label (hidden in mini mode)
        text_lbl = tk.Label(f, text=text, font=("Calibri", 10),
                            bg=bg, fg=fg, pady=12, anchor="w")
        if self._sb_expanded:
            text_lbl.pack(side="left", fill="x", expand=True)

        nav_btns[tab_id] = (f, stripe, icon_lbl, text_lbl)
        self._sb_nav_items.append((f, icon_lbl, text_lbl))

        def _click(e, t=tab_id): switch_fn(t)
        for w in [f, stripe, icon_lbl, text_lbl]:
            w.bind("<Button-1>", _click)
        return f

    def _set_active_nav(self, nav_btns, active_id):
        for tid, (frame, stripe, icon_lbl, text_lbl) in nav_btns.items():
            is_sel = (tid == active_id)
            bg = SIDEBAR_SEL if is_sel else SIDEBAR_BG
            fg = SIDEBAR_FG  if is_sel else SIDEBAR_ICN
            frame.config(bg=bg)
            stripe.config(bg=ACCENT_GOLD if is_sel else SIDEBAR_BG)
            icon_lbl.config(bg=bg, fg=fg)
            text_lbl.config(bg=bg, fg=fg)

    def _topbar(self, parent, title, subtitle="") -> tk.Frame:
        tb = tk.Frame(parent, bg=PANEL_BG, height=60,
                      highlightthickness=1, highlightbackground=BORDER_CLR)
        tb.pack(fill="x")
        tb.pack_propagate(False)
        tk.Label(tb, text=title, font=("Georgia", 16, "bold"),
                 bg=PANEL_BG, fg=TEXT_DARK, padx=24).pack(side="left", pady=14)
        if subtitle:
            tk.Label(tb, text=subtitle, font=("Calibri", 9),
                     bg=PANEL_BG, fg=TEXT_MID, padx=4).pack(side="left")
        return tb

    # ── LOGIN ────────────────────────────────────────────────────────────────────

    def _show_login(self):
        self._clear()
        self.configure(bg=SIDEBAR_BG)

        # Left branding
        left = tk.Frame(self, bg=SIDEBAR_BG)
        left.place(relx=0, rely=0, relwidth=0.4, relheight=1)
        tk.Frame(left, bg=ACCENT_GOLD, height=4).pack(fill="x")

        brand = tk.Frame(left, bg=SIDEBAR_BG)
        brand.place(relx=0.5, rely=0.44, anchor="center")
        c = tk.Canvas(brand, width=80, height=80, bg=SIDEBAR_BG, highlightthickness=0)
        c.pack(pady=(0, 16))
        c.create_oval(2, 2, 78, 78, fill="#03A3D0", outline="")
        c.create_text(40, 40, text="JAA", fill=SIDEBAR_BG, font=("Georgia", 18, "bold"))
        tk.Label(brand, text="JOSHI APTE\n& ASSOCIATES",
                 font=("Georgia", 20, "bold"), bg=SIDEBAR_BG, fg=SIDEBAR_FG,
                 justify="center").pack(pady=(0, 6))
        tk.Frame(brand, bg=ACCENT_GOLD, height=2, width=120).pack(pady=(0, 10))
        tk.Label(brand, text="Daily Timesheet System  v2.0",
                 font=("Calibri", 11), bg=SIDEBAR_BG, fg=SIDEBAR_ICN).pack()
        tk.Label(left, text="v2.0  ·  Confidential",
                 font=("Calibri", 8), bg=SIDEBAR_BG, fg=SIDEBAR_ICN).place(
            relx=0.5, rely=0.96, anchor="center")

        # Right form
        right = tk.Frame(self, bg=PAGE_BG)
        right.place(relx=0.4, rely=0, relwidth=0.6, relheight=1)

        form = tk.Frame(right, bg=PAGE_BG)
        form.place(relx=0.5, rely=0.5, anchor="center", width=400)

        tk.Label(form, text="Welcome back",
                 font=("Calibri", 13), bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w")
        tk.Label(form, text="Sign in to your timesheet",
                 font=("Georgia", 22, "bold"), bg=PAGE_BG, fg=TEXT_DARK).pack(anchor="w", pady=(2, 28))

        # Employee card
        emp_card = _card(form)
        emp_card.pack(fill="x", pady=(0, 12))
        inner = tk.Frame(emp_card, bg=PANEL_BG, padx=24, pady=22)
        inner.pack(fill="x")

        tk.Label(inner, text="SELECT YOUR NAME", font=("Calibri", 8, "bold"),
                 bg=PANEL_BG, fg=TEXT_LIGHT).pack(fill="x", pady=(0, 6))

        all_names = [e["full_name"] for e in get_all_employees()]
        self._login_name_var = tk.StringVar()
        cf = tk.Frame(inner, bg=PANEL_BG, highlightthickness=1,
                      highlightbackground=BORDER_CLR)
        cf.pack(fill="x", pady=(0, 14))
        combo = ttk.Combobox(cf, textvariable=self._login_name_var,
                             values=sorted(all_names), state="readonly",
                             font=("Calibri", 11))
        combo.pack(fill="x", ipady=5)
        _style_combo(combo)

        tk.Button(inner, text="Sign In  →", font=("Calibri", 11, "bold"),
                  bg=ACCENT_BLUE, fg="white", relief="flat", cursor="hand2",
                  pady=12, activebackground="#042D50",
                  command=self._login_employee).pack(fill="x")

        # Divider
        div = tk.Frame(form, bg=PAGE_BG, pady=4)
        div.pack(fill="x")
        tk.Frame(div, bg=BORDER_CLR, height=1).pack(side="left", fill="x", expand=True, pady=10)
        tk.Label(div, text="  or  ", font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_LIGHT).pack(side="left")
        tk.Frame(div, bg=BORDER_CLR, height=1).pack(side="left", fill="x", expand=True, pady=10)

        tk.Button(form, text="🔐   Admin / Manager Access",
                  font=("Calibri", 10, "bold"), bg=PANEL_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", pady=12,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=self._login_admin).pack(fill="x")

    def _login_employee(self):
        name = self._login_name_var.get().strip()
        if not name:
            messagebox.showwarning("Select Name", "Please select your name.")
            return
        emp = get_employee_by_name(name)
        if not emp:
            messagebox.showerror("Error", "Employee not found.")
            return
        if emp["status"] != "Active":
            messagebox.showwarning("Access Denied",
                f"Your account is {emp['status']}.\nPlease contact Admin.")
            return
        self._ask_pin(emp)

    def _ask_pin(self, emp: dict):
        win = tk.Toplevel(self)
        win.title("PIN Verification")
        win.geometry("380x290")
        win.configure(bg=PAGE_BG)
        win.resizable(False, False)
        win.grab_set()
        win.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 380) // 2
        py = self.winfo_y() + (self.winfo_height() - 290) // 2
        win.geometry(f"380x290+{px}+{py}")

        tk.Frame(win, bg=ACCENT_GOLD, height=4).pack(fill="x")
        hdr = tk.Frame(win, bg=SIDEBAR_BG, pady=14); hdr.pack(fill="x")
        tk.Label(hdr, text="PIN Verification", font=("Georgia", 13, "bold"),
                 bg=SIDEBAR_BG, fg=SIDEBAR_FG, padx=24).pack(anchor="w")
        tk.Label(hdr, text=emp["full_name"], font=("Calibri", 9),
                 bg=SIDEBAR_BG, fg=SIDEBAR_ICN, padx=24).pack(anchor="w")

        frame = tk.Frame(win, bg=PAGE_BG, padx=32, pady=20)
        frame.pack(fill="both", expand=True)

        first_time = not has_pin(emp["emp_id"])
        msg = "First login — set a 4-digit PIN:" if first_time else "Enter your 4-digit PIN:"
        tk.Label(frame, text=msg, font=("Calibri", 10), bg=PAGE_BG,
                 fg=TEXT_MID, wraplength=300).pack(anchor="w", pady=(0, 10))

        pin_var = tk.StringVar()
        pf = tk.Frame(frame, bg=PAGE_BG, highlightthickness=1,
                      highlightbackground=BORDER_CLR); pf.pack(fill="x", pady=(0, 6))
        pin_e = tk.Entry(pf, textvariable=pin_var, font=("Calibri", 16),
                         bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                         show="●", justify="center")
        pin_e.pack(fill="x", ipady=10)
        pin_e.focus_set()

        err_lbl = tk.Label(frame, text="", font=("Calibri", 9),
                           bg=PAGE_BG, fg=DANGER)
        err_lbl.pack(anchor="w", pady=(0, 8))

        def attempt(event=None):
            pin = pin_var.get().strip()
            if not re.fullmatch(r"\d{4}", pin):
                err_lbl.config(text="⚠  PIN must be exactly 4 digits.")
                return
            if verify_pin(emp["emp_id"], pin):
                win.destroy()
                self.current_emp  = emp
                self.current_role = "user"
                # Record login session
                self.session_id = record_login(emp["emp_id"])
                self._show_main()
                self.after(400, self._check_and_show_birthday_popups)
            else:
                err_lbl.config(text="⚠  Incorrect PIN. Try again.")
                pin_var.set("")

        pin_e.bind("<Return>", attempt)
        tk.Button(frame, text="Verify & Sign In  →",
                  font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", pady=10,
                  command=attempt).pack(fill="x")

    def _login_admin(self):
        win = tk.Toplevel(self)
        win.title("Admin Login")
        win.geometry("380x280")
        win.configure(bg=PAGE_BG)
        win.resizable(False, False)
        win.grab_set()
        win.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 380) // 2
        py = self.winfo_y() + (self.winfo_height() - 280) // 2
        win.geometry(f"380x280+{px}+{py}")

        tk.Frame(win, bg=ACCENT_GOLD, height=4).pack(fill="x")
        hdr = tk.Frame(win, bg=SIDEBAR_BG, pady=14); hdr.pack(fill="x")
        tk.Label(hdr, text="Admin / Manager Access",
                 font=("Georgia", 13, "bold"), bg=SIDEBAR_BG, fg=SIDEBAR_FG,
                 padx=24).pack(anchor="w")
        tk.Label(hdr, text=f"Default PIN: {DEFAULT_ADMIN} (change on first login)",
                 font=("Calibri", 9), bg=SIDEBAR_BG, fg=SIDEBAR_ICN, padx=24).pack(anchor="w")

        frame = tk.Frame(win, bg=PAGE_BG, padx=32, pady=20)
        frame.pack(fill="both", expand=True)
        tk.Label(frame, text="Admin PIN:", font=("Calibri", 10, "bold"),
                 bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w", pady=(0, 6))
        pin_var = tk.StringVar()
        pf = tk.Frame(frame, bg=PAGE_BG, highlightthickness=1,
                      highlightbackground=BORDER_CLR); pf.pack(fill="x", pady=(0, 8))
        pe = tk.Entry(pf, textvariable=pin_var, font=("Calibri", 16),
                      bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0, show="●", justify="center")
        pe.pack(fill="x", ipady=10)
        pe.focus_set()
        err = tk.Label(frame, text="", font=("Calibri", 9), bg=PAGE_BG, fg=DANGER)
        err.pack(anchor="w", pady=(0, 8))

        def attempt(event=None):
            if verify_pin(ADMIN_PIN_KEY, pin_var.get()):
                if not has_pin(ADMIN_PIN_KEY):
                    change_pin(ADMIN_PIN_KEY, DEFAULT_ADMIN)
                win.destroy()
                self.current_role = "admin"
                self.current_emp  = None
                self._show_admin()
                self.after(400, self._check_and_show_birthday_popups)
            else:
                err.config(text="⚠  Incorrect admin PIN.")
                pin_var.set("")

        pe.bind("<Return>", attempt)
        tk.Button(frame, text="Enter Admin Panel  →",
                  font=("Calibri", 10, "bold"), bg=SIDEBAR_BG, fg="white",
                  relief="flat", cursor="hand2", pady=10,
                  command=attempt).pack(fill="x")

    # ── MAIN EMPLOYEE VIEW ───────────────────────────────────────────────────────

    def _show_main(self):
        self._clear()
        self.configure(bg=PAGE_BG)
        emp   = self.current_emp
        today = date.today()
        on_leave, leave_reason = is_on_leave(emp["emp_id"], today)

        nav_btns = {}

        # Sidebar
        sb = self._sidebar(self, "Joshi Apte & Associates", emp["full_name"])

        # User info pill — pack before the nav canvas so it stays at the top
        user_card = tk.Frame(sb, bg=SIDEBAR_SEL, padx=14, pady=12)
        user_card.pack(fill="x", padx=12, pady=(0, 14), before=self._sb_nav_canvas)
        code = emp.get("emp_id", "")
        tk.Label(user_card, text=emp["full_name"],
                 font=("Calibri", 10, "bold"), bg=SIDEBAR_SEL, fg=SIDEBAR_FG,
                 wraplength=175, justify="left").pack(anchor="w")
        tk.Label(user_card, text=f"{emp['role']}  ·  {code}",
                 font=("Calibri", 8), bg=SIDEBAR_SEL, fg=SIDEBAR_ICN).pack(anchor="w", pady=(2, 0))

        content_holder = {"area": None}
        main_body = tk.Frame(self, bg=PAGE_BG)
        main_body.pack(side="left", fill="both", expand=True)

        topbar = self._topbar(main_body, "Daily Timesheet",
                              today.strftime("  %A, %d %B %Y"))
        if on_leave:
            tk.Label(topbar, text="🏖 On Leave", font=("Calibri", 9, "bold"),
                     bg="#EBF5FB", fg=DANGER, padx=10, pady=4).pack(side="left", padx=8)

        content_area = tk.Frame(main_body, bg=PAGE_BG)
        content_area.pack(fill="both", expand=True)
        content_holder["area"] = content_area

        def switch(tab_id):
            self._set_active_nav(nav_btns, tab_id)
            for w in content_holder["area"].winfo_children():
                w.destroy()
            if tab_id == "timesheet":
                self._build_timesheet_tab(content_holder["area"], emp, today, on_leave, leave_reason)
            elif tab_id == "entries":
                self._build_entries_tab(content_holder["area"], emp["emp_id"])
            elif tab_id == "leaves":
                self._build_my_leaves_tab(content_holder["area"], emp)
            elif tab_id == "documents":
                self._build_documents_tab(content_holder["area"], emp)
            elif tab_id == "learning_hub":
                self._build_learning_hub(content_holder["area"], emp, is_admin=False)
            elif tab_id == "birthdays":
                self._build_employee_birthdays_tab(content_holder["area"])
            elif tab_id == "pin":
                self._build_change_pin_tab(content_holder["area"], emp["emp_id"])

        self._nav_btn(sb, "Timesheet",    "📝", "timesheet",    nav_btns, switch, True)
        self._nav_btn(sb, "My Entries",   "📋", "entries",      nav_btns, switch)
        self._nav_btn(sb, "My Leaves",    "🏖", "leaves",       nav_btns, switch)
        # Documents tab only for Interns
        if emp.get("role", "").lower() == "intern":
            self._nav_btn(sb, "Documents",    "📁", "documents",    nav_btns, switch)
        self._nav_btn(sb, "Learning Hub",  "🎓", "learning_hub", nav_btns, switch)
        self._nav_btn(sb, "Birthdays",     "🎂", "birthdays",    nav_btns, switch)
        self._nav_btn(sb, "Change PIN",    "🔑", "pin",          nav_btns, switch)

        tk.Frame(self._sb_bottom, bg=SIDEBAR_SEL, height=1).pack(fill="x", padx=10, pady=8)

        def sign_out():
            if self.session_id:
                record_logout(self.session_id)
                self.session_id = None
            self.current_emp = None
            self._show_login()

        so = tk.Frame(self._sb_bottom, bg=SIDEBAR_BG, cursor="hand2")
        so.pack(fill="x", pady=(0, 10))
        tk.Label(so, text="⟵", font=("Calibri", 12),
                 bg=SIDEBAR_BG, fg=SIDEBAR_ICN, pady=12, padx=8).pack(side="left")
        self._so_text = tk.Label(so, text="Sign Out", font=("Calibri", 10),
                 bg=SIDEBAR_BG, fg=SIDEBAR_ICN, pady=12, anchor="w")
        if self._sb_expanded:
            self._so_text.pack(side="left")
        self._sb_nav_items.append((so, so.winfo_children()[0], self._so_text))
        for w in [so] + list(so.winfo_children()):
            w.bind("<Button-1>", lambda e: sign_out())

        switch("timesheet")

    # ── TIMESHEET TAB ────────────────────────────────────────────────────────────

    def _build_timesheet_tab(self, parent, emp, today, on_leave, leave_reason):
        # ── Page-level vertical scroll ───────────────────────────────────────────
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True)

        v_canvas = tk.Canvas(outer, bg=PAGE_BG, highlightthickness=0)
        v_sb = ttk.Scrollbar(outer, orient="vertical", command=v_canvas.yview)
        v_canvas.configure(yscrollcommand=v_sb.set)
        v_sb.pack(side="right", fill="y")
        v_canvas.pack(side="left", fill="both", expand=True)

        page_frame = tk.Frame(v_canvas, bg=PAGE_BG)
        page_win = v_canvas.create_window((0, 0), window=page_frame, anchor="nw",
                                           width=900)   # initial width; Configure updates it

        def _on_page_cfg(e):
            v_canvas.configure(scrollregion=v_canvas.bbox("all"))

        def _on_vcv_cfg(e):
            v_canvas.itemconfig(page_win, width=e.width)

        page_frame.bind("<Configure>", _on_page_cfg)
        v_canvas.bind("<Configure>",   _on_vcv_cfg)

        def _vert_scroll(e):
            v_canvas.yview_scroll(-1 * (e.delta // 120), "units")
        # Bind to canvas and its content frame only, not globally
        v_canvas.bind("<MouseWheel>", _vert_scroll)
        page_frame.bind("<MouseWheel>", _vert_scroll)

        pad = tk.Frame(page_frame, bg=PAGE_BG, padx=18, pady=16)
        pad.pack(fill="both", expand=True)

        # ── Employee Info Card ──────────────────────────────────────────────────
        info_card = _card(pad); info_card.pack(fill="x", pady=(0, 12))
        _section_hdr(info_card, "Employee Information", "👤")
        ib = tk.Frame(info_card, bg=PANEL_BG, padx=20)
        ib.pack(fill="x", pady=(0, 14))

        def info_col(parent, lbl, val, color=TEXT_DARK):
            col = tk.Frame(parent, bg=PANEL_BG)
            col.pack(side="left", padx=(0, 32))
            tk.Label(col, text=lbl, font=("Calibri", 8, "bold"),
                     bg=PANEL_BG, fg=TEXT_LIGHT).pack(anchor="w")
            tk.Label(col, text=val, font=("Calibri", 11, "bold"),
                     bg=PANEL_BG, fg=color).pack(anchor="w")

        ra = tk.Frame(ib, bg=PANEL_BG); ra.pack(fill="x", pady=(0, 10))
        info_col(ra, "EMPLOYEE",    emp["full_name"],    ACCENT_BLUE)
        info_col(ra, "EMP ID",      emp.get("emp_id", ""))
        info_col(ra, "DESIGNATION", emp.get("role", ""))
        info_col(ra, "DATE",        today.strftime("%d-%m-%Y"))
        info_col(ra, "DAY",         today.strftime("%A"))

        rb = tk.Frame(ib, bg=PANEL_BG); rb.pack(fill="x", pady=(0, 10))
        tk.Label(rb, text="REPORTING TO", font=("Calibri", 8, "bold"),
                 bg=PANEL_BG, fg=TEXT_LIGHT).pack(anchor="w")
        self.reporting_var = tk.StringVar(value=emp.get("reporting_to") or REPORTING_OPT[0])
        rff = tk.Frame(rb, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        rff.pack(anchor="w", pady=(2, 0))
        rc = ttk.Combobox(rff, textvariable=self.reporting_var,
                          values=REPORTING_OPT, state="readonly",
                          font=("Calibri", 10), width=24)
        rc.pack(ipady=4); _style_combo(rc)

        # ── Manual Login / Logout Time ──────────────────────────────────────────
        rc_row = tk.Frame(ib, bg=PANEL_BG); rc_row.pack(fill="x", pady=(0, 12))

        def _time_field(parent, label, var, side="left"):
            col = tk.Frame(parent, bg=PANEL_BG)
            col.pack(side=side, padx=(0, 32))
            tk.Label(col, text=label, font=("Calibri", 8, "bold"),
                     bg=PANEL_BG, fg=TEXT_LIGHT).pack(anchor="w")
            ff = tk.Frame(col, bg=PANEL_BG, highlightthickness=1,
                          highlightbackground=BORDER_CLR)
            ff.pack(anchor="w", pady=(2, 0))
            e = tk.Entry(ff, textvariable=var, font=("Calibri", 11, "bold"),
                         bg=ENTRY_BG, fg=ACCENT_BLUE, relief="flat", bd=0,
                         width=8, justify="center",
                         insertbackground=FOCUS_CLR)
            e.pack(ipady=6, padx=6)
            return e

        self.day_login_var  = tk.StringVar()
        self.day_logout_var = tk.StringVar()
        self.day_workhrs_lbl = None

        login_e  = _time_field(rc_row, "LOGIN TIME  (HH:MM)",  self.day_login_var)
        logout_e = _time_field(rc_row, "LOGOUT TIME  (HH:MM)", self.day_logout_var)

        hrs_col = tk.Frame(rc_row, bg=PANEL_BG); hrs_col.pack(side="left", padx=(0, 0))
        tk.Label(hrs_col, text="TOTAL WORK HRS", font=("Calibri", 8, "bold"),
                 bg=PANEL_BG, fg=TEXT_LIGHT).pack(anchor="w")
        self.day_workhrs_lbl = tk.Label(hrs_col, text="—",
                                        font=("Calibri", 13, "bold"),
                                        bg=PANEL_BG, fg=ACCENT_GOLD)
        self.day_workhrs_lbl.pack(anchor="w", pady=(4, 0))

        def _calc_day_hrs(*_):
            try:
                li = datetime.strptime(self.day_login_var.get().strip(),  "%H:%M")
                lo = datetime.strptime(self.day_logout_var.get().strip(), "%H:%M")
                if lo < li:
                    self.day_workhrs_lbl.config(text="⚠ Invalid", fg=DANGER)
                    return
                diff = (lo - li).seconds / 3600
                self.day_workhrs_lbl.config(
                    text=f"{diff:.1f} hrs",
                    fg=SUCCESS if diff >= 8 else WARNING)  # 8h raw = 7.5h net (incl. break)
            except Exception:
                self.day_workhrs_lbl.config(text="—", fg=ACCENT_GOLD)

        def _validate_logout(*_):
            try:
                li = datetime.strptime(self.day_login_var.get().strip(),  "%H:%M")
                lo = datetime.strptime(self.day_logout_var.get().strip(), "%H:%M")
                if lo < li:
                    logout_e.config(bg="#FDE8E8")
                    self.day_workhrs_lbl.config(text="⚠ Logout < Login", fg=DANGER)
                    return
                logout_e.config(bg=ENTRY_BG)
            except Exception:
                logout_e.config(bg=ENTRY_BG)
            _calc_day_hrs()

        login_e.bind("<FocusOut>",  _calc_day_hrs)
        logout_e.bind("<FocusOut>", _validate_logout)
        login_e.bind("<Return>",    lambda e: logout_e.focus_set())
        logout_e.bind("<Return>",   _validate_logout)

        tk.Label(ib, text="💡  Login/Logout times are for attendance. Enter manually. Logout cannot be before Login.",
                 font=("Calibri", 8), bg=PANEL_BG, fg=TEXT_LIGHT,
                 anchor="w").pack(fill="x", pady=(0, 4))

        # Leave banner
        if on_leave:
            lb = tk.Frame(pad, bg="#EBF5FB", highlightthickness=1,
                          highlightbackground="#A8C4DC")
            lb.pack(fill="x", pady=(0, 12))
            tk.Label(lb, text=f"🏖  Approved leave — timesheet locked.  {leave_reason}",
                     font=("Calibri", 10, "bold"), bg="#EBF5FB", fg=DANGER,
                     padx=20, pady=12, anchor="w").pack(fill="x")


        # ══════════════════════════════════════════════════════════════════════
        # TASK ENTRY TABLE  — v6 clean rebuild
        # Architecture: plain tk.Frame rows packed in a scrollable canvas.
        # Each row owns all its widgets + StringVars independently.
        # Canvas is ONLY for horizontal scrolling; rows pack naturally tall.
        # ══════════════════════════════════════════════════════════════════════

        task_card = _card(pad)
        task_card.pack(fill="x", pady=(0, 12))
        _section_hdr(task_card, "Task Entries  (up to 10 rows)", "📝")

        # ── Load master data once ─────────────────────────────────────────────
        _co_all      = get_all_companies()
        _co_names    = [c["full_name"] for c in _co_all]
        _co_by_name  = {c["full_name"]: c for c in _co_all}
        _cats        = get_work_categories()

        # ── Column definitions ─────────────────────────────────────────────────
        # (header_label, pixel_width, col_id)
        _COLS = [
            ("#",               38,  "num"),
            ("Co. Code",        90,  "cocode"),
            ("Client/Company", 195,  "company"),
            ("Work Category",  165,  "cat"),
            ("Op. Area",       165,  "area"),
            ("Sub-Category",   145,  "sub"),
            ("Location",       108,  "loc"),
            ("Status",         118,  "stat"),
            ("Task Desc.",     200,  "task"),
            ("Start",           58,  "tstart"),
            ("End",             58,  "tend"),
            ("Hrs",             46,  "hrs"),
            ("Notes",          140,  "notes"),
            ("",                36,  "del"),
        ]
        _TOTAL_W  = sum(w for _, w, _ in _COLS)   # natural width of one row
        _ROW_H    = 34
        _HDR_H    = 34
        _N_ROWS   = 10

        # ── Outer container: horizontal scrollbar + canvas ─────────────────────
        tbl_wrap = tk.Frame(task_card, bg=PAGE_BG)
        tbl_wrap.pack(fill="x", padx=10, pady=(0, 6))

        _h_bar = ttk.Scrollbar(tbl_wrap, orient="horizontal")
        _h_bar.pack(side="bottom", fill="x")

        _h_can = tk.Canvas(tbl_wrap, bg=PAGE_BG, highlightthickness=0,
                           height=_HDR_H + _N_ROWS * _ROW_H + 4)
        _h_can.pack(side="top", fill="x")
        _h_can.configure(xscrollcommand=_h_bar.set)
        _h_bar.configure(command=_h_can.xview)

        # Inner frame — rows live here.
        # Width is set to _TOTAL_W initially so fill="x" children have a real width.
        # The Configure binding updates it when the canvas resizes.
        _inner = tk.Frame(_h_can, bg=PAGE_BG)
        _inner_id = _h_can.create_window((0, 0), window=_inner, anchor="nw",
                                          width=_TOTAL_W)

        def _sr(*_):
            _h_can.configure(scrollregion=_h_can.bbox("all"))

        def _update_inner_width(e):
            # Only set width — NEVER height (that clips rows)
            canvas_w = e.width
            # Use max(canvas_w, _TOTAL_W) so horizontal scroll works
            _h_can.itemconfig(_inner_id, width=max(canvas_w, _TOTAL_W))
            _sr()

        _inner.bind("<Configure>", _sr)
        _h_can.bind("<Configure>", _update_inner_width)

        # Shift-scroll → horizontal
        def _hscroll(e):
            _h_can.xview_scroll(-1 * (e.delta // 120), "units")
        _h_can.bind("<Shift-MouseWheel>", _hscroll)
        _inner.bind("<Shift-MouseWheel>", _hscroll)

        # ── Header row ─────────────────────────────────────────────────────────
        _hdr_frame = tk.Frame(_inner, bg=TBL_HDR_BG, height=_HDR_H)
        _hdr_frame.pack(side="top", fill="x")
        _hdr_frame.pack_propagate(False)

        for _ht, _hw, _ in _COLS:
            _hcell = tk.Frame(_hdr_frame, bg=TBL_HDR_BG,
                              width=_hw, height=_HDR_H)
            _hcell.pack(side="left")
            _hcell.pack_propagate(False)
            tk.Label(_hcell, text=_ht,
                     font=("Calibri", 8, "bold"),
                     bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                     anchor="center", wraplength=_hw - 4
                     ).pack(fill="both", expand=True)

        # ── Row container ──────────────────────────────────────────────────────
        _row_area = tk.Frame(_inner, bg=PAGE_BG)
        _row_area.pack(side="top", fill="x")

        self.task_rows = []    # list of _TRow instances

        # ── Shared per-row ttk.Style ───────────────────────────────────────────
        # We create ONE style per row index and reuse it.
        # configure() is idempotent so safe to call on every tab rebuild.
        for _ri in range(_N_ROWS):
            _sn = f"TR{_ri}.TCombobox"
            _st = ttk.Style()
            _st.configure(_sn,
                fieldbackground=ENTRY_BG, background=ENTRY_BG,
                foreground=ENTRY_FG,
                selectbackground=ENTRY_BG, selectforeground=ENTRY_FG,
                arrowcolor=TEXT_MID, bordercolor=BORDER_CLR,
                lightcolor=BORDER_CLR, darkcolor=BORDER_CLR,
                relief="flat", padding=(2, 1))
            _st.map(_sn,
                fieldbackground=[("readonly", ENTRY_BG),
                                  ("disabled", PANEL_ALT)],
                foreground=[("readonly", ENTRY_FG),
                             ("disabled", TEXT_LIGHT)],
                selectbackground=[("readonly", ENTRY_BG)],
                selectforeground=[("readonly", ENTRY_FG)])

        # ── _TRow: fully independent row object ────────────────────────────────
        class _TRow:
            """One complete, self-contained timesheet entry row."""

            def __init__(self_r, parent, idx, on_del, on_change):
                self_r.idx       = idx
                self_r._on_change = on_change
                self_r._popup      = None
                self_r._poplb      = None
                self_r._ac_selecting = False   # guard: suppress trace during pick
                self_r._sname    = f"TR{idx}.TCombobox"
                _bg = ROW_ODD if idx % 2 == 0 else ROW_EVEN
                self_r._bg = _bg

                # ── Row frame ──────────────────────────────────────────────────
                self_r.frame = tk.Frame(parent, bg=_bg,
                                        highlightthickness=1,
                                        highlightbackground=BORDER_CLR,
                                        height=_ROW_H)
                self_r.frame.pack(side="top", fill="x")
                self_r.frame.pack_propagate(False)

                # ── StringVars ────────────────────────────────────────────────
                self_r.v_code   = tk.StringVar()
                self_r.v_client = tk.StringVar()
                self_r.v_cat    = tk.StringVar()
                self_r.v_area   = tk.StringVar()
                self_r.v_sub    = tk.StringVar()
                self_r.v_loc    = tk.StringVar()
                self_r.v_stat   = tk.StringVar(value="")  # empty until user picks
                self_r.v_task   = tk.StringVar()
                self_r.v_start  = tk.StringVar()
                self_r.v_end    = tk.StringVar()
                self_r.v_notes  = tk.StringVar()
                self_r._area_map = {}   # area_name → area_id

                # ── Build every cell ─────────────────────────────────────────
                for _label, _w, _cid in _COLS:
                    self_r._cell(_w, _cid, on_del)

                # ── Wire events ──────────────────────────────────────────────
                # Cascading dropdowns
                self_r.cb_cat.bind("<<ComboboxSelected>>",
                    lambda e, r=self_r: r._cat_sel())
                self_r.cb_area.bind("<<ComboboxSelected>>",
                    lambda e, r=self_r: r._area_sel())

                # Hours calc
                self_r.e_start.bind("<FocusOut>",
                    lambda e, r=self_r: r._calc_hrs())
                self_r.e_end.bind("<FocusOut>",
                    lambda e, r=self_r: r._calc_hrs())
                self_r.e_start.bind("<Return>",
                    lambda e, r=self_r: r.e_end.focus_set())

                # Alt+Down to open combos; MouseWheel blocked to prevent accidental changes
                for _cb in (self_r.cb_cat, self_r.cb_area,
                             self_r.cb_sub, self_r.cb_loc, self_r.cb_stat):
                    _cb.bind("<Alt-Down>",
                        lambda e, c=_cb: c.event_generate("<ButtonPress-1>"))
                    # Prevent mousewheel scroll from changing readonly combo values
                    _cb.bind("<MouseWheel>", lambda e: "break")
                    _cb.bind("<Button-4>",   lambda e: "break")
                    _cb.bind("<Button-5>",   lambda e: "break")

                # Client autocomplete
                self_r.v_client.trace_add("write",
                    lambda *_, r=self_r: r._ac_type())
                self_r.e_client.bind("<FocusOut>",
                    lambda e, r=self_r: r._ac_focusout())
                self_r.e_client.bind("<Down>",
                    lambda e, r=self_r: r._ac_down())
                self_r.e_client.bind("<Return>",
                    lambda e, r=self_r: r._ac_enter())
                self_r.e_client.bind("<Escape>",
                    lambda e, r=self_r: r._ac_hide())

            # ── Cell factory ──────────────────────────────────────────────────
            def _cell(self_r, w, cid, on_del):
                cell = tk.Frame(self_r.frame, bg=self_r._bg,
                                width=w, height=_ROW_H)
                cell.pack(side="left")
                cell.pack_propagate(False)

                def _e(var=None, mono=False, ro=False, **kw):
                    st = "readonly" if ro else "normal"
                    fnt = ("Courier", 8, "bold") if mono else ("Calibri", 9)
                    ent = tk.Entry(cell, textvariable=var, font=fnt,
                                   bg=PANEL_ALT if ro else ENTRY_BG,
                                   fg=ACCENT_BLUE if ro else ENTRY_FG,
                                   disabledbackground=PANEL_ALT,
                                   relief="flat", bd=0, state=st,
                                   highlightthickness=1,
                                   highlightbackground=BORDER_CLR,
                                   highlightcolor=FOCUS_CLR,
                                   insertbackground=FOCUS_CLR,
                                   readonlybackground=PANEL_ALT, **kw)
                    ent.pack(fill="both", expand=True, padx=2, pady=2)
                    return ent

                def _cb(values, var):
                    combo = ttk.Combobox(cell, textvariable=var,
                                         values=values, state="readonly",
                                         font=("Calibri", 9),
                                         style=self_r._sname)
                    combo.pack(fill="both", expand=True, padx=1, pady=2)
                    combo.configure(
                        postcommand=lambda c=combo, v=var: c.set(v.get()))
                    return combo

                if cid == "num":
                    tk.Label(cell, text=str(self_r.idx + 1),
                             font=("Calibri", 9, "bold"),
                             bg=self_r._bg, fg=TEXT_LIGHT,
                             anchor="center"
                             ).pack(fill="both", expand=True)

                elif cid == "cocode":
                    self_r.e_code = _e(var=self_r.v_code, mono=True, ro=True)

                elif cid == "company":
                    self_r.e_client = _e(var=self_r.v_client)

                elif cid == "cat":
                    self_r.cb_cat = _cb(_cats, self_r.v_cat)

                elif cid == "area":
                    self_r.cb_area = _cb([], self_r.v_area)

                elif cid == "sub":
                    self_r.cb_sub = _cb([], self_r.v_sub)

                elif cid == "loc":
                    self_r.cb_loc = _cb(WORK_LOCATION, self_r.v_loc)

                elif cid == "stat":
                    self_r.cb_stat = _cb(STATUSES, self_r.v_stat)

                elif cid == "task":
                    self_r.e_task = _e(var=self_r.v_task)

                elif cid == "tstart":
                    self_r.e_start = _e(var=self_r.v_start, justify="center")

                elif cid == "tend":
                    self_r.e_end = _e(var=self_r.v_end, justify="center")

                elif cid == "hrs":
                    self_r.lbl_hrs = tk.Label(cell, text="—",
                                               font=("Calibri", 9, "bold"),
                                               bg=self_r._bg, fg=TEXT_LIGHT,
                                               anchor="center")
                    self_r.lbl_hrs.pack(fill="both", expand=True)

                elif cid == "notes":
                    self_r.e_notes = _e(var=self_r.v_notes)

                elif cid == "del":
                    tk.Button(cell, text="✕",
                              font=("Calibri", 10, "bold"),
                              bg=self_r._bg, fg=DANGER,
                              relief="flat", bd=0,
                              cursor="hand2",
                              activebackground="#FFE8E8",
                              command=lambda r=self_r: on_del(r)
                              ).pack(fill="both", expand=True)

            # ── Autocomplete popup ─────────────────────────────────────────────
            def _ac_type(self_r):
                # Skip if triggered by _ac_pick's programmatic set
                if getattr(self_r, "_ac_selecting", False):
                    return
                q = self_r.v_client.get().strip().lower()
                if not q:
                    self_r.v_code.set(""); self_r._ac_hide(); return
                hits = [n for n in _co_names if q in n.lower()][:12]
                # Auto-fill co_code for exact match (case-insensitive)
                typed = self_r.v_client.get()
                exact = _co_by_name.get(typed)
                if exact:
                    self_r._ac_selecting = True
                    self_r.v_code.set(exact["unique_code"])
                    self_r._ac_selecting = False
                    self_r._ac_hide()
                    return
                if hits:
                    self_r._ac_show(hits)
                else:
                    self_r._ac_hide()
                    if typed not in _co_by_name:
                        self_r.v_code.set("")

            def _ac_show(self_r, hits):
                self_r._ac_hide()
                try:
                    rx = self_r.e_client.winfo_rootx()
                    ry = self_r.e_client.winfo_rooty() + \
                         self_r.e_client.winfo_height()
                except Exception:
                    return
                import tkinter.font as tkf
                try:
                    f = tkf.Font(family="Calibri", size=9)
                    pw = max(f.measure(h) for h in hits) + 28
                except Exception:
                    pw = 220
                pw = max(pw, 200)
                ph = min(len(hits), 8) * 22
                p = tk.Toplevel(self_r.e_client)
                p.wm_overrideredirect(True)
                p.wm_geometry(f"{pw}x{ph}+{rx}+{ry}")
                p.configure(bg=BORDER_CLR)
                p.attributes("-topmost", True)
                lb = tk.Listbox(p, font=("Calibri", 9),
                                bg=PANEL_BG, fg=TEXT_DARK,
                                selectbackground=ACCENT_BLUE,
                                selectforeground="white",
                                relief="flat", bd=0,
                                activestyle="none")
                lb.pack(fill="both", expand=True, padx=1, pady=1)
                for h in hits:
                    lb.insert("end", h)

                # On mouse hover, highlight the item under cursor
                def _lb_motion(e, _lb=lb):
                    idx = _lb.nearest(e.y)
                    _lb.selection_clear(0, "end")
                    _lb.selection_set(idx)
                lb.bind("<Motion>", _lb_motion)

                # Button-1 press: capture item immediately (before focus moves)
                def _lb_click(e, r=self_r, _lb=lb):
                    idx = _lb.nearest(e.y)
                    _lb.selection_clear(0, "end")
                    _lb.selection_set(idx)
                    r._ac_pick()
                    return "break"
                lb.bind("<Button-1>", _lb_click)
                lb.bind("<ButtonRelease-1>", lambda e: "break")  # suppress duplicate
                lb.bind("<Return>",
                    lambda e, r=self_r: r._ac_pick())
                lb.bind("<Escape>",
                    lambda e, r=self_r: r._ac_hide())

                # Only close on FocusOut if not mid-pick
                def _popup_focusout(e, r=self_r):
                    if not getattr(r, "_ac_selecting", False):
                        r.e_client.after(300, r._ac_hide)
                p.bind("<FocusOut>", _popup_focusout)

                self_r._popup = p
                self_r._poplb = lb

            def _ac_pick(self_r):
                if not self_r._poplb:
                    return
                sel = self_r._poplb.curselection()
                # Fallback: use the active (hovered) index if nothing is selected
                if not sel:
                    try:
                        active = self_r._poplb.index("active")
                        if active >= 0:
                            sel = (active,)
                    except Exception:
                        pass
                if sel:
                    name = self_r._poplb.get(sel[0])
                    # Guard: suppress trace during programmatic set
                    self_r._ac_selecting = True
                    self_r.v_client.set(name)
                    rec = _co_by_name.get(name)
                    if rec:
                        self_r.v_code.set(rec["unique_code"])
                    # Hide popup immediately
                    self_r._ac_hide()
                    # Move focus to next field
                    try:
                        self_r.cb_cat.focus_set()
                    except Exception:
                        pass
                    # Keep guard True briefly so focusout doesn't clear co_code
                    def _clear_guard(r=self_r):
                        r._ac_selecting = False
                    self_r.e_client.after(350, _clear_guard)
                else:
                    self_r._ac_hide()

            def _ac_down(self_r, *_):
                if self_r._poplb:
                    self_r._poplb.focus_set()
                    self_r._poplb.selection_clear(0, "end")
                    self_r._poplb.selection_set(0)
                    self_r._poplb.activate(0)
                    # Arrow keys navigate the list
                    def _lb_key(e, r=self_r):
                        lb = r._poplb
                        if lb is None:
                            return
                        sel = lb.curselection()
                        cur = sel[0] if sel else 0
                        if e.keysym == "Down":
                            nxt = min(cur + 1, lb.size() - 1)
                        elif e.keysym == "Up":
                            nxt = max(cur - 1, 0)
                        else:
                            return
                        lb.selection_clear(0, "end")
                        lb.selection_set(nxt)
                        lb.activate(nxt)
                        lb.see(nxt)
                        return "break"
                    self_r._poplb.bind("<Down>", _lb_key)
                    self_r._poplb.bind("<Up>",   _lb_key)
                    self_r._poplb.bind("<Return>",
                        lambda e, r=self_r: r._ac_pick())
                    self_r._poplb.bind("<Escape>",
                        lambda e, r=self_r: (r._ac_hide(),
                                             r.e_client.focus_set()))

            def _ac_enter(self_r, *_):
                if self_r._poplb:
                    self_r._ac_pick()
                else:
                    self_r._ac_hide()
                    self_r.cb_cat.focus_set()

            def _ac_focusout(self_r, *_):
                # Use a longer delay so _ac_pick (triggered by Button-1) runs first
                def _do_hide(r=self_r):
                    if not getattr(r, "_ac_selecting", False):
                        r._ac_hide()
                self_r.e_client.after(400, _do_hide)
                # Only clear co_code if name truly not recognised and not mid-pick
                def _maybe_clear_code(r=self_r):
                    if getattr(r, "_ac_selecting", False):
                        return
                    n = r.v_client.get().strip()
                    if n and n not in _co_by_name:
                        r.v_code.set("")
                self_r.e_client.after(420, _maybe_clear_code)

            def _ac_hide(self_r, *_):
                if self_r._popup:
                    try:
                        self_r._popup.destroy()
                    except Exception:
                        pass
                    self_r._popup = None
                    self_r._poplb = None

            # ── Cascading dropdowns ────────────────────────────────────────────
            def _cat_sel(self_r):
                cat = self_r.v_cat.get()
                if not cat:
                    self_r._reset_area_sub(); return
                areas = get_operational_areas_full(cat)
                self_r._area_map = {a["name"]: a["id"] for a in areas}
                self_r.cb_area.configure(values=[a["name"] for a in areas])
                self_r.v_area.set("")
                self_r.cb_sub.configure(values=[])
                self_r.v_sub.set("")

            def _area_sel(self_r):
                area = self_r.v_area.get()
                if not area:
                    self_r.cb_sub.configure(values=[])
                    self_r.v_sub.set(""); return
                aid = self_r._area_map.get(area)
                subs = ([s["name"] for s in get_subs_by_area_id(aid)]
                        if aid else
                        get_sub_categories(self_r.v_cat.get(), area))
                self_r.cb_sub.configure(values=subs)
                self_r.v_sub.set("")

            def _reset_area_sub(self_r):
                self_r._area_map = {}
                self_r.cb_area.configure(values=[]); self_r.v_area.set("")
                self_r.cb_sub.configure(values=[]);  self_r.v_sub.set("")

            # ── Hours calculation ──────────────────────────────────────────────
            def _calc_hrs(self_r):
                try:
                    li = datetime.strptime(
                        self_r.v_start.get().strip(), "%H:%M")
                    lo = datetime.strptime(
                        self_r.v_end.get().strip(),   "%H:%M")
                    diff = max(0, (lo - li).seconds / 3600)
                    self_r.lbl_hrs.config(
                        text=f"{diff:.1f}h",
                        fg=SUCCESS if diff >= 1 else WARNING)
                except Exception:
                    self_r.lbl_hrs.config(text="—", fg=TEXT_LIGHT)
                self_r._on_change()

            # ── Data read-out ──────────────────────────────────────────────────
            def get_data(self_r):
                hrs = 0.0
                try:
                    t = self_r.lbl_hrs.cget("text").replace("h", "").strip()
                    hrs = float(t) if t not in ("—", "", "?") else 0.0
                except Exception:
                    pass
                return {
                    "co_code":          self_r.v_code.get().strip(),
                    "client":           self_r.v_client.get().strip(),
                    "work_category":    self_r.v_cat.get().strip(),
                    "operational_area": self_r.v_area.get().strip(),
                    "sub_category":     self_r.v_sub.get().strip(),
                    "work_location":    self_r.v_loc.get().strip(),
                    "status":           self_r.v_stat.get().strip()
                                        or "In Progress",
                    "task_desc":        self_r.v_task.get().strip(),
                    "start_time":       self_r.v_start.get().strip(),
                    "end_time":         self_r.v_end.get().strip(),
                    "total_hrs":        hrs,
                    "notes":            self_r.v_notes.get().strip(),
                }

            # ── Utilities ─────────────────────────────────────────────────────
            def clear(self_r):
                for v in (self_r.v_code, self_r.v_client,
                          self_r.v_cat, self_r.v_area, self_r.v_sub,
                          self_r.v_loc, self_r.v_task,
                          self_r.v_start, self_r.v_end, self_r.v_notes):
                    v.set("")
                self_r.v_stat.set("")
                self_r.cb_area.configure(values=[])
                self_r.cb_sub.configure(values=[])
                self_r._area_map = {}
                self_r.lbl_hrs.config(text="—", fg=TEXT_LIGHT)
                self_r._ac_hide()

            def disable(self_r, yes=True):
                st = "disabled" if yes else "normal"
                ro = "disabled" if yes else "readonly"
                for e in (self_r.e_client, self_r.e_start,
                          self_r.e_end, self_r.e_task, self_r.e_notes):
                    try: e.config(state=st)
                    except Exception: pass
                for cb in (self_r.cb_cat, self_r.cb_area, self_r.cb_sub,
                           self_r.cb_loc, self_r.cb_stat):
                    try: cb.config(state=ro)
                    except Exception: pass

        # ── Instantiate 10 rows ────────────────────────────────────────────────
        def _row_del(r):
            r.clear()
            _refresh_totals()

        def _row_change():
            _refresh_totals()

        for _ri2 in range(_N_ROWS):
            _r = _TRow(_row_area, _ri2, _row_del, _row_change)
            if on_leave:
                _r.disable(True)
            self.task_rows.append(_r)

        # Force geometry manager to process all pending requests
        # so the canvas scrollregion reflects the actual row content
        _row_area.update_idletasks()
        _h_can.update_idletasks()
        _sr()
        # Re-run Configure handler to ensure width is set correctly
        _h_can.event_generate("<Configure>")

        # ── Summary bar ───────────────────────────────────────────────────────
        sbar = tk.Frame(task_card, bg=PANEL_ALT, padx=14, pady=8,
                        highlightthickness=1, highlightbackground=BORDER_CLR)
        sbar.pack(fill="x", padx=10, pady=(4, 10))

        tk.Label(sbar, text="Break:", font=("Calibri", 9),
                 bg=PANEL_ALT, fg=TEXT_MID).pack(side="left")
        self.break_var = tk.StringVar(value="30")
        _bf = tk.Frame(sbar, bg=PANEL_ALT, highlightthickness=1,
                       highlightbackground=BORDER_CLR)
        _bf.pack(side="left", padx=4)
        _be = tk.Entry(_bf, textvariable=self.break_var,
                       font=("Calibri", 9), bg=ENTRY_BG, fg=ENTRY_FG,
                       relief="flat", bd=0, width=4, justify="center")
        _be.pack(ipady=4)
        tk.Label(sbar, text="mins", font=("Calibri", 9),
                 bg=PANEL_ALT, fg=TEXT_MID).pack(side="left", padx=(0, 12))

        tk.Frame(sbar, bg=BORDER_CLR, width=1).pack(
            side="left", fill="y", padx=8)
        tk.Label(sbar, text="Raw hrs:", font=("Calibri", 9),
                 bg=PANEL_ALT, fg=TEXT_MID).pack(side="left", padx=(8, 4))
        self.raw_total_lbl = tk.Label(sbar, text="0.0 h",
                                      font=("Calibri", 10, "bold"),
                                      bg=PANEL_ALT, fg=TEXT_DARK)
        self.raw_total_lbl.pack(side="left", padx=(0, 12))

        tk.Frame(sbar, bg=BORDER_CLR, width=1).pack(
            side="left", fill="y", padx=8)
        tk.Label(sbar, text="Net billable:", font=("Calibri", 9),
                 bg=PANEL_ALT, fg=TEXT_MID).pack(side="left", padx=(8, 4))
        self.net_total_lbl = tk.Label(sbar, text="0.0 h",
                                      font=("Calibri", 11, "bold"),
                                      bg=PANEL_ALT, fg=SUCCESS, padx=4)
        self.net_total_lbl.pack(side="left")

        # ── Live 8h progress badge ────────────────────────────────────────────
        tk.Frame(sbar, bg=BORDER_CLR, width=1).pack(side="left", fill="y", padx=12)
        badge_frame = tk.Frame(sbar, bg=PANEL_ALT)
        badge_frame.pack(side="left", fill="y")
        self._hrs_badge_lbl = tk.Label(badge_frame,
                                       text="🔴  0.0 / 8 h  Incomplete",
                                       font=("Calibri", 9, "bold"),
                                       bg=PANEL_ALT, fg="#B52A2A")
        self._hrs_badge_lbl.pack(anchor="w")
        self._hrs_bar_outer = tk.Frame(badge_frame, bg="#D0D8E4",
                                       height=5, width=120)
        self._hrs_bar_outer.pack(anchor="w", pady=(2, 0))
        self._hrs_bar_outer.pack_propagate(False)
        self._hrs_bar_fill  = tk.Frame(self._hrs_bar_outer, bg="#B52A2A",
                                       width=0, height=5)
        self._hrs_bar_fill.place(x=0, y=0, width=0, height=5)

        def _refresh_totals(*_):
            raw = 0.0
            for _row in self.task_rows:
                try:
                    t = _row.lbl_hrs.cget("text").replace("h","").strip()
                    raw += float(t) if t not in ("—","?","") else 0.0
                except Exception:
                    pass
            try:
                brk = int(self.break_var.get() or 0)
            except ValueError:
                brk = 0
            net = max(0.0, raw - brk / 60)
            self.raw_total_lbl.config(text=f"{raw:.1f} h")
            self.net_total_lbl.config(
                text=f"{net:.1f} h",
                fg=SUCCESS if net >= DAILY_TARGET else WARNING)
            # Update live progress bar
            if hasattr(self, "_hrs_bar_fill") and hasattr(self, "_hrs_bar_outer"):
                pct = min(net / DAILY_TARGET, 1.0)
                clr = "#1A6B45" if net >= DAILY_TARGET else ("#B45309" if pct >= 0.5 else "#B52A2A")
                w = self._hrs_bar_outer.winfo_width()
                fill_w = max(int(w * pct), 2 if pct > 0 else 0)
                try:
                    self._hrs_bar_fill.place(x=0, y=0, width=fill_w, height=6)
                    self._hrs_bar_fill.config(bg=clr)
                    self._hrs_badge_lbl.config(
                        text=f"{'🟢' if net>=DAILY_TARGET else '🔴'}  {net:.1f} / {DAILY_TARGET} h  {'✔ Complete' if net>=DAILY_TARGET else 'Incomplete'}",
                        fg=clr)
                except Exception:
                    pass

        self.refresh_totals = _refresh_totals
        _be.bind("<KeyRelease>", _refresh_totals)

        # ── Action buttons ────────────────────────────────────────────────────
        btn_card = _card(pad)
        btn_card.pack(fill="x", pady=(0, 12))
        bi = tk.Frame(btn_card, bg=PANEL_BG, padx=20, pady=14)
        bi.pack(fill="x")

        role   = getattr(self, "current_role", "user")
        locked = is_timesheet_locked(entry_date=str(today), role=role)

        if locked:
            tk.Label(bi,
                     text=f"  🔒  Timesheet locked from the {LOCK_DAY}th — contact Admin to make changes.",
                     font=("Calibri", 9, "bold"), bg="#F5E6C8", fg="#7A4F00",
                     padx=12, pady=8, relief="flat").pack(side="right")
        else:
            tk.Button(bi, text="✔   Save Today's Timesheet",
                      font=("Calibri", 11, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=28, pady=11,
                      state="disabled" if on_leave else "normal",
                      activebackground="#155C38",
                      command=lambda: self._save_entries(emp, today)
                      ).pack(side="right")

        tk.Button(bi, text="Clear All Rows",
                  font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=16, pady=11,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=self._clear_rows).pack(side="right", padx=8)

        # ── Today's saved entries preview ─────────────────────────────────────
        prev_card = _card(pad)
        prev_card.pack(fill="x", pady=(0, 12))
        _section_hdr(prev_card, "Today's Saved Entries", "📋")
        self.preview_frame = tk.Frame(prev_card, bg=PANEL_BG, padx=14)
        self.preview_frame.pack(fill="x", pady=(0, 14))
        self._refresh_preview(emp["emp_id"], today)

        # ── Recent Entries — 7-day grace period ──────────────────────────────────
        role_now     = getattr(self, "current_role", "user")
        grace_cutoff = today - timedelta(days=GRACE_DAYS)
        full_locked  = is_timesheet_locked(role=role_now)   # blanket 27th lock

        recent_card = _card(pad)
        recent_card.pack(fill="x", pady=(0, 20))
        _section_hdr(recent_card,
                     f"Recent Entries  (editable up to {GRACE_DAYS} days)", "📅")

        recent_frame = tk.Frame(recent_card, bg=PANEL_BG, padx=14)
        recent_frame.pack(fill="x", pady=(0, 14))

        def _auto_fill_leave(for_date: date):
            """Auto-insert an On Leave entry for a past date if none exists."""
            existing = load_entries(emp_id=emp["emp_id"], filter_date=for_date)
            if existing:
                messagebox.showinfo("Already filled",
                    f"Entries already exist for {for_date.strftime('%d %b %Y')}.")
                return
            leave_entry = [{
                "date":             str(for_date),
                "day":              for_date.strftime("%A"),
                "emp_id":           emp["emp_id"],
                "reporting_to":     emp.get("reporting_to", ""),
                "company_code":     "",
                "client":           "NA",
                "work_category":    "Others",
                "operational_area": "General",
                "sub_category":     "Other",
                "start_time":       "",
                "end_time":         "",
                "break_mins":       0,
                "total_hrs":        0,
                "task_desc":        "Unplanned Leave",
                "notes":            "Auto-filled via 7-day grace period",
                "work_location":    "",
                "status":           "On Leave",
            }]
            ok, err = save_timesheet_entries(leave_entry)
            if ok:
                _refresh_recent()
            else:
                messagebox.showerror("Error", f"Could not auto-fill leave:\n{err}")

        def _refresh_recent():
            for w in recent_frame.winfo_children():
                w.destroy()

            if full_locked:
                tk.Label(recent_frame,
                         text=f"  🔒  Timesheet locked from the {LOCK_DAY}th — contact Admin to make changes.",
                         font=("Calibri", 8, "bold"), bg="#F5E6C8", fg="#7A4F00",
                         anchor="w", pady=6).pack(fill="x", padx=6)
                return

            # Build a day-by-day view for the last GRACE_DAYS days (excluding today)
            cols     = ["Date", "Co. Code", "Client", "Work Category", "Op. Area",
                        "Start", "End", "Hrs", "Task", "Status", ""]
            col_w    = [12, 10, 14, 16, 14, 6, 6, 5, 18, 10, 8]
            col_flex = [0,  0,  1,  1,  1,  0, 0, 0,  1,  0,  0]
            for ci, flex in enumerate(col_flex):
                recent_frame.columnconfigure(ci, weight=flex)
            for ci, (col, cw) in enumerate(zip(cols, col_w)):
                tk.Label(recent_frame, text=col, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=cw if not col_flex[ci] else 1,
                         anchor="w", padx=6, pady=6
                         ).grid(row=0, column=ci, padx=1, pady=(0,1), sticky="nsew")

            grid_row  = 1
            total_hrs = 0.0
            any_shown = False

            for day_offset in range(1, GRACE_DAYS + 1):
                past_date = today - timedelta(days=day_offset)
                day_rows  = load_entries(emp_id=emp["emp_id"], filter_date=past_date)
                day_locked = is_timesheet_locked(entry_date=str(past_date), role=role_now)

                if day_rows:
                    any_shown = True
                    for r in day_rows:
                        vals = [past_date.strftime("%d %b"),
                                r.get("company_code",""), r.get("client_name",""),
                                r.get("work_category",""), r.get("operational_area",""),
                                r.get("start_time",""), r.get("end_time",""),
                                str(r.get("total_hrs","")), r.get("task_desc",""),
                                r.get("status","")]
                        bg = ROW_ODD if grid_row % 2 else ROW_EVEN
                        recent_frame.rowconfigure(grid_row, weight=0)
                        for ci, (v, cw, flex) in enumerate(zip(vals, col_w[:-1], col_flex[:-1])):
                            lbl = tk.Label(recent_frame, text=str(v or ""),
                                           font=("Calibri", 9), bg=bg, fg=TEXT_DARK,
                                           anchor="nw", padx=6, pady=5, justify="left",
                                           wraplength=0 if not flex else 1)
                            lbl.grid(row=grid_row, column=ci, padx=1, sticky="nsew")
                            if flex:
                                lbl.bind("<Configure>",
                                         lambda e, lb=lbl: lb.config(wraplength=max(1, e.width-12)))
                        if day_locked:
                            tk.Label(recent_frame, text="🔒",
                                     font=("Calibri", 9), bg=bg, fg=TEXT_LIGHT,
                                     anchor="center"
                                     ).grid(row=grid_row, column=len(cols)-1,
                                            padx=2, pady=1, sticky="ew")
                        else:
                            tk.Button(recent_frame, text="✏ Edit",
                                      font=("Calibri", 8, "bold"),
                                      bg=WARNING, fg="white", relief="flat",
                                      cursor="hand2", padx=4, pady=3,
                                      command=lambda row=r, pd=past_date: [
                                          self._open_edit_entry_dialog(row, emp["emp_id"], pd),
                                          self.after(300, _refresh_recent)
                                      ]).grid(row=grid_row, column=len(cols)-1,
                                              padx=2, pady=1, sticky="ew")
                        try:
                            total_hrs += float(r.get("total_hrs") or 0)
                        except Exception:
                            pass
                        grid_row += 1

                    # ── "➕ Add Entry" row after existing entries ─────────────────
                    if not day_locked:
                        add_row = tk.Frame(recent_frame, bg=PANEL_ALT)
                        add_row.grid(row=grid_row, column=0, columnspan=len(cols),
                                     sticky="ew", pady=(1, 4))
                        tk.Button(add_row,
                                  text=f"➕  Add another entry for {past_date.strftime('%d %b')}",
                                  font=("Calibri", 8, "bold"), bg=PANEL_ALT,
                                  fg=ACCENT_BLUE, relief="flat", cursor="hand2",
                                  padx=10, pady=4,
                                  command=lambda pd=past_date: [
                                      self._open_add_entry_dialog(emp, pd),
                                      self.after(300, _refresh_recent)
                                  ]).pack(side="left", padx=6)
                        grid_row += 1

                else:
                    # Empty day — show "Mark as Leave" AND "Add Entry" buttons
                    any_shown = True
                    bg = ROW_ODD if grid_row % 2 else ROW_EVEN
                    recent_frame.rowconfigure(grid_row, weight=0)
                    tk.Label(recent_frame,
                             text=f"  {past_date.strftime('%d %b')}",
                             font=("Calibri", 9, "bold"), bg=bg, fg=TEXT_LIGHT,
                             padx=6, pady=5, anchor="w"
                             ).grid(row=grid_row, column=0, padx=1, sticky="nsew")
                    if day_locked:
                        tk.Label(recent_frame, text="No entries  🔒",
                                 font=("Calibri", 9), bg=bg, fg=TEXT_LIGHT,
                                 padx=6, pady=5, anchor="w"
                                 ).grid(row=grid_row, column=1, columnspan=len(cols)-1,
                                        padx=1, sticky="nsew")
                    else:
                        notice = tk.Frame(recent_frame, bg=bg)
                        notice.grid(row=grid_row, column=1, columnspan=len(cols)-1,
                                    padx=1, sticky="nsew")
                        tk.Label(notice, text="  No entries",
                                 font=("Calibri", 9), bg=bg,
                                 fg=TEXT_LIGHT, pady=5).pack(side="left")
                        tk.Button(notice, text="➕  Add Entry",
                                  font=("Calibri", 8, "bold"), bg=SUCCESS,
                                  fg="white", relief="flat", cursor="hand2",
                                  padx=8, pady=3,
                                  command=lambda pd=past_date: [
                                      self._open_add_entry_dialog(emp, pd),
                                      self.after(300, _refresh_recent)
                                  ]).pack(side="left", padx=4)
                        tk.Button(notice, text="🏖  Mark as Leave",
                                  font=("Calibri", 8, "bold"), bg="#D2E1F1",
                                  fg=ACCENT_BLUE, relief="flat", cursor="hand2",
                                  padx=8, pady=3,
                                  command=lambda pd=past_date: _auto_fill_leave(pd)
                                  ).pack(side="left", padx=4)
                    grid_row += 1

            if not any_shown:
                tk.Label(recent_frame, text="  No entries in the last 7 days.",
                         font=("Calibri", 9), bg=PANEL_BG,
                         fg=TEXT_LIGHT, pady=10).pack(anchor="w")
                return

            foot = tk.Frame(recent_frame, bg="#D2E1F1")
            foot.grid(row=grid_row, column=0, columnspan=len(cols),
                      sticky="ew", pady=(4,0))
            badge = _hrs_badge(foot, total_hrs, target=DAILY_TARGET * GRACE_DAYS,
                               show_bar=True)
            badge.config(bg="#D2E1F1")
            for w in badge.winfo_children():
                w.config(bg="#D2E1F1")
            badge.pack(side="left", padx=6, pady=4)
            tk.Label(foot,
                     text=f"  Total (last {GRACE_DAYS} days): {total_hrs:.1f} hrs",
                     font=("Calibri", 9), bg="#D2E1F1",
                     fg=SUCCESS if total_hrs >= DAILY_TARGET * GRACE_DAYS else WARNING,
                     pady=6).pack(side="left")

        _refresh_recent()


    def _clear_rows(self):
        """Clear all TimesheetRow objects in the entry table."""
        for row_obj in self.task_rows:
            try:
                row_obj.clear()
            except Exception:
                pass
        if hasattr(self, "refresh_totals"):
            self.refresh_totals()

    def _quick_export_entries(self, emp_id=None, filter_date=None):
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"JAA_Entries_{date.today():%Y%m%d}.xlsx")
        if not fp: return
        try:
            rows = load_entries(emp_id=emp_id, filter_date=filter_date, limit=10000)
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Timesheet Entries"
            hf    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            hfill = PatternFill("solid", fgColor="06355E")
            thin  = Side(style="thin", color="CCCCCC")
            bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdrs  = ["Date","Day","Employee","EMP ID","Role","Reporting To",
                     "Co. Code","Client","Work Category","Op. Area","Sub-Category",
                     "Start","End","Break(min)","Total Hrs","Task","Notes","Location","Status"]
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font=hf; c.fill=hfill; c.border=bdr
                c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            ws.row_dimensions[1].height = 28
            of = PatternFill("solid", fgColor="D2E1F1")
            ef = PatternFill("solid", fgColor="FFFFFF")
            for ri, r in enumerate(rows, 2):
                fill = of if ri%2==0 else ef
                vals = [r.get("entry_date"), r.get("day_name"), r.get("full_name"),
                        r.get("employee_id"), r.get("role"), r.get("reporting_to"),
                        r.get("company_code"), r.get("client_name"),
                        r.get("work_category"), r.get("operational_area"),
                        r.get("sub_category"), r.get("start_time"), r.get("end_time"),
                        r.get("break_mins"), r.get("total_hrs"),
                        r.get("task_desc"), r.get("notes"),
                        r.get("work_location"), r.get("status")]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.fill=fill; c.border=bdr
                    c.font=Font(name="Calibri", size=9)
                    c.alignment=Alignment(
                        horizontal="center" if ci in (1,2,4,5,12,13,14,15) else "left",
                        vertical="center")
            for col in ws.columns:
                mx = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4,45)
            ws.freeze_panes = "A2"
            wb.save(fp)
            messagebox.showinfo("Exported ✔", f"{len(rows)} entries saved to:\n{fp}")
        except Exception as ex:
            messagebox.showerror("Export Error", str(ex))

    def _save_entries(self, emp, today):
        """Collect data from all TimesheetRow objects and save."""
        # ── Lock check ───────────────────────────────────────────────────────────
        role = getattr(self, "current_role", "user")
        if is_timesheet_locked(entry_date=str(today), role=role):
            messagebox.showwarning(
                "Timesheet Locked 🔒",
                f"The timesheet for {date.today().strftime('%B %Y')} is locked.\n\n"
                f"Timesheets lock on the {LOCK_DAY}th of each month.\n"
                f"Please contact Admin if you need to make changes.")
            return
        entries = []
        time_ranges = []

        for row_obj in self.task_rows:
            d = row_obj.get_data()
            cat = d.get("work_category", "").strip()
            if not cat:
                continue   # skip empty rows

            start_s = d.get("start_time", "").strip()
            end_s   = d.get("end_time",   "").strip()
            if start_s and end_s:
                try:
                    li = datetime.strptime(start_s, "%H:%M")
                    lo = datetime.strptime(end_s,   "%H:%M")
                    time_ranges.append((li, lo, cat))
                except Exception:
                    pass

            entries.append({
                "date":             today.isoformat(),
                "day":              today.strftime("%A"),
                "emp_id":           emp["emp_id"],
                "reporting_to":     self.reporting_var.get(),
                "company_code":     d.get("co_code", ""),
                "client":           d.get("client", ""),
                "work_category":    cat,
                "operational_area": d.get("operational_area", ""),
                "sub_category":     d.get("sub_category", ""),
                "start_time":       start_s,
                "end_time":         end_s,
                "break_mins":       0,
                "total_hrs":        d.get("total_hrs", 0.0),
                "task_desc":        d.get("task_desc", ""),
                "notes":            d.get("notes", ""),
                "work_location":    d.get("work_location", ""),
                "status":           d.get("status", "In Progress"),
            })

        if not entries:
            messagebox.showwarning("No Data",
                "Fill at least one Work Category row before saving.")
            return

        # ── Time overlap validation ─────────────────────────────────────────────
        overlaps = []
        for i in range(len(time_ranges)):
            for j in range(i + 1, len(time_ranges)):
                li1, lo1, cat1 = time_ranges[i]
                li2, lo2, cat2 = time_ranges[j]
                if li1 < lo2 and li2 < lo1:
                    overlaps.append(
                        f"Row {i+1} ({cat1}) overlaps Row {j+1} ({cat2})")
        if overlaps:
            if not messagebox.askyesno("Time Overlap Warning",
                    "⚠  Overlapping time entries detected:\n\n"
                    + "\n".join(overlaps)
                    + "\n\nProceed anyway?"):
                return

        try:
            brk = int(self.break_var.get() or 0)
        except ValueError:
            brk = 30
        entries[0]["break_mins"] = brk
        if entries[0]["total_hrs"]:
            entries[0]["total_hrs"] = round(
                max(0, entries[0]["total_hrs"] - brk / 60), 2)

        # ── Record manual attendance ────────────────────────────────────────────
        day_login_v  = getattr(self, "day_login_var",  None)
        day_logout_v = getattr(self, "day_logout_var", None)
        if day_login_v and day_logout_v:
            ls = day_login_v.get().strip()
            lo = day_logout_v.get().strip()
            if ls and lo:
                try:
                    li_t = datetime.strptime(ls, "%H:%M")
                    lo_t = datetime.strptime(lo, "%H:%M")
                    if lo_t >= li_t:
                        dur    = round((lo_t - li_t).seconds / 60, 1)
                        li_iso = datetime.combine(today, li_t.time()).isoformat()
                        lo_iso = datetime.combine(today, lo_t.time()).isoformat()
                        with get_conn() as conn:
                            conn.execute("""
                                INSERT INTO attendance_log
                                (emp_id, login_dt, logout_dt, duration_mins, session_date)
                                VALUES (?,?,?,?,?)
                            """, (emp["emp_id"], li_iso, lo_iso,
                                  dur, today.isoformat()))
                except Exception as ae:
                    log.warning(f"Attendance record warn: {ae}")

        ok, err_msg = save_timesheet_entries(entries)
        if ok:
            threading.Thread(
                target=lambda: append_entries_to_excel(entries), daemon=True
            ).start()
            total_hrs_saved = sum(e.get("total_hrs", 0) for e in entries)
            cfg = load_config()
            try:
                upsert_submission(
                    emp["emp_id"], today.isoformat(),
                    len(entries), total_hrs_saved,
                    cfg.get("submission_cutoff", "20:30"))
            except Exception as se:
                log.warning(f"Submission log error: {se}")
            messagebox.showinfo("Saved ✔",
                f"{len(entries)} row(s) saved successfully.")
            self._clear_rows()
            self._refresh_preview(emp["emp_id"], today)
        else:
            messagebox.showerror("Save Failed",
                f"Could not save entries.\n\nReason: {err_msg}\n\n"
                f"Tip: If this mentions 'FOREIGN KEY', make sure the Company Code "
                f"entered matches a code in the Companies master list, "
                f"or leave it blank.")


    def _refresh_preview(self, emp_id, today):
        for w in self.preview_frame.winfo_children():
            w.destroy()
        rows = load_entries(emp_id=emp_id, filter_date=today)
        if not rows:
            tk.Label(self.preview_frame, text="No entries saved yet today.",
                     font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_LIGHT, pady=14).pack()
            return

        # ── Lock state ───────────────────────────────────────────────────────────
        role   = getattr(self, "current_role", "user")
        locked = is_timesheet_locked(entry_date=str(today), role=role)

        # Show lock banner if locked
        if locked:
            lock_bar = tk.Frame(self.preview_frame, bg="#F5E6C8")
            lock_bar.grid(row=0, column=0, columnspan=12, sticky="ew", pady=(0, 4))
            tk.Label(lock_bar,
                     text=f"  🔒  Timesheet locked from the {LOCK_DAY}th — contact Admin to make changes.",
                     font=("Calibri", 8, "bold"), bg="#F5E6C8", fg="#7A4F00",
                     anchor="w", pady=5).pack(fill="x", padx=6)
            hdr_row = 1
        else:
            hdr_row = 0

        cols    = ["Co. Code", "Client", "Work Category", "Op. Area", "Sub-Cat",
                   "Start", "End", "Hrs", "Task", "Status", ""]
        col_w   = [10, 18, 16, 14, 12, 6, 6, 5, 18, 10, 8]
        col_flex = [0, 1, 1, 1, 1, 0, 0, 0, 1, 0, 0]

        for ci, flex in enumerate(col_flex):
            self.preview_frame.columnconfigure(ci, weight=flex)

        for ci, (col, cw) in enumerate(zip(cols, col_w)):
            tk.Label(self.preview_frame, text=col, font=("Calibri", 8, "bold"),
                     bg=TBL_HDR_BG, fg=TBL_HDR_FG, width=cw if not col_flex[ci] else 1,
                     anchor="w", padx=6, pady=6
                     ).grid(row=hdr_row, column=ci, padx=1, pady=(0, 1), sticky="nsew")

        total_hrs = 0
        for ri, r in enumerate(rows, hdr_row + 1):
            vals = [r.get("company_code", ""), r.get("client_name", ""),
                    r.get("work_category", ""), r.get("operational_area", ""),
                    r.get("sub_category", ""), r.get("start_time", ""),
                    r.get("end_time", ""), str(r.get("total_hrs", "")),
                    r.get("task_desc", ""), r.get("status", "")]
            bg = ROW_ODD if ri % 2 else ROW_EVEN
            self.preview_frame.rowconfigure(ri, weight=0)
            for ci, (v, cw, flex) in enumerate(zip(vals, col_w[:-1], col_flex[:-1])):
                lbl = tk.Label(self.preview_frame, text=str(v or ""),
                               font=("Calibri", 9),
                               bg=bg, fg=TEXT_DARK, anchor="nw", padx=6, pady=5,
                               justify="left",
                               wraplength=0 if not flex else 1)
                lbl.grid(row=ri, column=ci, padx=1, sticky="nsew")
                if flex:
                    lbl.bind("<Configure>",
                             lambda e, lb=lbl: lb.config(wraplength=max(1, e.width - 12)))
            # Edit button — disabled when locked
            if locked:
                tk.Label(self.preview_frame, text="🔒",
                         font=("Calibri", 9), bg=bg, fg=TEXT_LIGHT,
                         anchor="center"
                         ).grid(row=ri, column=len(cols)-1, padx=2, pady=1, sticky="ew")
            else:
                tk.Button(self.preview_frame, text="✏ Edit",
                          font=("Calibri", 8, "bold"),
                          bg=WARNING, fg="white", relief="flat", cursor="hand2",
                          padx=4, pady=3,
                          command=lambda row=r: self._open_edit_entry_dialog(row, emp_id, today)
                          ).grid(row=ri, column=len(cols)-1, padx=2, pady=1, sticky="ew")
            try:
                total_hrs += float(r.get("total_hrs") or 0)
            except Exception:
                pass

        foot = tk.Frame(self.preview_frame, bg="#D2E1F1")
        foot.grid(row=len(rows)+hdr_row+1, column=0, columnspan=len(cols), sticky="ew", pady=(4, 0))
        badge = _hrs_badge(foot, total_hrs, target=DAILY_TARGET, show_bar=True)
        badge.config(bg="#D2E1F1")
        for w in badge.winfo_children():
            w.config(bg="#D2E1F1")
        badge.pack(side="left", padx=6, pady=4)
        tk.Label(foot, text=f"  Total today: {total_hrs:.1f} hrs",
                 font=("Calibri", 9), bg="#D2E1F1",
                 fg=SUCCESS if total_hrs >= DAILY_TARGET else WARNING,
                 pady=6).pack(side="left")

    def _open_edit_entry_dialog(self, row: dict, emp_id: str, today):
        """Pop-up dialog to edit a single saved timesheet entry."""
        role = getattr(self, "current_role", "user")
        if is_timesheet_locked(entry_date=row.get("entry_date", str(today)), role=role):
            messagebox.showwarning(
                "Timesheet Locked 🔒",
                f"This entry is locked from the {LOCK_DAY}th onwards.\n"
                f"Please contact Admin if changes are needed.")
            return
        dlg = tk.Toplevel(self)
        dlg.title("Edit Entry")
        dlg.geometry("540x560")
        dlg.configure(bg=PAGE_BG)
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 540) // 2
        py = self.winfo_y() + (self.winfo_height() - 560) // 2
        dlg.geometry(f"540x560+{px}+{py}")

        # Header
        tk.Frame(dlg, bg=ACCENT_GOLD, height=4).pack(fill="x")
        hdr = tk.Frame(dlg, bg=SIDEBAR_BG, padx=18, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="✏  Edit Timesheet Entry",
                 font=("Georgia", 12, "bold"), bg=SIDEBAR_BG, fg=SIDEBAR_FG).pack(anchor="w")
        tk.Label(hdr, text=f"Row saved on {row.get('entry_date','')}",
                 font=("Calibri", 8), bg=SIDEBAR_BG, fg=SIDEBAR_ICN).pack(anchor="w")

        # Scrollable body
        sc = tk.Canvas(dlg, bg=PAGE_BG, highlightthickness=0)
        sb = ttk.Scrollbar(dlg, orient="vertical", command=sc.yview)
        sc.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        sc.pack(side="left", fill="both", expand=True)
        fm = tk.Frame(sc, bg=PAGE_BG, padx=22, pady=16)
        fm_id = sc.create_window((0, 0), window=fm, anchor="nw")
        fm.bind("<Configure>", lambda e: sc.configure(scrollregion=sc.bbox("all")))
        sc.bind("<Configure>", lambda e: sc.itemconfig(fm_id, width=e.width))
        sc.bind("<MouseWheel>", lambda e: sc.yview_scroll(-1*(e.delta//120), "units"))

        def _lbl(t):
            tk.Label(fm, text=t, font=("Calibri", 9, "bold"),
                     bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w", pady=(10, 2))

        def _ef(var, width=None):
            f = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1,
                         highlightbackground=BORDER_CLR)
            f.pack(fill="x")
            kw = dict(textvariable=var, font=("Calibri", 10), bg=PANEL_BG,
                      fg=ENTRY_FG, relief="flat", bd=0, insertbackground=FOCUS_CLR)
            if width:
                kw["width"] = width
            tk.Entry(f, **kw).pack(fill="x", ipady=7, padx=6)

        def _cb(var, values):
            f = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1,
                         highlightbackground=BORDER_CLR)
            f.pack(fill="x")
            c = ttk.Combobox(f, textvariable=var, values=values,
                             state="readonly", font=("Calibri", 10))
            c.pack(fill="x", ipady=5, padx=2)
            _style_combo(c)

        # Load companies for the dropdown
        all_cos   = get_all_companies()
        co_names  = [c["full_name"] for c in all_cos]
        co_by_name = {c["full_name"]: c for c in all_cos}

        # Find company name from code stored in row
        stored_code = row.get("company_code") or ""
        stored_co_name = ""
        for c in all_cos:
            if c["unique_code"] == stored_code:
                stored_co_name = c["full_name"]
                break

        v_co    = tk.StringVar(value=stored_co_name)
        v_cat   = tk.StringVar(value=row.get("work_category", ""))
        v_area  = tk.StringVar(value=row.get("operational_area", ""))
        v_sub   = tk.StringVar(value=row.get("sub_category", ""))
        v_start = tk.StringVar(value=row.get("start_time", ""))
        v_end   = tk.StringVar(value=row.get("end_time", ""))
        v_break = tk.StringVar(value=str(row.get("break_mins", 0)))
        v_task  = tk.StringVar(value=row.get("task_desc", ""))
        v_notes = tk.StringVar(value=row.get("notes", ""))
        v_loc   = tk.StringVar(value=row.get("work_location", ""))
        v_stat  = tk.StringVar(value=row.get("status", "In Progress"))

        _lbl("Client / Company")
        _cb(v_co, co_names)

        _lbl("Work Category")
        cats = get_work_categories()
        _cb(v_cat, cats)

        _lbl("Operational Area")
        area_var_frame = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1,
                                  highlightbackground=BORDER_CLR)
        area_var_frame.pack(fill="x")
        area_cb = ttk.Combobox(area_var_frame, textvariable=v_area,
                               values=get_operational_areas(v_cat.get()),
                               state="readonly", font=("Calibri", 10))
        area_cb.pack(fill="x", ipady=5, padx=2)
        _style_combo(area_cb)

        _lbl("Sub-Category")
        sub_var_frame = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1,
                                 highlightbackground=BORDER_CLR)
        sub_var_frame.pack(fill="x")
        sub_cb = ttk.Combobox(sub_var_frame, textvariable=v_sub,
                              values=get_sub_categories(v_cat.get(), v_area.get()),
                              state="readonly", font=("Calibri", 10))
        sub_cb.pack(fill="x", ipady=5, padx=2)
        _style_combo(sub_cb)

        def _update_areas(e=None):
            area_cb["values"] = get_operational_areas(v_cat.get())
            v_area.set(""); v_sub.set("")
            sub_cb["values"] = []

        def _update_subs(e=None):
            sub_cb["values"] = get_sub_categories(v_cat.get(), v_area.get())
            v_sub.set("")

        area_cb.bind("<<ComboboxSelected>>", _update_subs)

        # Re-bind category cb after creation
        for child in area_var_frame.master.winfo_children():
            pass  # category cb is already packed; bind via variable trace
        v_cat.trace_add("write", lambda *_: _update_areas())

        time_row = tk.Frame(fm, bg=PAGE_BG)
        time_row.pack(fill="x", pady=(10, 0))
        for lbl_t, var in [("Start (HH:MM)", v_start), ("End (HH:MM)", v_end), ("Break (mins)", v_break)]:
            col = tk.Frame(time_row, bg=PAGE_BG)
            col.pack(side="left", padx=(0, 14))
            tk.Label(col, text=lbl_t, font=("Calibri", 9, "bold"),
                     bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w", pady=(0, 2))
            f = tk.Frame(col, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            f.pack()
            tk.Entry(f, textvariable=var, font=("Calibri", 10), width=10,
                     bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                     insertbackground=FOCUS_CLR).pack(ipady=6, padx=6)

        _lbl("Task Description")
        _ef(v_task)

        _lbl("Notes")
        _ef(v_notes)

        _lbl("Work Location")
        _cb(v_loc, WORK_LOCATION)

        _lbl("Status")
        _cb(v_stat, STATUSES)

        err_lbl = tk.Label(fm, text="", font=("Calibri", 9), bg=PAGE_BG, fg=DANGER)
        err_lbl.pack(anchor="w", pady=(8, 0))

        def _save():
            # Compute total_hrs
            try:
                s = datetime.strptime(v_start.get().strip(), "%H:%M")
                e = datetime.strptime(v_end.get().strip(),   "%H:%M")
                if e < s:
                    err_lbl.config(text="⚠  End time cannot be before Start time.")
                    return
                brk  = int(v_break.get().strip() or 0)
                hrs  = round(max((e - s).seconds / 3600 - brk / 60, 0), 2)
            except Exception:
                err_lbl.config(text="⚠  Enter valid Start / End times (HH:MM).")
                return

            # Resolve company code
            co_name  = v_co.get().strip()
            co_code  = co_by_name[co_name]["unique_code"] if co_name in co_by_name else None

            ok = update_timesheet_entry(
                row["id"],
                company_code=co_code,
                client_name=co_name,
                work_category=v_cat.get().strip(),
                operational_area=v_area.get().strip(),
                sub_category=v_sub.get().strip(),
                start_time=v_start.get().strip(),
                end_time=v_end.get().strip(),
                break_mins=int(v_break.get().strip() or 0),
                total_hrs=hrs,
                task_desc=v_task.get().strip(),
                notes=v_notes.get().strip(),
                work_location=v_loc.get().strip(),
                status=v_stat.get().strip(),
            )
            if ok:
                dlg.destroy()
                self._refresh_preview(emp_id, today)
            else:
                err_lbl.config(text="⚠  Nothing changed or save failed.")

        btn_row = tk.Frame(fm, bg=PAGE_BG)
        btn_row.pack(anchor="w", pady=(14, 0))
        tk.Button(btn_row, text="💾  Save Changes",
                  font=("Calibri", 10, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=18, pady=9,
                  command=_save).pack(side="left")
        tk.Button(btn_row, text="Cancel",
                  font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=12, pady=9,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=dlg.destroy).pack(side="left", padx=(8, 0))

    def _open_add_entry_dialog(self, emp: dict, for_date: date):
        """Pop-up dialog to add a brand-new entry for a past date (grace period)."""
        role = getattr(self, "current_role", "user")
        if is_timesheet_locked(entry_date=str(for_date), role=role):
            messagebox.showwarning("Timesheet Locked 🔒",
                f"Entries for {for_date.strftime('%d %b %Y')} are locked.\n"
                f"Please contact Admin if changes are needed.")
            return

        dlg = tk.Toplevel(self)
        dlg.title(f"Add Entry — {for_date.strftime('%d %b %Y')}")
        dlg.geometry("540x580")
        dlg.configure(bg=PAGE_BG)
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 540) // 2
        py = self.winfo_y() + (self.winfo_height() - 580) // 2
        dlg.geometry(f"540x580+{px}+{py}")

        tk.Frame(dlg, bg=SUCCESS, height=4).pack(fill="x")
        hdr = tk.Frame(dlg, bg=SIDEBAR_BG, padx=18, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="➕  Add Timesheet Entry",
                 font=("Georgia", 12, "bold"), bg=SIDEBAR_BG, fg=SIDEBAR_FG).pack(anchor="w")
        tk.Label(hdr, text=f"{for_date.strftime('%A, %d %B %Y')}  (grace period edit)",
                 font=("Calibri", 8), bg=SIDEBAR_BG, fg=SIDEBAR_ICN).pack(anchor="w")

        sc = tk.Canvas(dlg, bg=PAGE_BG, highlightthickness=0)
        sb = ttk.Scrollbar(dlg, orient="vertical", command=sc.yview)
        sc.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        sc.pack(side="left", fill="both", expand=True)
        fm = tk.Frame(sc, bg=PAGE_BG, padx=22, pady=16)
        fm_id = sc.create_window((0, 0), window=fm, anchor="nw")
        fm.bind("<Configure>", lambda e: sc.configure(scrollregion=sc.bbox("all")))
        sc.bind("<Configure>", lambda e: sc.itemconfig(fm_id, width=e.width))
        sc.bind("<MouseWheel>", lambda e: sc.yview_scroll(-1*(e.delta//120), "units"))

        def _lbl(t):
            tk.Label(fm, text=t, font=("Calibri", 9, "bold"),
                     bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w", pady=(10, 2))

        def _ef(var, width=None):
            f = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            f.pack(fill="x")
            kw = dict(textvariable=var, font=("Calibri", 10), bg=PANEL_BG,
                      fg=ENTRY_FG, relief="flat", bd=0, insertbackground=FOCUS_CLR)
            if width: kw["width"] = width
            tk.Entry(f, **kw).pack(fill="x", ipady=7, padx=6)

        def _cb(var, values):
            f = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            f.pack(fill="x")
            c = ttk.Combobox(f, textvariable=var, values=values,
                             state="readonly", font=("Calibri", 10))
            c.pack(fill="x", ipady=5, padx=2); _style_combo(c)

        all_cos    = get_all_companies()
        co_names   = [c["full_name"] for c in all_cos]
        co_by_name = {c["full_name"]: c for c in all_cos}

        v_co    = tk.StringVar()
        v_cat   = tk.StringVar()
        v_area  = tk.StringVar()
        v_sub   = tk.StringVar()
        v_start = tk.StringVar()
        v_end   = tk.StringVar()
        v_break = tk.StringVar(value="0")
        v_task  = tk.StringVar()
        v_notes = tk.StringVar()
        v_loc   = tk.StringVar()
        v_stat  = tk.StringVar(value="Completed")

        _lbl("Client / Company")
        _cb(v_co, co_names)

        _lbl("Work Category")
        cats = get_work_categories()
        _cb(v_cat, cats)

        _lbl("Operational Area")
        area_f = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        area_f.pack(fill="x")
        area_cb = ttk.Combobox(area_f, textvariable=v_area,
                               values=[], state="readonly", font=("Calibri", 10))
        area_cb.pack(fill="x", ipady=5, padx=2); _style_combo(area_cb)

        _lbl("Sub-Category")
        sub_f = tk.Frame(fm, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        sub_f.pack(fill="x")
        sub_cb = ttk.Combobox(sub_f, textvariable=v_sub,
                              values=[], state="readonly", font=("Calibri", 10))
        sub_cb.pack(fill="x", ipady=5, padx=2); _style_combo(sub_cb)

        def _update_areas(*_):
            area_cb["values"] = get_operational_areas(v_cat.get())
            v_area.set(""); v_sub.set(""); sub_cb["values"] = []
        def _update_subs(*_):
            sub_cb["values"] = get_sub_categories(v_cat.get(), v_area.get())
            v_sub.set("")

        v_cat.trace_add("write",  lambda *_: _update_areas())
        area_cb.bind("<<ComboboxSelected>>", _update_subs)

        time_row = tk.Frame(fm, bg=PAGE_BG); time_row.pack(fill="x", pady=(10, 0))
        for lbl_t, var in [("Start (HH:MM)", v_start), ("End (HH:MM)", v_end), ("Break (mins)", v_break)]:
            col = tk.Frame(time_row, bg=PAGE_BG); col.pack(side="left", padx=(0, 14))
            tk.Label(col, text=lbl_t, font=("Calibri", 9, "bold"),
                     bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w", pady=(0, 2))
            f = tk.Frame(col, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            f.pack()
            tk.Entry(f, textvariable=var, font=("Calibri", 10), width=10,
                     bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                     insertbackground=FOCUS_CLR).pack(ipady=6, padx=6)

        _lbl("Task Description"); _ef(v_task)
        _lbl("Notes");            _ef(v_notes)
        _lbl("Work Location");    _cb(v_loc, WORK_LOCATION)
        _lbl("Status");           _cb(v_stat, STATUSES)

        err_lbl = tk.Label(fm, text="", font=("Calibri", 9), bg=PAGE_BG, fg=DANGER)
        err_lbl.pack(anchor="w", pady=(8, 0))

        def _save():
            if not v_cat.get().strip():
                err_lbl.config(text="⚠  Work Category is required."); return
            try:
                s   = datetime.strptime(v_start.get().strip(), "%H:%M")
                e   = datetime.strptime(v_end.get().strip(),   "%H:%M")
                if e < s:
                    err_lbl.config(text="⚠  End time cannot be before Start time."); return
                brk = int(v_break.get().strip() or 0)
                hrs = round(max((e - s).seconds / 3600 - brk / 60, 0), 2)
            except Exception:
                err_lbl.config(text="⚠  Enter valid Start / End times (HH:MM)."); return

            co_name = v_co.get().strip()
            co_code = co_by_name[co_name]["unique_code"] if co_name in co_by_name else ""

            new_entry = [{
                "date":             str(for_date),
                "day":              for_date.strftime("%A"),
                "emp_id":           emp["emp_id"],
                "reporting_to":     emp.get("reporting_to", ""),
                "company_code":     co_code,
                "client":           co_name,
                "work_category":    v_cat.get().strip(),
                "operational_area": v_area.get().strip(),
                "sub_category":     v_sub.get().strip(),
                "start_time":       v_start.get().strip(),
                "end_time":         v_end.get().strip(),
                "break_mins":       brk,
                "total_hrs":        hrs,
                "task_desc":        v_task.get().strip(),
                "notes":            v_notes.get().strip(),
                "work_location":    v_loc.get().strip(),
                "status":           v_stat.get().strip(),
            }]
            ok, err = save_timesheet_entries(new_entry)
            if ok:
                dlg.destroy()
            else:
                err_lbl.config(text=f"⚠  Save failed: {err}")

        btn_row = tk.Frame(fm, bg=PAGE_BG); btn_row.pack(anchor="w", pady=(14, 0))
        tk.Button(btn_row, text="💾  Save Entry",
                  font=("Calibri", 10, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=18, pady=9,
                  command=_save).pack(side="left")
        tk.Button(btn_row, text="Cancel",
                  font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=12, pady=9,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=dlg.destroy).pack(side="left", padx=(8, 0))

    # ── MY LEAVES TAB (Employee) ─────────────────────────────────────────────────

    def _build_my_leaves_tab(self, parent, emp: dict):
        LEAVE_TYPES = ["Casual", "Sick", "Earned", "Compensatory", "Other"]
        STATUS_CLR  = {
            "Pending":  "#7C3AED",
            "Approved": "#1A6B45",
            "Rejected": "#B52A2A",
        }
        STATUS_BG = {
            "Pending":  "#EDE9FE",
            "Approved": "#D1FAE5",
            "Rejected": "#FEE2E2",
        }

        tb = self._topbar(parent, "My Leaves",
                          "  Apply for leave · Track your requests")

        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=16)

        # ── Left column: request form ────────────────────────────────────────────
        left = tk.Frame(outer, bg=PANEL_BG,
                        highlightthickness=1, highlightbackground=BORDER_CLR,
                        width=310)
        left.pack(side="left", fill="y", padx=(0, 14))
        left.pack_propagate(False)

        # header stripe
        tk.Frame(left, bg="#7C3AED", height=4).pack(fill="x")
        tk.Label(left, text="Apply for Leave",
                 font=("Calibri", 11, "bold"),
                 bg=PANEL_BG, fg=TEXT_DARK, padx=18, pady=12,
                 anchor="w").pack(fill="x")
        tk.Frame(left, bg=BORDER_CLR, height=1).pack(fill="x")

        form = tk.Frame(left, bg=PANEL_BG, padx=18, pady=14)
        form.pack(fill="x")

        def _lbl(text):
            tk.Label(form, text=text, font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(8, 2))

        def _entry_field(var, width=28):
            wrap = tk.Frame(form, bg=BORDER_CLR, padx=1, pady=1)
            wrap.pack(fill="x")
            tk.Entry(wrap, textvariable=var, font=("Calibri", 10),
                     bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                     width=width, insertbackground=FOCUS_CLR).pack(
                     fill="x", ipady=6, padx=4)

        _lbl("Leave Type")
        type_var = tk.StringVar(value="Casual")
        tw = tk.Frame(form, bg=BORDER_CLR, padx=1, pady=1); tw.pack(fill="x")
        tc = ttk.Combobox(tw, textvariable=type_var, values=LEAVE_TYPES,
                          state="readonly", font=("Calibri", 10), width=26)
        tc.pack(fill="x", ipady=4, padx=2); _style_combo(tc)

        _lbl("From Date  (YYYY-MM-DD)")
        start_var = tk.StringVar()
        _entry_field(start_var)

        _lbl("To Date  (YYYY-MM-DD)")
        end_var = tk.StringVar()
        _entry_field(end_var)

        _lbl("Reason")
        reason_var = tk.StringVar()
        reason_wrap = tk.Frame(form, bg=BORDER_CLR, padx=1, pady=1)
        reason_wrap.pack(fill="x")
        tk.Entry(reason_wrap, textvariable=reason_var, font=("Calibri", 10),
                 bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                 width=28, insertbackground=FOCUS_CLR).pack(
                 fill="x", ipady=6, padx=4)

        status_lbl = tk.Label(form, text="", font=("Calibri", 9),
                              bg=PANEL_BG, wraplength=270)
        status_lbl.pack(anchor="w", pady=(10, 0))

        # days-count label
        days_lbl = tk.Label(form, text="", font=("Calibri", 9, "bold"),
                            bg=PANEL_BG, fg="#7C3AED")
        days_lbl.pack(anchor="w")

        def _calc_days(*_):
            try:
                s = datetime.strptime(start_var.get().strip(), "%Y-%m-%d").date()
                e = datetime.strptime(end_var.get().strip(),   "%Y-%m-%d").date()
                if e < s:
                    days_lbl.config(text="⚠ End before start", fg=DANGER); return
                days = (e - s).days + 1
                days_lbl.config(text=f"📅  {days} day{'s' if days!=1 else ''}",
                                fg="#7C3AED")
            except Exception:
                days_lbl.config(text="")

        start_var.trace_add("write", _calc_days)
        end_var.trace_add("write",   _calc_days)

        def _submit():
            try:
                s = datetime.strptime(start_var.get().strip(), "%Y-%m-%d").date()
                e = datetime.strptime(end_var.get().strip(),   "%Y-%m-%d").date()
                if e < s: raise ValueError("end before start")
            except Exception:
                status_lbl.config(
                    text="⚠  Invalid dates. Use YYYY-MM-DD.", fg=DANGER); return
            if not reason_var.get().strip():
                status_lbl.config(text="⚠  Please enter a reason.", fg=DANGER)
                return
            request_leave(emp["emp_id"],
                          start_var.get().strip(), end_var.get().strip(),
                          reason_var.get().strip(), type_var.get())
            start_var.set(""); end_var.set(""); reason_var.set("")
            days_lbl.config(text="")
            status_lbl.config(
                text="✔  Request submitted — awaiting admin approval.",
                fg=SUCCESS)
            _refresh_list()

        tk.Frame(form, bg=BORDER_CLR, height=1).pack(fill="x", pady=(14, 0))
        sub_btn = tk.Label(form, text="Submit Request →",
                           font=("Calibri", 10, "bold"),
                           bg="#7C3AED", fg="white",
                           padx=16, pady=9, cursor="hand2", anchor="center")
        sub_btn.pack(fill="x", pady=(10, 0))
        sub_btn.bind("<Button-1>", lambda e: _submit())

        # ── Right column: my leave history ───────────────────────────────────────
        right = tk.Frame(outer, bg=PAGE_BG)
        right.pack(side="left", fill="both", expand=True)

        list_card = tk.Frame(right, bg=PANEL_BG,
                             highlightthickness=1, highlightbackground=BORDER_CLR)
        list_card.pack(fill="both", expand=True)

        tk.Frame(list_card, bg="#7C3AED", height=4).pack(fill="x")
        list_hdr_row = tk.Frame(list_card, bg=PANEL_BG, padx=16, pady=10)
        list_hdr_row.pack(fill="x")
        tk.Label(list_hdr_row, text="My Leave History",
                 font=("Calibri", 11, "bold"),
                 bg=PANEL_BG, fg=TEXT_DARK).pack(side="left")
        tk.Frame(list_card, bg=BORDER_CLR, height=1).pack(fill="x")

        scroll_wrap = tk.Frame(list_card, bg=PANEL_BG)
        scroll_wrap.pack(fill="both", expand=True)

        lv_canvas = tk.Canvas(scroll_wrap, bg=PANEL_BG, highlightthickness=0)
        lv_vsb = ttk.Scrollbar(scroll_wrap, orient="vertical",
                               command=lv_canvas.yview)
        lv_canvas.configure(yscrollcommand=lv_vsb.set)
        lv_vsb.pack(side="right", fill="y")
        lv_canvas.pack(side="left", fill="both", expand=True)
        lv_canvas.bind("<MouseWheel>",
                       lambda e: lv_canvas.yview_scroll(-1*(e.delta//120), "units"))

        inner = tk.Frame(lv_canvas, bg=PANEL_BG)
        lv_win = lv_canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: lv_canvas.configure(
                       scrollregion=lv_canvas.bbox("all")))
        lv_canvas.bind("<Configure>",
                       lambda e: lv_canvas.itemconfig(lv_win, width=e.width))

        def _refresh_list():
            for w in inner.winfo_children():
                w.destroy()
            leaves = get_leaves(emp_id=emp["emp_id"])
            if not leaves:
                tk.Label(inner, text="No leave records yet.",
                         font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_MID,
                         pady=24).pack()
                return

            # column header
            hdr = tk.Frame(inner, bg=TBL_HDR_BG)
            hdr.pack(fill="x", padx=10, pady=(6, 0))
            for txt, w in [("Type", 12), ("From", 13), ("To", 13),
                           ("Days", 6), ("Reason", 20), ("Status", 10),
                           ("Note", 18)]:
                tk.Label(hdr, text=txt, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=w, anchor="w", padx=6, pady=5).pack(side="left")

            rows_f = tk.Frame(inner, bg=PANEL_BG)
            rows_f.pack(fill="x", padx=10, pady=(0, 10))
            for i, lv in enumerate(leaves):
                st = lv.get("status", "Pending")
                row_bg = ROW_ODD if i % 2 == 0 else ROW_EVEN
                try:
                    s = datetime.strptime(lv["start_date"], "%Y-%m-%d").date()
                    e = datetime.strptime(lv["end_date"],   "%Y-%m-%d").date()
                    days = str((e - s).days + 1)
                except Exception:
                    days = "—"
                rf = tk.Frame(rows_f, bg=row_bg); rf.pack(fill="x")
                for txt, w, fg in [
                    (lv.get("leave_type","Casual"), 12, TEXT_DARK),
                    (lv["start_date"],              13, TEXT_DARK),
                    (lv["end_date"],                13, TEXT_DARK),
                    (days,                           6, TEXT_DARK),
                    (lv.get("reason","") or "—",   20, TEXT_DARK),
                    (st,                            10, STATUS_CLR.get(st, TEXT_DARK)),
                    (lv.get("admin_note","") or "—",18, TEXT_MID),
                ]:
                    tk.Label(rf, text=str(txt), font=("Calibri", 9),
                             bg=row_bg, fg=fg, width=w, anchor="w",
                             padx=6, pady=4).pack(side="left")

                # cancel button for Pending requests
                if st == "Pending":
                    cancel = tk.Label(rf, text="✕ Cancel",
                                      font=("Calibri", 8), bg=row_bg,
                                      fg=DANGER, cursor="hand2", padx=6)
                    cancel.pack(side="left")
                    cancel.bind("<Button-1>",
                                lambda e, lid=lv["id"]: [
                                    remove_leave(lid), _refresh_list()])

        _refresh_list()

    # ── MY ENTRIES TAB ───────────────────────────────────────────────────────────

    def _build_entries_tab(self, parent, emp_id):
        tb = self._topbar(parent, "My Timesheet History")
        tk.Button(tb, text="📥  Export .xlsx",
                  font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  activebackground="#155C38",
                  command=lambda: self._quick_export_entries(emp_id)
                  ).pack(side="right", padx=16, pady=10)
        container = tk.Frame(parent, bg=PAGE_BG)
        container.pack(fill="both", expand=True, padx=20, pady=16)
        self._entries_table(container, emp_id=emp_id)

    def _entries_table(self, parent, emp_id=None, all_users=False):
        ctrl_card = _card(parent); ctrl_card.pack(fill="x", pady=(0, 8))
        ctrl = tk.Frame(ctrl_card, bg=PANEL_BG, padx=16, pady=12); ctrl.pack(fill="x")

        tk.Label(ctrl, text="Filter by Date:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        date_var = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        df = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        df.pack(side="left", padx=(4, 12))
        tk.Entry(df, textvariable=date_var, font=("Calibri", 10),
                 bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0, width=14,
                 insertbackground=ACCENT_BLUE).pack(ipady=6)

        if all_users:
            tk.Label(ctrl, text="Employee:", font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(side="left", padx=(0, 4))
            names = ["All"] + [e["full_name"] for e in get_all_employees(include_inactive=True)]
            ef_var = tk.StringVar(value="All")
            eff = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            eff.pack(side="left", padx=(0, 12))
            ec = ttk.Combobox(eff, textvariable=ef_var, values=names, state="readonly",
                              font=("Calibri", 10), width=24)
            ec.pack(ipady=4)
            _style_combo(ec)
        else:
            ef_var = tk.StringVar(value="All")

        container = tk.Frame(parent, bg=PAGE_BG)
        container.pack(fill="both", expand=True)

        def _load():
            for w in container.winfo_children():
                w.destroy()
            try:
                fd = datetime.strptime(date_var.get(), "%Y-%m-%d").date() if date_var.get() else None
            except Exception:
                fd = None

            eid = None if all_users else emp_id
            rows = load_entries(emp_id=eid, filter_date=fd)

            if all_users and ef_var.get() != "All":
                rows = [r for r in rows if r.get("full_name", "").lower() == ef_var.get().lower()]

            if not rows:
                empty = _card(container); empty.pack(fill="x", pady=8)
                tk.Label(empty, text="No entries found.", font=("Calibri", 10),
                         bg=PANEL_BG, fg=TEXT_MID, pady=20).pack()
                return

            cols = (["Date", "Employee", "Co.Code", "Client", "Work Category",
                     "Op. Area", "Sub-Cat", "Start", "End", "Hrs", "Task", "Status"]
                    if all_users else
                    ["Date", "Co.Code", "Client", "Work Category", "Op. Area",
                     "Sub-Cat", "Start", "End", "Hrs", "Task", "Status"])

            style = ttk.Style(); style.theme_use("clam")
            style.configure("JAA.Treeview", background=PANEL_BG, fieldbackground=PANEL_BG,
                            foreground=TEXT_DARK, rowheight=26, font=("Calibri", 9))
            style.configure("JAA.Treeview.Heading", background=TBL_HDR_BG,
                            foreground=TBL_HDR_FG, font=("Calibri", 9, "bold"), relief="flat")
            style.map("JAA.Treeview", background=[("selected", "#C5DCEE")])

            tree_card = _card(container); tree_card.pack(fill="both", expand=True)
            tf = tk.Frame(tree_card, bg=PANEL_BG)
            tf.pack(fill="both", expand=True, padx=1, pady=1)
            tree = ttk.Treeview(tf, columns=cols, show="headings",
                                style="JAA.Treeview", selectmode="browse")
            vsb = ttk.Scrollbar(tf, orient="vertical",   command=tree.yview)
            hsb = ttk.Scrollbar(tf, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
            tree.pack(side="left", fill="both", expand=True)

            cw_map = {"Date":80,"Employee":140,"Co.Code":80,"Client":160,
                      "Work Category":150,"Op. Area":150,"Sub-Cat":120,
                      "Start":55,"End":55,"Hrs":45,"Task":200,"Status":90}
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=cw_map.get(c, 100), minwidth=50, anchor="w")

            total_hrs = 0
            for i, r in enumerate(rows):
                if all_users:
                    vals = (r.get("entry_date"), r.get("full_name"),
                            r.get("company_code"), r.get("client_name"),
                            r.get("work_category"), r.get("operational_area"),
                            r.get("sub_category"), r.get("start_time"),
                            r.get("end_time"), r.get("total_hrs"),
                            r.get("task_desc"), r.get("status"))
                else:
                    vals = (r.get("entry_date"), r.get("company_code"),
                            r.get("client_name"), r.get("work_category"),
                            r.get("operational_area"), r.get("sub_category"),
                            r.get("start_time"), r.get("end_time"),
                            r.get("total_hrs"), r.get("task_desc"), r.get("status"))
                tree.insert("", "end", values=vals,
                            tags=("odd" if i%2 else "even",))
                try:
                    total_hrs += float(r.get("total_hrs") or 0)
                except Exception:
                    pass

            tree.tag_configure("odd",  background=ROW_ODD)
            tree.tag_configure("even", background=ROW_EVEN)

            foot = tk.Frame(container, bg=PANEL_ALT, padx=16, pady=10,
                            highlightthickness=1, highlightbackground=BORDER_CLR)
            foot.pack(fill="x", pady=(4, 0))
            tk.Label(foot, text=f"✔  {len(rows)} entries  ·  {total_hrs:.1f} hrs total",
                     font=("Calibri", 9, "bold"), bg=PANEL_ALT, fg=SUCCESS).pack(side="left")

        tk.Button(ctrl, text="🔍  Refresh",
                  font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  command=_load).pack(side="left")
        tk.Button(ctrl, text="All Dates",
                  font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID, relief="flat",
                  cursor="hand2", padx=10, pady=6,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=lambda: [date_var.set(""), _load()]).pack(side="left", padx=6)

        if all_users:
            def _export():
                fp = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV", "*.csv"), ("All", "*.*")],
                    initialfile=f"JAA_Export_{date.today():%Y%m%d}.csv")
                if not fp:
                    return
                try:
                    fd = datetime.strptime(date_var.get(), "%Y-%m-%d").date() if date_var.get() else None
                except Exception:
                    fd = None
                eid = None if ef_var.get() == "All" else \
                    get_employee_by_name(ef_var.get())
                eid = eid["emp_id"] if isinstance(eid, dict) else None
                count = export_to_csv(fp, emp_id=eid, filter_date=fd)
                messagebox.showinfo("Exported ✔", f"{count} rows → {fp}")

            tk.Button(ctrl, text="📥  Export CSV",
                      font=("Calibri", 9), bg="#D2E1F1", fg=SUCCESS, relief="flat",
                      cursor="hand2", padx=10, pady=6,
                      highlightthickness=1, highlightbackground="#A8C4DC",
                      command=_export).pack(side="left", padx=6)
        _load()

    # ── DOCUMENTS TAB ────────────────────────────────────────────────────────────

    def _build_documents_tab(self, parent, emp: dict):
        self._topbar(parent, "My Documents", f"  T1 / T5 Forms — {emp['full_name']}")
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=24, pady=18)

        # T1/T5 is only applicable to Interns
        if emp.get("role", "").lower() != "intern":
            na = _card(outer); na.pack(fill="x")
            tk.Label(na, text="📄  T1 / T5 documents are only applicable to Interns.",
                     font=("Calibri", 11), bg=PANEL_BG, fg=TEXT_MID,
                     padx=24, pady=24, anchor="w").pack(fill="x")
            return

        doc_info = get_document_info(emp["emp_id"])

        def _doc_card(title, form_type, path_key, upload_fn):
            card = _card(outer); card.pack(fill="x", pady=(0, 14))
            _section_hdr(card, title)
            inner = tk.Frame(card, bg=PANEL_BG, padx=24, pady=16); inner.pack(fill="x")

            exists = doc_info.get(f"{form_type}_exists", False)
            path   = doc_info.get(f"{form_type}_path", "")
            status_color = SUCCESS if exists else TEXT_LIGHT
            status_text  = f"✔  Uploaded: {os.path.basename(path)}" if exists else "—  Not uploaded yet"
            tk.Label(inner, text=status_text, font=("Calibri", 10),
                     bg=PANEL_BG, fg=status_color).pack(anchor="w", pady=(0, 10))

            btn_row = tk.Frame(inner, bg=PANEL_BG); btn_row.pack(anchor="w")
            err_lbl = tk.Label(inner, text="", font=("Calibri", 9), bg=PANEL_BG, fg=DANGER)
            err_lbl.pack(anchor="w", pady=(6, 0))

            def do_upload(fn=upload_fn):
                fp = filedialog.askopenfilename(
                    title=f"Select {title} PDF",
                    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
                if not fp:
                    return
                ok, msg = fn(emp["emp_id"], fp)
                if ok:
                    messagebox.showinfo("Uploaded ✔", f"{title} uploaded successfully.")
                    self._build_documents_tab(parent, get_employee(emp["emp_id"]))
                else:
                    err_lbl.config(text=f"⚠  {msg}")

            def do_view():
                ok, msg = open_document(path)
                if not ok:
                    messagebox.showerror("Error", msg)

            tk.Button(btn_row, text=f"📤  Upload {form_type.upper()} PDF",
                      font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white",
                      relief="flat", cursor="hand2", padx=18, pady=8,
                      command=do_upload).pack(side="left")
            if exists:
                tk.Button(btn_row, text="👁  View",
                          font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_MID,
                          relief="flat", cursor="hand2", padx=14, pady=8,
                          highlightthickness=1, highlightbackground=BORDER_CLR,
                          command=do_view).pack(side="left", padx=8)

        _doc_card("T1 Form — Joining Document", "t1", "t1_path", upload_t1)
        _doc_card("T5 Form — Exit / Completion Document", "t5", "t5_path", upload_t5)

        if doc_info.get("t5_exists"):
            note = _card(outer); note.pack(fill="x")
            tk.Label(note, text="⚠  T5 uploaded — your timesheet access has been deactivated. "
                     "Historical records are still viewable.",
                     font=("Calibri", 10), bg="#FEF9C3", fg=WARNING,
                     padx=20, pady=14, anchor="w", wraplength=700).pack(fill="x")

    # ── CHANGE PIN TAB ───────────────────────────────────────────────────────────

    def _build_change_pin_tab(self, parent, emp_id: str):
        self._topbar(parent, "Change PIN")
        wrapper = tk.Frame(parent, bg=PAGE_BG)
        wrapper.pack(fill="both", expand=True)
        card = _card(wrapper, width=420)
        card.place(relx=0.5, rely=0.4, anchor="center", width=420)
        _section_hdr(card, "Update Your PIN", "🔑")
        frame = tk.Frame(card, bg=PANEL_BG, padx=28, pady=16); frame.pack(fill="x")
        frame.columnconfigure(1, weight=1)

        def field(label, r):
            tk.Label(frame, text=label, font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(8, 2))
            v = tk.StringVar()
            f = tk.Frame(frame, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            f.pack(fill="x")
            tk.Entry(f, textvariable=v, font=("Calibri", 11), show="●",
                     bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0, justify="center").pack(fill="x", ipady=8)
            return v

        v_old     = field("Current PIN", 0)
        v_new     = field("New 4-digit PIN", 1)
        v_confirm = field("Confirm New PIN", 2)
        err = tk.Label(frame, text="", font=("Calibri", 9), bg=PANEL_BG, fg=DANGER)
        err.pack(anchor="w", pady=(8, 0))

        def do_change():
            if not verify_pin(emp_id, v_old.get().strip()):
                err.config(text="⚠  Current PIN is incorrect.", fg=DANGER); return
            new_pin = v_new.get().strip()
            if not re.fullmatch(r"\d{4}", new_pin):
                err.config(text="⚠  New PIN must be exactly 4 digits.", fg=DANGER); return
            if new_pin != v_confirm.get().strip():
                err.config(text="⚠  PINs do not match.", fg=DANGER); return
            change_pin(emp_id, new_pin)
            err.config(text="✔  PIN updated successfully.", fg=SUCCESS)

        tk.Button(frame, text="Update PIN", font=("Calibri", 10, "bold"),
                  bg=ACCENT_BLUE, fg="white", relief="flat", cursor="hand2", pady=10,
                  command=do_change).pack(fill="x", pady=(12, 0))

    # ═══════════════════════════════════════════════════════════════════════════
    #  ADMIN VIEW
    # ═══════════════════════════════════════════════════════════════════════════

    def _show_admin(self):
        self._clear()
        self.configure(bg=PAGE_BG)
        nav_btns = {}

        sb = self._sidebar(self, "Admin / Manager Panel")

        # ── helper: section label injected into the scrollable nav inner frame ──
        def _nav_section(label_text):
            nav_parent = getattr(self, "_sb_nav_inner", sb)
            lf = tk.Frame(nav_parent, bg=SIDEBAR_BG)
            lf.pack(fill="x", pady=(10, 2))
            # accent line
            tk.Frame(lf, bg="#03A3D0", height=1).pack(fill="x", padx=14, pady=(0, 3))
            self._lbl_section = tk.Label(
                lf, text=label_text.upper(),
                font=("Calibri", 7, "bold"),
                bg=SIDEBAR_BG, fg="#7AAFD4", anchor="w", padx=16)
            if self._sb_expanded:
                self._lbl_section.pack(fill="x")
            # track for collapse/expand
            self._sb_nav_items.append((lf, tk.Label(lf, bg=SIDEBAR_BG), self._lbl_section))

        main_body = tk.Frame(self, bg=PAGE_BG)
        main_body.pack(side="left", fill="both", expand=True)
        topbar = self._topbar(main_body, "Admin Panel")
        tk.Label(topbar, text="MANAGER", font=("Calibri", 8, "bold"),
                 bg="#03A3D0", fg="white", padx=10, pady=4).pack(side="left", padx=4)

        content_area = tk.Frame(main_body, bg=PAGE_BG)
        content_area.pack(fill="both", expand=True)

        def switch(tab_id):
            self._set_active_nav(nav_btns, tab_id)
            for w in content_area.winfo_children():
                w.destroy()
            dispatch = {
                "dashboard":    lambda: self._build_admin_dashboard(content_area),
                "entries":      lambda: self._entries_table(content_area, all_users=True),
                "attendance":   lambda: self._build_attendance_log(content_area),
                "approval":     lambda: self._build_approval_manager(content_area),
                "reports":      lambda: self._build_reports_tab(content_area),
                "staff":        lambda: self._build_staff_manager(content_area),
                "leaves":       lambda: self._build_leave_manager(content_area),
                "companies":    lambda: self._build_company_manager(content_area),
                "work_map":     lambda: self._build_work_mapping_manager(content_area),
                "birthdays":    lambda: self._build_birthdays_tab(content_area),
                "learning_hub": lambda: self._build_learning_hub(content_area, None, is_admin=True),
                "settings":     lambda: self._build_settings_tab(content_area),
            }
            dispatch.get(tab_id, lambda: None)()

        # ── Overview ─────────────────────────────────────────────────────────────
        _nav_section("Overview")
        self._nav_btn(sb, "Dashboard",      "📊", "dashboard",  nav_btns, switch, True)

        # ── Reporting ────────────────────────────────────────────────────────────
        _nav_section("Reporting")
        self._nav_btn(sb, "All Entries",    "📋", "entries",    nav_btns, switch)
        self._nav_btn(sb, "Attendance Log", "🕐", "attendance", nav_btns, switch)
        self._nav_btn(sb, "Approvals",      "✅", "approval",   nav_btns, switch)
        self._nav_btn(sb, "Reports",        "📈", "reports",    nav_btns, switch)

        # ── Administration ───────────────────────────────────────────────────────
        _nav_section("Administration")
        self._nav_btn(sb, "Manage Staff",   "👥", "staff",      nav_btns, switch)
        self._nav_btn(sb, "Manage Leaves",  "🏖", "leaves",     nav_btns, switch)
        self._nav_btn(sb, "Birthdays",      "🎂", "birthdays",  nav_btns, switch)
        self._nav_btn(sb, "Companies",      "🏢", "companies",  nav_btns, switch)
        self._nav_btn(sb, "Work Mapping",   "⚙",  "work_map",  nav_btns, switch)

        # ── Tools ────────────────────────────────────────────────────────────────
        _nav_section("Tools")
        self._nav_btn(sb, "Learning Hub",   "🎓", "learning_hub", nav_btns, switch)
        self._nav_btn(sb, "Settings",       "🔧", "settings",     nav_btns, switch)

        tk.Frame(self._sb_bottom, bg=SIDEBAR_SEL, height=1).pack(fill="x", padx=10, pady=8)
        back = tk.Frame(self._sb_bottom, bg=SIDEBAR_BG, cursor="hand2")
        back.pack(fill="x", pady=(0, 10))
        tk.Label(back, text="⟵", font=("Calibri", 12),
                 bg=SIDEBAR_BG, fg=SIDEBAR_ICN, pady=12, padx=8).pack(side="left")
        back_txt = tk.Label(back, text="Back to Login", font=("Calibri", 10),
                 bg=SIDEBAR_BG, fg=SIDEBAR_ICN, pady=12, anchor="w")
        if self._sb_expanded:
            back_txt.pack(side="left")
        self._sb_nav_items.append((back, back.winfo_children()[0], back_txt))
        for w in [back] + list(back.winfo_children()):
            w.bind("<Button-1>", lambda e: self._show_login())

        switch("dashboard")

    # ── STAFF MANAGER ────────────────────────────────────────────────────────────

    def _build_staff_manager(self, parent):
        # ── Top bar with Export button ──────────────────────────────────────────
        tb = self._topbar(parent, "Manage Staff", "  Add / Edit / Remove employees")

        def export_staff():
            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx"),("All","*.*")],
                initialfile=f"JAA_Staff_{date.today():%Y%m%d}.xlsx")
            if not fp: return
            try:
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Staff"
                hf    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                hfill = PatternFill("solid", fgColor="06355E")
                thin  = Side(style="thin", color="CCCCCC")
                bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
                hdrs  = ["EMP ID","Full Name","Role","Status","Join Date","T1 Uploaded","T5 Uploaded"]
                for ci, h in enumerate(hdrs, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = hf; c.fill = hfill; c.border = bdr
                    c.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 22
                of = PatternFill("solid", fgColor="D2E1F1")
                ef = PatternFill("solid", fgColor="FFFFFF")
                for ri, e in enumerate(sorted(get_all_employees(include_inactive=True),
                                              key=lambda x: x["full_name"]), 2):
                    fill = of if ri % 2 == 0 else ef
                    vals = [e["emp_id"], e["full_name"], e["role"], e["status"],
                            e.get("join_date",""),
                            "Yes" if e.get("t1_path") else "No",
                            "Yes" if e.get("t5_path") else "No"]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=ri, column=ci, value=v)
                        c.fill = fill; c.border = bdr
                        c.font = Font(name="Calibri", size=9)
                        c.alignment = Alignment(horizontal="left", vertical="center")
                for col in ws.columns:
                    mx = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4, 50)
                ws.freeze_panes = "A2"
                wb.save(fp)
                messagebox.showinfo("Exported ✔", f"Staff list saved to:\n{fp}")
            except Exception as ex:
                messagebox.showerror("Export Error", str(ex))

        tk.Button(tb, text="📥  Export .xlsx",
                  font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  activebackground="#155C38",
                  command=export_staff).pack(side="right", padx=16, pady=10)

        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=14)
        body = tk.Frame(outer, bg=PAGE_BG)
        body.pack(fill="both", expand=True)

        # ── Left: employee list ─────────────────────────────────────────────────
        left = _card(body)
        left.pack(side="left", fill="both", expand=True, padx=(0, 12))
        tk.Label(left, text="  Current Staff", font=("Calibri", 10, "bold"),
                 bg=TBL_HDR_BG, fg=TBL_HDR_FG, anchor="w", padx=6, pady=10).pack(fill="x")

        # Lifecycle note
        from datetime import timedelta
        cutoff_display = (date.today() - timedelta(days=INTERN_PERIOD_DAYS)).strftime("%d %b %Y")
        tk.Label(left,
                 text=(f"ℹ  Interns joined on/before {cutoff_display} are auto-marked Completed (15-month rule). "
                       f"Use '📅 Extend Internship' to override for individual interns."),
                 font=("Calibri", 8), bg=PANEL_ALT, fg=TEXT_MID,
                 padx=10, pady=5, anchor="w", wraplength=700).pack(fill="x")

        flt = tk.Frame(left, bg=PANEL_BG, padx=10, pady=8); flt.pack(fill="x")
        tk.Label(flt, text="🔍", font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_LIGHT).pack(side="left")
        fv = tk.StringVar()
        ff = tk.Frame(flt, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        ff.pack(side="left", fill="x", expand=True, padx=(4,0))
        tk.Entry(ff, textvariable=fv, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="flat", bd=0, insertbackground=ACCENT_BLUE).pack(fill="x", ipady=5)

        col_hdr = tk.Frame(left, bg=SIDEBAR_SEL); col_hdr.pack(fill="x")
        for txt, w in [("EMP ID",20),("Full Name",28),("Role",16),("Status",12)]:
            tk.Label(col_hdr, text=txt, font=("Calibri", 8, "bold"), width=w,
                     bg=SIDEBAR_SEL, fg=TBL_HDR_FG, anchor="w",
                     padx=6, pady=5).pack(side="left")

        lf = tk.Frame(left, bg=PANEL_BG); lf.pack(fill="both", expand=True, padx=6, pady=(2,8))
        vsb = ttk.Scrollbar(lf); vsb.pack(side="right", fill="y")
        listbox = tk.Listbox(lf, font=("Courier", 9), bg=PANEL_BG, fg=TEXT_DARK,
                             selectbackground=ACCENT_BLUE, selectforeground="white",
                             relief="flat", bd=0, highlightthickness=0,
                             yscrollcommand=vsb.set, activestyle="none", height=18)
        listbox.pack(side="left", fill="both", expand=True)
        vsb.config(command=listbox.yview)

        _lb_emp_ids = []   # parallel list — index matches listbox index exactly

        def refresh_list(q=""):
            listbox.delete(0, "end")
            _lb_emp_ids.clear()
            emps = get_all_employees(include_inactive=True)
            filtered = [e for e in sorted(emps, key=lambda x: x["full_name"])
                        if not q or q.lower() in e["full_name"].lower()
                        or q.lower() in e["emp_id"].lower()]
            for i, e in enumerate(filtered):
                dot = "🟢" if e["status"] == "Active" else "🔴"
                is_intern = e["role"].lower() == "intern"
                t1_tag = " 📄T1" if (is_intern and e.get("t1_path")) else ""
                t5_tag = " 📄T5" if (is_intern and e.get("t5_path")) else ""

                # For interns: show end date + days left / overdue
                if e["role"].lower() == "intern":
                    end_dt = get_intern_end_date(e)
                    if end_dt:
                        days = (end_dt - date.today()).days
                        ext_tag = " 📅" if e.get("internship_end_date") else ""
                        if days > 0:
                            tenure = f"  ⏳{days}d left{ext_tag}"
                        else:
                            tenure = f"  ⚠ Expired {abs(days)}d ago"
                    else:
                        tenure = ""
                else:
                    tenure = ""

                line = (f"  {e['emp_id']:<18}{e['full_name']:<26}"
                        f"{e['role']:<14}{dot} {e['status']}{t1_tag}{t5_tag}{tenure}")
                listbox.insert("end", line)
                listbox.itemconfig(i, bg=ROW_ODD if i % 2 == 0 else ROW_EVEN)
                _lb_emp_ids.append(e["emp_id"])

        fv.trace_add("write", lambda *_: refresh_list(fv.get()))
        refresh_list()

        # ── Right: add/edit form ────────────────────────────────────────────────
        right = _card(body, width=340)
        right.pack(side="left", fill="y"); right.pack_propagate(False)

        editing = {"emp_id": None}
        panel_title = tk.Label(right, text="  ➕  Add New Employee",
                               font=("Calibri", 10, "bold"),
                               bg=TBL_HDR_BG, fg=TBL_HDR_FG, anchor="w", padx=6, pady=10)
        panel_title.pack(fill="x")

        # ── Scrollable canvas for the form ──────────────────────────────────────
        r_canvas = tk.Canvas(right, bg=PANEL_BG, highlightthickness=0)
        r_vsb = ttk.Scrollbar(right, orient="vertical", command=r_canvas.yview)
        r_canvas.configure(yscrollcommand=r_vsb.set)
        r_vsb.pack(side="right", fill="y")
        r_canvas.pack(side="left", fill="both", expand=True)

        form = tk.Frame(r_canvas, bg=PANEL_BG, padx=18, pady=14)
        r_canvas_win = r_canvas.create_window((0, 0), window=form, anchor="nw")

        def _on_form_cfg(e):
            r_canvas.configure(scrollregion=r_canvas.bbox("all"))
        def _on_rcanvas_cfg(e):
            r_canvas.itemconfig(r_canvas_win, width=e.width)
        form.bind("<Configure>", _on_form_cfg)
        r_canvas.bind("<Configure>", _on_rcanvas_cfg)

        def _form_scroll(e):
            r_canvas.yview_scroll(-1 * (e.delta // 120), "units")
        r_canvas.bind("<MouseWheel>", _form_scroll)
        form.bind("<MouseWheel>", _form_scroll)

        def _lbl(t):
            tk.Label(form, text=t, font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(10, 2))
        def _ef(var):
            f = tk.Frame(form, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            f.pack(fill="x")
            e = tk.Entry(f, textvariable=var, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                         relief="flat", bd=0, insertbackground=FOCUS_CLR)
            e.pack(fill="x", ipady=7)
            return e

        _lbl("EMP ID  *  (alphanumeric  e.g. JAA00000055)")
        id_var = tk.StringVar(); _ef(id_var)

        _lbl("Full Name  *")
        name_var = tk.StringVar(); _ef(name_var)

        _lbl("Role  *")
        role_var = tk.StringVar(value="Intern")
        rf = tk.Frame(form, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR); rf.pack(fill="x")
        rc = ttk.Combobox(rf, textvariable=role_var,
                          values=["Partner","Associate","Employee","Admin","Intern"],
                          state="readonly", font=("Calibri", 10)); rc.pack(fill="x", ipady=4)
        _style_combo(rc)

        _lbl("Status")
        status_var = tk.StringVar(value="Active")
        stf = tk.Frame(form, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR); stf.pack(fill="x")
        stc = ttk.Combobox(stf, textvariable=status_var,
                           values=["Active","Inactive","Completed"],
                           state="readonly", font=("Calibri", 10)); stc.pack(fill="x", ipady=4)
        _style_combo(stc)

        _lbl("Join Date  (DD-MM-YYYY)")
        jd_var = tk.StringVar(); _ef(jd_var)

        def _parse_join_date(raw: str) -> str:
            """Accept DD-MM-YYYY or YYYY-MM-DD, always store as YYYY-MM-DD."""
            raw = raw.strip()
            if not raw:
                return ""
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return raw  # return as-is if unparseable (validation will catch it)

        status_lbl = tk.Label(form, text="", font=("Calibri", 9),
                              bg=PANEL_BG, fg=SUCCESS, wraplength=290)
        status_lbl.pack(anchor="w", pady=(10, 4))

        # T1 / T5 document status display (read-only, shown when editing)
        doc_frame = tk.Frame(form, bg=PANEL_BG)
        doc_frame.pack(fill="x", pady=(0, 4))
        tk.Label(doc_frame, text="Documents:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(0, 4))
        t1_lbl = tk.Label(doc_frame, text="T1: —", font=("Calibri", 9),
                          bg=PANEL_BG, fg=TEXT_LIGHT, anchor="w", wraplength=280)
        t1_lbl.pack(fill="x")
        t2_lbl = tk.Label(doc_frame, text="T5: —", font=("Calibri", 9),
                          bg=PANEL_BG, fg=TEXT_LIGHT, anchor="w", wraplength=280)
        t2_lbl.pack(fill="x")
        doc_frame.pack_forget()   # hidden until an employee is selected

        def on_select(event=None):
            sel = listbox.curselection()
            if not sel: return
            eid = _lb_emp_ids[sel[0]]
            emp = get_employee(eid)
            if not emp: return
            editing["emp_id"] = eid
            id_var.set(emp["emp_id"]); name_var.set(emp["full_name"])
            role_var.set(emp["role"]); status_var.set(emp["status"])
            # Display stored YYYY-MM-DD as DD-MM-YYYY for user friendliness
            raw_jd = emp.get("join_date") or ""
            try:
                display_jd = datetime.strptime(raw_jd, "%Y-%m-%d").strftime("%d-%m-%Y") if raw_jd else ""
            except ValueError:
                display_jd = raw_jd
            jd_var.set(display_jd)
            panel_title.config(text=f"  ✏  {emp['full_name'][:22]}")
            btn_add.config(text="💾  Save Changes", bg=WARNING,
                           activebackground="#6A5200", command=save_edit)
            btn_cancel.pack(fill="x", pady=(6, 0))
            status_lbl.config(text="Edit fields then Save.", fg=TEXT_MID)
            # Show T1/T5 document status — Interns only
            if emp.get("role", "").lower() == "intern":
                t1_path = emp.get("t1_path") or ""
                t5_path = emp.get("t5_path") or ""
                t5_date = emp.get("t5_upload_date") or ""
                t1_lbl.config(
                    text=f"T1: ✔  {os.path.basename(t1_path)}" if t1_path and os.path.exists(t1_path)
                         else ("T1: ✔  Uploaded (file moved)" if t1_path else "T1: —  Not uploaded"),
                    fg=SUCCESS if t1_path else TEXT_LIGHT)
                t2_lbl.config(
                    text=f"T5: ✔  {os.path.basename(t5_path)}" + (f"  ({t5_date})" if t5_date else "")
                         if t5_path and os.path.exists(t5_path)
                         else ("T5: ✔  Uploaded (file moved)" if t5_path else "T5: —  Not uploaded"),
                    fg=SUCCESS if t5_path else TEXT_LIGHT)
                doc_frame.pack(fill="x", pady=(0, 4), before=status_lbl)
            else:
                doc_frame.pack_forget()

        listbox.bind("<<ListboxSelect>>", on_select)

        def add_emp():
            eid = id_var.get().strip().upper()
            nm  = name_var.get().strip()
            if not eid or not nm:
                status_lbl.config(text="⚠  EMP ID and Name required.", fg=DANGER)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0)); return
            if not re.fullmatch(r"[A-Za-z0-9\-]{3,30}", eid):
                status_lbl.config(text="⚠  EMP ID: letters/digits/hyphens, 3-30 chars.\ne.g. JAA00000055", fg=DANGER)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0)); return
            try:
                ok = add_employee(eid, nm, role_var.get(), join_date=_parse_join_date(jd_var.get()) or None)
                if ok:
                    refresh_list(fv.get()); reset_form()
                    status_lbl.config(text=f"✔  Added: {nm}  [{eid}]", fg=SUCCESS)
                else:
                    status_lbl.config(text=f"⚠  EMP ID '{eid}' already exists.", fg=WARNING)
            except Exception as ex:
                status_lbl.config(text=f"⚠  Error: {ex}", fg=DANGER)
            r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))

        def save_edit():
            old_eid  = editing["emp_id"]
            new_eid  = id_var.get().strip().upper()
            nm       = name_var.get().strip()

            if not old_eid:
                status_lbl.config(text="⚠  No employee selected. Click a name in the list first.", fg=DANGER)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))
                return
            if not new_eid or not nm:
                status_lbl.config(text="⚠  EMP ID and Name are required.", fg=DANGER)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))
                return
            if not re.fullmatch(r"[A-Za-z0-9\-]{3,30}", new_eid):
                status_lbl.config(
                    text="⚠  EMP ID: letters/digits/hyphens only, 3-30 chars.",
                    fg=DANGER)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))
                return

            # Check if new EMP ID conflicts with another employee
            if new_eid != old_eid.strip().upper():
                existing = get_employee(new_eid)
                if existing:
                    status_lbl.config(
                        text=f"⚠  EMP ID '{new_eid}' already belongs to another employee.",
                        fg=DANGER)
                    r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))
                    return

            try:
                result = update_employee(
                    old_eid,
                    new_emp_id=new_eid,
                    full_name=nm,
                    role=role_var.get(),
                    status=status_var.get(),
                    join_date=_parse_join_date(jd_var.get()) or None,
                )
                if result:
                    # Update editing state to new EMP ID in case user saves again
                    editing["emp_id"] = new_eid
                    refresh_list(fv.get())
                    status_lbl.config(
                        text=f"✔  Saved successfully:  {nm}  [{new_eid}]",
                        fg=SUCCESS)
                else:
                    status_lbl.config(
                        text="⚠  Nothing was changed. Check fields and try again.",
                        fg=WARNING)
            except Exception as ex:
                status_lbl.config(text=f"⚠  Save failed: {ex}", fg=DANGER)

            # Scroll the panel down so the status message is always visible
            r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))

        def remove_selected():
            sel = listbox.curselection()
            if not sel:
                status_lbl.config(text="⚠  Select an employee first.", fg=WARNING)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0)); return
            eid = _lb_emp_ids[sel[0]]
            emp = get_employee(eid)
            nm  = emp["full_name"] if emp else eid
            if messagebox.askyesno("Confirm Remove",
                    f"Remove '{nm}'?\n\nThis will also delete all their timesheet entries, "
                    f"attendance records and leave records.\nThis cannot be undone."):
                try:
                    with get_conn() as conn:
                        conn.execute("DELETE FROM timesheet_entries WHERE emp_id=?", (eid,))
                        conn.execute("DELETE FROM attendance_log    WHERE emp_id=?", (eid,))
                        conn.execute("DELETE FROM leave_records     WHERE emp_id=?", (eid,))
                        conn.execute("DELETE FROM admin_pins        WHERE key=?",    (eid.lower(),))
                        conn.execute("DELETE FROM employees         WHERE emp_id=?", (eid,))
                    refresh_list(fv.get())
                    reset_form()
                    status_lbl.config(text=f"✔  Removed '{nm}'.", fg=SUCCESS)
                except Exception as ex:
                    status_lbl.config(text=f"⚠  Remove failed: {ex}", fg=DANGER)
                r_canvas.after_idle(lambda: r_canvas.yview_moveto(1.0))

        def reset_form():
            editing["emp_id"] = None
            for v in (id_var, name_var, jd_var): v.set("")
            role_var.set("Intern"); status_var.set("Active")
            panel_title.config(text="  ➕  Add New Employee")
            btn_add.config(text="➕  Add Employee", bg=ACCENT_BLUE,
                           activebackground="#042D50", command=add_emp)
            btn_cancel.pack_forget()
            doc_frame.pack_forget()
            status_lbl.config(text="", fg=SUCCESS)
            listbox.selection_clear(0, "end")

        btn_add = tk.Button(form, text="➕  Add Employee",
                            font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white",
                            relief="flat", cursor="hand2", pady=9,
                            activebackground="#042D50", command=add_emp)
        btn_add.pack(fill="x", pady=(4, 0))
        btn_cancel = tk.Button(form, text="✕  Cancel",
                               font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                               relief="flat", cursor="hand2", pady=6,
                               highlightthickness=1, highlightbackground=BORDER_CLR,
                               command=reset_form)
        tk.Frame(form, bg=BORDER_CLR, height=1).pack(fill="x", pady=10)
        tk.Button(form, text="🗑  Remove Selected",
                  font=("Calibri", 10), bg="#EEE8E0", fg=DANGER,
                  relief="flat", cursor="hand2", pady=8,
                  command=remove_selected).pack(fill="x")

        tk.Frame(form, bg=BORDER_CLR, height=1).pack(fill="x", pady=10)

        def extend_internship():
            """Open dialog to set a custom internship end date for the selected intern."""
            sel = listbox.curselection()
            if not sel:
                status_lbl.config(text="⚠  Select an intern to extend.", fg=WARNING)
                return
            eid = _lb_emp_ids[sel[0]]
            emp = get_employee(eid)
            if not emp:
                return
            if emp["role"].lower() != "intern":
                status_lbl.config(
                    text="⚠  Only Interns can have their internship extended.",
                    fg=WARNING)
                return

            # ── Extension dialog ────────────────────────────────────────────────
            win = tk.Toplevel(parent)
            win.title(f"Extend Internship — {emp['full_name']}")
            win.geometry("460x380")
            win.configure(bg=PAGE_BG)
            win.resizable(False, False)
            win.grab_set()
            win.update_idletasks()

            tk.Frame(win, bg=ACCENT_GOLD, height=4).pack(fill="x")
            hdr = tk.Frame(win, bg=SIDEBAR_BG, pady=14)
            hdr.pack(fill="x")
            tk.Label(hdr, text="Extend Internship Period",
                     font=("Georgia", 13, "bold"),
                     bg=SIDEBAR_BG, fg=SIDEBAR_FG, padx=20).pack(anchor="w")
            tk.Label(hdr, text=f"{emp['full_name']}  ·  {emp['emp_id']}",
                     font=("Calibri", 9),
                     bg=SIDEBAR_BG, fg=SIDEBAR_ICN, padx=20).pack(anchor="w")

            frame = tk.Frame(win, bg=PAGE_BG, padx=28, pady=20)
            frame.pack(fill="both", expand=True)

            # Show current / default end date
            from datetime import timedelta
            eff_end = get_intern_end_date(emp)
            default_end_str = eff_end.strftime("%d %b %Y") if eff_end else "Not set"
            has_extension = bool(emp.get("internship_end_date"))

            info_bg = "#EBF5FB" if not has_extension else "#FFFBEB"
            info_frame = tk.Frame(frame, bg=info_bg,
                                  highlightthickness=1,
                                  highlightbackground=BORDER_CLR)
            info_frame.pack(fill="x", pady=(0, 16))

            tk.Label(info_frame,
                     text=("📅  Current effective end date: " + default_end_str +
                           ("\n🔖  (Extension already set)" if has_extension else
                            "\n📌  (Default: Join date + 15 months)")),
                     font=("Calibri", 9), bg=info_bg, fg=TEXT_MID,
                     padx=14, pady=10, justify="left", anchor="w").pack(fill="x")

            tk.Label(frame, text="New Internship End Date  (YYYY-MM-DD)  *",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG,
                     fg=TEXT_MID).pack(anchor="w", pady=(0, 4))

            end_var = tk.StringVar(
                value=emp.get("internship_end_date") or
                      (eff_end.isoformat() if eff_end else ""))
            ef = tk.Frame(frame, bg=PAGE_BG, highlightthickness=1,
                          highlightbackground=BORDER_CLR)
            ef.pack(fill="x")
            end_entry = tk.Entry(ef, textvariable=end_var,
                                 font=("Calibri", 12, "bold"),
                                 bg=ENTRY_BG, fg=ACCENT_BLUE,
                                 relief="flat", bd=0, justify="center",
                                 insertbackground=FOCUS_CLR)
            end_entry.pack(fill="x", ipady=10)
            end_entry.focus_set()

            tk.Label(frame, text="Reason / Note for extension  (optional)",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG,
                     fg=TEXT_MID).pack(anchor="w", pady=(12, 4))

            note_var = tk.StringVar(value=emp.get("extension_note") or "")
            nf = tk.Frame(frame, bg=PAGE_BG, highlightthickness=1,
                          highlightbackground=BORDER_CLR)
            nf.pack(fill="x")
            tk.Entry(nf, textvariable=note_var, font=("Calibri", 10),
                     bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                     insertbackground=FOCUS_CLR).pack(fill="x", ipady=7)

            err_lbl = tk.Label(frame, text="", font=("Calibri", 9),
                               bg=PAGE_BG, fg=DANGER)
            err_lbl.pack(anchor="w", pady=(8, 0))

            def save_extension():
                new_end = end_var.get().strip()
                note    = note_var.get().strip()
                if not new_end:
                    err_lbl.config(text="⚠  Please enter a new end date."); return
                try:
                    new_end_dt = datetime.strptime(new_end, "%Y-%m-%d").date()
                except ValueError:
                    err_lbl.config(text="⚠  Invalid date. Use YYYY-MM-DD."); return
                if new_end_dt <= date.today():
                    err_lbl.config(
                        text="⚠  New end date must be in the future."); return

                # If intern was Completed by the old rule, re-activate them
                if emp["status"] == "Completed":
                    update_employee(eid, status="Active")

                update_employee(eid,
                                internship_end_date=new_end,
                                extension_note=note)
                win.destroy()
                refresh_list(fv.get())
                days_left = (new_end_dt - date.today()).days
                status_lbl.config(
                    text=f"✔  Extended: {emp['full_name']} active until "
                         f"{new_end_dt.strftime('%d %b %Y')}  "
                         f"({days_left} days remaining)",
                    fg=SUCCESS)

            def clear_extension():
                """Remove extension — revert to default 15-month rule."""
                update_employee(eid,
                                internship_end_date=None,
                                extension_note=None)
                win.destroy()
                refresh_list(fv.get())
                status_lbl.config(
                    text=f"✔  Extension cleared. Default 15-month rule applies.",
                    fg=WARNING)

            btn_row = tk.Frame(frame, bg=PAGE_BG)
            btn_row.pack(fill="x", pady=(12, 0))

            tk.Button(btn_row, text="✔  Save Extension",
                      font=("Calibri", 10, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", pady=9,
                      activebackground="#155C38",
                      command=save_extension).pack(side="left", fill="x",
                                                   expand=True, padx=(0, 6))
            if has_extension:
                tk.Button(btn_row, text="✕  Clear Extension",
                          font=("Calibri", 9), bg="#EEE8E0", fg=WARNING,
                          relief="flat", cursor="hand2", pady=9,
                          command=clear_extension).pack(side="left",
                                                        padx=(0, 0))

        tk.Button(form, text="📅  Extend Internship",
                  font=("Calibri", 10, "bold"),
                  bg="#EBF5FB", fg=ACCENT_BLUE,
                  relief="flat", cursor="hand2", pady=9,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  activebackground="#D2E1F1",
                  command=extend_internship).pack(fill="x")

        # Bind mousewheel on every child widget inside the scrollable form
        def _bind_scroll(widget):
            widget.bind("<MouseWheel>", _form_scroll, add="+")
            for child in widget.winfo_children():
                _bind_scroll(child)
        form.after(100, lambda: _bind_scroll(form))

    def _build_attendance_log(self, parent):
        tb = self._topbar(parent, "Attendance Log", "  Manual login/logout records per employee")

        def _export_att():
            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
                initialfile=f"JAA_Attendance_{date.today():%Y%m%d}.xlsx")
            if not fp: return
            try:
                logs = get_attendance_log(limit=5000)
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Attendance"
                hf    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                hfill = PatternFill("solid", fgColor="06355E")
                thin  = Side(style="thin", color="CCCCCC")
                bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
                hdrs  = ["ID","Employee ID","Login Time","Logout Time","Duration (mins)","Date"]
                for ci, h in enumerate(hdrs, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font=hf; c.fill=hfill; c.border=bdr
                    c.alignment=Alignment(horizontal="center",vertical="center")
                of = PatternFill("solid", fgColor="D2E1F1")
                ef = PatternFill("solid", fgColor="FFFFFF")
                date_fmt = "DD-MMM-YYYY HH:MM"
                for ri, r in enumerate(logs, 2):
                    fill = of if ri%2==0 else ef
                    vals = [r["id"], r["emp_id"], r["login_dt"],
                            r.get("logout_dt",""), r.get("duration_mins",""), r["session_date"]]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=ri, column=ci, value=v)
                        c.fill=fill; c.border=bdr
                        c.font=Font(name="Calibri", size=9)
                        c.alignment=Alignment(horizontal="center", vertical="center")
                for col in ws.columns:
                    mx = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4,40)
                ws.freeze_panes = "A2"
                wb.save(fp)
                messagebox.showinfo("Exported ✔", f"Attendance log saved to:\n{fp}")
            except Exception as ex:
                messagebox.showerror("Export Error", str(ex))

        tk.Button(tb, text="📥  Export .xlsx",
                  font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  activebackground="#155C38",
                  command=_export_att).pack(side="right", padx=16, pady=10)

        # Info note
        note = tk.Frame(parent, bg="#EBF5FB", highlightthickness=1,
                        highlightbackground="#A8C4DC")
        note.pack(fill="x", padx=20, pady=(10,0))
        tk.Label(note,
                 text="ℹ  Attendance is recorded when employees manually enter Login/Logout times in their Timesheet form and save.",
                 font=("Calibri", 9), bg="#EBF5FB", fg=TEXT_MID,
                 padx=14, pady=8, anchor="w", wraplength=900).pack(fill="x")
        ctrl_card = _card(parent); ctrl_card.pack(fill="x", padx=20, pady=(16, 8))
        ctrl = tk.Frame(ctrl_card, bg=PANEL_BG, padx=16, pady=12); ctrl.pack(fill="x")

        tk.Label(ctrl, text="Date:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        date_var = tk.StringVar(value=date.today().isoformat())
        df = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        df.pack(side="left", padx=(4, 12))
        tk.Entry(df, textvariable=date_var, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="flat", bd=0, width=14).pack(ipady=6)

        container = tk.Frame(parent, bg=PAGE_BG)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        def _load():
            for w in container.winfo_children():
                w.destroy()
            try:
                fd = datetime.strptime(date_var.get(), "%Y-%m-%d").date()
            except Exception:
                fd = None
            logs = get_attendance_log(filter_date=fd)
            if not logs:
                emp = _card(container); emp.pack(fill="x")
                tk.Label(emp, text="No attendance records found.", font=("Calibri", 10),
                         bg=PANEL_BG, fg=TEXT_MID, pady=20).pack(); return

            cols = ["ID", "Employee ID", "Login Time", "Logout Time", "Duration (mins)", "Date"]
            style = ttk.Style(); style.theme_use("clam")
            style.configure("JAA.Treeview", background=PANEL_BG, fieldbackground=PANEL_BG,
                            foreground=TEXT_DARK, rowheight=26, font=("Calibri", 9))
            style.configure("JAA.Treeview.Heading", background=TBL_HDR_BG,
                            foreground=TBL_HDR_FG, font=("Calibri", 9, "bold"), relief="flat")

            tc = _card(container); tc.pack(fill="both", expand=True)
            tf = tk.Frame(tc, bg=PANEL_BG); tf.pack(fill="both", expand=True, padx=1, pady=1)
            tree = ttk.Treeview(tf, columns=cols, show="headings",
                                style="JAA.Treeview", selectmode="browse")
            vsb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y"); tree.pack(side="left", fill="both", expand=True)
            cw = {"ID":50,"Employee ID":120,"Login Time":160,"Logout Time":160,"Duration (mins)":130,"Date":100}
            for c in cols:
                tree.heading(c, text=c); tree.column(c, width=cw.get(c, 120), anchor="center")
            for i, r in enumerate(logs):
                tree.insert("", "end",
                            values=(r["id"], r["emp_id"], r["login_dt"],
                                    r.get("logout_dt") or "— Active —",
                                    r.get("duration_mins") or "", r["session_date"]),
                            tags=("odd" if i%2 else "even",))
            tree.tag_configure("odd",  background=ROW_ODD)
            tree.tag_configure("even", background=ROW_EVEN)

        tk.Button(ctrl, text="🔍  Load", font=("Calibri", 9, "bold"),
                  bg=ACCENT_BLUE, fg="white", relief="flat", cursor="hand2",
                  padx=14, pady=6, command=_load).pack(side="left")
        tk.Button(ctrl, text="Today", font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=10, pady=6,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=lambda: [date_var.set(date.today().isoformat()), _load()]).pack(side="left", padx=6)
        _load()

    # ── LEAVE MANAGER ────────────────────────────────────────────────────────────

    def _build_leave_manager(self, parent):
        LEAVE_TYPES = ["Casual", "Sick", "Earned", "Compensatory", "Other"]
        STATUS_CLR  = {"Pending": "#7C3AED", "Approved": "#1A6B45", "Rejected": "#B52A2A"}

        tb = self._topbar(parent, "Manage Leaves",
                          "  Approve / reject employee requests · Add leaves directly")

        # ── Tab strip: Pending Requests | All Records | Add Leave ────────────────
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=(10, 16))

        tab_bar = tk.Frame(outer, bg=PAGE_BG)
        tab_bar.pack(fill="x", pady=(0, 12))

        content_area = tk.Frame(outer, bg=PAGE_BG)
        content_area.pack(fill="both", expand=True)

        tab_btns = {}
        active_tab = {"id": "pending"}

        def _tab_btn(label, tab_id):
            btn = tk.Label(tab_bar, text=label,
                           font=("Calibri", 10, "bold"),
                           padx=18, pady=7, cursor="hand2")
            btn.pack(side="left", padx=(0, 4))
            tab_btns[tab_id] = btn
            btn.bind("<Button-1>", lambda e, t=tab_id: _switch_tab(t))

        def _switch_tab(tab_id):
            active_tab["id"] = tab_id
            for tid, btn in tab_btns.items():
                if tid == tab_id:
                    btn.config(bg="#7C3AED", fg="white")
                else:
                    btn.config(bg=PANEL_BG, fg=TEXT_MID,
                               highlightthickness=1,
                               highlightbackground=BORDER_CLR)
            for w in content_area.winfo_children():
                w.destroy()
            {"pending": _build_pending,
             "all":     _build_all,
             "add":     _build_add}[tab_id]()

        _tab_btn("🕐  Pending Requests", "pending")
        _tab_btn("📋  All Records",      "all")
        _tab_btn("➕  Add Leave",        "add")

        # ── PENDING REQUESTS ─────────────────────────────────────────────────────
        def _build_pending():
            leaves = get_leaves(status_filter="Pending")

            scroll_card = tk.Frame(content_area, bg=PANEL_BG,
                                   highlightthickness=1, highlightbackground=BORDER_CLR)
            scroll_card.pack(fill="both", expand=True)
            tk.Frame(scroll_card, bg="#7C3AED", height=4).pack(fill="x")

            if not leaves:
                tk.Label(scroll_card,
                         text="✅  No pending leave requests.",
                         font=("Calibri", 11), bg=PANEL_BG, fg=SUCCESS,
                         pady=32).pack()
                return

            # column headers
            hdr = tk.Frame(scroll_card, bg=TBL_HDR_BG)
            hdr.pack(fill="x", padx=12, pady=(8, 0))
            for txt, w in [("Employee", 18), ("Type", 11), ("From", 13),
                           ("To", 13), ("Days", 6), ("Reason", 22)]:
                tk.Label(hdr, text=txt, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=w, anchor="w", padx=6, pady=5).pack(side="left")
            tk.Label(hdr, text="Action", font=("Calibri", 8, "bold"),
                     bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                     padx=6, pady=5).pack(side="left")

            rows_f = tk.Frame(scroll_card, bg=PANEL_BG)
            rows_f.pack(fill="both", expand=True, padx=12, pady=(0, 12))

            note_var = tk.StringVar()

            for i, lv in enumerate(leaves):
                row_bg = ROW_ODD if i % 2 == 0 else ROW_EVEN
                try:
                    s = datetime.strptime(lv["start_date"], "%Y-%m-%d").date()
                    e = datetime.strptime(lv["end_date"],   "%Y-%m-%d").date()
                    days = str((e - s).days + 1)
                except Exception:
                    days = "—"

                rf = tk.Frame(rows_f, bg=row_bg)
                rf.pack(fill="x", pady=(0, 1))

                for txt, w in [
                    (lv.get("full_name", lv["emp_id"]), 18),
                    (lv.get("leave_type", "Casual"),     11),
                    (lv["start_date"],                   13),
                    (lv["end_date"],                     13),
                    (days,                                6),
                    (lv.get("reason","") or "—",         22),
                ]:
                    tk.Label(rf, text=str(txt), font=("Calibri", 9),
                             bg=row_bg, fg=TEXT_DARK, width=w, anchor="w",
                             padx=6, pady=6).pack(side="left")

                # approve / reject buttons
                def _approve(lid=lv["id"], nm=lv.get("full_name","")):
                    review_leave(lid, "Approved")
                    _build_pending()

                def _reject(lid=lv["id"], nm=lv.get("full_name","")):
                    # small inline note entry
                    def _confirm_reject():
                        review_leave(lid, "Rejected",
                                     note=note_popup_var.get().strip())
                        popup.destroy()
                        _build_pending()

                    popup = tk.Toplevel(self)
                    popup.title("Reject Leave")
                    popup.geometry("360x180")
                    popup.configure(bg=PAGE_BG)
                    popup.grab_set()
                    popup.resizable(False, False)
                    tk.Frame(popup, bg=DANGER, height=4).pack(fill="x")
                    tk.Label(popup, text=f"Reject leave for {nm}?",
                             font=("Georgia", 12, "bold"),
                             bg=PAGE_BG, fg=TEXT_DARK, pady=12).pack()
                    tk.Label(popup, text="Optional note to employee:",
                             font=("Calibri", 9), bg=PAGE_BG,
                             fg=TEXT_MID).pack()
                    note_popup_var = tk.StringVar()
                    nw = tk.Frame(popup, bg=BORDER_CLR, padx=1, pady=1,
                                  padx_=16)
                    nw = tk.Frame(popup, bg=BORDER_CLR, padx=1, pady=1)
                    nw.pack(fill="x", padx=24, pady=6)
                    tk.Entry(nw, textvariable=note_popup_var,
                             font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                             relief="flat", bd=0).pack(fill="x", ipady=6, padx=4)
                    btn_row = tk.Frame(popup, bg=PAGE_BG)
                    btn_row.pack(pady=8)
                    tk.Button(btn_row, text="Confirm Reject",
                              font=("Calibri", 10, "bold"),
                              bg=DANGER, fg="white", relief="flat",
                              cursor="hand2", padx=14, pady=7,
                              command=_confirm_reject).pack(side="left", padx=6)
                    tk.Button(btn_row, text="Cancel",
                              font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_MID,
                              relief="flat", cursor="hand2", padx=14, pady=7,
                              highlightthickness=1, highlightbackground=BORDER_CLR,
                              command=popup.destroy).pack(side="left")

                ap_btn = tk.Label(rf, text="✔ Approve",
                                  font=("Calibri", 8, "bold"),
                                  bg="#D1FAE5", fg="#1A6B45",
                                  padx=8, pady=4, cursor="hand2")
                ap_btn.pack(side="left", padx=(4, 2))
                ap_btn.bind("<Button-1>", lambda e, fn=_approve: fn())

                rj_btn = tk.Label(rf, text="✕ Reject",
                                  font=("Calibri", 8, "bold"),
                                  bg="#FEE2E2", fg="#B52A2A",
                                  padx=8, pady=4, cursor="hand2")
                rj_btn.pack(side="left", padx=(0, 4))
                rj_btn.bind("<Button-1>", lambda e, fn=_reject: fn())

        # ── ALL RECORDS ──────────────────────────────────────────────────────────
        def _build_all():
            ctrl_row = tk.Frame(content_area, bg=PANEL_BG,
                                highlightthickness=1, highlightbackground=BORDER_CLR)
            ctrl_row.pack(fill="x", pady=(0, 10))
            ctrl = tk.Frame(ctrl_row, bg=PANEL_BG, padx=14, pady=10)
            ctrl.pack(fill="x")

            tk.Label(ctrl, text="Filter by status:", font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
            fv = tk.StringVar(value="All")
            fw = tk.Frame(ctrl, bg=BORDER_CLR, padx=1, pady=1)
            fw.pack(side="left", padx=(6, 12))
            fc = ttk.Combobox(fw, textvariable=fv,
                              values=["All", "Pending", "Approved", "Rejected"],
                              state="readonly", font=("Calibri", 10), width=14)
            fc.pack(ipady=4); _style_combo(fc)

            list_card = tk.Frame(content_area, bg=PANEL_BG,
                                 highlightthickness=1, highlightbackground=BORDER_CLR)
            list_card.pack(fill="both", expand=True)

            def _load_all():
                for w in list_card.winfo_children():
                    w.destroy()
                sf = fv.get()
                leaves = get_leaves(status_filter=sf if sf != "All" else None)

                if not leaves:
                    tk.Label(list_card, text="No leave records found.",
                             font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_MID,
                             pady=20).pack()
                    return

                hdr = tk.Frame(list_card, bg=TBL_HDR_BG)
                hdr.pack(fill="x", padx=12, pady=(8, 0))
                for txt, w in [("Employee", 18), ("Type", 11), ("From", 13),
                               ("To", 13), ("Days", 6), ("Reason", 18),
                               ("Status", 11), ("Note", 18)]:
                    tk.Label(hdr, text=txt, font=("Calibri", 8, "bold"),
                             bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                             width=w, anchor="w", padx=6, pady=5).pack(side="left")

                rows_f = tk.Frame(list_card, bg=PANEL_BG)
                rows_f.pack(fill="both", expand=True, padx=12, pady=(0, 10))
                for i, lv in enumerate(leaves):
                    row_bg = ROW_ODD if i % 2 == 0 else ROW_EVEN
                    st = lv.get("status", "—")
                    try:
                        s = datetime.strptime(lv["start_date"], "%Y-%m-%d").date()
                        e = datetime.strptime(lv["end_date"],   "%Y-%m-%d").date()
                        days = str((e - s).days + 1)
                    except Exception:
                        days = "—"
                    rf = tk.Frame(rows_f, bg=row_bg); rf.pack(fill="x", pady=(0,1))
                    for txt, w, fg in [
                        (lv.get("full_name", lv["emp_id"]), 18, TEXT_DARK),
                        (lv.get("leave_type","Casual"),     11, TEXT_DARK),
                        (lv["start_date"],                  13, TEXT_DARK),
                        (lv["end_date"],                    13, TEXT_DARK),
                        (days,                               6, TEXT_DARK),
                        (lv.get("reason","") or "—",        18, TEXT_DARK),
                        (st,                                11, STATUS_CLR.get(st, TEXT_DARK)),
                        (lv.get("admin_note","") or "—",    18, TEXT_MID),
                    ]:
                        tk.Label(rf, text=str(txt), font=("Calibri", 9),
                                 bg=row_bg, fg=fg, width=w, anchor="w",
                                 padx=6, pady=4).pack(side="left")

                    # delete button
                    del_btn = tk.Label(rf, text="🗑",
                                       font=("Calibri", 9), bg=row_bg,
                                       fg=DANGER, cursor="hand2", padx=6)
                    del_btn.pack(side="left")
                    del_btn.bind("<Button-1>",
                                 lambda e, lid=lv["id"]: [
                                     remove_leave(lid), _load_all()])

            tk.Button(ctrl, text="↺  Load",
                      font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                      relief="flat", cursor="hand2", padx=12, pady=6,
                      command=_load_all).pack(side="left")
            _load_all()

        # ── ADD LEAVE (admin direct-add) ─────────────────────────────────────────
        def _build_add():
            add_card = tk.Frame(content_area, bg=PANEL_BG,
                                highlightthickness=1, highlightbackground=BORDER_CLR,
                                width=380)
            add_card.pack(side="left", fill="y")
            add_card.pack_propagate(False)

            tk.Frame(add_card, bg=ACCENT_BLUE, height=4).pack(fill="x")
            tk.Label(add_card, text="Add Leave Directly",
                     font=("Calibri", 11, "bold"),
                     bg=PANEL_BG, fg=TEXT_DARK, padx=18, pady=12,
                     anchor="w").pack(fill="x")
            tk.Frame(add_card, bg=BORDER_CLR, height=1).pack(fill="x")

            form = tk.Frame(add_card, bg=PANEL_BG, padx=18, pady=14)
            form.pack(fill="x")

            def _lbl(t):
                tk.Label(form, text=t, font=("Calibri", 9, "bold"),
                         bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(8, 2))

            def _ef(var, w=28):
                wrap = tk.Frame(form, bg=BORDER_CLR, padx=1, pady=1)
                wrap.pack(fill="x")
                tk.Entry(wrap, textvariable=var, font=("Calibri", 10),
                         bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                         width=w, insertbackground=FOCUS_CLR).pack(
                         fill="x", ipady=6, padx=4)

            names   = [e["full_name"] for e in get_all_employees()]
            emp_var = tk.StringVar()
            _lbl("Employee *")
            ew = tk.Frame(form, bg=BORDER_CLR, padx=1, pady=1); ew.pack(fill="x")
            ec = ttk.Combobox(ew, textvariable=emp_var, values=sorted(names),
                              state="readonly", font=("Calibri", 10), width=26)
            ec.pack(fill="x", ipady=4, padx=2); _style_combo(ec)

            type_var = tk.StringVar(value="Casual")
            _lbl("Leave Type")
            tw = tk.Frame(form, bg=BORDER_CLR, padx=1, pady=1); tw.pack(fill="x")
            tc = ttk.Combobox(tw, textvariable=type_var, values=LEAVE_TYPES,
                              state="readonly", font=("Calibri", 10), width=26)
            tc.pack(fill="x", ipady=4, padx=2); _style_combo(tc)

            start_var  = tk.StringVar()
            end_var    = tk.StringVar()
            reason_var = tk.StringVar()
            _lbl("From (YYYY-MM-DD) *"); _ef(start_var)
            _lbl("To (YYYY-MM-DD) *");   _ef(end_var)
            _lbl("Reason");              _ef(reason_var)

            msg_lbl = tk.Label(form, text="", font=("Calibri", 9),
                               bg=PANEL_BG, wraplength=320)
            msg_lbl.pack(anchor="w", pady=(10, 0))

            def _do_add():
                nm = emp_var.get().strip()
                if not nm:
                    msg_lbl.config(text="⚠  Select an employee.", fg=DANGER); return
                emp = get_employee_by_name(nm)
                if not emp:
                    msg_lbl.config(text="⚠  Employee not found.", fg=DANGER); return
                try:
                    s = datetime.strptime(start_var.get().strip(), "%Y-%m-%d").date()
                    e = datetime.strptime(end_var.get().strip(),   "%Y-%m-%d").date()
                    if e < s: raise ValueError
                except ValueError:
                    msg_lbl.config(text="⚠  Invalid dates. Use YYYY-MM-DD.", fg=DANGER)
                    return
                add_leave(emp["emp_id"],
                          start_var.get().strip(), end_var.get().strip(),
                          reason_var.get().strip(), type_var.get(),
                          approved_by="Admin", status="Approved")
                start_var.set(""); end_var.set(""); reason_var.set("")
                msg_lbl.config(text=f"✔  Leave added for {nm}", fg=SUCCESS)

            tk.Frame(form, bg=BORDER_CLR, height=1).pack(fill="x", pady=(14, 0))
            sub = tk.Label(form, text="➕  Add Leave",
                           font=("Calibri", 10, "bold"),
                           bg=ACCENT_BLUE, fg="white",
                           padx=16, pady=9, cursor="hand2", anchor="center")
            sub.pack(fill="x", pady=(10, 0))
            sub.bind("<Button-1>", lambda e: _do_add())

            tk.Label(content_area,
                     text="ℹ  Leaves added here are immediately Approved.\n"
                          "Employees can also request leaves from their own panel.",
                     font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                     justify="left").pack(side="left", padx=18, anchor="n", pady=4)

        _switch_tab("pending")

    # ── COMPANY MANAGER ──────────────────────────────────────────────────────────

    def _build_company_manager(self, parent):
        tb = self._topbar(parent, "Company Directory")

        def _export_co():
            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
                initialfile=f"JAA_Companies_{date.today():%Y%m%d}.xlsx")
            if not fp: return
            try:
                cos = get_all_companies()
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Companies"
                hf    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                hfill = PatternFill("solid", fgColor="06355E")
                thin  = Side(style="thin", color="CCCCCC")
                bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
                for ci, h in enumerate(["Short Code","Company Name","Unique Code"], 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font=hf; c.fill=hfill; c.border=bdr
                    c.alignment=Alignment(horizontal="center",vertical="center")
                of = PatternFill("solid", fgColor="D2E1F1")
                ef = PatternFill("solid", fgColor="FFFFFF")
                for ri, co in enumerate(cos, 2):
                    fill = of if ri%2==0 else ef
                    for ci, v in enumerate([co["short_code"],co["full_name"],co["unique_code"]],1):
                        c = ws.cell(row=ri, column=ci, value=v)
                        c.fill=fill; c.border=bdr
                        c.font=Font(name="Calibri", size=9)
                        c.alignment=Alignment(horizontal="left" if ci==2 else "center",vertical="center")
                for col in ws.columns:
                    mx = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4,60)
                ws.freeze_panes = "A2"
                wb.save(fp)
                messagebox.showinfo("Exported ✔", f"Companies saved to:\n{fp}")
            except Exception as ex:
                messagebox.showerror("Export Error", str(ex))

        tk.Button(tb, text="📥  Export .xlsx",
                  font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  activebackground="#155C38",
                  command=_export_co).pack(side="right", padx=16, pady=10)
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=16)
        body = tk.Frame(outer, bg=PAGE_BG); body.pack(fill="both", expand=True)

        left = _card(body); left.pack(side="left", fill="both", expand=True, padx=(0, 12))
        tk.Label(left, text="  Companies", font=("Calibri", 10, "bold"),
                 bg=TBL_HDR_BG, fg=TBL_HDR_FG, anchor="w", padx=4, pady=10).pack(fill="x")

        sv = tk.StringVar()
        sf = tk.Frame(left, bg=PANEL_BG, padx=8, pady=6); sf.pack(fill="x")
        sff = tk.Frame(sf, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR); sff.pack(fill="x")
        tk.Entry(sff, textvariable=sv, font=("Calibri", 9), bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="flat", bd=0).pack(fill="x", ipady=5)

        lf = tk.Frame(left, bg=PANEL_BG); lf.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        vsb = ttk.Scrollbar(lf); vsb.pack(side="right", fill="y")
        listbox = tk.Listbox(lf, font=("Courier", 9), bg=PANEL_BG, fg=TEXT_DARK,
                             selectbackground=ACCENT_BLUE, selectforeground="white",
                             relief="flat", bd=0, highlightthickness=0,
                             yscrollcommand=vsb.set, activestyle="none", height=20)
        listbox.pack(side="left", fill="both", expand=True)
        vsb.config(command=listbox.yview)

        def refresh_list(q=""):
            listbox.delete(0, "end")
            companies = search_companies(q, limit=200) if q else get_all_companies()
            for i, c in enumerate(companies):
                listbox.insert("end", f"  {c['short_code']:<8} {c['unique_code']:<13} {c['full_name']}")
                listbox.itemconfig(listbox.size()-1, bg=ROW_ODD if i % 2 == 0 else ROW_EVEN)

        sv.trace_add("write", lambda *_: refresh_list(sv.get()))
        refresh_list()

        right = _card(body, width=340); right.pack(side="left", fill="y"); right.pack_propagate(False)
        editing_sc = {"code": None}
        panel_title = tk.Label(right, text="  Add / Lookup Company",
                               font=("Calibri", 10, "bold"), bg=TBL_HDR_BG,
                               fg=TBL_HDR_FG, anchor="w", padx=4, pady=10)
        panel_title.pack(fill="x")
        form = tk.Frame(right, bg=PANEL_BG, padx=16, pady=14); form.pack(fill="x")

        def _lbl(t): tk.Label(form, text=t, font=("Calibri", 9, "bold"),
                               bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(8, 2))
        def _ef(var):
            f = tk.Frame(form, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR); f.pack(fill="x")
            tk.Entry(f, textvariable=var, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                     relief="flat", bd=0, insertbackground=ACCENT_BLUE).pack(fill="x", ipady=7)

        _lbl("Short Code"); sc_var = tk.StringVar(); _ef(sc_var)
        _lbl("Full Company Name"); nm_var = tk.StringVar(); _ef(nm_var)
        _lbl("Unique Code (auto if blank)"); cd_var = tk.StringVar(); _ef(cd_var)

        status_lbl = tk.Label(form, text="", font=("Calibri", 9), bg=PANEL_BG, fg=SUCCESS, wraplength=300)
        status_lbl.pack(anchor="w", pady=(8, 4))

        def on_sc_change(*_):
            sc = sc_var.get().strip().upper()
            if len(sc) >= 3:
                co = get_company_by_shortcode(sc)
                if co:
                    nm_var.set(co["full_name"]); cd_var.set(co["unique_code"])
                    status_lbl.config(text=f"✔  Found: {co['unique_code']}", fg=SUCCESS)
                else:
                    cd_var.set(f"{sc}{next_company_number():03d}")
                    status_lbl.config(text="New company — enter name.", fg=TEXT_MID)
            else:
                cd_var.set(""); status_lbl.config(text="", fg=TEXT_MID)

        sc_var.trace_add("write", on_sc_change)

        def do_add_company():
            sc = sc_var.get().strip().upper()
            nm = nm_var.get().strip()
            cd = cd_var.get().strip().upper() or f"{sc}{next_company_number():03d}"
            if not sc or not nm:
                status_lbl.config(text="⚠  Short code and name required.", fg=DANGER); return
            ok = add_company(sc, nm, cd)
            if ok:
                refresh_list(sv.get()); reset_co()
                status_lbl.config(text=f"✔  Added: {cd}", fg=SUCCESS)
            else:
                status_lbl.config(text=f"⚠  Code '{sc}' or '{cd}' already exists.", fg=WARNING)

        def remove_co():
            sel = listbox.curselection()
            if not sel:
                status_lbl.config(text="⚠  Select a company.", fg=WARNING); return
            raw = listbox.get(sel[0]).strip().split()
            if not raw: return
            cd = raw[1]  # unique_code is second token
            remove_company(cd)
            refresh_list(sv.get())
            status_lbl.config(text="✔  Removed.", fg=SUCCESS)

        def on_list_select(event=None):
            sel = listbox.curselection()
            if not sel: return
            raw = listbox.get(sel[0]).strip().split()
            if not raw: return
            sc = raw[0]
            co = get_company_by_shortcode(sc)
            if co:
                sc_var.set(co["short_code"]); nm_var.set(co["full_name"]); cd_var.set(co["unique_code"])
                editing_sc["code"] = co["unique_code"]
                panel_title.config(text=f"  ✏  Editing: {co['short_code']}")
                btn_add.config(text="💾  Save", bg=WARNING, command=save_co_edit)

        listbox.bind("<<ListboxSelect>>", on_list_select)

        def save_co_edit():
            update_company(editing_sc["code"], short_code=sc_var.get().strip(),
                              full_name=nm_var.get().strip())
            refresh_list(sv.get()); reset_co()
            status_lbl.config(text="✔  Updated.", fg=SUCCESS)

        def reset_co():
            editing_sc["code"] = None
            sc_var.set(""); nm_var.set(""); cd_var.set("")
            panel_title.config(text="  Add / Lookup Company")
            btn_add.config(text="➕  Add Company", bg=ACCENT_BLUE, command=do_add_company)
            listbox.selection_clear(0, "end")

        def export_co():
            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
                initialfile=f"JAA_Companies_{date.today():%Y%m%d}.xlsx")
            if not fp: return
            try:
                import openpyxl as ox
                from openpyxl.styles import Font as F, PatternFill as PF, Alignment as A, Border as B, Side as S
                from openpyxl.utils import get_column_letter as gcl
                wb = ox.Workbook(); ws = wb.active; ws.title = "Companies"
                hf = F(name="Calibri", bold=True, color="FFFFFF", size=10)
                hfill = PF("solid", fgColor="06355E")
                thin = S(style="thin", color="CCCCCC")
                border = B(left=thin, right=thin, top=thin, bottom=thin)
                for ci, (h, wd) in enumerate(zip(["Short Code","Company Name","Unique Code"],[14,50,16]),1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = hf; c.fill = hfill; c.border = border
                    c.alignment = A(horizontal="center"); ws.column_dimensions[gcl(ci)].width = wd
                of = PF("solid", fgColor="D2E1F1"); ef = PF("solid", fgColor="FFFFFF")
                for ri, co in enumerate(get_all_companies(), 2):
                    fill = of if ri%2==0 else ef
                    for ci, v in enumerate([co["short_code"], co["full_name"], co["unique_code"]], 1):
                        c = ws.cell(row=ri, column=ci, value=v)
                        c.fill = fill; c.border = border; c.font = F(name="Calibri", size=9)
                        c.alignment = A(horizontal="center" if ci in (1,3) else "left")
                ws.freeze_panes = "A2"; wb.save(fp)
                messagebox.showinfo("Exported ✔", f"Saved to:\n{fp}")
            except Exception as ex:
                messagebox.showerror("Error", str(ex))

        btn_add = tk.Button(form, text="➕  Add Company",
                            font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white",
                            relief="flat", cursor="hand2", pady=8, command=do_add_company)
        btn_add.pack(fill="x", pady=(4, 0))
        tk.Frame(form, bg=BORDER_CLR, height=1).pack(fill="x", pady=8)
        tk.Button(form, text="🗑  Remove Selected",
                  font=("Calibri", 10), bg="#EBF5FB", fg=DANGER, relief="flat",
                  cursor="hand2", pady=7, command=remove_co).pack(fill="x")
        tk.Button(form, text="↺  Reset Form",
                  font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID, relief="flat",
                  cursor="hand2", pady=6, highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=reset_co).pack(fill="x", pady=(4, 0))
        tk.Frame(form, bg=BORDER_CLR, height=1).pack(fill="x", pady=8)
        tk.Button(form, text="📥  Export to Excel",
                  font=("Calibri", 10, "bold"), bg="#D2E1F1", fg=SUCCESS, relief="flat",
                  cursor="hand2", pady=8, highlightthickness=1,
                  highlightbackground="#A8C4DC", command=export_co).pack(fill="x")

    # ── WORK MAPPING MANAGER ─────────────────────────────────────────────────────

    def _build_work_mapping_manager(self, parent):
        """
        Work Category Mapping Manager — Round 18 complete rewrite.
        Layout:
          LEFT  — Tree preview (click to select & load into right panel)
          RIGHT — Split: top = Detail/Edit panel for selected item
                         bottom = Tab-style CRUD manager (Add rows)
        All operations use ID-based SQLite helpers. Active/Inactive toggle
        controls Timesheet dropdown visibility without deleting history.
        """
        tb = self._topbar(parent, "Work Category Mapping",
                     "  Category → Area → Sub-Category  |  Select item in tree to edit")

        # ── Export button on topbar ──────────────────────────────────────────
        def _export_wm():
            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
                initialfile=f"JAA_WorkMapping_{date.today():%Y%m%d}.xlsx")
            if not fp: return
            try:
                mapping = get_full_mapping()
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Work Mapping"
                hf    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                hfill = PatternFill("solid", fgColor="06355E")
                thin  = Side(style="thin", color="CCCCCC")
                bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
                for ci, h in enumerate(["Work Category","Operational Area","Sub-Category"], 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font=hf; c.fill=hfill; c.border=bdr
                    c.alignment=Alignment(horizontal="center",vertical="center")
                fills = [PatternFill("solid", fgColor=x) for x in
                         ["EBF2FA","F0F8FF","FFFCF0","F5FFF5","FFF0F5","F5F0FF"]]
                ri = 2
                for ci_map, (cat, areas) in enumerate(mapping.items()):
                    fill = fills[ci_map % len(fills)]
                    if not areas:
                        for ci2, v in enumerate([cat,"",""], 1):
                            c = ws.cell(row=ri, column=ci2, value=v)
                            c.fill=fill; c.border=bdr; c.font=Font(name="Calibri",size=9)
                        ri += 1; continue
                    for area, subs in areas.items():
                        for sub in (subs or [""]):
                            for ci2, v in enumerate([cat, area, sub or ""], 1):
                                c = ws.cell(row=ri, column=ci2, value=v)
                                c.fill=fill; c.border=bdr; c.font=Font(name="Calibri",size=9)
                            ri += 1
                for col in ws.columns:
                    mx = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4,50)
                ws.freeze_panes = "A2"
                wb.save(fp)
                messagebox.showinfo("Exported \u2714", f"Work Mapping saved:\n{fp}")
            except Exception as ex:
                messagebox.showerror("Export Error", str(ex))

        tk.Button(tb, text="\u2795  Export .xlsx",
                  font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  command=_export_wm).pack(side="right", padx=16, pady=10)

        # ═══════════════════════════════════════════════════════════════════════
        # MAIN 2-COLUMN LAYOUT
        # ═══════════════════════════════════════════════════════════════════════
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=14)

        # ─── LEFT PANEL: Tree Preview ────────────────────────────────────────
        left = _card(outer)
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))

        lhdr = tk.Frame(left, bg=TBL_HDR_BG, padx=12, pady=8)
        lhdr.pack(fill="x")
        lhdr_lbl = tk.Label(lhdr, text="\U0001f4cb  Mapping Preview",
                            font=("Calibri", 10, "bold"), bg=TBL_HDR_BG,
                            fg=TBL_HDR_FG)
        lhdr_lbl.pack(side="left")
        tk.Label(lhdr, text="Click any item to edit \u2192",
                 font=("Calibri", 8), bg=TBL_HDR_BG,
                 fg="#A0C8E8").pack(side="right")

        # Treeview style
        _wm_style = ttk.Style()
        _wm_style.theme_use("clam")
        _wm_style.configure("WM.Treeview", background=PANEL_BG,
                            fieldbackground=PANEL_BG, foreground=TEXT_DARK,
                            rowheight=24, font=("Calibri", 9))
        _wm_style.configure("WM.Treeview.Heading", background=TBL_HDR_BG,
                            foreground=TBL_HDR_FG, font=("Calibri", 9, "bold"), relief="flat")
        _wm_style.map("WM.Treeview",
                      background=[("selected", ACCENT_BLUE)],
                      foreground=[("selected", "white")])

        tree_f = tk.Frame(left, bg=PANEL_BG)
        tree_f.pack(fill="both", expand=True, padx=6, pady=6)
        tree_vsb = ttk.Scrollbar(tree_f, orient="vertical")
        tree_vsb.pack(side="right", fill="y")
        tree = ttk.Treeview(tree_f, yscrollcommand=tree_vsb.set,
                            selectmode="browse", show="tree",
                            style="WM.Treeview")
        tree.pack(side="left", fill="both", expand=True)
        tree_vsb.config(command=tree.yview)

        # ── node_meta: maps tree iid → {level, id, name, parent_id, ...} ───
        node_meta = {}   # iid -> dict with DB info

        def refresh_tree():
            """Rebuild the full tree from DB, preserving expand state."""
            # Remember which cats are open
            open_cats = {iid for iid in tree.get_children()
                         if tree.item(iid, "open")}
            open_names = {node_meta[iid]["name"] for iid in open_cats
                          if iid in node_meta}

            tree.delete(*tree.get_children())
            node_meta.clear()

            with get_conn() as conn:
                cats = conn.execute(
                    "SELECT id, name, is_active FROM work_categories ORDER BY sort_order, name"
                ).fetchall()
                for cat in cats:
                    is_active = bool(cat["is_active"])
                    icon = "\U0001f4c1" if is_active else "\U0001f4c1"
                    disp = f"  {icon}  {cat['name']}" + ("" if is_active else "  [Inactive]")
                    cn = tree.insert("", "end", text=disp,
                                     open=(cat["name"] in open_names),
                                     tags=("cat", "inactive_cat" if not is_active else "cat_active"))
                    node_meta[cn] = {
                        "level": "cat", "id": cat["id"],
                        "name": cat["name"], "is_active": is_active
                    }
                    areas = conn.execute(
                        "SELECT id, name, is_active FROM operational_areas "
                        "WHERE category_id=? ORDER BY sort_order, name", (cat["id"],)
                    ).fetchall()
                    for area in areas:
                        a_active = bool(area["is_active"])
                        adisp = f"    \U0001f4c2  {area['name']}" + ("" if a_active else "  [Inactive]")
                        an = tree.insert(cn, "end", text=adisp,
                                         tags=("area", "inactive_area" if not a_active else "area_active"))
                        node_meta[an] = {
                            "level": "area", "id": area["id"],
                            "name": area["name"], "is_active": a_active,
                            "cat_id": cat["id"], "cat_name": cat["name"]
                        }
                        subs = conn.execute(
                            "SELECT id, name, is_active FROM sub_categories "
                            "WHERE area_id=? ORDER BY sort_order, name", (area["id"],)
                        ).fetchall()
                        for sub in subs:
                            s_active = bool(sub["is_active"])
                            sdisp = f"      \u25b8  {sub['name']}" + ("" if s_active else "  [Inactive]")
                            sn = tree.insert(an, "end", text=sdisp,
                                             tags=("sub", "inactive_sub" if not s_active else "sub_active"))
                            node_meta[sn] = {
                                "level": "sub", "id": sub["id"],
                                "name": sub["name"], "is_active": s_active,
                                "area_id": area["id"], "area_name": area["name"],
                                "cat_name": cat["name"]
                            }

            # Tag colours
            tree.tag_configure("cat_active",   foreground=ACCENT_BLUE)
            tree.tag_configure("cat",          foreground=ACCENT_BLUE)
            tree.tag_configure("inactive_cat", foreground=TEXT_LIGHT)
            tree.tag_configure("area_active",  foreground=TEXT_MID)
            tree.tag_configure("area",         foreground=TEXT_MID)
            tree.tag_configure("inactive_area",foreground=TEXT_LIGHT)
            tree.tag_configure("sub_active",   foreground=TEXT_DARK)
            tree.tag_configure("sub",          foreground=TEXT_DARK)
            tree.tag_configure("inactive_sub", foreground=TEXT_LIGHT)

        refresh_tree()

        # ═══════════════════════════════════════════════════════════════════════
        # RIGHT PANEL — fixed width, split vertically:
        #   top    → Detail / Edit panel (bound to tree selection)
        #   bottom → Tab CRUD manager (Add operations)
        # ═══════════════════════════════════════════════════════════════════════
        right = tk.Frame(outer, bg=PAGE_BG, width=470)
        right.pack(side="left", fill="both", expand=False)
        right.pack_propagate(False)

        # ── QUICK ADD ROW (always visible at top of right panel) ─────────────
        qadd_card = _card(right)
        qadd_card.pack(fill="x", pady=(0, 6))
        qa_hdr = tk.Frame(qadd_card, bg=ACCENT_BLUE, padx=12, pady=7)
        qa_hdr.pack(fill="x")
        tk.Label(qa_hdr, text="➕  Quick Add New Category",
                 font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white").pack(side="left")

        qa_body = tk.Frame(qadd_card, bg=PANEL_BG, padx=14, pady=10)
        qa_body.pack(fill="x")
        qa_row = tk.Frame(qa_body, bg=PANEL_BG); qa_row.pack(fill="x")
        tk.Label(qa_row, text="Category Name:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        qa_var = tk.StringVar()
        qa_ef = tk.Frame(qa_row, bg=ENTRY_BG, highlightthickness=1,
                         highlightbackground=BORDER_CLR)
        qa_ef.pack(side="left", fill="x", expand=True, padx=(6, 8))
        qa_entry = tk.Entry(qa_ef, textvariable=qa_var, font=("Calibri", 10),
                            bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0,
                            insertbackground=FOCUS_CLR)
        qa_entry.pack(fill="x", ipady=6, padx=6)
        qa_status = tk.Label(qa_body, text="", font=("Calibri", 8),
                             bg=PANEL_BG, fg=SUCCESS, anchor="w")
        qa_status.pack(fill="x", pady=(4, 0))

        # Deferred refresh ref — filled in after _full_refresh is defined below
        _qa_refresh_ref = {"fn": None}

        def _qa_add_cat(event=None):
            n = qa_var.get().strip()
            if not n:
                qa_status.config(text="⚠  Enter a category name.", fg=DANGER); return
            if add_work_category(n):
                qa_var.set("")
                qa_status.config(text=f"✔  Category '{n}' added.", fg=SUCCESS)
                if _qa_refresh_ref["fn"]:
                    _qa_refresh_ref["fn"]()
                else:
                    refresh_tree()
            else:
                qa_status.config(text=f"⚠  '{n}' already exists.", fg=DANGER)

        qa_entry.bind("<Return>", _qa_add_cat)
        tk.Button(qa_row, text="＋ Add",
                  font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=12, pady=6,
                  command=_qa_add_cat).pack(side="left")

        # ── DETAIL / EDIT PANEL (top of right) ──────────────────────────────
        detail_card = _card(right)
        detail_card.pack(fill="x", pady=(0, 8))

        det_hdr = tk.Frame(detail_card, bg=ACCENT_BLUE, padx=12, pady=8)
        det_hdr.pack(fill="x")
        det_hdr_lbl = tk.Label(det_hdr, text="\U0001f4dd  Selected Item",
                               font=("Calibri", 10, "bold"),
                               bg=ACCENT_BLUE, fg="white")
        det_hdr_lbl.pack(side="left")
        det_level_lbl = tk.Label(det_hdr, text="",
                                  font=("Calibri", 8), bg=ACCENT_BLUE, fg="#A0D0F0")
        det_level_lbl.pack(side="right")

        det_body = tk.Frame(detail_card, bg=PANEL_BG, padx=14, pady=10)
        det_body.pack(fill="x")

        # Breadcrumb label (shows Cat > Area > Sub)
        breadcrumb_lbl = tk.Label(det_body, text="No item selected",
                                   font=("Calibri", 8), bg=PANEL_BG,
                                   fg=TEXT_LIGHT, anchor="w")
        breadcrumb_lbl.pack(fill="x", pady=(0, 8))

        # Name field
        det_name_row = tk.Frame(det_body, bg=PANEL_BG)
        det_name_row.pack(fill="x", pady=(0, 6))
        tk.Label(det_name_row, text="Name:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID, width=7, anchor="w").pack(side="left")
        det_name_var = tk.StringVar()
        det_name_frame = tk.Frame(det_name_row, bg=PANEL_BG,
                                   highlightthickness=1, highlightbackground=BORDER_CLR)
        det_name_frame.pack(side="left", fill="x", expand=True, padx=(4, 0))
        det_name_entry = tk.Entry(det_name_frame, textvariable=det_name_var,
                                   font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                                   relief="flat", bd=0, insertbackground=FOCUS_CLR,
                                   state="disabled")
        det_name_entry.pack(fill="x", ipady=6, padx=6)

        # Status row
        det_status_row = tk.Frame(det_body, bg=PANEL_BG)
        det_status_row.pack(fill="x", pady=(0, 8))
        tk.Label(det_status_row, text="Status:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID, width=7, anchor="w").pack(side="left")
        det_status_lbl = tk.Label(det_status_row, text="—",
                                   font=("Calibri", 9, "bold"), bg=PANEL_BG,
                                   fg=TEXT_LIGHT)
        det_status_lbl.pack(side="left", padx=(4, 0))

        # Action buttons row
        det_btn_row = tk.Frame(det_body, bg=PANEL_BG)
        det_btn_row.pack(fill="x")

        det_save_btn   = tk.Button(det_btn_row, text="\U0001f4be  Save Name",
                                    font=("Calibri", 9, "bold"),
                                    bg=SUCCESS, fg="white",
                                    relief="flat", cursor="hand2", padx=12, pady=6,
                                    state="disabled")
        det_save_btn.pack(side="left", padx=(0, 6))

        det_toggle_btn = tk.Button(det_btn_row, text="\u23f8 Toggle Active",
                                    font=("Calibri", 9, "bold"),
                                    bg="#D2E1F1", fg=TEXT_MID,
                                    relief="flat", cursor="hand2", padx=12, pady=6,
                                    state="disabled")
        det_toggle_btn.pack(side="left", padx=(0, 6))

        det_delete_btn = tk.Button(det_btn_row, text="\U0001f5d1  Delete",
                                    font=("Calibri", 9, "bold"),
                                    bg=DANGER, fg="white",
                                    relief="flat", cursor="hand2", padx=12, pady=6,
                                    state="disabled")
        det_delete_btn.pack(side="left")

        # ── Contextual "Add Child" section (shown below action buttons) ─────
        det_child_frame = tk.Frame(det_body, bg=PANEL_BG)
        # (packed/unpacked dynamically by _load_detail)

        det_child_sep = tk.Frame(det_child_frame, bg=BORDER_CLR, height=1)
        det_child_sep.pack(fill="x", pady=(10, 6))

        det_child_lbl = tk.Label(det_child_frame, text="",
                                  font=("Calibri", 9, "bold"), bg=PANEL_BG, fg=ACCENT_BLUE,
                                  anchor="w")
        det_child_lbl.pack(fill="x")

        det_child_row = tk.Frame(det_child_frame, bg=PANEL_BG)
        det_child_row.pack(fill="x", pady=(4, 0))

        # For "cat" selected → add Area dropdown is the category (fixed) + area name entry
        # For "area" selected → add Sub entry only
        det_child_var1 = tk.StringVar()   # area combo (when adding sub from area node)
        det_child_var2 = tk.StringVar()   # new name entry

        det_child_ef = tk.Frame(det_child_row, bg=ENTRY_BG, highlightthickness=1,
                                 highlightbackground=BORDER_CLR)
        det_child_ef.pack(side="left", fill="x", expand=True, padx=(0, 8))
        det_child_entry = tk.Entry(det_child_ef, textvariable=det_child_var2,
                                    font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                                    relief="flat", bd=0, insertbackground=FOCUS_CLR)
        det_child_entry.pack(fill="x", ipady=6, padx=6)

        det_child_add_btn = tk.Button(det_child_row, text="\u2795 Add",
                                       font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                                       relief="flat", cursor="hand2", padx=12, pady=6)
        det_child_add_btn.pack(side="left")

        det_child_status = tk.Label(det_child_frame, text="", font=("Calibri", 8),
                                     bg=PANEL_BG, fg=SUCCESS, anchor="w")
        det_child_status.pack(fill="x", pady=(3, 0))

        # Status bar (bottom of detail card)
        det_status_bar = tk.Frame(detail_card, bg="#D2E1F1", height=28)
        det_status_bar.pack(fill="x")
        det_status_bar.pack_propagate(False)
        status_lbl = tk.Label(det_status_bar, text="  Click an item in the tree to edit",
                               font=("Calibri", 8), bg="#D2E1F1",
                               fg=TEXT_MID, anchor="w")
        status_lbl.pack(fill="x", padx=8, pady=5)

        def _status(msg, ok=True):
            status_lbl.config(text=f"  {msg}", fg=SUCCESS if ok else DANGER)

        # ── Current selection state (holds mutable reference) ────────────────
        _sel = {"meta": None}   # meta = dict from node_meta

        def _clear_detail():
            det_hdr_lbl.config(text="\U0001f4dd  Selected Item")
            det_level_lbl.config(text="")
            breadcrumb_lbl.config(text="No item selected", fg=TEXT_LIGHT)
            det_name_var.set("")
            det_name_entry.config(state="disabled", bg=PANEL_BG)
            det_name_frame.config(highlightbackground=BORDER_CLR)
            det_status_lbl.config(text="—", fg=TEXT_LIGHT)
            det_save_btn.config(state="disabled", command=lambda: None)
            det_toggle_btn.config(state="disabled", command=lambda: None, text="\u23f8 Toggle Active")
            det_delete_btn.config(state="disabled", command=lambda: None)
            det_child_frame.pack_forget()
            _sel["meta"] = None

        def _load_detail(meta: dict):
            """Populate the detail/edit panel from a node_meta dict."""
            _sel["meta"] = meta
            level   = meta["level"]
            name    = meta["name"]
            is_act  = meta["is_active"]
            item_id = meta["id"]

            # Header
            level_labels = {"cat": "Category", "area": "Area", "sub": "Sub-Category"}
            icons         = {"cat": "\U0001f4c1", "area": "\U0001f4c2", "sub": "\u25b8"}
            det_hdr_lbl.config(text=f"{icons[level]}  {name}")
            det_level_lbl.config(text=level_labels[level])

            # Breadcrumb
            if level == "cat":
                bc = name
            elif level == "area":
                bc = f"{meta.get('cat_name','')} \u203a {name}"
            else:
                bc = f"{meta.get('cat_name','')} \u203a {meta.get('area_name','')} \u203a {name}"
            breadcrumb_lbl.config(text=bc, fg=TEXT_MID)

            # Name entry
            det_name_var.set(name)
            det_name_entry.config(state="normal", bg=ENTRY_BG)
            det_name_frame.config(highlightbackground=FOCUS_CLR)

            # Status badge
            if is_act:
                det_status_lbl.config(text="\u25cf  Active", fg=SUCCESS)
            else:
                det_status_lbl.config(text="\u25cf  Inactive", fg=DANGER)

            # ── Save Name ──────────────────────────────────────────────────
            def _do_save():
                new_name = det_name_var.get().strip()
                if not new_name:
                    _status("\u26a0  Name cannot be empty.", False); return
                if level == "cat":
                    ok, msg = rename_work_category(name, new_name)
                elif level == "area":
                    ok, msg = rename_operational_area(item_id, new_name)
                else:
                    ok, msg = rename_sub_category(item_id, new_name)
                if ok:
                    _status(f"\u2714  {msg}")
                    _full_refresh()
                    # Re-select the renamed item by looking for new name in tree
                    _reselect_by_name(level, new_name,
                                      meta.get("cat_name"), meta.get("area_name"))
                else:
                    _status(f"\u26a0  {msg}", False)

            det_save_btn.config(state="normal", command=_do_save)

            # ── Toggle Active ──────────────────────────────────────────────
            tog_label = "\u23f8 Deactivate" if is_act else "\u25b6 Activate"
            tog_bg    = "#D2E1F1" if is_act else "#C8F0D8"

            def _do_toggle():
                if level == "cat":
                    toggle_work_category(name, not is_act)
                    msg = f"\u2714  '{name}' {'deactivated' if is_act else 'activated'}."
                elif level == "area":
                    toggle_operational_area(item_id, not is_act)
                    msg = f"\u2714  '{name}' {'deactivated' if is_act else 'activated'}."
                else:
                    toggle_sub_category(item_id, not is_act)
                    msg = f"\u2714  '{name}' {'deactivated' if is_act else 'activated'}."
                _status(msg)
                _full_refresh()

            det_toggle_btn.config(state="normal", text=tog_label,
                                   bg=tog_bg, command=_do_toggle)

            # ── Delete ─────────────────────────────────────────────────────
            level_names = {"cat": "category (and all its areas & sub-categories)",
                           "area": "area (and all its sub-categories)",
                           "sub": "sub-category"}

            def _do_delete():
                if not messagebox.askyesno("Confirm Delete",
                    f"Delete {level_names[level]}:\n\n  '{name}'?\n\n"
                    "Historical timesheet data using this item will be unaffected."):
                    return
                if level == "cat":
                    ok, msg = delete_work_category_cascade(name)
                elif level == "area":
                    ok, msg = delete_operational_area_cascade(item_id)
                else:
                    ok, msg = delete_sub_category_by_id(item_id)
                _status(f"{'\u2714' if ok else '\u26a0'}  {msg}", ok)
                _full_refresh()
                _clear_detail()

            det_delete_btn.config(state="normal", command=_do_delete)

            # ── Contextual "Add Child" section ─────────────────────────────
            # Reset the child row widgets
            for w in det_child_row.winfo_children():
                w.destroy()
            det_child_var2.set("")
            det_child_status.config(text="")

            if level == "cat":
                # Show "Add Area under <name>"
                det_child_lbl.config(text=f"\u2795  Add Operational Area under '{name}'")
                det_child_frame.pack(fill="x", pady=(6, 0))

                new_area_ef = tk.Frame(det_child_row, bg=ENTRY_BG,
                                        highlightthickness=1, highlightbackground=BORDER_CLR)
                new_area_ef.pack(side="left", fill="x", expand=True, padx=(0, 8))
                new_area_e = tk.Entry(new_area_ef, textvariable=det_child_var2,
                                       font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                                       relief="flat", bd=0, insertbackground=FOCUS_CLR)
                new_area_e.pack(fill="x", ipady=6, padx=6)

                def _add_area_inline(event=None, _cat=name):
                    area = det_child_var2.get().strip()
                    if not area:
                        det_child_status.config(text="\u26a0  Enter an area name.", fg=DANGER)
                        return
                    if add_operational_area(_cat, area):
                        det_child_var2.set("")
                        det_child_status.config(
                            text=f"\u2714  Area \u2018{area}\u2019 added. Click it in the tree to add Sub-Categories.",
                            fg=SUCCESS)
                        _full_refresh()
                        # Select the new area node directly → shows Sub-Category add box
                        _reselect_by_name("area", area, _cat)
                    else:
                        det_child_status.config(
                            text=f"\u26a0  Area \u2018{area}\u2019 already exists.", fg=DANGER)

                new_area_e.bind("<Return>", _add_area_inline)
                tk.Button(det_child_row, text="\u2795 Add",
                          font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                          relief="flat", cursor="hand2", padx=12, pady=6,
                          command=_add_area_inline).pack(side="left")

            elif level == "area":  # noqa
                # Show "Add Sub-Category under <name>"
                _area_id   = item_id
                _area_name = name
                _cat_name  = meta.get("cat_name", "")
                det_child_lbl.config(
                    text=f"\u2795  Add Sub-Category under '{_area_name}'")
                det_child_frame.pack(fill="x", pady=(6, 0))

                new_sub_ef = tk.Frame(det_child_row, bg=ENTRY_BG,
                                       highlightthickness=1, highlightbackground=BORDER_CLR)
                new_sub_ef.pack(side="left", fill="x", expand=True, padx=(0, 8))
                new_sub_e = tk.Entry(new_sub_ef, textvariable=det_child_var2,
                                      font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                                      relief="flat", bd=0, insertbackground=FOCUS_CLR)
                new_sub_e.pack(fill="x", ipady=6, padx=6)

                def _add_sub_inline(event=None, _cat=_cat_name, _area=_area_name):
                    sub = det_child_var2.get().strip()
                    if not sub:
                        det_child_status.config(text="\u26a0  Enter a sub-category name.", fg=DANGER)
                        return
                    if add_sub_category(_cat, _area, sub):
                        det_child_var2.set("")
                        det_child_status.config(
                            text=f"\u2714  Sub-Category \u2018{sub}\u2019 added.", fg=SUCCESS)
                        _full_refresh()
                        _reselect_by_name("area", _area, _cat)
                    else:
                        det_child_status.config(
                            text=f"\u26a0  \u2018{sub}\u2019 already exists in this area.", fg=DANGER)

                new_sub_e.bind("<Return>", _add_sub_inline)
                tk.Button(det_child_row, text="\u2795 Add",
                          font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                          relief="flat", cursor="hand2", padx=12, pady=6,
                          command=_add_sub_inline).pack(side="left")

            else:
                # Sub-category selected — no children to add
                det_child_frame.pack_forget()

        def _reselect_by_name(level, name, cat_name=None, area_name=None):
            """After refresh, find & re-select the node with the given name.
            Expands all ancestor nodes so the item is visible in the tree."""
            for iid, meta in node_meta.items():
                if meta["level"] != level: continue
                if meta["name"] != name: continue
                if level == "area" and meta.get("cat_name") != cat_name: continue
                if level == "sub" and (meta.get("cat_name") != cat_name or
                                       meta.get("area_name") != area_name): continue
                # Expand all ancestors so the node is visible
                parent = tree.parent(iid)
                while parent:
                    tree.item(parent, open=True)
                    parent = tree.parent(parent)
                tree.selection_set(iid)
                tree.see(iid)
                _load_detail(meta)
                return

        # ── Tree selection binding ───────────────────────────────────────────
        def _on_tree_select(event=None):
            sel = tree.selection()
            if not sel:
                _clear_detail(); return
            iid = sel[0]
            meta = node_meta.get(iid)
            if meta is None:
                _clear_detail(); return
            _load_detail(meta)

        tree.bind("<<TreeviewSelect>>", _on_tree_select)

        # ═══════════════════════════════════════════════════════════════════════
        # TAB CRUD MANAGER (bottom of right panel)
        # ═══════════════════════════════════════════════════════════════════════
        crud_card = _card(right)
        crud_card.pack(fill="both", expand=True)

        # Tab bar
        tab_bar = tk.Frame(crud_card, bg=ACCENT_BLUE, height=34)
        tab_bar.pack(fill="x")
        tab_bar.pack_propagate(False)

        # Status bar (bottom of crud card)
        crud_status_bar = tk.Frame(crud_card, bg="#D2E1F1", height=28)
        crud_status_bar.pack(side="bottom", fill="x")
        crud_status_bar.pack_propagate(False)
        crud_status_lbl = tk.Label(crud_status_bar, text="  Ready",
                                    font=("Calibri", 8), bg="#D2E1F1",
                                    fg=TEXT_MID, anchor="w")
        crud_status_lbl.pack(fill="x", padx=8, pady=5)

        def _crud_status(msg, ok=True):
            crud_status_lbl.config(text=f"  {msg}", fg=SUCCESS if ok else DANGER)

        # Notebook area
        notebook = tk.Frame(crud_card, bg=PAGE_BG)
        notebook.pack(fill="both", expand=True)

        tab_frames   = {}
        tab_btns     = {}
        current_tab  = {"frame": None}

        def _switch_tab(name):
            if current_tab["frame"] is not None:
                current_tab["frame"].place_forget()
            for n, btn in tab_btns.items():
                is_sel = (n == name)
                btn.config(
                    bg="#EBF2FA" if is_sel else ACCENT_BLUE,
                    fg=ACCENT_BLUE if is_sel else "white",
                    font=("Calibri", 8, "bold")
                )
            frm = tab_frames[name]
            frm.place(in_=notebook, x=0, y=0, relwidth=1.0, relheight=1.0)
            frm.lift()
            current_tab["frame"] = frm

        def _full_refresh():
            refresh_tree()
            self._load_mapping()
            _load_cats_tab()
            _load_areas_tab()
            _load_subs_tab()

        # Now that _full_refresh exists, wire up the Quick Add button
        _qa_refresh_ref["fn"] = _full_refresh

        # ─────────────────────────────────────────────────────────────────────
        # TAB 1 — CATEGORIES (Add only; edit/delete via tree)
        # ─────────────────────────────────────────────────────────────────────
        tab_frames["categories"] = tk.Frame(notebook, bg=PAGE_BG)

        def _build_cats_tab():
            frm = tab_frames["categories"]
            for w in frm.winfo_children(): w.destroy()

            # Header hint
            tk.Label(frm, text="Add a new category, or click a category in the tree to edit/delete.",
                     font=("Calibri", 8), bg=PAGE_BG, fg=TEXT_LIGHT,
                     wraplength=440, justify="left", anchor="w",
                     padx=14, pady=(8,0)).pack(fill="x")

            add_row = tk.Frame(frm, bg=PAGE_BG, padx=14, pady=8); add_row.pack(fill="x")
            tk.Label(add_row, text="New Category:",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG,
                     fg=TEXT_MID).pack(side="left")
            new_cat_v = tk.StringVar()
            ef = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                          highlightbackground=BORDER_CLR)
            ef.pack(side="left", fill="x", expand=True, padx=(6, 8))
            e = tk.Entry(ef, textvariable=new_cat_v, font=("Calibri", 10),
                         bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                         insertbackground=FOCUS_CLR)
            e.pack(fill="x", ipady=6, padx=4)

            def _add_cat(event=None):
                n = new_cat_v.get().strip()
                if not n:
                    _crud_status("\u26a0  Enter a category name.", False); return
                if add_work_category(n):
                    new_cat_v.set("")
                    _full_refresh()
                    _crud_status(f"\u2714  Category \u2018{n}\u2019 added.")
                else:
                    _crud_status(f"\u26a0  \u2018{n}\u2019 already exists.", False)

            e.bind("<Return>", _add_cat)
            tk.Button(add_row, text="\u2795 Add",
                      font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=12, pady=6,
                      command=_add_cat).pack(side="left")

            # Table header
            hdr = tk.Frame(frm, bg=TBL_HDR_BG, padx=4, pady=1)
            hdr.pack(fill="x", padx=14)
            for txt, w in [("Category Name", 28), ("Status", 10)]:
                tk.Label(hdr, text=txt, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=w, anchor="w", padx=4, pady=5).pack(side="left")

            # Scrollable list
            sc = tk.Canvas(frm, bg=PAGE_BG, highlightthickness=0)
            sb2 = ttk.Scrollbar(frm, orient="vertical", command=sc.yview)
            sc.configure(yscrollcommand=sb2.set)
            sb2.pack(side="right", fill="y", padx=(0,14))
            sc.pack(fill="both", expand=True, padx=(14,0))
            inner = tk.Frame(sc, bg=PAGE_BG)
            sc_wid = sc.create_window((0,0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                       lambda e2: sc.configure(scrollregion=sc.bbox("all")))
            sc.bind("<Configure>",
                    lambda e2, _sc=sc, _wid=sc_wid: _sc.itemconfig(_wid, width=e2.width))

            frm._cats_inner = inner
            frm._cats_sc    = sc
            _load_cats_tab()

        def _load_cats_tab():
            frm = tab_frames["categories"]
            inner = getattr(frm, "_cats_inner", None)
            if inner is None: return
            for w in inner.winfo_children(): w.destroy()
            with get_conn() as conn:
                rows = conn.execute(
                    "SELECT id, name, is_active FROM work_categories ORDER BY sort_order, name"
                ).fetchall()
            for ri, r in enumerate(rows):
                bg = ROW_ODD if ri%2 else ROW_EVEN
                row_f = tk.Frame(inner, bg=bg); row_f.pack(fill="x")
                fg_name = TEXT_DARK if r["is_active"] else TEXT_LIGHT
                tk.Label(row_f, text=r["name"], font=("Calibri", 9),
                         bg=bg, fg=fg_name, width=28, anchor="w",
                         padx=6, pady=5).pack(side="left")
                st_txt = "\u25cf Active" if r["is_active"] else "\u25cb Inactive"
                st_clr = SUCCESS if r["is_active"] else TEXT_LIGHT
                tk.Label(row_f, text=st_txt, font=("Calibri", 8),
                         bg=bg, fg=st_clr, width=10, anchor="w").pack(side="left")

        # ─────────────────────────────────────────────────────────────────────
        # TAB 2 — AREAS
        # ─────────────────────────────────────────────────────────────────────
        tab_frames["areas"] = tk.Frame(notebook, bg=PAGE_BG)

        def _build_areas_tab():
            frm = tab_frames["areas"]
            for w in frm.winfo_children(): w.destroy()

            tk.Label(frm, text="Add a new area under a category, or click an area in the tree to edit/delete.",
                     font=("Calibri", 8), bg=PAGE_BG, fg=TEXT_LIGHT,
                     wraplength=440, justify="left", anchor="w",
                     padx=14, pady=(8,0)).pack(fill="x")

            add_row = tk.Frame(frm, bg=PAGE_BG, padx=14, pady=8); add_row.pack(fill="x")
            cat_sel_v  = tk.StringVar()
            area_new_v = tk.StringVar()

            cats = get_work_categories()
            tk.Label(add_row, text="Category:",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            cf2 = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                           highlightbackground=BORDER_CLR)
            cf2.pack(side="left", padx=(4,8))
            cat_cb = ttk.Combobox(cf2, textvariable=cat_sel_v, values=cats,
                                  state="readonly", font=("Calibri", 9), width=16)
            cat_cb.pack(ipady=4, padx=2); _style_combo(cat_cb)

            tk.Label(add_row, text="Area:",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            af2 = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                           highlightbackground=BORDER_CLR)
            af2.pack(side="left", fill="x", expand=True, padx=(4,8))
            area_e = tk.Entry(af2, textvariable=area_new_v, font=("Calibri", 9),
                              bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                              insertbackground=FOCUS_CLR)
            area_e.pack(fill="x", ipady=6, padx=4)

            cat_filter_v = tk.StringVar(value="All")

            def _add_area(event=None):
                cat  = cat_sel_v.get().strip()
                area = area_new_v.get().strip()
                if not cat or not area:
                    _crud_status("\u26a0  Select category and enter area name.", False); return
                if add_operational_area(cat, area):
                    area_new_v.set("")
                    _full_refresh()
                    _crud_status(f"\u2714  Area \u2018{area}\u2019 added under \u2018{cat}\u2019.")
                else:
                    _crud_status(f"\u26a0  Area already exists in \u2018{cat}\u2019.", False)

            area_e.bind("<Return>", _add_area)
            tk.Button(add_row, text="\u2795 Add",
                      font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      command=_add_area).pack(side="left")

            # Filter bar
            flt = tk.Frame(frm, bg=PAGE_BG, padx=14, pady=(0,4)); flt.pack(fill="x")
            tk.Label(flt, text="Filter:",
                     font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            flt_f = tk.Frame(flt, bg=PANEL_BG, highlightthickness=1,
                             highlightbackground=BORDER_CLR)
            flt_f.pack(side="left", padx=(4,0))
            flt_cb = ttk.Combobox(flt_f, textvariable=cat_filter_v,
                                  values=["All"] + cats,
                                  state="readonly", font=("Calibri", 9), width=18)
            flt_cb.pack(ipady=4, padx=2); _style_combo(flt_cb)
            flt_cb.bind("<<ComboboxSelected>>", lambda e2: _load_areas_tab())

            # Table header
            hdr = tk.Frame(frm, bg=TBL_HDR_BG, padx=4, pady=1); hdr.pack(fill="x", padx=14)
            for txt, w in [("Category", 16), ("Area Name", 22), ("Status", 9)]:
                tk.Label(hdr, text=txt, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=w, anchor="w", padx=4, pady=5).pack(side="left")

            sc = tk.Canvas(frm, bg=PAGE_BG, highlightthickness=0)
            sb2 = ttk.Scrollbar(frm, orient="vertical", command=sc.yview)
            sc.configure(yscrollcommand=sb2.set)
            sb2.pack(side="right", fill="y", padx=(0,14))
            sc.pack(fill="both", expand=True, padx=(14,0))
            inner = tk.Frame(sc, bg=PAGE_BG)
            sc_wid2 = sc.create_window((0,0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                       lambda e2: sc.configure(scrollregion=sc.bbox("all")))
            sc.bind("<Configure>",
                    lambda e2, _sc=sc, _wid=sc_wid2: _sc.itemconfig(_wid, width=e2.width))

            frm._areas_inner      = inner
            frm._areas_sc         = sc
            frm._areas_cat_filter = cat_filter_v
            _load_areas_tab()

        def _load_areas_tab():
            frm = tab_frames["areas"]
            inner = getattr(frm, "_areas_inner", None)
            if inner is None: return
            cf_val = getattr(frm, "_areas_cat_filter", None)
            cf_val = cf_val.get() if cf_val else "All"
            for w in inner.winfo_children(): w.destroy()
            with get_conn() as conn:
                if cf_val and cf_val != "All":
                    rows = conn.execute("""
                        SELECT oa.id, oa.name, oa.is_active, wc.name AS cat_name
                        FROM operational_areas oa
                        JOIN work_categories wc ON oa.category_id=wc.id
                        WHERE wc.name=? ORDER BY wc.sort_order, wc.name, oa.sort_order, oa.name
                    """, (cf_val,)).fetchall()
                else:
                    rows = conn.execute("""
                        SELECT oa.id, oa.name, oa.is_active, wc.name AS cat_name
                        FROM operational_areas oa
                        JOIN work_categories wc ON oa.category_id=wc.id
                        ORDER BY wc.sort_order, wc.name, oa.sort_order, oa.name
                    """).fetchall()
            for ri, r in enumerate(rows):
                bg = ROW_ODD if ri%2 else ROW_EVEN
                row_f = tk.Frame(inner, bg=bg); row_f.pack(fill="x")
                tk.Label(row_f, text=r["cat_name"], font=("Calibri", 8),
                         bg=bg, fg=TEXT_LIGHT, width=16, anchor="w",
                         padx=4, pady=5).pack(side="left")
                fg_n = TEXT_DARK if r["is_active"] else TEXT_LIGHT
                tk.Label(row_f, text=r["name"], font=("Calibri", 9),
                         bg=bg, fg=fg_n, width=22, anchor="w",
                         padx=4, pady=5).pack(side="left")
                st_txt = "\u25cf Active" if r["is_active"] else "\u25cb Inactive"
                st_clr = SUCCESS if r["is_active"] else TEXT_LIGHT
                tk.Label(row_f, text=st_txt, font=("Calibri", 8),
                         bg=bg, fg=st_clr, width=9, anchor="w").pack(side="left")

        # ─────────────────────────────────────────────────────────────────────
        # TAB 3 — SUB-CATEGORIES
        # ─────────────────────────────────────────────────────────────────────
        tab_frames["subcats"] = tk.Frame(notebook, bg=PAGE_BG)

        def _build_subs_tab():
            frm = tab_frames["subcats"]
            for w in frm.winfo_children(): w.destroy()

            tk.Label(frm, text="Add a sub-category, or click a sub-category in the tree to edit/delete.",
                     font=("Calibri", 8), bg=PAGE_BG, fg=TEXT_LIGHT,
                     wraplength=440, justify="left", anchor="w",
                     padx=14, pady=(8,0)).pack(fill="x")

            add_row = tk.Frame(frm, bg=PAGE_BG, padx=14, pady=8); add_row.pack(fill="x")
            add_cat_v  = tk.StringVar()
            add_area_v = tk.StringVar()
            add_sub_v  = tk.StringVar()

            cats = get_work_categories()
            tk.Label(add_row, text="Cat:",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            cf3 = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                           highlightbackground=BORDER_CLR)
            cf3.pack(side="left", padx=(4,6))
            add_cat_cb = ttk.Combobox(cf3, textvariable=add_cat_v, values=cats,
                                      state="readonly", font=("Calibri", 9), width=13)
            add_cat_cb.pack(ipady=4, padx=2); _style_combo(add_cat_cb)

            tk.Label(add_row, text="Area:",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            af3 = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                           highlightbackground=BORDER_CLR)
            af3.pack(side="left", padx=(4,6))
            add_area_cb = ttk.Combobox(af3, textvariable=add_area_v, values=[],
                                       state="readonly", font=("Calibri", 9), width=13)
            add_area_cb.pack(ipady=4, padx=2); _style_combo(add_area_cb)

            def _cat_sel3(*_):
                areas = get_operational_areas(add_cat_v.get())
                add_area_cb["values"] = areas; add_area_v.set("")
            add_cat_cb.bind("<<ComboboxSelected>>", _cat_sel3)

            tk.Label(add_row, text="Sub:",
                     font=("Calibri", 9, "bold"), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            sf3 = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                           highlightbackground=BORDER_CLR)
            sf3.pack(side="left", fill="x", expand=True, padx=(4,8))
            sub_e = tk.Entry(sf3, textvariable=add_sub_v, font=("Calibri", 9),
                             bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                             insertbackground=FOCUS_CLR)
            sub_e.pack(fill="x", ipady=6, padx=4)

            def _add_sub(event=None):
                cat  = add_cat_v.get().strip()
                area = add_area_v.get().strip()
                sub  = add_sub_v.get().strip()
                if not all([cat, area, sub]):
                    _crud_status("\u26a0  Select Category, Area and enter Sub-Category.", False)
                    return
                if add_sub_category(cat, area, sub):
                    add_sub_v.set(""); _full_refresh()
                    _crud_status(f"\u2714  Sub \u2018{sub}\u2019 added.")
                else:
                    _crud_status(f"\u26a0  \u2018{sub}\u2019 already exists in this area.", False)

            sub_e.bind("<Return>", _add_sub)
            tk.Button(add_row, text="\u2795 Add",
                      font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      command=_add_sub).pack(side="left")

            # Filter bars
            flt = tk.Frame(frm, bg=PAGE_BG, padx=14, pady=(0,4)); flt.pack(fill="x")
            flt_cat_v  = tk.StringVar(value="All")
            flt_area_v = tk.StringVar(value="All")

            tk.Label(flt, text="Cat:",
                     font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            flt_cf = tk.Frame(flt, bg=PANEL_BG, highlightthickness=1,
                              highlightbackground=BORDER_CLR)
            flt_cf.pack(side="left", padx=(4,8))
            flt_cat_cb = ttk.Combobox(flt_cf, textvariable=flt_cat_v,
                                      values=["All"] + cats,
                                      state="readonly", font=("Calibri", 9), width=14)
            flt_cat_cb.pack(ipady=4, padx=2); _style_combo(flt_cat_cb)

            tk.Label(flt, text="Area:",
                     font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            flt_af = tk.Frame(flt, bg=PANEL_BG, highlightthickness=1,
                              highlightbackground=BORDER_CLR)
            flt_af.pack(side="left", padx=(4,0))
            flt_area_cb = ttk.Combobox(flt_af, textvariable=flt_area_v,
                                       values=["All"], state="readonly",
                                       font=("Calibri", 9), width=14)
            flt_area_cb.pack(ipady=4, padx=2); _style_combo(flt_area_cb)

            def _flt_cat_chg(*_):
                c = flt_cat_v.get()
                areas = get_operational_areas(c) if c != "All" else []
                flt_area_cb["values"] = ["All"] + areas
                flt_area_v.set("All"); _load_subs_tab()

            flt_cat_cb.bind("<<ComboboxSelected>>", _flt_cat_chg)
            flt_area_cb.bind("<<ComboboxSelected>>", lambda e2: _load_subs_tab())

            # Table header
            hdr = tk.Frame(frm, bg=TBL_HDR_BG, padx=4, pady=1); hdr.pack(fill="x", padx=14)
            for txt, w in [("Category", 12), ("Area", 14), ("Sub-Category", 16), ("Status", 9)]:
                tk.Label(hdr, text=txt, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=w, anchor="w", padx=4, pady=5).pack(side="left")

            sc = tk.Canvas(frm, bg=PAGE_BG, highlightthickness=0)
            sb2 = ttk.Scrollbar(frm, orient="vertical", command=sc.yview)
            sc.configure(yscrollcommand=sb2.set)
            sb2.pack(side="right", fill="y", padx=(0,14))
            sc.pack(fill="both", expand=True, padx=(14,0))
            inner = tk.Frame(sc, bg=PAGE_BG)
            sc_wid3 = sc.create_window((0,0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                       lambda e2: sc.configure(scrollregion=sc.bbox("all")))
            sc.bind("<Configure>",
                    lambda e2, _sc=sc, _wid=sc_wid3: _sc.itemconfig(_wid, width=e2.width))

            frm._subs_inner    = inner
            frm._subs_sc       = sc
            frm._subs_flt_cat  = flt_cat_v
            frm._subs_flt_area = flt_area_v
            _load_subs_tab()

        def _load_subs_tab():
            frm = tab_frames["subcats"]
            inner = getattr(frm, "_subs_inner", None)
            if inner is None: return
            flt_cat  = getattr(frm, "_subs_flt_cat",  None)
            flt_area = getattr(frm, "_subs_flt_area", None)
            cat_v  = flt_cat.get()  if flt_cat  else "All"
            area_v = flt_area.get() if flt_area else "All"
            for w in inner.winfo_children(): w.destroy()
            with get_conn() as conn:
                q = """
                    SELECT sc.id, sc.name, sc.is_active,
                           oa.name AS area_name, wc.name AS cat_name
                    FROM sub_categories sc
                    JOIN operational_areas oa ON sc.area_id=oa.id
                    JOIN work_categories wc ON oa.category_id=wc.id
                    WHERE 1=1
                """
                params = []
                if cat_v and cat_v != "All":
                    q += " AND wc.name=?"; params.append(cat_v)
                if area_v and area_v != "All":
                    q += " AND oa.name=?"; params.append(area_v)
                q += " ORDER BY wc.sort_order, wc.name, oa.sort_order, oa.name, sc.sort_order, sc.name"
                rows = conn.execute(q, params).fetchall()
            for ri, r in enumerate(rows):
                bg = ROW_ODD if ri%2 else ROW_EVEN
                row_f = tk.Frame(inner, bg=bg); row_f.pack(fill="x")
                tk.Label(row_f, text=r["cat_name"], font=("Calibri", 8),
                         bg=bg, fg=TEXT_LIGHT, width=12, anchor="w", padx=4, pady=5).pack(side="left")
                tk.Label(row_f, text=r["area_name"], font=("Calibri", 8),
                         bg=bg, fg=TEXT_MID, width=14, anchor="w", padx=4, pady=5).pack(side="left")
                fg_n = TEXT_DARK if r["is_active"] else TEXT_LIGHT
                tk.Label(row_f, text=r["name"], font=("Calibri", 9),
                         bg=bg, fg=fg_n, width=16, anchor="w", padx=4, pady=5).pack(side="left")
                st_txt = "\u25cf Active" if r["is_active"] else "\u25cb Inactive"
                st_clr = SUCCESS if r["is_active"] else TEXT_LIGHT
                tk.Label(row_f, text=st_txt, font=("Calibri", 8),
                         bg=bg, fg=st_clr, width=9, anchor="w").pack(side="left")

        # ─────────────────────────────────────────────────────────────────────
        # TAB 4 — IMPORT / EXPORT
        # ─────────────────────────────────────────────────────────────────────
        tab_frames["import"] = tk.Frame(notebook, bg=PAGE_BG)

        def _build_import_tab():
            frm = tab_frames["import"]
            for w in frm.winfo_children(): w.destroy()
            pad = tk.Frame(frm, bg=PAGE_BG, padx=20, pady=16); pad.pack(fill="x")

            tk.Label(pad, text="Import from Excel",
                     font=("Georgia", 12, "bold"), bg=PAGE_BG, fg=TEXT_DARK).pack(anchor="w")
            tk.Label(pad,
                     text="Excel must have 3 columns: 'Work Category' | 'Operational Area' | 'Sub Category'",
                     font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_LIGHT,
                     wraplength=400, justify="left").pack(anchor="w", pady=(4,12))

            imp_lbl = tk.Label(pad, text="", font=("Calibri", 9),
                               bg=PAGE_BG, fg=SUCCESS, wraplength=400, justify="left")
            imp_lbl.pack(anchor="w", pady=(0,8))

            def import_from_excel():
                fp = filedialog.askopenfilename(
                    title="Select Work Mapping Excel",
                    filetypes=[("Excel","*.xlsx *.xls"), ("All","*.*")])
                if not fp: return
                try:
                    import pandas as pd
                    df = pd.read_excel(fp)
                    added = {"cat": 0, "area": 0, "sub": 0}
                    for _, row in df.iterrows():
                        cat  = str(row.get("Work Category","")).strip()
                        area = str(row.get("Operational Area","")).strip()
                        sub  = str(row.get("Sub Category",
                                   row.get("Sub-Category",""))).strip()
                        if not cat or cat == "nan": continue
                        if add_work_category(cat): added["cat"] += 1
                        if area and area != "nan":
                            if add_operational_area(cat, area): added["area"] += 1
                        if sub and sub != "nan" and area and area != "nan":
                            if add_sub_category(cat, area, sub): added["sub"] += 1
                    _full_refresh()
                    imp_lbl.config(
                        text=(f"✔  Imported: {added['cat']} categories, "
                              f"{added['area']} areas, {added['sub']} sub-categories."),
                        fg=SUCCESS)
                except ImportError:
                    imp_lbl.config(
                        text="⚠  pandas not installed. Run: pip install pandas",
                        fg=DANGER)
                except Exception as ex:
                    imp_lbl.config(text=f"⚠  {ex}", fg=DANGER)

            tk.Button(pad, text="📂  Browse & Import Excel",
                      font=("Calibri", 11, "bold"), bg=ACCENT_BLUE, fg="white",
                      relief="flat", cursor="hand2", padx=20, pady=11,
                      command=import_from_excel).pack(anchor="w", pady=(0, 20))

            tk.Frame(pad, bg=BORDER_CLR, height=1).pack(fill="x", pady=(0,12))
            tk.Label(pad, text="Export Work Mapping",
                     font=("Georgia", 12, "bold"), bg=PAGE_BG, fg=TEXT_DARK).pack(anchor="w")
            tk.Button(pad, text="📥  Export to Excel",
                      font=("Calibri", 11, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=20, pady=11,
                      command=_export_wm).pack(anchor="w", pady=(8,0))

        # ── Build all tab contents ────────────────────────────────────────────
        _build_cats_tab()
        _build_areas_tab()
        _build_subs_tab()
        _build_import_tab()

        # ── Create tab buttons ────────────────────────────────────────────────
        TAB_CFG = [
            ("categories", "📁 Categories"),
            ("areas",      "📂 Areas"),
            ("subcats",    "▸ Sub-Cats"),
            ("import",     "📥 Import"),
        ]
        for tab_id, label in TAB_CFG:
            btn = tk.Button(tab_bar, text=label,
                            font=("Calibri", 8, "bold"),
                            bg=ACCENT_BLUE, fg="white",
                            relief="flat", cursor="hand2",
                            padx=10, pady=7, bd=0,
                            activebackground="#024277",
                            activeforeground="white",
                            command=lambda t=tab_id: _switch_tab(t))
            btn.pack(side="left", padx=(0, 1))
            tab_btns[tab_id] = btn

        # Switch to first tab after geometry is settled
        crud_card.after(20, lambda: _switch_tab("categories"))


    # ── ADMIN DASHBOARD ──────────────────────────────────────────────────────────

    def _build_admin_dashboard(self, parent):
        CARD_STAFF   = "#06355E"
        CARD_ONTIME  = "#1A6B45"
        CARD_LATE    = "#B45309"
        CARD_MISS    = "#B52A2A"
        CARD_PARTIAL = "#0369A1"
        CARD_LEAVE   = "#7C3AED"
        today_str    = date.today().isoformat()
        date_var     = tk.StringVar(value=today_str)

        # ── Custom topbar (taller, holds date + buttons) ─────────────────────────
        tb = tk.Frame(parent, bg=PANEL_BG,
                      highlightthickness=1, highlightbackground=BORDER_CLR)
        tb.pack(fill="x")

        left_tb = tk.Frame(tb, bg=PANEL_BG)
        left_tb.pack(side="left", fill="y", padx=(20, 0))
        tk.Label(left_tb, text="Admin Dashboard",
                 font=("Georgia", 15, "bold"),
                 bg=PANEL_BG, fg=TEXT_DARK).pack(side="left", pady=12)
        tk.Label(left_tb,
                 text=f"  {date.today().strftime('%A, %d %B %Y')}",
                 font=("Calibri", 9), bg=PANEL_BG,
                 fg=TEXT_MID).pack(side="left")

        right_tb = tk.Frame(tb, bg=PANEL_BG)
        right_tb.pack(side="right", padx=16, pady=10)

        # Date picker
        de = tk.Frame(right_tb, bg=BORDER_CLR, padx=1, pady=1)
        de.pack(side="left", padx=(0, 8))
        de2 = tk.Frame(de, bg=ENTRY_BG); de2.pack()
        tk.Label(de2, text="📅", font=("Calibri", 9),
                 bg=ENTRY_BG, fg=TEXT_MID, padx=4).pack(side="left")
        tk.Entry(de2, textvariable=date_var, font=("Calibri", 9),
                 bg=ENTRY_BG, fg=ENTRY_FG, relief="flat",
                 bd=0, width=11).pack(side="left", ipady=5, padx=(0, 4))

        def _load_detail():
            ds = date_var.get().strip() or today_str
            sm = get_daily_submission_summary(ds)
            ms = get_missing_timesheets(ds)
            pending = get_pending_leave_count()

            # ── Rebuild KPI cards ────────────────────────────────────────────
            for w in kpi_frame.winfo_children():
                w.destroy()

            def _kpi(parent, icon, label, val, color):
                f = tk.Frame(parent, bg=color, padx=2, pady=2)
                f.pack(side="left", fill="x", expand=True, padx=(0, 8))
                inner = tk.Frame(f, bg=PANEL_BG, padx=12, pady=10)
                inner.pack(fill="both", expand=True)
                tk.Label(inner, text=str(val),
                         font=("Georgia", 26, "bold"),
                         bg=PANEL_BG, fg=color).pack(anchor="w")
                bot = tk.Frame(inner, bg=PANEL_BG); bot.pack(fill="x")
                tk.Label(bot, text=icon, font=("Calibri", 12),
                         bg=PANEL_BG, fg=color).pack(side="left")
                tk.Label(bot, text=f"  {label}",
                         font=("Calibri", 8), bg=PANEL_BG,
                         fg=TEXT_MID, wraplength=100,
                         justify="left").pack(side="left")

            _kpi(kpi_frame, "👥", "Active Staff",    sm["total_active"], CARD_STAFF)
            _kpi(kpi_frame, "✅", "On Time",          sm["on_time"],      CARD_ONTIME)
            _kpi(kpi_frame, "⚠️", "Late",             sm["late"],         CARD_LATE)
            _kpi(kpi_frame, "❌", "Not Submitted",    sm["not_submitted"],CARD_MISS)
            _kpi(kpi_frame, "📝", "Partial",          sm["partial"],      CARD_PARTIAL)
            _kpi(kpi_frame, "🏖", "Pending Leaves",   pending,            CARD_LEAVE)

            # ── Breakdown bar ────────────────────────────────────────────────
            def _draw():
                bar_cv.delete("all")
                bar_cv.update_idletasks()
                W = bar_cv.winfo_width()
                if W < 20: W = 700
                total = max(sm["total_active"], 1)
                segs  = [(sm["on_time"], CARD_ONTIME),
                         (sm["late"],    CARD_LATE),
                         (sm["partial"], CARD_PARTIAL),
                         (sm["not_submitted"], CARD_MISS)]
                x = 0
                for cnt, col in segs:
                    if cnt <= 0: continue
                    w = max(int(cnt / total * W), 2)
                    bar_cv.create_rectangle(x, 0, x+w, 24,
                                            fill=col, outline="")
                    if w > 30:
                        bar_cv.create_text(
                            x + w//2, 12,
                            text=f"{round(cnt/total*100)}%",
                            font=("Calibri", 8, "bold"), fill="white")
                    x += w
            bar_cv.after(120, _draw)

            # ── Status table ────────────────────────────────────────────────
            for w in st_body.winfo_children(): w.destroy()
            st_date_lbl.config(text=f"  {ds}")
            col_w = [20, 10, 13, 13, 6, 14]
            hdrs  = ["Employee","Role","Submitted","Approval","Hrs","8h Status"]
            hf = tk.Frame(st_body, bg=TBL_HDR_BG); hf.pack(fill="x")
            for h, cw in zip(hdrs, col_w):
                tk.Label(hf, text=h, font=("Calibri", 8, "bold"),
                         bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         width=cw, anchor="w",
                         padx=6, pady=5).pack(side="left")
            if not sm["records"]:
                tk.Label(st_body,
                         text="  No submissions recorded for this date.",
                         font=("Calibri", 9), bg=PANEL_BG,
                         fg=TEXT_MID, pady=10).pack(anchor="w")
            else:
                for ri, r in enumerate(sm["records"]):
                    bg = ROW_ODD if ri%2==0 else ROW_EVEN
                    sf = SUB_COLORS.get(r.get("submission_status",""), TEXT_DARK)
                    af = APR_COLORS.get(r.get("approval_status",""), TEXT_DARK)
                    hrs = float(r.get("total_hrs") or 0)
                    done = hrs >= DAILY_TARGET
                    hrs_clr = "#1A6B45" if done else ("#B45309" if hrs >= DAILY_TARGET/2 else "#B52A2A")
                    rf = tk.Frame(st_body, bg=bg); rf.pack(fill="x")
                    for v, cw, fg in zip(
                        [r.get("full_name",""), r.get("role",""),
                         r.get("submission_status","—"),
                         r.get("approval_status","—"),
                         f"{hrs:.1f}"],
                        col_w[:-1],
                        [TEXT_DARK, TEXT_DARK, sf, af, hrs_clr]
                    ):
                        tk.Label(rf, text=str(v or "—"),
                                 font=("Calibri", 9), bg=bg, fg=fg,
                                 width=cw, anchor="w",
                                 padx=6, pady=4).pack(side="left")
                    # 8h badge cell
                    badge_txt = f"{'🟢 ✔ Complete' if done else f'🔴 {hrs:.1f}/{DAILY_TARGET}h'}"
                    tk.Label(rf, text=badge_txt,
                             font=("Calibri", 8, "bold"),
                             bg=bg, fg=hrs_clr,
                             width=col_w[-1], anchor="w",
                             padx=6, pady=4).pack(side="left")

            # ── Missing list ─────────────────────────────────────────────────
            for w in ms_body.winfo_children(): w.destroy()
            ms_cnt_lbl.config(text=f"  {len(ms)} staff")
            if not ms:
                tk.Label(ms_body,
                         text="  ✅  All staff submitted.",
                         font=("Calibri", 9), bg=PANEL_BG,
                         fg=SUCCESS, pady=10).pack(anchor="w")
            else:
                mh = tk.Frame(ms_body, bg=TBL_HDR_BG); mh.pack(fill="x")
                for h, cw in zip(["Employee","Role","Reports To"],[22,10,16]):
                    tk.Label(mh, text=h, font=("Calibri", 8, "bold"),
                             bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                             width=cw, anchor="w",
                             padx=6, pady=5).pack(side="left")
                for ri, m in enumerate(ms):
                    bg = ROW_ODD if ri%2==0 else ROW_EVEN
                    mrf = tk.Frame(ms_body, bg=bg); mrf.pack(fill="x")
                    for v, cw in zip(
                        [m["full_name"], m["role"],
                         m.get("reporting_to","—")],
                        [22, 10, 16]
                    ):
                        tk.Label(mrf, text=str(v or "—"),
                                 font=("Calibri", 9), bg=bg, fg=DANGER,
                                 width=cw, anchor="w",
                                 padx=6, pady=4).pack(side="left")

        # Action buttons — defined after _load_detail so lambda captures it
        tk.Button(right_tb, text="↺  Refresh",
                  font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", padx=12, pady=5,
                  command=_load_detail).pack(side="left", padx=(0, 6))
        tk.Button(right_tb, text="Today",
                  font=("Calibri", 9), bg="#D2E1F1", fg=TEXT_DARK,
                  relief="flat", cursor="hand2", padx=10, pady=5,
                  command=lambda: [date_var.set(today_str),
                                   _load_detail()]).pack(side="left", padx=(0, 6))
        tk.Button(right_tb, text="📧  Email",
                  font=("Calibri", 9), bg="#E0F2FE", fg="#0369A1",
                  relief="flat", cursor="hand2", padx=10, pady=5,
                  command=lambda: self._send_dashboard_email(
                      date_var.get() or today_str)
                  ).pack(side="left")

        # ── Body (no canvas — direct pack) ────────────────────────────────────────
        body = tk.Frame(parent, bg=PAGE_BG)
        body.pack(fill="both", expand=True, padx=20, pady=14)

        # KPI row
        kpi_frame = tk.Frame(body, bg=PAGE_BG)
        kpi_frame.pack(fill="x", pady=(0, 12))

        # Breakdown bar card
        bar_card = tk.Frame(body, bg=PANEL_BG,
                            highlightthickness=1,
                            highlightbackground=BORDER_CLR)
        bar_card.pack(fill="x", pady=(0, 12))
        bar_top = tk.Frame(bar_card, bg=PANEL_BG, padx=14, pady=8)
        bar_top.pack(fill="x")
        tk.Label(bar_top, text="Submission Breakdown",
                 font=("Calibri", 10, "bold"),
                 bg=PANEL_BG, fg=TEXT_DARK).pack(side="left")
        # Legend
        leg = tk.Frame(bar_top, bg=PANEL_BG); leg.pack(side="left", padx=20)
        for col, lbl in [(CARD_ONTIME,"On Time"),(CARD_LATE,"Late"),
                         (CARD_PARTIAL,"Partial"),(CARD_MISS,"Not Submitted")]:
            f = tk.Frame(leg, bg=PANEL_BG); f.pack(side="left", padx=(0,12))
            tk.Frame(f, bg=col, width=10, height=10).pack(side="left", padx=(0,3))
            tk.Label(f, text=lbl, font=("Calibri", 8),
                     bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        bar_wrap = tk.Frame(bar_card, bg=PANEL_BG, padx=14, pady=(0,10))
        bar_wrap.pack(fill="x")
        bar_cv = tk.Canvas(bar_wrap, bg="#E8EDF2", height=24,
                           highlightthickness=0)
        bar_cv.pack(fill="x")

        # Bottom two columns
        bot = tk.Frame(body, bg=PAGE_BG)
        bot.pack(fill="both", expand=True)
        bot.columnconfigure(0, weight=1)
        bot.columnconfigure(1, weight=1)

        # Submission status (left)
        st_card = tk.Frame(bot, bg=PANEL_BG,
                           highlightthickness=1,
                           highlightbackground=BORDER_CLR)
        st_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        st_hdr = tk.Frame(st_card, bg=TBL_HDR_BG); st_hdr.pack(fill="x")
        tk.Label(st_hdr, text="  📋  Submission Status",
                 font=("Calibri", 10, "bold"), bg=TBL_HDR_BG,
                 fg=TBL_HDR_FG, pady=9).pack(side="left")
        st_date_lbl = tk.Label(st_hdr, text="",
                               font=("Calibri", 8),
                               bg=TBL_HDR_BG, fg=TBL_HDR_FG)
        st_date_lbl.pack(side="left")
        st_body = tk.Frame(st_card, bg=PANEL_BG)
        st_body.pack(fill="both", expand=True)

        # Missing list (right)
        ms_card = tk.Frame(bot, bg=PANEL_BG,
                           highlightthickness=1,
                           highlightbackground=BORDER_CLR)
        ms_card.grid(row=0, column=1, sticky="nsew")
        ms_hdr = tk.Frame(ms_card, bg=TBL_HDR_BG); ms_hdr.pack(fill="x")
        tk.Label(ms_hdr, text="  ❌  Not Submitted",
                 font=("Calibri", 10, "bold"), bg=TBL_HDR_BG,
                 fg=TBL_HDR_FG, pady=9).pack(side="left")
        ms_cnt_lbl = tk.Label(ms_hdr, text="",
                              font=("Calibri", 8),
                              bg=TBL_HDR_BG, fg=TBL_HDR_FG)
        ms_cnt_lbl.pack(side="left")
        ms_body = tk.Frame(ms_card, bg=PANEL_BG)
        ms_body.pack(fill="both", expand=True)

        # Load data
        _load_detail()

    # ── APPROVAL MANAGER ─────────────────────────────────────────────────────────

    def _send_dashboard_email(self, ds: str):
        cfg = load_config()
        if not cfg.get("email_enabled"):
            messagebox.showinfo("Email Disabled",
                "Enable email in Settings → Email Configuration.")
            return
        ok, msg = send_daily_summary_email(cfg, ds)
        if ok:
            messagebox.showinfo("Sent ✔", f"Summary sent to {cfg['manager_email']}")
        else:
            messagebox.showerror("Failed", msg)

    def _build_approval_manager(self, parent):
        tb = self._topbar(parent, "Timesheet Approvals",
                          "  Review, approve or reject submitted timesheets")

        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=14)

        ctrl_card = _card(outer); ctrl_card.pack(fill="x", pady=(0, 10))
        ctrl = tk.Frame(ctrl_card, bg=PANEL_BG, padx=16, pady=10); ctrl.pack(fill="x")

        tk.Label(ctrl, text="Date:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        date_var = tk.StringVar(value=date.today().isoformat())
        df = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        df.pack(side="left", padx=(4, 12))
        tk.Entry(df, textvariable=date_var, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="flat", bd=0, width=14).pack(ipady=5)

        tk.Label(ctrl, text="Status:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        status_fv = tk.StringVar(value="All")
        sf = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        sf.pack(side="left", padx=(4, 12))
        status_c = ttk.Combobox(sf, textvariable=status_fv,
                                values=["All"] + APPROVAL_STATUSES,
                                state="readonly", font=("Calibri", 10), width=16)
        status_c.pack(ipady=4); _style_combo(status_c)

        # ── Reported To filter ────────────────────────────────────────────────
        tk.Label(ctrl, text="Reported To:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        reporter_fv = tk.StringVar(value="All")
        rf = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        rf.pack(side="left", padx=(4, 12))
        reporter_c = ttk.Combobox(rf, textvariable=reporter_fv,
                                  values=["All"] + REPORTING_OPT,
                                  state="readonly", font=("Calibri", 10), width=18)
        reporter_c.pack(ipady=4); _style_combo(reporter_c)

        list_frame = tk.Frame(outer, bg=PAGE_BG)
        list_frame.pack(fill="both", expand=True)
        sub_data = []

        def _load():
            for w in list_frame.winfo_children():
                w.destroy()
            sub_data.clear()

            ds = date_var.get().strip() or date.today().isoformat()
            records = get_all_submissions(filter_date=ds if ds else None)
            sf_val = status_fv.get()
            if sf_val != "All":
                records = [r for r in records if r.get("approval_status") == sf_val]
            # ── Filter by Reported To ─────────────────────────────────────────
            rep_val = reporter_fv.get()
            if rep_val != "All":
                records = [r for r in records if r.get("reporting_to", "") == rep_val]
            sub_data.extend(records)

            if not records:
                ec = _card(list_frame); ec.pack(fill="x")
                tk.Label(ec, text="No submissions found for this filter.",
                         font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_MID, pady=20).pack()
                return

            cols = ["Employee", "Reported To", "Role", "Date", "Submitted At",
                    "Submission", "Approval", "Hrs", "Entries", "Comment"]
            col_widths = [18, 16, 12, 12, 18, 14, 16, 7, 8, 22]

            style = ttk.Style(); style.theme_use("clam")
            style.configure("JAA.Treeview", background=PANEL_BG,
                            fieldbackground=PANEL_BG, foreground=TEXT_DARK,
                            rowheight=26, font=("Calibri", 9))
            style.configure("JAA.Treeview.Heading", background=TBL_HDR_BG,
                            foreground=TBL_HDR_FG, font=("Calibri", 9, "bold"), relief="flat")
            style.map("JAA.Treeview", background=[("selected", "#C5DCEE")])

            tc = _card(list_frame); tc.pack(fill="both", expand=True)
            tf = tk.Frame(tc, bg=PANEL_BG); tf.pack(fill="both", expand=True, padx=1, pady=1)
            tree = ttk.Treeview(tf, columns=cols, show="headings",
                                style="JAA.Treeview", selectmode="browse")
            vsb = ttk.Scrollbar(tf, orient="vertical",   command=tree.yview)
            hsb = ttk.Scrollbar(tf, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
            tree.pack(side="left", fill="both", expand=True)

            cw_map = dict(zip(cols, [cw*7 for cw in col_widths]))
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=cw_map.get(c, 100), minwidth=50, anchor="w")

            for i, r in enumerate(records):
                sub_clr = r.get("submission_status","")
                apr_clr = r.get("approval_status","")
                tree.insert("", "end", iid=str(r["id"]),
                            values=(r.get("full_name",""),
                                    r.get("reporting_to","—") or "—",
                                    r.get("role",""),
                                    r.get("submit_date",""),
                                    (r.get("submitted_at","")[:16] if r.get("submitted_at") else "—"),
                                    r.get("submission_status","—"),
                                    r.get("approval_status","—"),
                                    r.get("total_hrs",0),
                                    r.get("entry_count",0),
                                    r.get("reviewer_comment","") or ""),
                            tags=(f"apr_{apr_clr.lower().replace(' ','_')}",
                                  "odd" if i%2 else "even"))

            tree.tag_configure("odd",  background=ROW_ODD)
            tree.tag_configure("even", background=ROW_EVEN)

            # ── Action panel under tree ────────────────────────────────────────
            action_card = _card(list_frame); action_card.pack(fill="x", pady=(8, 0))
            af = tk.Frame(action_card, bg=PANEL_BG, padx=16, pady=12); af.pack(fill="x")

            tk.Label(af, text="Selected entry action:", font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(side="left")

            comment_var = tk.StringVar()
            tk.Label(af, text=" Comment:", font=("Calibri", 9),
                     bg=PANEL_BG, fg=TEXT_MID).pack(side="left", padx=(12, 0))
            cmf = tk.Frame(af, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
            cmf.pack(side="left", padx=(4, 12))
            tk.Entry(cmf, textvariable=comment_var, font=("Calibri", 9),
                     bg=ENTRY_BG, fg=ENTRY_FG, relief="flat", bd=0, width=28).pack(ipady=5)

            status_lbl = tk.Label(action_card, text="", font=("Calibri", 9),
                                  bg=PANEL_BG, fg=SUCCESS, padx=16, pady=4)
            status_lbl.pack(anchor="w")

            def _get_selected_id():
                sel = tree.selection()
                if not sel:
                    status_lbl.config(text="⚠  Select a row first.", fg=WARNING)
                    return None
                return int(sel[0])

            def _approve():
                sid = _get_selected_id()
                if sid is None: return
                update_approval(sid, "Approved", "Admin", comment_var.get())
                status_lbl.config(text="✔  Approved.", fg=SUCCESS)
                _load()

            def _reject():
                sid = _get_selected_id()
                if sid is None: return
                update_approval(sid, "Rejected", "Admin", comment_var.get())
                status_lbl.config(text="✔  Rejected.", fg=DANGER)
                _load()

            def _revise():
                sid = _get_selected_id()
                if sid is None: return
                update_approval(sid, "Revision Required", "Admin", comment_var.get())
                status_lbl.config(text="✔  Marked: Revision Required.", fg=WARNING)
                _load()

            for txt, bg, cmd in [
                ("✅  Approve",          SUCCESS,  _approve),
                ("❌  Reject",           DANGER,   _reject),
                ("🔄  Revision Needed",  WARNING,  _revise),
            ]:
                tk.Button(af, text=txt, font=("Calibri", 9, "bold"),
                          bg=bg, fg="white", relief="flat", cursor="hand2",
                          padx=14, pady=7, command=cmd).pack(side="left", padx=(0, 6))

        tk.Button(ctrl, text="🔍  Load",
                  font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  command=_load).pack(side="left")
        tk.Button(ctrl, text="All Dates",
                  font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=10, pady=6,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=lambda: [date_var.set(""), _load()]).pack(side="left", padx=6)
        _load()

    # ── REPORTS TAB ──────────────────────────────────────────────────────────────

    def _build_reports_tab(self, parent):
        self._topbar(parent, "Reports & Analytics",
                     "  Productivity, Client Effort, Submission Analysis")

        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True)

        v_canvas = tk.Canvas(outer, bg=PAGE_BG, highlightthickness=0)
        v_sb = ttk.Scrollbar(outer, orient="vertical", command=v_canvas.yview)
        v_canvas.configure(yscrollcommand=v_sb.set)
        v_sb.pack(side="right", fill="y")
        v_canvas.pack(side="left", fill="both", expand=True)
        page = tk.Frame(v_canvas, bg=PAGE_BG)
        pw = v_canvas.create_window((0, 0), window=page, anchor="nw")
        page.bind("<Configure>", lambda e: v_canvas.configure(scrollregion=v_canvas.bbox("all")))
        v_canvas.bind("<Configure>", lambda e: v_canvas.itemconfig(pw, width=e.width))
        v_canvas.bind("<MouseWheel>",
            lambda e: v_canvas.yview_scroll(-1*(e.delta//120), "units"))

        pad = tk.Frame(page, bg=PAGE_BG, padx=20, pady=16)
        pad.pack(fill="both", expand=True)

        # Date range controls
        ctrl_card = _card(pad); ctrl_card.pack(fill="x", pady=(0, 14))
        ctrl = tk.Frame(ctrl_card, bg=PANEL_BG, padx=16, pady=12); ctrl.pack(fill="x")

        tk.Label(ctrl, text="From:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        from_var = tk.StringVar(value=(date.today() - timedelta(days=30)).isoformat())
        ff = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        ff.pack(side="left", padx=(4, 12))
        tk.Entry(ff, textvariable=from_var, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="flat", bd=0, width=14).pack(ipady=5)

        tk.Label(ctrl, text="To:", font=("Calibri", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MID).pack(side="left")
        to_var = tk.StringVar(value=date.today().isoformat())
        tf2 = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        tf2.pack(side="left", padx=(4, 12))
        tk.Entry(tf2, textvariable=to_var, font=("Calibri", 10), bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="flat", bd=0, width=14).pack(ipady=5)

        report_type_var = tk.StringVar(value="Employee Productivity")
        rf = tk.Frame(ctrl, bg=PANEL_BG, highlightthickness=1, highlightbackground=BORDER_CLR)
        rf.pack(side="left", padx=(0, 12))
        rc = ttk.Combobox(rf, textvariable=report_type_var,
                          values=["Employee Productivity", "Client-wise Effort",
                                  "Late Submission Tracking", "Missing Timesheets",
                                  "Work Category Analysis"],
                          state="readonly", font=("Calibri", 10), width=26)
        rc.pack(ipady=4); _style_combo(rc)

        result_frame = tk.Frame(pad, bg=PAGE_BG)
        result_frame.pack(fill="both", expand=True)

        def _run_report():
            for w in result_frame.winfo_children():
                w.destroy()
            fd = from_var.get().strip() or "2020-01-01"
            td = to_var.get().strip()   or date.today().isoformat()
            rtype = report_type_var.get()

            def _treeview_report(cols, col_widths, rows_data, value_fn,
                                  title, export_fn=None):
                rc2 = _card(result_frame); rc2.pack(fill="both", expand=True, pady=(0,8))
                _section_hdr(rc2, title)
                if export_fn:
                    exp_btn = tk.Button(rc2, text="📥  Export .xlsx",
                                        font=("Calibri", 9, "bold"), bg=SUCCESS, fg="white",
                                        relief="flat", cursor="hand2", padx=12, pady=5,
                                        command=export_fn)
                    exp_btn.pack(anchor="e", padx=12, pady=(0,4))
                style = ttk.Style(); style.theme_use("clam")
                style.configure("JAA.Treeview", background=PANEL_BG,
                                fieldbackground=PANEL_BG, foreground=TEXT_DARK,
                                rowheight=26, font=("Calibri", 9))
                style.configure("JAA.Treeview.Heading", background=TBL_HDR_BG,
                                foreground=TBL_HDR_FG, font=("Calibri",9,"bold"), relief="flat")
                tf3 = tk.Frame(rc2, bg=PANEL_BG); tf3.pack(fill="both", expand=True, padx=4, pady=4)
                tree = ttk.Treeview(tf3, columns=cols, show="headings",
                                    style="JAA.Treeview", selectmode="browse")
                vsb = ttk.Scrollbar(tf3, orient="vertical", command=tree.yview)
                hsb = ttk.Scrollbar(tf3, orient="horizontal", command=tree.xview)
                tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
                vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
                tree.pack(side="left", fill="both", expand=True)
                for c, cw in zip(cols, col_widths):
                    tree.heading(c, text=c)
                    tree.column(c, width=cw, minwidth=50, anchor="w")
                for i, r in enumerate(rows_data):
                    tree.insert("", "end", values=value_fn(r),
                                tags=("odd" if i%2 else "even",))
                tree.tag_configure("odd",  background=ROW_ODD)
                tree.tag_configure("even", background=ROW_EVEN)
                foot = tk.Frame(result_frame, bg=PANEL_ALT, padx=14, pady=8)
                foot.pack(fill="x", pady=(0,8))
                tk.Label(foot, text=f"✔  {len(rows_data)} records",
                         font=("Calibri", 9, "bold"), bg=PANEL_ALT, fg=SUCCESS).pack(side="left")

            if rtype == "Employee Productivity":
                rows = get_employee_productivity(fd, td)

                def _export_prod():
                    fp = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel","*.xlsx")],
                        initialfile=f"JAA_Productivity_{fd}_{td}.xlsx")
                    if not fp: return
                    try:
                        wb = openpyxl.Workbook(); ws = wb.active
                        ws.title = "Employee Productivity"
                        hdrs = ["Employee","Role","Reporting To","Days Worked",
                                "Entries","Total Hrs","Avg Hrs/Entry"]
                        hf = Font(name="Calibri",bold=True,color="FFFFFF",size=10)
                        hfill = PatternFill("solid",fgColor="06355E")
                        thin = Side(style="thin",color="CCCCCC")
                        bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
                        for ci, h in enumerate(hdrs, 1):
                            c = ws.cell(row=1, column=ci, value=h)
                            c.font=hf; c.fill=hfill; c.border=bdr
                            c.alignment=Alignment(horizontal="center",vertical="center")
                        of = PatternFill("solid",fgColor="D2E1F1")
                        ef = PatternFill("solid",fgColor="FFFFFF")
                        for ri, r in enumerate(rows, 2):
                            fill = of if ri%2==0 else ef
                            vals = [r["full_name"],r["role"],r.get("reporting_to",""),
                                    r["days_worked"],r["entry_count"],
                                    r["total_hrs"],round(r.get("avg_hrs_per_entry",0) or 0, 2)]
                            for ci, v in enumerate(vals, 1):
                                c = ws.cell(row=ri, column=ci, value=v)
                                c.fill=fill; c.border=bdr; c.font=Font(name="Calibri",size=9)
                                c.alignment=Alignment(horizontal="left",vertical="center")
                        ws.freeze_panes="A2"
                        for col in ws.columns:
                            mx = max((len(str(cell.value or "")) for cell in col), default=8)
                            ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx+4,45)
                        wb.save(fp)
                        messagebox.showinfo("Exported ✔", f"Saved to:\n{fp}")
                    except Exception as ex:
                        messagebox.showerror("Error", str(ex))

                cols = ["Employee","Role","Reporting To","Days","Entries","Total Hrs","Avg Hrs"]
                widths = [160, 100, 140, 60, 70, 90, 90]
                _treeview_report(cols, widths, rows,
                    lambda r: (r["full_name"], r["role"], r.get("reporting_to",""),
                               r["days_worked"], r["entry_count"],
                               r["total_hrs"], round(r.get("avg_hrs_per_entry",0) or 0, 2)),
                    f"Employee Productivity — {fd} to {td}", _export_prod)

            elif rtype == "Client-wise Effort":
                rows = get_client_effort(fd, td)

                def _export_client():
                    fp = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel","*.xlsx")],
                        initialfile=f"JAA_ClientEffort_{fd}_{td}.xlsx")
                    if not fp: return
                    try:
                        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Client Effort"
                        hdrs=["Client","Co.Code","Work Category","Entries","Staff","Total Hrs"]
                        hf = Font(name="Calibri",bold=True,color="FFFFFF",size=10)
                        hfill = PatternFill("solid",fgColor="06355E")
                        thin = Side(style="thin",color="CCCCCC")
                        bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
                        for ci, h in enumerate(hdrs, 1):
                            c = ws.cell(row=1,column=ci,value=h)
                            c.font=hf; c.fill=hfill; c.border=bdr
                            c.alignment=Alignment(horizontal="center")
                        of = PatternFill("solid",fgColor="D2E1F1")
                        ef = PatternFill("solid",fgColor="FFFFFF")
                        for ri, r in enumerate(rows, 2):
                            fill = of if ri%2==0 else ef
                            vals=[r.get("client_name",""),r.get("company_code",""),
                                  r.get("work_category",""),r["entry_count"],
                                  r["staff_count"],r["total_hrs"]]
                            for ci, v in enumerate(vals, 1):
                                c=ws.cell(row=ri,column=ci,value=v)
                                c.fill=fill; c.border=bdr; c.font=Font(name="Calibri",size=9)
                        ws.freeze_panes="A2"
                        for col in ws.columns:
                            mx=max((len(str(cell.value or "")) for cell in col),default=8)
                            ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx+4,50)
                        wb.save(fp)
                        messagebox.showinfo("Exported ✔",f"Saved to:\n{fp}")
                    except Exception as ex:
                        messagebox.showerror("Error",str(ex))

                cols = ["Client","Co.Code","Work Category","Entries","Staff","Total Hrs"]
                widths = [200, 90, 160, 70, 60, 90]
                _treeview_report(cols, widths, rows,
                    lambda r: (r.get("client_name",""), r.get("company_code",""),
                               r.get("work_category",""), r["entry_count"],
                               r["staff_count"], r["total_hrs"]),
                    f"Client-wise Effort — {fd} to {td}", _export_client)

            elif rtype == "Late Submission Tracking":
                rows = get_all_submissions()
                late = [r for r in rows if r.get("submission_status") in
                        ("Late", "Partially Filled", "Not Submitted")]
                if fd:
                    late = [r for r in late if r.get("submit_date","") >= fd]
                if td:
                    late = [r for r in late if r.get("submit_date","") <= td]
                cols = ["Date","Employee","Role","Status","Approval","Hrs","Submitted At"]
                widths = [100, 160, 100, 130, 140, 70, 160]
                _treeview_report(cols, widths, late,
                    lambda r: (r.get("submit_date",""), r.get("full_name",""),
                               r.get("role",""), r.get("submission_status",""),
                               r.get("approval_status",""),
                               r.get("total_hrs",0),
                               (r.get("submitted_at","")[:16] if r.get("submitted_at") else "—")),
                    f"Late Submissions — {fd} to {td}")

            elif rtype == "Missing Timesheets":
                rows = get_missing_timesheets(td)
                cols = ["Employee","Role","Reporting To"]
                widths = [200, 120, 160]
                _treeview_report(cols, widths, rows,
                    lambda r: (r["full_name"], r["role"], r.get("reporting_to","—")),
                    f"Missing Timesheets — {td}")

            elif rtype == "Work Category Analysis":
                with get_conn() as conn:
                    rows = conn.execute("""
                        SELECT work_category,
                               COUNT(*) as entries,
                               COUNT(DISTINCT emp_id) as staff_count,
                               ROUND(SUM(total_hrs),2) as total_hrs,
                               COUNT(DISTINCT entry_date) as days_active
                        FROM timesheet_entries
                        WHERE entry_date BETWEEN ? AND ?
                          AND work_category IS NOT NULL AND work_category != ''
                        GROUP BY work_category
                        ORDER BY total_hrs DESC
                    """, (fd, td)).fetchall()
                    rows = [dict(r) for r in rows]
                cols = ["Work Category","Entries","Staff","Total Hrs","Active Days"]
                widths = [200, 80, 70, 100, 100]
                _treeview_report(cols, widths, rows,
                    lambda r: (r.get("work_category",""), r["entries"],
                               r["staff_count"], r["total_hrs"], r["days_active"]),
                    f"Work Category Analysis — {fd} to {td}")

        tk.Button(ctrl, text="▶  Run Report",
                  font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=6,
                  command=_run_report).pack(side="left")

        # Quick range buttons
        def _set_range(days):
            to_var.set(date.today().isoformat())
            from_var.set((date.today() - timedelta(days=days)).isoformat())
        for label, days in [("7 Days", 7), ("30 Days", 30), ("90 Days", 90)]:
            tk.Button(ctrl, text=label,
                      font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      highlightthickness=1, highlightbackground=BORDER_CLR,
                      command=lambda d=days: _set_range(d)).pack(side="left", padx=4)

    # ── SETTINGS TAB ─────────────────────────────────────────────────────────────

    def _build_settings_tab(self, parent):
        self._topbar(parent, "Settings",
                     "  Submission cutoff, reminders, email configuration")

        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True)
        v_canvas = tk.Canvas(outer, bg=PAGE_BG, highlightthickness=0)
        v_sb = ttk.Scrollbar(outer, orient="vertical", command=v_canvas.yview)
        v_canvas.configure(yscrollcommand=v_sb.set)
        v_sb.pack(side="right", fill="y")
        v_canvas.pack(side="left", fill="both", expand=True)
        page = tk.Frame(v_canvas, bg=PAGE_BG)
        pw = v_canvas.create_window((0, 0), window=page, anchor="nw")
        page.bind("<Configure>", lambda e: v_canvas.configure(scrollregion=v_canvas.bbox("all")))
        v_canvas.bind("<Configure>", lambda e: v_canvas.itemconfig(pw, width=e.width))

        pad = tk.Frame(page, bg=PAGE_BG, padx=24, pady=18)
        pad.pack(fill="both", expand=True)

        cfg = load_config()

        def _section(title, icon="⚙"):
            c = _card(pad); c.pack(fill="x", pady=(0, 14))
            _section_hdr(c, title, icon)
            return tk.Frame(c, bg=PANEL_BG, padx=20, pady=14)

        def _row(frame, label, var, width=24, show=None):
            tk.Label(frame, text=label, font=("Calibri", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MID).pack(anchor="w", pady=(8, 2))
            ff = tk.Frame(frame, bg=PANEL_BG, highlightthickness=1,
                          highlightbackground=BORDER_CLR)
            ff.pack(fill="x")
            kw = dict(textvariable=var, font=("Calibri", 10), bg=ENTRY_BG,
                      fg=ENTRY_FG, relief="flat", bd=0, width=width,
                      insertbackground=FOCUS_CLR)
            if show:
                kw["show"] = show
            tk.Entry(ff, **kw).pack(fill="x", ipady=7)

        # ── Submission Timing ────────────────────────────────────────────────
        s1 = _section("Submission Timing", "⏰"); s1.pack(fill="x")
        cutoff_var = tk.StringVar(value=cfg.get("submission_cutoff","20:30"))
        _row(s1, "Submission Cutoff Time (HH:MM) — after this = Late",
             cutoff_var, width=10)
        reminder_var = tk.StringVar(value=cfg.get("reminder_time","19:00"))
        _row(s1, "Popup Reminder Time (HH:MM)", reminder_var, width=10)
        rem_enabled_var = tk.BooleanVar(value=cfg.get("reminder_enabled", True))
        tk.Checkbutton(s1, text="Enable auto popup reminder",
                       variable=rem_enabled_var,
                       font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_DARK,
                       activebackground=PANEL_BG, cursor="hand2").pack(anchor="w", pady=(8,0))

        # ── Email Configuration ──────────────────────────────────────────────
        s2 = _section("Email Configuration (SMTP)", "📧"); s2.pack(fill="x")
        email_enabled_var = tk.BooleanVar(value=cfg.get("email_enabled", False))
        tk.Checkbutton(s2, text="Enable email notifications",
                       variable=email_enabled_var,
                       font=("Calibri", 10), bg=PANEL_BG, fg=TEXT_DARK,
                       activebackground=PANEL_BG, cursor="hand2").pack(anchor="w", pady=(0,8))
        smtp_host_var = tk.StringVar(value=cfg.get("smtp_host","smtp.gmail.com"))
        _row(s2, "SMTP Host", smtp_host_var)
        smtp_port_var = tk.StringVar(value=str(cfg.get("smtp_port",587)))
        _row(s2, "SMTP Port", smtp_port_var, width=8)
        smtp_user_var = tk.StringVar(value=cfg.get("smtp_user",""))
        _row(s2, "Gmail / SMTP Username (email address)", smtp_user_var)
        smtp_pass_var = tk.StringVar(value=cfg.get("smtp_pass",""))
        _row(s2, "App Password (Gmail App Password, not login password)",
             smtp_pass_var, show="●")
        mgr_email_var = tk.StringVar(value=cfg.get("manager_email",""))
        _row(s2, "Manager Email (for daily summary)", mgr_email_var)

        tk.Label(s2,
                 text="💡 For Gmail: enable 2FA and create an App Password at myaccount.google.com/apppasswords",
                 font=("Calibri", 8), bg=PANEL_BG, fg=TEXT_LIGHT,
                 wraplength=600, justify="left").pack(anchor="w", pady=(8,0))

        status_lbl = tk.Label(pad, text="", font=("Calibri", 10),
                              bg=PAGE_BG, fg=SUCCESS)
        status_lbl.pack(anchor="w", pady=(4, 0))

        def save_settings():
            try:
                new_cfg = {
                    "submission_cutoff": cutoff_var.get().strip() or "20:30",
                    "reminder_time":     reminder_var.get().strip() or "19:00",
                    "reminder_enabled":  rem_enabled_var.get(),
                    "smtp_host":         smtp_host_var.get().strip(),
                    "smtp_port":         int(smtp_port_var.get().strip() or 587),
                    "smtp_user":         smtp_user_var.get().strip(),
                    "smtp_pass":         smtp_pass_var.get().strip(),
                    "manager_email":     mgr_email_var.get().strip(),
                    "email_enabled":     email_enabled_var.get(),
                }
                save_config(new_cfg)
                # Restart reminder thread with new settings
                self._start_reminder_thread(new_cfg)
                status_lbl.config(text="✔  Settings saved successfully.", fg=SUCCESS)
                log.info("Settings updated.")
            except Exception as ex:
                status_lbl.config(text=f"⚠  Error: {ex}", fg=DANGER)

        def test_email():
            cfg_now = {
                "email_enabled": email_enabled_var.get(),
                "smtp_host": smtp_host_var.get().strip(),
                "smtp_port": int(smtp_port_var.get().strip() or 587),
                "smtp_user": smtp_user_var.get().strip(),
                "smtp_pass": smtp_pass_var.get().strip(),
                "manager_email": mgr_email_var.get().strip(),
                "submission_cutoff": cutoff_var.get().strip(),
            }
            if not cfg_now["manager_email"]:
                messagebox.showwarning("No Email", "Enter manager email first.")
                return
            status_lbl.config(text="Sending test email…", fg=TEXT_MID)
            pad.update()
            ok, msg = send_daily_summary_email(cfg_now)
            if ok:
                messagebox.showinfo("Test ✔", f"Test email sent to {cfg_now['manager_email']}")
                status_lbl.config(text="✔  Test email sent.", fg=SUCCESS)
            else:
                messagebox.showerror("Failed", f"Email failed:\n{msg}")
                status_lbl.config(text=f"⚠  {msg}", fg=DANGER)

        btn_row = tk.Frame(pad, bg=PAGE_BG); btn_row.pack(anchor="w", pady=(8,0))
        tk.Button(btn_row, text="💾  Save Settings",
                  font=("Calibri", 11, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=24, pady=11,
                  command=save_settings).pack(side="left")
        tk.Button(btn_row, text="📧  Test Email",
                  font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=16, pady=11,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=test_email).pack(side="left", padx=10)

        # ── Data Management ──────────────────────────────────────────────────────
        s3 = _section("Data Management", "🗄️"); s3.pack(fill="x")

        # Backup info
        info_frame = tk.Frame(s3, bg=PANEL_BG,
                              highlightthickness=1, highlightbackground="#D0D8E4")
        info_frame.pack(fill="x", pady=(0, 12))
        tk.Label(info_frame,
                 text="💡  Monthly auto-backup runs on first startup of each month.\n"
                      f"     Backups saved to:  {BACKUP_DIR}\n"
                      "     Last 12 monthly backups are kept automatically.",
                 font=("Calibri", 9), bg="#F0F5FB", fg=TEXT_MID,
                 justify="left", padx=14, pady=10, anchor="w").pack(fill="x")

        # Manual backup button
        def _manual_backup():
            try:
                BACKUP_DIR.mkdir(parents=True, exist_ok=True)
                ts   = datetime.now().strftime("%Y_%m_%d_%H%M%S")
                dest = BACKUP_DIR / f"jaa_manual_backup_{ts}.db"
                shutil.copy2(str(DB_PATH), str(dest))
                messagebox.showinfo("Backup ✔",
                    f"Backup saved to:\n{dest}")
                status_lbl.config(text="✔  Manual backup created.", fg=SUCCESS)
            except Exception as ex:
                messagebox.showerror("Backup Failed", str(ex))

        # DB size info
        def _refresh_db_info():
            try:
                size_mb = DB_PATH.stat().st_size / (1024 * 1024)
                backups = list(BACKUP_DIR.glob("jaa_*backup_*.db"))
                db_info_lbl.config(
                    text=f"  Database size: {size_mb:.2f} MB  ·  "
                         f"{len(backups)} backup(s) in backups/ folder")
            except Exception:
                db_info_lbl.config(text="  Could not read DB info.")

        db_info_lbl = tk.Label(s3, text="", font=("Calibri", 9),
                               bg=PANEL_BG, fg=TEXT_MID)
        db_info_lbl.pack(anchor="w", pady=(0, 8))
        _refresh_db_info()

        bk_row = tk.Frame(s3, bg=PANEL_BG); bk_row.pack(anchor="w", pady=(0, 16))
        tk.Button(bk_row, text="💾  Backup Now",
                  font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", padx=16, pady=8,
                  command=_manual_backup).pack(side="left")
        tk.Button(bk_row, text="📂  Open Backup Folder",
                  font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=14, pady=8,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=lambda: os.startfile(str(BACKUP_DIR))
                  ).pack(side="left", padx=8)
        tk.Button(bk_row, text="🔄  Refresh Info",
                  font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                  relief="flat", cursor="hand2", padx=10, pady=8,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  command=_refresh_db_info).pack(side="left")

        # ── Clear Trial Data ───────────────────────────────────────────────────
        tk.Frame(s3, bg=BORDER_CLR, height=1).pack(fill="x", pady=(4, 12))
        tk.Label(s3, text="⚠️  Clear Trial / Test Data",
                 font=("Calibri", 10, "bold"), bg=PANEL_BG, fg=DANGER).pack(anchor="w")
        tk.Label(s3,
                 text="Permanently deletes all timesheet entries, submissions, "
                      "attendance logs and email logs.\n"
                      "Employees, companies and work mapping are preserved by default.",
                 font=("Calibri", 9), bg=PANEL_BG, fg=TEXT_MID,
                 justify="left", wraplength=560).pack(anchor="w", pady=(4, 10))

        keep_emp_var = tk.BooleanVar(value=True)
        keep_co_var  = tk.BooleanVar(value=True)
        tk.Checkbutton(s3, text="Keep employees & PINs",
                       variable=keep_emp_var,
                       font=("Calibri", 9), bg=PANEL_BG, fg=TEXT_DARK,
                       activebackground=PANEL_BG, cursor="hand2").pack(anchor="w")
        tk.Checkbutton(s3, text="Keep companies master list",
                       variable=keep_co_var,
                       font=("Calibri", 9), bg=PANEL_BG, fg=TEXT_DARK,
                       activebackground=PANEL_BG, cursor="hand2").pack(anchor="w", pady=(2, 10))

        def _do_clear():
            confirm = messagebox.askyesno(
                "⚠️ Confirm Clear",
                "This will permanently delete all trial data.\n\n"
                "A backup will be created first.\n\n"
                "Are you sure you want to continue?",
                icon="warning")
            if not confirm:
                return
            # Auto-backup before clearing
            try:
                ts   = datetime.now().strftime("%Y_%m_%d_%H%M%S")
                dest = BACKUP_DIR / f"jaa_pre_clear_backup_{ts}.db"
                shutil.copy2(str(DB_PATH), str(dest))
            except Exception as ex:
                messagebox.showerror("Backup Failed",
                    f"Could not create backup before clearing:\n{ex}\n\nAborting.")
                return
            counts = clear_trial_data(
                keep_employees=keep_emp_var.get(),
                keep_companies=keep_co_var.get())
            total = sum(counts.values())
            messagebox.showinfo("Cleared ✔",
                f"Trial data cleared successfully.\n\n"
                f"{total} records removed.\n"
                f"Backup saved to:\n{dest}")
            status_lbl.config(text="✔  Trial data cleared. Backup saved.", fg=SUCCESS)
            _refresh_db_info()

        tk.Button(s3, text="🗑️  Clear Trial Data",
                  font=("Calibri", 10, "bold"), bg="#EEE8E0", fg=DANGER,
                  relief="flat", cursor="hand2", padx=18, pady=9,
                  command=_do_clear).pack(anchor="w")


    # ════════════════════════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════════════════════════
    # LEARNING HUB  v3 — Collaborative Knowledge Space
    # All employees: upload, view, download, search.
    # Admin/Manager: pin, archive, delete.
    # ════════════════════════════════════════════════════════════════════════════

    def _build_learning_hub(self, parent, emp, is_admin=False):
        """
        Full Learning Hub / JAA Learning Space module.
        is_admin=True  → upload / edit / delete controls visible
        is_admin=False → view / download only (unless role is Partner/Associate)
        """
        # Determine effective permission level
        can_upload = True   # all employees can upload
        if emp and emp.get("role", "") in ("Partner", "Associate"):
            can_upload = True

        # ── Top bar ──────────────────────────────────────────────────────────
        self._topbar(parent, "🎓  JAA Learning Space",
                     "  Notice Board · SOPs · Guidelines · Training · Reference Materials")

        # ── Pinned notices banner (shown if any pinned items exist) ───────────
        pinned_banner_frame = tk.Frame(parent, bg=PAGE_BG)
        pinned_banner_frame.pack(fill="x")

        def _refresh_pinned_banner():
            for w in pinned_banner_frame.winfo_children():
                w.destroy()
            pinned = [m for m in get_learning_materials(active_only=True)
                      if m.get("is_pinned")]
            # Filter out expired
            today_str = date.today().isoformat()
            pinned = [m for m in pinned
                      if not m.get("expiry_date") or m["expiry_date"] >= today_str]
            if not pinned:
                return
            banner = tk.Frame(pinned_banner_frame, bg="#FFF4CC",
                              highlightthickness=1, highlightbackground="#E0C850",
                              padx=14, pady=8)
            banner.pack(fill="x", padx=16, pady=(8,0))
            tk.Label(banner, text="📌  PINNED NOTICES",
                     font=("Calibri", 8, "bold"), bg="#FFF4CC",
                     fg="#7A5A00").pack(side="left")
            for pm in pinned[:5]:
                pri_color = "#B52A2A" if pm.get("priority") == "High" else "#7A5A00"
                lbl = tk.Label(banner,
                               text=f"  •  {pm['title']}" + (
                                   f"  ⚠" if pm.get("priority") == "High" else ""),
                               font=("Calibri", 9, "bold" if pm.get("priority")=="High" else "normal"),
                               bg="#FFF4CC", fg=pri_color, cursor="hand2")
                lbl.pack(side="left", padx=(8,0))
                lbl.bind("<Button-1>", lambda e, m=pm: _open_notice_viewer(m))
            if len(pinned) > 5:
                tk.Label(banner, text=f"  +{len(pinned)-5} more",
                         font=("Calibri", 8), bg="#FFF4CC",
                         fg="#7A5A00").pack(side="left", padx=(6,0))

        # ── Two-column layout: left category sidebar + right content ─────────
        body = tk.Frame(parent, bg=PAGE_BG)
        body.pack(fill="both", expand=True)

        # ── LEFT: Category sidebar ───────────────────────────────────────────
        cat_sidebar = tk.Frame(body, bg=SIDEBAR_BG, width=190)
        cat_sidebar.pack(side="left", fill="y")
        cat_sidebar.pack_propagate(False)

        tk.Label(cat_sidebar, text="CATEGORIES",
                 font=("Calibri", 8, "bold"), bg=SIDEBAR_BG,
                 fg=SIDEBAR_ICN, anchor="w").pack(fill="x", padx=14, pady=(14, 6))

        # Scrollable list of categories
        cat_canvas = tk.Canvas(cat_sidebar, bg=SIDEBAR_BG, highlightthickness=0)
        cat_scroll = ttk.Scrollbar(cat_sidebar, orient="vertical",
                                   command=cat_canvas.yview)
        cat_canvas.configure(yscrollcommand=cat_scroll.set)
        cat_scroll.pack(side="right", fill="y")
        cat_canvas.pack(side="left", fill="both", expand=True)
        cat_inner = tk.Frame(cat_canvas, bg=SIDEBAR_BG)
        cat_cw_id = cat_canvas.create_window((0, 0), window=cat_inner, anchor="nw")
        cat_inner.bind("<Configure>",
            lambda e: cat_canvas.configure(scrollregion=cat_canvas.bbox("all")))
        cat_canvas.bind("<Configure>",
            lambda e: cat_canvas.itemconfig(cat_cw_id, width=e.width))

        # ── RIGHT: Main content area ─────────────────────────────────────────
        right = tk.Frame(body, bg=PAGE_BG)
        right.pack(side="left", fill="both", expand=True)

        # Search bar at top of right panel
        search_bar = tk.Frame(right, bg=PAGE_BG, padx=16, pady=10)
        search_bar.pack(fill="x")

        tk.Label(search_bar, text="🔍", font=("Calibri", 12),
                 bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
        search_var = tk.StringVar()
        sf = tk.Frame(search_bar, bg=PANEL_BG, highlightthickness=1,
                      highlightbackground=BORDER_CLR)
        sf.pack(side="left", fill="x", expand=True, padx=(6, 10))
        search_entry = tk.Entry(sf, textvariable=search_var,
                                font=("Calibri", 11), bg=PANEL_BG, fg=ENTRY_FG,
                                relief="flat", bd=0, insertbackground=FOCUS_CLR)
        search_entry.pack(fill="x", ipady=8, padx=8)
        search_entry.insert(0, "Search title, category, keywords…")
        search_entry.config(fg=TEXT_LIGHT)

        def _search_focus_in(e):
            if search_entry.get() == "Search title, category, keywords…":
                search_entry.delete(0, "end")
                search_entry.config(fg=ENTRY_FG)
        def _search_focus_out(e):
            if not search_entry.get():
                search_entry.insert(0, "Search title, category, keywords…")
                search_entry.config(fg=TEXT_LIGHT)
        search_entry.bind("<FocusIn>",  _search_focus_in)
        search_entry.bind("<FocusOut>", _search_focus_out)

        type_var = tk.StringVar(value="All")
        tf = tk.Frame(search_bar, bg=PAGE_BG, highlightthickness=1,
                      highlightbackground=BORDER_CLR)
        tf.pack(side="left", padx=(0, 10))
        type_combo = ttk.Combobox(tf, textvariable=type_var,
                                  values=["All","Notice","PDF","PPTX","DOCX","Excel",
                                          "YouTube","Article","Other"],
                                  state="readonly", font=("Calibri", 10), width=10)
        type_combo.pack(ipady=6)
        _style_combo(type_combo)

        tk.Button(search_bar, text="Search",
                  font=("Calibri", 10, "bold"), bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=7,
                  command=lambda: _load_content()).pack(side="left")

        if can_upload:
            tk.Button(search_bar, text="＋  Add Material",
                      font=("Calibri", 10, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=14, pady=7,
                      command=lambda: _open_upload_dialog()).pack(side="left", padx=(10, 0))
            if is_admin:
                tk.Button(search_bar, text="⚙  Manage Categories",
                          font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                          relief="flat", cursor="hand2", padx=10, pady=7,
                          highlightthickness=1, highlightbackground=BORDER_CLR,
                          command=lambda: _open_category_manager()).pack(side="left", padx=(8, 0))

        # Scrollable content canvas
        content_outer = tk.Frame(right, bg=PAGE_BG)
        content_outer.pack(fill="both", expand=True)
        content_canvas = tk.Canvas(content_outer, bg=PAGE_BG, highlightthickness=0)
        content_vsb   = ttk.Scrollbar(content_outer, orient="vertical",
                                      command=content_canvas.yview)
        content_canvas.configure(yscrollcommand=content_vsb.set)
        content_vsb.pack(side="right", fill="y")
        content_canvas.pack(side="left", fill="both", expand=True)
        content_frame = tk.Frame(content_canvas, bg=PAGE_BG)
        cfid = content_canvas.create_window((0, 0), window=content_frame, anchor="nw")
        content_frame.bind("<Configure>",
            lambda e: content_canvas.configure(scrollregion=content_canvas.bbox("all")))
        content_canvas.bind("<Configure>",
            lambda e: content_canvas.itemconfig(cfid, width=e.width))
        content_canvas.bind_all("<MouseWheel>",
            lambda e: content_canvas.yview_scroll(-1*(e.delta//120), "units"))

        # State tracking
        state = {"category": "All", "materials": []}

        # ── Material type icon map ────────────────────────────────────────────
        _TYPE_ICON = {
            "Notice":  "📢",
            "PDF":     "🔴",
            "PPTX":    "🟠",
            "DOCX":    "🔵",
            "Excel":   "🟢",
            "YouTube": "▶",
            "Article": "📝",
            "Other":   "📎",
        }
        _TYPE_COLOR = {
            "Notice":  "#06355E",
            "PDF":     "#B52A2A",
            "PPTX":    "#8A3A00",
            "DOCX":    "#024277",
            "Excel":   "#1A6B45",
            "YouTube": "#CC0000",
            "Article": "#46278A",
            "Other":   "#5A5A5A",
        }

        # ── Load and render material cards ────────────────────────────────────
        def _load_content(category=None):
            if category is not None:
                state["category"] = category
            for w in content_frame.winfo_children():
                w.destroy()

            raw_search = search_var.get().strip()
            kw = raw_search if raw_search != "Search title, category, keywords…" else None
            mats = get_learning_materials(
                category=state["category"],
                material_type=type_var.get() if type_var.get() != "All" else None,
                search=kw
            )
            state["materials"] = mats

            pad = tk.Frame(content_frame, bg=PAGE_BG, padx=16, pady=12)
            pad.pack(fill="both", expand=True)

            # Section heading
            hdr_txt = state["category"] if state["category"] != "All" else "All Materials"
            hdr_row = tk.Frame(pad, bg=PAGE_BG)
            hdr_row.pack(fill="x", pady=(0, 12))
            tk.Label(hdr_row, text=hdr_txt,
                     font=("Georgia", 14, "bold"), bg=PAGE_BG, fg=TEXT_DARK).pack(side="left")
            tk.Label(hdr_row, text=f"  {len(mats)} item(s)",
                     font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_LIGHT).pack(side="left")

            if not mats:
                empty = _card(pad); empty.pack(fill="x", pady=20)
                tk.Label(empty,
                         text="📭  No materials found. Use '＋ Add Material' to upload content.",
                         font=("Calibri", 11), bg=PANEL_BG, fg=TEXT_LIGHT,
                         pady=30).pack()
                return

            # Card grid — 2 columns when wide enough
            grid = tk.Frame(pad, bg=PAGE_BG)
            grid.pack(fill="both", expand=True)
            grid.columnconfigure(0, weight=1)
            grid.columnconfigure(1, weight=1)

            for idx, mat in enumerate(mats):
                row_i = idx // 2
                col_i = idx  % 2
                _make_material_card(grid, mat, row_i, col_i, can_upload, emp)

        def _make_material_card(grid, mat, row_i, col_i, can_edit, emp_ctx):
            """Render a single material as a card tile."""
            mtype   = mat.get("material_type", "Other")
            icon    = _TYPE_ICON.get(mtype, "📎")
            tcolor  = _TYPE_COLOR.get(mtype, "#5A5A5A")
            is_pinned  = bool(mat.get("is_pinned", 0))
            priority   = mat.get("priority", "Normal")
            expiry     = mat.get("expiry_date", "")

            # Notice cards get a yellow highlight bg; high priority get a red tint
            card_bg = PANEL_BG
            if mtype == "Notice":
                card_bg = "#FFFFF4" if priority != "High" else "#FFF4F4"

            outer = tk.Frame(grid, bg=card_bg, highlightthickness=1,
                             highlightbackground=(
                                 "#E0C850" if mtype == "Notice" else BORDER_CLR),
                             cursor="hand2")
            outer.grid(row=row_i, column=col_i, padx=(0,8), pady=(0,8), sticky="nsew")

            # Colour stripe on left by type; wider for pinned
            stripe_w = 7 if is_pinned else 5
            stripe = tk.Frame(outer, bg=tcolor, width=stripe_w)
            stripe.pack(side="left", fill="y")

            body = tk.Frame(outer, bg=card_bg, padx=12, pady=10)
            body.pack(side="left", fill="both", expand=True)

            # Type badge + category pill + pin/priority badges
            top_row = tk.Frame(body, bg=card_bg); top_row.pack(fill="x")
            tk.Label(top_row, text=f"{icon}  {mtype}",
                     font=("Calibri", 8, "bold"), bg=tcolor, fg="white",
                     padx=6, pady=2).pack(side="left")
            if mat.get("category"):
                tk.Label(top_row, text=f"  {mat['category']}",
                         font=("Calibri", 8), bg=PANEL_ALT, fg=TEXT_MID,
                         padx=6, pady=2).pack(side="left", padx=(6,0))
            if is_pinned:
                tk.Label(top_row, text="📌",
                         font=("Calibri", 10), bg=card_bg,
                         fg="#7A5A00").pack(side="right")
            if priority == "High":
                tk.Label(top_row, text="⚠ HIGH",
                         font=("Calibri", 7, "bold"), bg="#B52A2A", fg="white",
                         padx=4, pady=1).pack(side="right", padx=(0,4))
            if mat.get("view_count", 0):
                tk.Label(top_row, text=f"👁 {mat['view_count']}",
                         font=("Calibri", 8), bg=card_bg, fg=TEXT_LIGHT).pack(side="right", padx=(0,4))

            # Title
            tk.Label(body, text=mat["title"],
                     font=("Calibri", 11, "bold"), bg=PANEL_BG, fg=TEXT_DARK,
                     anchor="w", wraplength=280, justify="left").pack(fill="x", pady=(6,2))

            # Description (trimmed)
            desc = (mat.get("description") or "").strip()
            if desc:
                tk.Label(body, text=desc[:120] + ("…" if len(desc)>120 else ""),
                         font=("Calibri", 9), bg=PANEL_BG, fg=TEXT_MID,
                         anchor="w", wraplength=280, justify="left").pack(fill="x")

            # Tags
            tags = (mat.get("tags") or "").strip()
            if tags:
                tag_row = tk.Frame(body, bg=PANEL_BG); tag_row.pack(fill="x", pady=(4,0))
                for t in tags.split(",")[:5]:
                    t = t.strip()
                    if t:
                        tk.Label(tag_row, text=f"#{t}",
                                 font=("Calibri", 8), bg=PAGE_BG, fg=ACCENT_GOLD,
                                 padx=4, pady=1).pack(side="left", padx=(0,4))

            # Meta row
            meta = tk.Frame(body, bg=PANEL_BG); meta.pack(fill="x", pady=(6,0))
            by   = mat.get("uploaded_by","")
            dt   = mat.get("upload_date","")
            tk.Label(meta, text=f"📅 {dt}  ·  👤 {by}",
                     font=("Calibri", 8), bg=PANEL_BG, fg=TEXT_LIGHT).pack(side="left")

            # Action buttons
            btn_row = tk.Frame(body, bg=PANEL_BG); btn_row.pack(fill="x", pady=(8,0))

            # Expiry label if set
            if expiry:
                today_s = date.today().isoformat()
                exp_color = DANGER if expiry < today_s else TEXT_LIGHT
                exp_txt   = f"⏳ Expires: {expiry}" + (" (EXPIRED)" if expiry < today_s else "")
                tk.Label(body, text=exp_txt, font=("Calibri", 8),
                         bg=card_bg, fg=exp_color).pack(anchor="w", pady=(2,0))

            if mtype == "Notice":
                tk.Button(btn_row, text="📢  Read Notice",
                          font=("Calibri", 9, "bold"), bg="#06355E", fg="white",
                          relief="flat", cursor="hand2", padx=10, pady=5,
                          command=lambda m=mat: _open_notice_viewer(m)).pack(side="left")
            elif mtype == "YouTube":
                tk.Button(btn_row, text="▶  Open Video",
                          font=("Calibri", 9, "bold"), bg="#CC0000", fg="white",
                          relief="flat", cursor="hand2", padx=10, pady=5,
                          command=lambda m=mat: _open_youtube(m)).pack(side="left")
            elif mtype == "Article":
                tk.Button(btn_row, text="📖  Read Article",
                          font=("Calibri", 9, "bold"), bg="#46278A", fg="white",
                          relief="flat", cursor="hand2", padx=10, pady=5,
                          command=lambda m=mat: _open_article_reader(m)).pack(side="left")
            elif mat.get("file_path"):
                tk.Button(btn_row, text="📥  Open / Download",
                          font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                          relief="flat", cursor="hand2", padx=10, pady=5,
                          command=lambda m=mat: _open_file(m)).pack(side="left")

            if can_edit:
                # Pin / Unpin toggle
                pin_txt = "📌 Unpin" if is_pinned else "📌 Pin"
                def _toggle_pin(m=mat):
                    with get_conn() as conn:
                        new_val = 0 if m.get("is_pinned") else 1
                        conn.execute(
                            "UPDATE learning_materials SET is_pinned=? WHERE id=?",
                            (new_val, m["id"])
                        )
                    _load_content(); _refresh_pinned_banner()
                tk.Button(btn_row, text=pin_txt,
                          font=("Calibri", 8), bg="#FFF4CC" if not is_pinned else "#E0C850",
                          fg="#7A5A00",
                          relief="flat", cursor="hand2", padx=8, pady=5,
                          highlightthickness=1, highlightbackground="#E0C850",
                          command=_toggle_pin).pack(side="left", padx=(0,4))
                tk.Button(btn_row, text="✏",
                          font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                          relief="flat", cursor="hand2", padx=8, pady=5,
                          highlightthickness=1, highlightbackground=BORDER_CLR,
                          command=lambda m=mat: _open_upload_dialog(edit_mat=m)
                          ).pack(side="left", padx=(0,4))
                tk.Button(btn_row, text="🗑",
                          font=("Calibri", 9), bg=PAGE_BG, fg=DANGER,
                          relief="flat", cursor="hand2", padx=8, pady=5,
                          highlightthickness=1, highlightbackground=BORDER_CLR,
                          command=lambda m=mat: _confirm_delete(m)
                          ).pack(side="left")

        # ── Open file (PDF, PPTX, DOCX, Excel) ───────────────────────────────
        def _open_file(mat):
            import subprocess
            fp = mat.get("file_path","")
            if not fp or not Path(fp).exists():
                messagebox.showerror("File Not Found",
                    f"The file could not be located:\n{fp}\n\n"
                    "It may have been moved or deleted.")
                return
            increment_view_count(mat["id"])
            try:
                if sys.platform == "win32":
                    os.startfile(fp)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", fp])
                else:
                    subprocess.Popen(["xdg-open", fp])
            except Exception as e:
                messagebox.showerror("Cannot Open", str(e))

        # ── Open YouTube in browser ───────────────────────────────────────────
        def _open_youtube(mat):
            import webbrowser
            url = mat.get("youtube_url","").strip()
            if not url:
                messagebox.showwarning("No URL", "No YouTube URL stored for this entry.")
                return
            increment_view_count(mat["id"])
            webbrowser.open(url)

        # ── Article reader popup ──────────────────────────────────────────────
        def _open_article_reader(mat):
            increment_view_count(mat["id"])
            win = tk.Toplevel(self)
            win.title(f"📖  {mat['title']}")
            win.geometry("760x580")
            win.configure(bg=PAGE_BG)
            win.resizable(True, True)

            # Topbar inside popup
            hdr = tk.Frame(win, bg=SIDEBAR_BG, padx=20, pady=12)
            hdr.pack(fill="x")
            tk.Label(hdr, text=f"📝  {mat['title']}",
                     font=("Georgia", 13, "bold"), bg=SIDEBAR_BG,
                     fg=SIDEBAR_FG).pack(side="left")
            cat = mat.get("category","")
            if cat:
                tk.Label(hdr, text=f"  [{cat}]",
                         font=("Calibri", 9), bg=SIDEBAR_BG,
                         fg=SIDEBAR_ICN).pack(side="left")

            body_frame = tk.Frame(win, bg=PANEL_BG, padx=24, pady=18)
            body_frame.pack(fill="both", expand=True, padx=10, pady=10)

            txt = scrolledtext.ScrolledText(body_frame,
                                            font=("Calibri", 11),
                                            bg=PANEL_BG, fg=TEXT_DARK,
                                            relief="flat", wrap="word",
                                            insertbackground=FOCUS_CLR)
            txt.pack(fill="both", expand=True)
            article_body = get_article_body(mat["id"])
            txt.insert("1.0", article_body or "(No content)")
            txt.config(state="disabled")

            # Edit button if admin
            if can_upload:
                btn_row = tk.Frame(win, bg=PAGE_BG, pady=8)
                btn_row.pack(fill="x", padx=10)
                tk.Button(btn_row, text="✏  Edit Article",
                          font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                          relief="flat", cursor="hand2", padx=12, pady=6,
                          command=lambda: [win.destroy(), _open_upload_dialog(edit_mat=mat)]
                          ).pack(side="left")

        # ── Notice viewer popup ──────────────────────────────────────────────
        def _open_notice_viewer(mat):
            increment_view_count(mat["id"])
            win = tk.Toplevel(self)
            win.title(f"📢  {mat['title']}")
            win.geometry("620x500")
            win.configure(bg=PAGE_BG)
            win.resizable(True, True)
            win.attributes("-topmost", True)
            win.after(200, lambda: win.attributes("-topmost", False))

            # Header stripe
            tk.Frame(win, bg="#06355E", height=6).pack(fill="x")
            hdr = tk.Frame(win, bg="#06355E", padx=22, pady=14)
            hdr.pack(fill="x")
            title_row = tk.Frame(hdr, bg="#06355E"); title_row.pack(fill="x")
            tk.Label(title_row, text="📢",
                     font=("Calibri", 18), bg="#06355E", fg="#FFF4CC").pack(side="left")
            tk.Label(title_row, text=f"  {mat['title']}",
                     font=("Georgia", 13, "bold"), bg="#06355E",
                     fg="#E0E9F4", wraplength=500, justify="left").pack(side="left")
            priority = mat.get("priority","Normal")
            if priority == "High":
                tk.Label(title_row, text="  ⚠ HIGH PRIORITY",
                         font=("Calibri", 9, "bold"), bg="#B52A2A",
                         fg="white", padx=6, pady=3).pack(side="left", padx=(10,0))

            meta_row = tk.Frame(hdr, bg="#06355E"); meta_row.pack(fill="x", pady=(6,0))
            by  = mat.get("uploaded_by","")
            dt  = mat.get("upload_date","")
            cat = mat.get("category","")
            exp = mat.get("expiry_date","")
            meta_txt = f"📅 {dt}  ·  👤 {by}"
            if cat: meta_txt += f"  ·  📂 {cat}"
            if exp: meta_txt += f"  ·  ⏳ Expires: {exp}"
            tk.Label(meta_row, text=meta_txt,
                     font=("Calibri", 9), bg="#06355E",
                     fg="#7AAFD4").pack(side="left")

            # Body
            body_frame = tk.Frame(win, bg=PANEL_BG, padx=24, pady=18)
            body_frame.pack(fill="both", expand=True, padx=12, pady=12)
            txt = scrolledtext.ScrolledText(body_frame,
                                            font=("Calibri", 11),
                                            bg=PANEL_BG, fg=TEXT_DARK,
                                            relief="flat", wrap="word",
                                            insertbackground=FOCUS_CLR)
            txt.pack(fill="both", expand=True)
            # Show description as notice body, or article body if linked
            body_text = (mat.get("description") or "").strip()
            article_body = get_article_body(mat["id"])
            if article_body:
                body_text = article_body
            txt.insert("1.0", body_text or "(No content)")
            txt.config(state="disabled")

            footer = tk.Frame(win, bg=PAGE_BG, pady=10)
            footer.pack(fill="x", padx=12)
            if can_upload:
                tk.Button(footer, text="✏  Edit",
                          font=("Calibri", 9, "bold"), bg=ACCENT_BLUE, fg="white",
                          relief="flat", cursor="hand2", padx=12, pady=6,
                          command=lambda: [win.destroy(), _open_upload_dialog(edit_mat=mat)]
                          ).pack(side="left")
            tk.Button(footer, text="✕  Close",
                      font=("Calibri", 9), bg=PAGE_BG, fg=TEXT_MID,
                      relief="flat", cursor="hand2", padx=12, pady=6,
                      highlightthickness=1, highlightbackground=BORDER_CLR,
                      command=win.destroy).pack(side="left", padx=(8,0))

        # ── Delete confirmation ───────────────────────────────────────────────
        def _confirm_delete(mat):
            if messagebox.askyesno("Delete Material",
                    f"Delete '{mat['title']}'?\n\n"
                    "The file will be removed from the Learning Hub "
                    "(the stored file will also be deleted)."):
                delete_learning_material(mat["id"], delete_file=True)
                _load_content()

        # ── Upload / Edit dialog ──────────────────────────────────────────────
        def _open_upload_dialog(edit_mat=None):
            is_edit = edit_mat is not None
            win = tk.Toplevel(self)
            win.title("Edit Material" if is_edit else "Add Learning Material")
            win.geometry("640x620")
            win.configure(bg=PAGE_BG)
            win.resizable(True, True)
            win.grab_set()

            # Header
            hdr = tk.Frame(win, bg=SIDEBAR_BG, padx=20, pady=12); hdr.pack(fill="x")
            tk.Label(hdr, text="✏  Edit Material" if is_edit else "＋  Add Learning Material",
                     font=("Georgia", 12, "bold"), bg=SIDEBAR_BG,
                     fg=SIDEBAR_FG).pack(side="left")

            # Scrollable form
            sc = tk.Canvas(win, bg=PAGE_BG, highlightthickness=0)
            sb2 = ttk.Scrollbar(win, orient="vertical", command=sc.yview)
            sc.configure(yscrollcommand=sb2.set)
            sb2.pack(side="right", fill="y")
            sc.pack(side="left", fill="both", expand=True)
            form = tk.Frame(sc, bg=PAGE_BG, padx=24, pady=18)
            sc_id = sc.create_window((0,0), window=form, anchor="nw")
            form.bind("<Configure>",
                lambda e: sc.configure(scrollregion=sc.bbox("all")))
            sc.bind("<Configure>",
                lambda e: sc.itemconfig(sc_id, width=e.width))

            def _lbl(text):
                tk.Label(form, text=text, font=("Calibri", 9, "bold"),
                         bg=PAGE_BG, fg=TEXT_MID).pack(anchor="w", pady=(10,2))

            def _entry_field(var, width=50):
                ff = tk.Frame(form, bg=PANEL_BG, highlightthickness=1,
                              highlightbackground=BORDER_CLR)
                ff.pack(fill="x")
                tk.Entry(ff, textvariable=var, font=("Calibri", 10),
                         bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                         insertbackground=FOCUS_CLR).pack(fill="x", ipady=7, padx=6)

            # Title
            _lbl("Title *")
            title_var = tk.StringVar(value=edit_mat["title"] if is_edit else "")
            _entry_field(title_var)

            # Material type (locked in edit mode)
            _lbl("Material Type *")
            mtype_var = tk.StringVar(value=edit_mat["material_type"] if is_edit else "PDF")
            if is_edit:
                tk.Label(form, text=edit_mat["material_type"],
                         font=("Calibri", 10, "bold"), bg=PAGE_BG, fg=TEXT_DARK).pack(anchor="w")
            else:
                mtf = tk.Frame(form, bg=PANEL_BG, highlightthickness=1,
                               highlightbackground=BORDER_CLR)
                mtf.pack(fill="x")
                mt_cb = ttk.Combobox(mtf, textvariable=mtype_var,
                                     values=["Notice","PDF","PPTX","DOCX","Excel",
                                             "YouTube","Article","Other"],
                                     state="readonly", font=("Calibri",10), width=18)
                mt_cb.pack(ipady=6, padx=4, anchor="w")
                _style_combo(mt_cb)

            # Category
            _lbl("Category")
            cats = [c["name"] for c in get_learning_categories()]
            cat_var = tk.StringVar(value=edit_mat.get("category","") if is_edit else (cats[0] if cats else ""))
            catf = tk.Frame(form, bg=PANEL_BG, highlightthickness=1,
                            highlightbackground=BORDER_CLR)
            catf.pack(fill="x")
            cat_cb = ttk.Combobox(catf, textvariable=cat_var, values=cats,
                                  font=("Calibri",10), width=28)
            cat_cb.pack(ipady=6, padx=4, anchor="w")
            _style_combo(cat_cb)

            # Description
            _lbl("Description")
            desc_var = tk.StringVar(value=edit_mat.get("description","") if is_edit else "")
            _entry_field(desc_var)

            # Tags
            _lbl("Tags (comma separated, e.g. gst, reconciliation, sop)")
            tags_var = tk.StringVar(value=edit_mat.get("tags","") if is_edit else "")
            _entry_field(tags_var)

            # ── Notice / Admin controls ──────────────────────────────────────────
            pin_var      = tk.BooleanVar(value=bool(edit_mat.get("is_pinned",0)) if is_edit else False)
            priority_var = tk.StringVar(value=edit_mat.get("priority","Normal") if is_edit else "Normal")
            expiry_var   = tk.StringVar(value=edit_mat.get("expiry_date","") if is_edit else "")

            if can_upload:
                ctrl_row = tk.Frame(form, bg=PAGE_BG); ctrl_row.pack(fill="x", pady=(10,0))

                # Pin toggle
                tk.Checkbutton(ctrl_row, text="📌  Pin this item to Notice Banner",
                               variable=pin_var,
                               font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_DARK,
                               activebackground=PAGE_BG, cursor="hand2"
                               ).pack(side="left")

                # Priority
                tk.Label(ctrl_row, text="  Priority:",
                         font=("Calibri", 9, "bold"), bg=PAGE_BG,
                         fg=TEXT_MID).pack(side="left", padx=(16,4))
                pf = tk.Frame(ctrl_row, bg=PANEL_BG, highlightthickness=1,
                              highlightbackground=BORDER_CLR)
                pf.pack(side="left")
                pc = ttk.Combobox(pf, textvariable=priority_var,
                                  values=["Normal","High","Low"],
                                  state="readonly", font=("Calibri",9), width=8)
                pc.pack(ipady=4, padx=2)
                _style_combo(pc)

                # Expiry date
                tk.Label(ctrl_row, text="  Expiry (YYYY-MM-DD):",
                         font=("Calibri", 9, "bold"), bg=PAGE_BG,
                         fg=TEXT_MID).pack(side="left", padx=(14,4))
                ef3 = tk.Frame(ctrl_row, bg=PANEL_BG, highlightthickness=1,
                               highlightbackground=BORDER_CLR)
                ef3.pack(side="left")
                tk.Entry(ef3, textvariable=expiry_var, font=("Calibri",9),
                         bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                         width=14, insertbackground=FOCUS_CLR).pack(ipady=4, padx=4)
                tk.Label(ctrl_row, text="(blank = no expiry)",
                         font=("Calibri", 8), bg=PAGE_BG,
                         fg=TEXT_LIGHT).pack(side="left", padx=(4,0))

            # --- Type-specific fields shown/hidden dynamically ---
            file_path_var   = tk.StringVar()
            yt_url_var      = tk.StringVar(value=edit_mat.get("youtube_url","") if is_edit else "")
            article_var     = tk.StringVar()
            article_text    = {"widget": None}

            type_fields = tk.Frame(form, bg=PAGE_BG)
            type_fields.pack(fill="x", pady=(8,0))

            def _render_type_fields(mtype=None):
                if mtype is None:
                    mtype = mtype_var.get()
                for w in type_fields.winfo_children():
                    w.destroy()

                if mtype == "Notice":
                    tk.Label(type_fields, text="Notice Body *",
                             font=("Calibri",9,"bold"), bg=PAGE_BG, fg=TEXT_MID
                             ).pack(anchor="w", pady=(10,2))
                    nf2 = tk.Frame(type_fields, bg=PANEL_BG, highlightthickness=1,
                                   highlightbackground=BORDER_CLR)
                    nf2.pack(fill="x")
                    ntxt = scrolledtext.ScrolledText(nf2, font=("Calibri",10),
                                                     bg=PANEL_BG, fg=ENTRY_FG,
                                                     relief="flat", height=7,
                                                     wrap="word",
                                                     insertbackground=FOCUS_CLR)
                    ntxt.pack(fill="x", padx=4, pady=4)
                    if is_edit and mat.get("id"):
                        existing = get_article_body(edit_mat["id"])
                        ntxt.insert("1.0", existing)
                    article_text["widget"] = ntxt

                elif mtype == "YouTube":
                    tk.Label(type_fields, text="YouTube URL *",
                             font=("Calibri",9,"bold"), bg=PAGE_BG, fg=TEXT_MID
                             ).pack(anchor="w", pady=(10,2))
                    ff = tk.Frame(type_fields, bg=PANEL_BG, highlightthickness=1,
                                  highlightbackground=BORDER_CLR)
                    ff.pack(fill="x")
                    tk.Entry(ff, textvariable=yt_url_var, font=("Calibri",10),
                             bg=PANEL_BG, fg=ENTRY_FG, relief="flat", bd=0,
                             insertbackground=FOCUS_CLR).pack(fill="x", ipady=7, padx=6)

                elif mtype == "Article":
                    tk.Label(type_fields, text="Article Body *",
                             font=("Calibri",9,"bold"), bg=PAGE_BG, fg=TEXT_MID
                             ).pack(anchor="w", pady=(10,2))
                    af = tk.Frame(type_fields, bg=PANEL_BG, highlightthickness=1,
                                  highlightbackground=BORDER_CLR)
                    af.pack(fill="x")
                    atxt = scrolledtext.ScrolledText(af, font=("Calibri",10),
                                                     bg=PANEL_BG, fg=ENTRY_FG,
                                                     relief="flat", height=10,
                                                     wrap="word",
                                                     insertbackground=FOCUS_CLR)
                    atxt.pack(fill="x", padx=4, pady=4)
                    if is_edit:
                        existing = get_article_body(edit_mat["id"])
                        atxt.insert("1.0", existing)
                    article_text["widget"] = atxt

                else:
                    # File upload
                    if is_edit and edit_mat.get("file_path"):
                        tk.Label(type_fields,
                                 text=f"Current file: {Path(edit_mat['file_path']).name}",
                                 font=("Calibri",9), bg=PAGE_BG, fg=SUCCESS
                                 ).pack(anchor="w", pady=(8,4))
                    tk.Label(type_fields, text="Upload File" + (" (leave blank to keep current)" if is_edit else " *"),
                             font=("Calibri",9,"bold"), bg=PAGE_BG, fg=TEXT_MID
                             ).pack(anchor="w", pady=(10,2))
                    file_row = tk.Frame(type_fields, bg=PAGE_BG); file_row.pack(fill="x")
                    ff2 = tk.Frame(file_row, bg=PANEL_BG, highlightthickness=1,
                                   highlightbackground=BORDER_CLR)
                    ff2.pack(side="left", fill="x", expand=True, padx=(0,8))
                    file_lbl = tk.Entry(ff2, textvariable=file_path_var,
                                        font=("Calibri",9), bg=PANEL_BG, fg=TEXT_MID,
                                        relief="flat", bd=0, state="readonly")
                    file_lbl.pack(fill="x", ipady=7, padx=6)

                    def _browse():
                        ext_map = {
                            "PDF":   [("PDF","*.pdf")],
                            "PPTX":  [("PowerPoint","*.pptx *.ppt")],
                            "DOCX":  [("Word Document","*.docx *.doc")],
                            "Excel": [("Excel","*.xlsx *.xls")],
                            "Other": [("All Files","*.*")],
                        }
                        filetypes = ext_map.get(mtype_var.get(), [("All Files","*.*")])
                        fp = filedialog.askopenfilename(filetypes=filetypes)
                        if fp:
                            file_path_var.set(fp)

                    tk.Button(file_row, text="Browse…",
                              font=("Calibri",9,"bold"), bg=ACCENT_BLUE, fg="white",
                              relief="flat", cursor="hand2", padx=12, pady=7,
                              command=_browse).pack(side="left")

            if not is_edit:
                mt_cb.bind("<<ComboboxSelected>>",
                           lambda e: _render_type_fields())
            _render_type_fields(mtype_var.get())

            # Status label
            status_lbl = tk.Label(form, text="", font=("Calibri",9),
                                  bg=PAGE_BG, fg=SUCCESS)
            status_lbl.pack(anchor="w", pady=(8,0))

            # Save button
            def _do_save():
                t = title_var.get().strip()
                if not t:
                    status_lbl.config(text="⚠  Title is required.", fg=DANGER)
                    return
                mtype = mtype_var.get()
                cat   = cat_var.get().strip()
                desc  = desc_var.get().strip()
                tags  = tags_var.get().strip()
                uploader = (emp or {}).get("full_name", "Admin")

                article_body = None
                if mtype in ("Article","Notice") and article_text.get("widget"):
                    article_body = article_text["widget"].get("1.0","end").strip()

                yt_url = yt_url_var.get().strip() if mtype == "YouTube" else None

                stored_path = None
                raw_path = file_path_var.get().strip()
                if raw_path and mtype not in ("YouTube","Article"):
                    try:
                        stored_path = _copy_to_learning_store(raw_path, mtype)
                    except Exception as ex:
                        status_lbl.config(text=f"⚠  File copy failed: {ex}", fg=DANGER)
                        return
                elif is_edit:
                    stored_path = edit_mat.get("file_path")

                try:
                    if is_edit:
                        update_learning_material(
                            edit_mat["id"], t, cat, desc, tags, article_body,
                            is_pinned=int(pin_var.get()) if can_upload else int(edit_mat.get("is_pinned",0)),
                            priority=priority_var.get() if can_upload else edit_mat.get("priority","Normal"),
                            expiry_date=expiry_var.get().strip() or None if can_upload else edit_mat.get("expiry_date"),
                        )
                        # Also update yt_url if YouTube type
                        if mtype == "YouTube":
                            with get_conn() as conn:
                                conn.execute(
                                    "UPDATE learning_materials SET youtube_url=? WHERE id=?",
                                    (yt_url, edit_mat["id"])
                                )
                        status_lbl.config(text="✔  Material updated.", fg=SUCCESS)
                    else:
                        if mtype not in ("YouTube","Article","Notice") and not stored_path:
                            status_lbl.config(text="⚠  Please select a file.", fg=DANGER)
                            return
                        if mtype == "YouTube" and not yt_url:
                            status_lbl.config(text="⚠  YouTube URL is required.", fg=DANGER)
                            return
                        add_learning_material(
                            t, cat, mtype, stored_path, yt_url,
                            desc, tags, uploader, article_body,
                            is_pinned=int(pin_var.get()) if can_upload else 0,
                            priority=priority_var.get() if can_upload else "Normal",
                            expiry_date=expiry_var.get().strip() or None if can_upload else None,
                        )
                        status_lbl.config(text="✔  Material added.", fg=SUCCESS)

                    form.after(800, lambda: [win.destroy(), _load_content()])
                except Exception as ex:
                    status_lbl.config(text=f"⚠  Error: {ex}", fg=DANGER)

            btn_area = tk.Frame(form, bg=PAGE_BG, pady=12); btn_area.pack(fill="x")
            tk.Button(btn_area,
                      text="💾  Save Changes" if is_edit else "＋  Add Material",
                      font=("Calibri", 11, "bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=24, pady=10,
                      command=_do_save).pack(side="left")
            tk.Button(btn_area, text="Cancel",
                      font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_MID,
                      relief="flat", cursor="hand2", padx=16, pady=10,
                      highlightthickness=1, highlightbackground=BORDER_CLR,
                      command=win.destroy).pack(side="left", padx=10)

        # ── Category manager popup (admin only) ───────────────────────────────
        def _open_category_manager():
            win = tk.Toplevel(self)
            win.title("Manage Learning Categories")
            win.geometry("420x500")
            win.configure(bg=PAGE_BG)
            win.grab_set()

            tk.Frame(win, bg=SIDEBAR_BG, height=4).pack(fill="x")
            tk.Label(win, text="⚙  Manage Categories",
                     font=("Georgia", 12, "bold"), bg=PAGE_BG,
                     fg=TEXT_DARK, pady=12).pack()

            # Add new
            add_row = tk.Frame(win, bg=PAGE_BG, padx=20); add_row.pack(fill="x")
            nv  = tk.StringVar(); iv = tk.StringVar(value="📂")
            tk.Label(add_row, text="Name:", font=("Calibri",9,"bold"),
                     bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            nf = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                          highlightbackground=BORDER_CLR)
            nf.pack(side="left", padx=(4,6))
            tk.Entry(nf, textvariable=nv, font=("Calibri",10), bg=PANEL_BG,
                     fg=ENTRY_FG, relief="flat", bd=0, width=18).pack(ipady=5)
            tk.Label(add_row, text="Icon:", font=("Calibri",9,"bold"),
                     bg=PAGE_BG, fg=TEXT_MID).pack(side="left")
            if2 = tk.Frame(add_row, bg=PANEL_BG, highlightthickness=1,
                           highlightbackground=BORDER_CLR)
            if2.pack(side="left", padx=(4,6))
            tk.Entry(if2, textvariable=iv, font=("Calibri",12), bg=PANEL_BG,
                     fg=ENTRY_FG, relief="flat", bd=0, width=3).pack(ipady=5)

            lbox_frame = tk.Frame(win, bg=PAGE_BG, padx=20, pady=10)
            lbox_frame.pack(fill="both", expand=True)
            lb = tk.Listbox(lbox_frame, font=("Calibri",10), bg=PANEL_BG,
                            fg=TEXT_DARK, selectbackground=SIDEBAR_SEL,
                            selectforeground="white", relief="flat",
                            highlightthickness=1, highlightbackground=BORDER_CLR)
            lb.pack(fill="both", expand=True)
            cat_ids = []

            def _refresh_lb():
                lb.delete(0, "end"); cat_ids.clear()
                for c in get_learning_categories():
                    lb.insert("end", f"  {c['icon']}  {c['name']}")
                    cat_ids.append(c["id"])

            def _add_cat():
                name = nv.get().strip(); icon = iv.get().strip() or "📂"
                if not name: return
                add_learning_category(name, icon)
                nv.set(""); _refresh_lb(); _rebuild_cat_sidebar()

            def _del_cat():
                sel = lb.curselection()
                if not sel: return
                cid = cat_ids[sel[0]]
                if messagebox.askyesno("Delete", "Remove this category?"):
                    delete_learning_category(cid)
                    _refresh_lb(); _rebuild_cat_sidebar()

            tk.Button(add_row, text="＋ Add",
                      font=("Calibri",9,"bold"), bg=SUCCESS, fg="white",
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      command=_add_cat).pack(side="left")
            tk.Button(lbox_frame, text="🗑  Delete Selected",
                      font=("Calibri",9), bg=DANGER, fg="white",
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      command=_del_cat).pack(side="right", pady=(6,0))
            _refresh_lb()

        # ── Build the category sidebar ────────────────────────────────────────
        cat_btn_refs = []

        def _rebuild_cat_sidebar():
            for w in cat_inner.winfo_children():
                w.destroy()
            cat_btn_refs.clear()
            cats = get_learning_categories()

            def _cat_btn(name, icon=""):
                is_sel = (state["category"] == name)
                bg = SIDEBAR_SEL if is_sel else SIDEBAR_BG
                frame = tk.Frame(cat_inner, bg=bg, cursor="hand2")
                frame.pack(fill="x")
                tk.Label(frame, text=f"  {icon}  {name}",
                         font=("Calibri", 9, "bold" if is_sel else "normal"),
                         bg=bg, fg=SIDEBAR_FG if is_sel else SIDEBAR_ICN,
                         anchor="w", pady=9).pack(fill="x", padx=6)
                for w in [frame] + list(frame.winfo_children()):
                    w.bind("<Button-1>", lambda e, n=name: [
                        _load_content(category=n), _rebuild_cat_sidebar()
                    ])
                cat_btn_refs.append(frame)

            _cat_btn("All", "📚")
            for c in cats:
                _cat_btn(c["name"], c.get("icon","📂"))

        _rebuild_cat_sidebar()
        _refresh_pinned_banner()
        _load_content("All")

    # ── AUTO REMINDER THREAD ─────────────────────────────────────────────────────

    def _start_reminder_thread(self, cfg: dict = None):
        """Start or restart the background reminder checker thread."""
        if cfg is None:
            cfg = load_config()
        if not cfg.get("reminder_enabled", True):
            return

        def _reminder_loop(app_ref, config):
            import time
            reminded_today = None
            while True:
                try:
                    now = datetime.now()
                    today = now.date().isoformat()
                    rt = config.get("reminder_time", "19:00")
                    rh, rm = map(int, rt.split(":"))
                    reminder_dt = now.replace(hour=rh, minute=rm, second=0, microsecond=0)

                    if (now >= reminder_dt and reminded_today != today):
                        reminded_today = today
                        # Only remind if running (app still alive)
                        try:
                            app_ref.after(0, lambda: app_ref._show_reminder_popup())
                        except Exception:
                            break
                    time.sleep(60)
                except Exception as e:
                    log.warning(f"Reminder thread error: {e}")
                    break

        t = threading.Thread(target=_reminder_loop,
                             args=(self, cfg), daemon=True)
        t.start()

    def _show_reminder_popup(self):
        """Show a non-blocking reminder popup."""
        # Only show if user is logged in and hasn't submitted today
        if not self.current_emp:
            return
        emp_id = self.current_emp["emp_id"]
        today = date.today().isoformat()
        sub = get_submission(emp_id, today)
        if sub and sub.get("submission_status") in ("On Time", "Late"):
            return   # already submitted

        popup = tk.Toplevel(self)
        popup.title("⏰ Timesheet Reminder")
        popup.geometry("380x200")
        popup.configure(bg=PAGE_BG)
        popup.resizable(False, False)
        popup.attributes("-topmost", True)
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 380) // 2
        py = self.winfo_y() + (self.winfo_height() - 200) // 2
        popup.geometry(f"380x200+{px}+{py}")

        tk.Frame(popup, bg=WARNING, height=4).pack(fill="x")
        tk.Label(popup, text="⏰  Timesheet Reminder",
                 font=("Georgia", 13, "bold"), bg=PAGE_BG,
                 fg=WARNING, pady=16).pack()
        tk.Label(popup,
                 text=f"Hi {self.current_emp['full_name']},\nyou haven't submitted your timesheet yet!\nCutoff time approaching.",
                 font=("Calibri", 10), bg=PAGE_BG, fg=TEXT_DARK,
                 justify="center").pack(pady=(0, 16))
        tk.Button(popup, text="OK — I'll submit now",
                  font=("Calibri", 10, "bold"), bg=WARNING, fg="white",
                  relief="flat", cursor="hand2", padx=20, pady=8,
                  command=popup.destroy).pack()

    # ── BIRTHDAY POPUPS & PANELS ─────────────────────────────────────────────────

    def _check_and_show_birthday_popups(self):
        today_birthdays = get_todays_birthdays()
        if not today_birthdays:
            return
        if self.current_emp:
            today_mm_dd = date.today().strftime("%m-%d")
            if self.current_emp.get("date_of_birth") == today_mm_dd:
                self._show_personal_birthday_popup(self.current_emp["full_name"])
        if self.current_role == "admin":
            notice_list = today_birthdays
        else:
            own = (self.current_emp or {}).get("full_name","").strip().lower()
            notice_list = [e for e in today_birthdays
                           if e["full_name"].strip().lower() != own]
        if notice_list:
            self.after(600, lambda: self._show_team_birthday_popup(notice_list))

    def _show_personal_birthday_popup(self, name: str):
        popup = tk.Toplevel(self)
        popup.title("🎂 Happy Birthday!")
        popup.geometry("420x260")
        popup.configure(bg="#FFF8E7")
        popup.resizable(False, False)
        popup.attributes("-topmost", True)
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 420) // 2
        py = self.winfo_y() + (self.winfo_height() - 260) // 2
        popup.geometry(f"420x260+{px}+{py}")
        tk.Frame(popup, bg=ACCENT_GOLD, height=5).pack(fill="x")
        tk.Label(popup, text="🎂", font=("Segoe UI Emoji", 38),
                 bg="#FFF8E7").pack(pady=(18, 0))
        tk.Label(popup, text="Happy Birthday!",
                 font=("Georgia", 16, "bold"), bg="#FFF8E7",
                 fg="#B8630A").pack()
        tk.Label(popup,
                 text=f"Wishing you a wonderful day, {name.split()[0]}! 🎉\n"
                      "Joshi Apte & Associates celebrates with you.",
                 font=("Calibri", 10), bg="#FFF8E7", fg=TEXT_DARK,
                 justify="center", pady=6).pack()
        tk.Button(popup, text="Thank you! 😊",
                  font=("Calibri", 10, "bold"), bg=ACCENT_GOLD, fg="white",
                  relief="flat", cursor="hand2", padx=24, pady=8,
                  command=popup.destroy).pack(pady=(8, 0))

    def _show_team_birthday_popup(self, birthday_people: list):
        popup = tk.Toplevel(self)
        popup.title("🎂 Today's Birthdays")
        h = min(160 + len(birthday_people) * 30, 420)
        popup.geometry(f"400x{h}")
        popup.configure(bg=PAGE_BG)
        popup.resizable(False, False)
        popup.attributes("-topmost", True)
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - h)   // 2
        popup.geometry(f"400x{h}+{px}+{py}")
        tk.Frame(popup, bg=ACCENT_GOLD, height=4).pack(fill="x")
        hdr = tk.Frame(popup, bg=SIDEBAR_BG, padx=18, pady=12); hdr.pack(fill="x")
        tk.Label(hdr, text="🎂  Today's Birthdays",
                 font=("Georgia", 12, "bold"), bg=SIDEBAR_BG,
                 fg=SIDEBAR_FG).pack(anchor="w")
        tk.Label(hdr, text=date.today().strftime("%d %B %Y"),
                 font=("Calibri", 8), bg=SIDEBAR_BG,
                 fg=SIDEBAR_ICN).pack(anchor="w")
        body = tk.Frame(popup, bg=PAGE_BG, padx=20, pady=10)
        body.pack(fill="both", expand=True)
        for i, emp in enumerate(birthday_people):
            bg = ROW_ODD if i%2==0 else ROW_EVEN
            row = tk.Frame(body, bg=bg); row.pack(fill="x", pady=1)
            tk.Label(row, text="🎉", font=("Segoe UI Emoji", 11),
                     bg=bg).pack(side="left", padx=(8,4), pady=4)
            tk.Label(row, text=emp["full_name"],
                     font=("Calibri", 10, "bold"), bg=bg,
                     fg=TEXT_DARK).pack(side="left", pady=4)
            if emp.get("role"):
                tk.Label(row, text=f"  ({emp['role']})",
                         font=("Calibri", 9), bg=bg,
                         fg=TEXT_LIGHT).pack(side="left")
        tk.Button(popup, text="🎊  Noted!",
                  font=("Calibri", 10, "bold"), bg=SUCCESS, fg="white",
                  relief="flat", cursor="hand2", padx=20, pady=8,
                  command=popup.destroy).pack(pady=(0, 12))

    def _build_employee_birthdays_tab(self, parent):
        self._topbar(parent, "🎂  Birthdays", "  Upcoming team birthdays")
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=14)
        self._render_birthdays_content(outer, admin_mode=False)

    def _build_birthdays_tab(self, parent):
        self._topbar(parent, "🎂  Birthdays",
                     "  Staff birthdays — upcoming & full list by month")
        outer = tk.Frame(parent, bg=PAGE_BG)
        outer.pack(fill="both", expand=True, padx=20, pady=14)
        self._render_birthdays_content(outer, admin_mode=True)

    def _render_birthdays_content(self, outer, admin_mode: bool):
        today_str = date.today().strftime("%m-%d")
        upcoming  = get_upcoming_birthdays(30)

        # ── Upcoming card ────────────────────────────────────────────────────
        up_card = _card(outer); up_card.pack(fill="x", pady=(0, 14))
        tk.Label(up_card, text="  🎉  Upcoming Birthdays  (next 30 days)",
                 font=("Calibri", 10, "bold"), bg=TBL_HDR_BG,
                 fg=TBL_HDR_FG, anchor="w", padx=6, pady=10).pack(fill="x")
        up_body = tk.Frame(up_card, bg=PANEL_BG)
        up_body.pack(fill="x", padx=8, pady=8)
        if not upcoming:
            tk.Label(up_body, text="  No birthdays in the next 30 days.",
                     font=("Calibri", 9), bg=PANEL_BG,
                     fg=TEXT_LIGHT, pady=10).pack(anchor="w")
        else:
            for i, emp in enumerate(upcoming):
                bg = "#FFF8E7" if emp["date_of_birth"] == today_str \
                     else (ROW_ODD if i%2==0 else ROW_EVEN)
                row = tk.Frame(up_body, bg=bg); row.pack(fill="x", pady=1)
                icon = "🎂" if emp["date_of_birth"] == today_str else "🎁"
                tk.Label(row, text=f"  {icon}",
                         font=("Segoe UI Emoji", 11),
                         bg=bg).pack(side="left", padx=(4,2), pady=5)
                tk.Label(row, text=emp["full_name"],
                         font=("Calibri", 10, "bold"), bg=bg,
                         fg=TEXT_DARK, anchor="w",
                         width=28).pack(side="left", pady=5)
                mm, dd = emp["date_of_birth"].split("-")
                tk.Label(row, text=f"{dd} {_MONTH_NAMES[int(mm)]}",
                         font=("Calibri", 9), bg=bg,
                         fg=ACCENT_BLUE, anchor="w",
                         width=14).pack(side="left")
                if emp["days_away"] == 0:
                    btxt, bbg = "TODAY 🎂", ACCENT_GOLD
                elif emp["days_away"] == 1:
                    btxt, bbg = "Tomorrow", SUCCESS
                else:
                    btxt, bbg = f"in {emp['days_away']} days", TEXT_MID
                tk.Label(row, text=btxt,
                         font=("Calibri", 8, "bold"),
                         bg=bbg, fg="white",
                         padx=8, pady=3).pack(side="left", padx=8)

        # ── Full list by month ───────────────────────────────────────────────
        full_card = _card(outer); full_card.pack(fill="both", expand=True)
        tk.Label(full_card, text="  📅  Full Birthday List — by Month",
                 font=("Calibri", 10, "bold"), bg=TBL_HDR_BG,
                 fg=TBL_HDR_FG, anchor="w", padx=6,
                 pady=10).pack(fill="x")
        sc = tk.Canvas(full_card, bg=PANEL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(full_card, orient="vertical", command=sc.yview)
        sc.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        sc.pack(side="left", fill="both", expand=True)
        fm = tk.Frame(sc, bg=PANEL_BG, padx=10, pady=8)
        fm_id = sc.create_window((0,0), window=fm, anchor="nw")
        fm.bind("<Configure>",
                lambda e: sc.configure(scrollregion=sc.bbox("all")))
        sc.bind("<Configure>",
                lambda e: sc.itemconfig(fm_id, width=e.width))
        sc.bind("<MouseWheel>",
                lambda e: sc.yview_scroll(-1*(e.delta//120), "units"))
        by_month = get_all_birthdays_by_month()
        if not by_month:
            tk.Label(fm, text="No birthday data available.",
                     font=("Calibri", 9), bg=PANEL_BG,
                     fg=TEXT_LIGHT).pack(anchor="w", pady=10)
        else:
            for mm in sorted(by_month.keys()):
                mhdr = tk.Frame(fm, bg=SIDEBAR_BG)
                mhdr.pack(fill="x", pady=(10,2))
                tk.Label(mhdr, text=f"  {_MONTH_NAMES[mm].upper()}",
                         font=("Calibri", 9, "bold"), bg=SIDEBAR_BG,
                         fg=SIDEBAR_FG, pady=5, padx=6).pack(anchor="w")
                for i, emp in enumerate(by_month[mm]):
                    is_today = emp["date_of_birth"] == today_str
                    bg = "#FFF8E7" if is_today \
                         else (ROW_ODD if i%2==0 else ROW_EVEN)
                    row = tk.Frame(fm, bg=bg); row.pack(fill="x", pady=1)
                    mm_s, dd_s = emp["date_of_birth"].split("-")
                    tk.Label(row,
                             text=f"  {dd_s} {_MONTH_NAMES[int(mm_s)]}",
                             font=("Calibri", 9, "bold"), bg=bg,
                             fg=ACCENT_BLUE, width=14,
                             anchor="w").pack(side="left", pady=4)
                    tk.Label(row,
                             text=f"{'🎂 ' if is_today else ''}{emp['full_name']}",
                             font=("Calibri", 9,
                                   "bold" if is_today else "normal"),
                             bg=bg, fg=TEXT_DARK,
                             anchor="w", width=30).pack(side="left")
                    if admin_mode:
                        tk.Label(row, text=emp.get("role",""),
                                 font=("Calibri", 8), bg=bg,
                                 fg=TEXT_LIGHT,
                                 anchor="w").pack(side="left", padx=6)


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    startup()
    app = TimesheetApp()
    app.mainloop()